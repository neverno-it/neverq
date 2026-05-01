from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import transaction, IntegrityError
from django.db.models import Count, Sum, Q
from django.utils import timezone
from django.utils.dateparse import parse_time
from datetime import timedelta, date
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from decimal import Decimal, InvalidOperation
from django.views.decorators.http import require_POST
import csv
import re
from pathlib import Path

from apps.core.models import Company, Building, Location, State, City, ORDER_DAY_CHOICES, Coupon, KioskConfig, WebViewConfig, DisplayBoardConfig
from apps.menu.models import Category, Product, Cafe, ProductGallery, CategoryGallery, OfferingGallery
from apps.orders.models import (
    Order,
    OrderItem,
    OrderStatus,
    OrderStatusChoices,
    PaymentModeChoices,
    CompanySettlement,
    ORDER_TYPE_KIOSK,
    ORDER_TYPE_WALLET_RECHARGE,
)
from apps.accounts.models import Customer, StaffUser, WalletTransaction
from apps.accounts.decorators import staff_role_required
from apps.core.access import (
    check_module_permission,
    create_pending_change,
    get_list_perms,
    get_locked_html_names,
    get_module_level,
    get_primary_staff_company,
    get_staff_site_companies,
    user_can_access_company,
    user_can_action,
)
from apps.accounts.forms import CustomerRegisterForm

CUSTOMER_BIRTH_DEFAULT_YEAR = 2000
CUSTOMER_BIRTH_DAY_OPTIONS = list(range(1, 32))
CUSTOMER_BIRTH_MONTH_OPTIONS = list(range(1, 13))


def _scope(user):
    if not user.is_superadmin:
        return {'company__in': get_staff_site_companies(user)}
    return {}


def _companies(user):
    return get_staff_site_companies(user)

def _bulk_post_ids(request):
    ids = []
    for raw in request.POST.getlist('ids'):
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    return ids


def _bulk_message(request, label, count, skipped=0):
    if count:
        msg = f'{count} {label}{"s" if count != 1 else ""} deleted.'
        if skipped:
            msg += f' {skipped} skipped.'
        messages.success(request, msg)
    elif skipped:
        messages.warning(request, f'No items deleted. {skipped} skipped.')
    else:
        messages.warning(request, 'No items selected.')


def _deny_dashboard_action(request, message='Permission denied.'):
    messages.error(request, message)
    return redirect('dashboard:no_access')


def _has_full_module_edit(user, module_key):
    return user.is_superadmin or get_module_level(user, module_key) == 'full_edit'


def _company_id_allowed_for_user(user, company_id):
    return bool(company_id and (user.is_superadmin or _companies(user).filter(pk=company_id).exists()))



def _cashier_display_order_number(order_number):
    value = str(order_number or '').strip()
    match = re.match(r'^([A-Za-z]+)-(\d{6}|\d{8})-(.+)$', value)
    return f"{match.group(1)}-{match.group(3)}" if match else value

def _customer_qs_for_user(user):
    qs = Customer.objects.filter(is_deleted=False)
    if not user.is_superadmin:
        qs = qs.filter(company__in=_companies(user))
    return qs


def _coerce_customer_birth_date(day_raw, month_raw):
    day_raw = (day_raw or '').strip()
    month_raw = (month_raw or '').strip()
    if not day_raw and not month_raw:
        return None, ''
    if not day_raw or not month_raw:
        return None, 'Select both birth date and month.'
    try:
        return date(CUSTOMER_BIRTH_DEFAULT_YEAR, int(month_raw), int(day_raw)), ''
    except (TypeError, ValueError):
        return None, 'Select a valid birth date and month.'


def _normalize_customer_meal_benefit(customer, raw_benefit):
    benefit = raw_benefit if raw_benefit in {
        Customer.MEAL_BENEFIT_NONE,
        Customer.MEAL_BENEFIT_SUBSIDY,
        Customer.MEAL_BENEFIT_COMPANY_PAY,
    } else Customer.MEAL_BENEFIT_NONE
    if getattr(getattr(customer, 'company', None), 'bill_company', None) != 2:
        return Customer.MEAL_BENEFIT_NONE
    return benefit


def _parse_customer_subsidy_override(raw_value, benefit):
    if benefit != Customer.MEAL_BENEFIT_SUBSIDY:
        return None
    raw_value = (raw_value or '').strip()
    if not raw_value:
        return None
    try:
        return max(Decimal('0'), Decimal(raw_value))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _safe_birthday_for_year(dob, year):
    try:
        return dob.replace(year=year)
    except ValueError:
        if dob.month == 2 and dob.day == 29:
            return date(year, 2, 28)
        raise


def _upcoming_customer_birthdays(user, *, window_days=30, limit=8):
    today = timezone.localdate()
    customers = _customer_qs_for_user(user).filter(
        is_active=True,
        date_of_birth__isnull=False,
    ).select_related('company', 'building').order_by('name')
    upcoming = []
    for customer in customers:
        dob = customer.date_of_birth
        if not dob:
            continue
        next_birthday = _safe_birthday_for_year(dob, today.year)
        if next_birthday < today:
            next_birthday = _safe_birthday_for_year(dob, today.year + 1)
        days_until = (next_birthday - today).days
        if days_until > window_days:
            continue
        customer.next_birthday = next_birthday
        customer.days_until_birthday = days_until
        customer.upcoming_age = None if dob.year == 2000 else (next_birthday.year - dob.year)
        upcoming.append(customer)
    upcoming.sort(key=lambda customer: (customer.days_until_birthday, customer.name.lower()))
    birthdays_today_count = sum(1 for customer in upcoming if customer.days_until_birthday == 0)
    return upcoming[:limit], birthdays_today_count


def _reports_redirect_url(request, from_date, to_date, company_filter='', sale_particular=''):
    from urllib.parse import urlencode
    params = {
        'from_date': from_date,
        'to_date': to_date,
    }
    if company_filter:
        params['company'] = company_filter
    if sale_particular:
        params['sale_particular'] = sale_particular
    return f"{request.path}?{urlencode(params)}"


def _company_outstanding_as_of(company, as_of_date):
    provisional = (
        Order.objects.filter(
            company=company,
            is_deleted=False,
            order_status__in=[2, 3, 4, 5],
            created_at__date__lte=as_of_date,
        ).aggregate(t=Sum('bill_to_company'))['t'] or Decimal('0.00')
    )

    received = (
        CompanySettlement.objects.filter(
            company=company,
            is_deleted=False,
            payment_date__lte=as_of_date,
        ).aggregate(t=Sum('amount_received'))['t'] or Decimal('0.00')
    )

    outstanding = provisional - received
    return outstanding if outstanding > 0 else Decimal('0.00')
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  ADMIN / SUPERADMIN DASHBOARD HOME
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin','admin','pos','cafeman','reports')
def no_access(request):
    return render(request, 'dashboard/no_access.html', {
        'page_title': 'No Access Assigned',
    })

@staff_role_required('superadmin','admin','reports')
def dashboard_home(request):
    user  = request.user
    today = date.today()
    cf    = _scope(user)

    # â”€â”€ 12 stat cards â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    total_customers   = Customer.objects.filter(is_active=True, is_deleted=False, **cf).count()
    total_companies   = Company.objects.filter(is_active=True, is_deleted=False).count()
    total_categories  = Category.objects.filter(is_deleted=False).count()
    total_subcats     = Category.objects.filter(is_deleted=False, parent__isnull=False).count()

    web_qs_today  = Order.objects.filter(created_at__date=today, is_deleted=False, **cf)
    web_qs_all    = Order.objects.filter(is_deleted=False, **cf)
    pos_qs_today  = None
    pos_qs_all    = None
    try:
        from apps.pos.models import POSOrder
        pos_cf = {'company__in': _companies(user)} if not user.is_superadmin else {}
        pos_qs_today = POSOrder.objects.filter(is_deleted=False, **pos_cf).exclude(
            created_at__isnull=True).filter(created_at__date=today)
        pos_qs_all   = POSOrder.objects.filter(is_deleted=False, **pos_cf).exclude(created_at__isnull=True)
    except (ImportError, AttributeError, LookupError):
        # POS module may not be installed or DB table may not exist yet
        pass

    stats = {
        'total_customers':      total_customers,
        'total_companies':      total_companies,
        'total_categories':     total_categories,
        'total_subcats':        total_subcats,
        'today_web_orders':     web_qs_today.count(),
        'today_web_sale':       web_qs_today.filter(order_status__in=[2,3,4,5]).aggregate(t=Sum('total_amount'))['t'] or 0,
        'total_web_orders':     web_qs_all.count(),
        'total_web_sale':       web_qs_all.filter(order_status__in=[2,3,4,5]).aggregate(t=Sum('total_amount'))['t'] or 0,
        'today_pos_orders':     pos_qs_today.count() if pos_qs_today is not None else 0,
        'today_pos_sale':       pos_qs_today.aggregate(t=Sum('total_amount'))['t'] or 0 if pos_qs_today is not None else 0,
        'total_pos_orders':     pos_qs_all.count() if pos_qs_all is not None else 0,
        'total_pos_sale':       pos_qs_all.aggregate(t=Sum('total_amount'))['t'] or 0 if pos_qs_all is not None else 0,
        'pending_orders':       web_qs_today.filter(order_status=1).count(),
    }

    # â”€â”€ Today Web Order Items breakdown â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    company_id_filter = request.GET.get('web_company', '')
    web_item_filters = {'order__created_at__date': today, 'order__is_deleted': False}
    if company_id_filter and _company_id_allowed_for_user(user, company_id_filter):
        web_item_filters['order__company_id'] = company_id_filter
    elif company_id_filter:
        company_id_filter = ''
    elif cf:
        for k, v in cf.items():
            web_item_filters['order__' + k] = v
    today_web_items_qs = OrderItem.objects.filter(
        **web_item_filters
    ).values('product__name').annotate(cnt=Count('id')).order_by('-cnt')

    # â”€â”€ Today POS Order Items breakdown â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    today_pos_items = []
    pos_company_filter = request.GET.get('pos_company', '')
    try:
        from apps.pos.models import POSOrderItem
        pos_item_qs = POSOrderItem.objects.filter(
            order__created_at__date=today, order__is_deleted=False
        )
        if pos_company_filter and _company_id_allowed_for_user(user, pos_company_filter):
            pos_item_qs = pos_item_qs.filter(order__company_id=pos_company_filter)
        elif pos_company_filter:
            pos_company_filter = ''
        elif not user.is_superadmin:
            pos_item_qs = pos_item_qs.filter(order__company__in=_companies(user))
        today_pos_items = pos_item_qs.values('product_name').annotate(cnt=Sum('qty')).order_by('-cnt')
    except (ImportError, AttributeError, LookupError):
        # POS module may not be installed or table may not exist
        pass

    # â”€â”€ Weekly revenue chart â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    weekly = []
    for i in range(6, -1, -1):
        d   = today - timedelta(days=i)
        rev = Order.objects.filter(
            created_at__date=d, is_deleted=False,
            order_status__in=[2,3,4,5], **cf
        ).aggregate(t=Sum('total_amount'))['t'] or 0
        weekly.append({'date': d.strftime('%a'), 'revenue': float(rev)})
    weekly_max_revenue = max((item['revenue'] for item in weekly), default=0)

    recent_orders = Order.objects.filter(
        is_deleted=False, **cf
    ).select_related('customer','company').order_by('-created_at')[:10]
    upcoming_birthdays, birthdays_today_count = _upcoming_customer_birthdays(user)

    return render(request, 'dashboard/home.html', {
        **stats,
        'today_web_items':    today_web_items_qs[:30],
        'today_pos_items':    today_pos_items[:30],
        'recent_orders':      recent_orders,
        'weekly_data':        weekly,
        'weekly_max_revenue': weekly_max_revenue,
        'companies':          _companies(user),
        'web_company_filter': company_id_filter,
        'pos_company_filter': pos_company_filter,
        'upcoming_birthdays': upcoming_birthdays,
        'birthdays_today_count': birthdays_today_count,
        'page_title':         'Dashboard',
    })

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  CASHIER / SUBADMIN DASHBOARD
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin','admin','pos','cafeman')
def cashier_dashboard(request):
    """Today's web orders for cashier / POS / cafe manager role â€” with inline status change."""
    user    = request.user
    companies = _companies(user)
    company = get_primary_staff_company(user)

    qs = Order.objects.filter(is_deleted=False).select_related(
        'customer', 'company', 'cafe'
    ).prefetch_related('items__product').order_by('-created_at')

    if not user.is_superadmin:
        qs = qs.filter(company__in=companies)

    # Site filter for multi-site users
    company_filter = request.GET.get('company', '')
    if company_filter and (user.is_superadmin or companies.filter(pk=company_filter).exists()):
        qs = qs.filter(company_id=company_filter)
        company = Company.objects.filter(pk=company_filter).first() or company

    # Date filter â€” default today
    date_filter = request.GET.get('date', str(date.today()))
    if date_filter:
        qs = qs.filter(scheduled_date__date=date_filter)
        if not qs.exists():
            qs = Order.objects.filter(is_deleted=False).select_related(
                'customer','company','cafe'
            ).prefetch_related('items__product').order_by('-created_at')
            if not user.is_superadmin:
                qs = qs.filter(company__in=companies)
            if company_filter and (user.is_superadmin or companies.filter(pk=company_filter).exists()):
                qs = qs.filter(company_id=company_filter)
            qs = qs.filter(created_at__date=date_filter)

    # Status filter
    status_filter = request.GET.get('status', '')
    if status_filter:
        qs = qs.filter(order_status=status_filter)

    # Search
    q = request.GET.get('q', '')
    if q:
        qs = qs.filter(
            Q(order_number__icontains=q) |
            Q(customer_name_snapshot__icontains=q) |
            Q(customer_phone_snapshot__icontains=q) |
            Q(customer__name__icontains=q) |
            Q(customer__phone__icontains=q)
        )

    paginator = Paginator(qs, 25)
    page_obj  = paginator.get_page(request.GET.get('page'))
    for order in page_obj.object_list:
        order.display_order_number = _cashier_display_order_number(order.order_number)

    # Last order timestamp for polling
    latest = Order.objects.filter(
        is_deleted=False,
        order_status__in=[OrderStatusChoices.PENDING, OrderStatusChoices.CONFIRMED],
        **({'company__in': companies} if not user.is_superadmin else {}),
    ).order_by('-created_at').first()
    last_order_time = latest.created_at.strftime('%Y-%m-%dT%H:%M:%S') if latest and latest.created_at else ''

    status_options = [
        (1, 'Pending',       '#F59E0B'),
        (2, 'Confirmed',     '#3B82F6'),
        (3, 'Preparing',     '#8B5CF6'),
        (4, 'Food is Ready', '#10B981'),
        (5, 'Delivered',     '#059669'),
        (6, 'Cancelled',     '#EF4444'),
    ]

    return render(request, 'dashboard/cashier.html', {
        'orders':          page_obj,
        'page_obj':        page_obj,
        'date_filter':     date_filter,
        'status_filter':   status_filter,
        'company_filter':  company_filter,
        'companies':       companies,
        'q':               q,
        'company':         company,
        'last_order_time': last_order_time,
        'status_options':  status_options,
        'is_delivery_mode': bool(company and getattr(company, 'is_packet_delivery', False)),
        'page_title':      'Today Web Orders',
    })


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  KITCHEN VIEW
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin','admin','cafeman')
def kitchen(request):
    user = request.user
    cf   = _scope(user)
    status_filter = request.GET.get('status', '')
    qs = Order.objects.filter(
        is_deleted=False,
        order_status__in=[1, 2, 3, 4],
        **cf
    ).select_related('customer', 'company').prefetch_related('items__product').order_by('created_at')
    if status_filter:
        qs = qs.filter(order_status=status_filter)
    return render(request, 'dashboard/kitchen.html', {
        'orders':        qs,
        'status_filter': status_filter,
        'statuses':      [(1,'Pending'),(2,'Confirmed'),(3,'Preparing'),(4,'Ready')],
        'companies':     _companies(user),
        'page_title':    'Kitchen Orders',
    })


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  STORE OPEN / CLOSE (superadmin manual control only)
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _set_company_store_status(company, action):
    action = (action or '').strip().lower()
    if action not in {'open', 'close'}:
        return False, 'Invalid store action.'
    company.store_status = action == 'open'
    company.save(update_fields=['store_status'])
    return True, f'Store is now {"Open" if company.store_status else "Closed"} for {company.name}.'


@staff_role_required('superadmin', 'admin')
def store_toggle(request):
    if request.method != 'POST':
        return redirect(request.META.get('HTTP_REFERER', 'dashboard:home'))
    user = request.user
    company = get_primary_staff_company(user)
    if not company:
        messages.error(request, 'No company is linked to this account.')
        return redirect(request.META.get('HTTP_REFERER', 'dashboard:home'))
    # FIX: When the company has a schedule, only superadmin can manually override.
    # Admin role is blocked from overriding scheduled stores.
    if not user.is_superadmin and company.has_order_schedule:
        err = ('This company has an automatic ordering schedule. '
               'Only a Super Admin can manually override it.')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': err}, status=403)
        messages.error(request, err)
        return redirect(request.META.get('HTTP_REFERER', 'dashboard:home'))
    ok, message = _set_company_store_status(company, request.POST.get('action'))
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': ok, 'store_status': company.store_status if ok else None, 'message': message}, status=200 if ok else 400)
    (messages.success if ok else messages.error)(request, message)
    return redirect(request.META.get('HTTP_REFERER', 'dashboard:home'))


@require_POST
@staff_role_required('superadmin')
def company_store_toggle(request, pk):
    company = get_object_or_404(Company, pk=pk, is_deleted=False)
    user = request.user
    if not user.is_superadmin and (not user.company or user.company_id != company.pk):
        messages.error(request, 'You can only change your own company.')
        return redirect('dashboard:home')
    # FIX: When the company has a schedule, only superadmin can manually override.
    if not user.is_superadmin and company.has_order_schedule:
        messages.error(request, (
            'This company has an automatic ordering schedule. '
            'Only a Super Admin can manually override it.'
        ))
        return redirect(request.META.get('HTTP_REFERER') or
                        f'/dashboard/companies/{company.pk}/')
    ok, message = _set_company_store_status(company, request.POST.get('action'))
    (messages.success if ok else messages.error)(request, message)
    return redirect(request.META.get('HTTP_REFERER') or
                    f'/dashboard/companies/{company.pk}/')


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  COMPANIES
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin')
def company_list(request):
    companies = Company.objects.filter(is_deleted=False).annotate(
        customer_count=Count('customers', distinct=True),
        order_count=Count('orders', distinct=True),
    ).order_by('name')
    return render(request, 'dashboard/companies/list.html', {
        'companies': companies, 'page_title': 'Companies'
    })


@staff_role_required('superadmin', 'admin')
def company_detail(request, pk):
    user = request.user
    if not user.is_superadmin and not user_can_access_company(user, pk):
        messages.error(request, 'You can only view your own company.')
        return redirect('dashboard:home')
    company = get_object_or_404(Company, pk=pk, is_deleted=False)
    customers = Customer.objects.filter(company=company, is_deleted=False)
    buildings = Building.objects.filter(company=company, is_deleted=False).select_related('state', 'city', 'location').order_by('name')
    cafes = Cafe.objects.filter(company=company, is_deleted=False).select_related('building').order_by('building__name', 'name')
    from apps.menu.models import Counter
    counters = Counter.objects.filter(company=company, is_deleted=False).select_related('cafe', 'cafe__building').order_by('cafe__building__name', 'cafe__name', 'position_order', 'name')
    cafes_by_building = {}
    for cafe in cafes:
        cafes_by_building.setdefault(cafe.building_id or 0, []).append(cafe)
    counters_by_cafe = {}
    for counter in counters:
        counters_by_cafe.setdefault(counter.cafe_id or 0, []).append(counter)
    return render(request, 'dashboard/companies/detail.html', {
        'company':       company,
        'customers':     customers,
        'buildings':     buildings,
        'cafes':         cafes,
        'cafes_by_building': cafes_by_building,
        'counters':      counters,
        'counters_by_cafe': counters_by_cafe,
        'recent_orders': Order.objects.filter(company=company, is_deleted=False).select_related('customer').order_by('-created_at')[:20],
        'subsidy_count': customers.exclude(meal_benefit='none').count(),
        'page_title': company.name,
    })


def _parse_meal_amount(raw):
    from decimal import Decimal, InvalidOperation
    try:
        val = Decimal(str(raw).strip() or '0')
        return max(Decimal('0'), val)
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0')


def _parse_positive_int(raw, default=1, minimum=1, maximum=99):
    try:
        value = int(raw or default)
    except (TypeError, ValueError):
        value = default
    return max(minimum, min(maximum, value))


def _parse_percent(raw, default='3.50'):
    try:
        value = Decimal(str(raw or default))
    except (InvalidOperation, TypeError, ValueError):
        value = Decimal(default)
    return max(Decimal('0.00'), min(Decimal('25.00'), value))


def _selected_order_days(post_data):
    valid_days = {day for day, _ in ORDER_DAY_CHOICES}
    selected = [day for day in post_data.getlist('order_open_days') if day in valid_days]
    return selected




@staff_role_required('superadmin')

@staff_role_required('superadmin')
def company_add(request):
    selected_days = [day for day, _ in ORDER_DAY_CHOICES]

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        selected_days = _selected_order_days(request.POST)

        if not name:
            messages.error(request, 'Name is required.')
        elif not selected_days:
            messages.error(request, 'Select at least one open day for ordering.')
        else:
            try:
                try:
                    from decimal import Decimal as _D
                    royalty_ppr = _D(request.POST.get('royalty_points_per_rupee','1') or '1')
                except (InvalidOperation, ValueError, TypeError):
                    royalty_ppr = _D('1')
                co = Company.objects.create(
                    name=name,
                    company_address=request.POST.get('company_address', ''),
                    company_gst=request.POST.get('company_gst', ''),
                    fssai_number=request.POST.get('fssai_number', ''),
                    phone=request.POST.get('phone', ''),
                    bill_company=int(request.POST.get('bill_company', 2) or 2),
                    company_meal_amount=_parse_meal_amount(request.POST.get('company_meal_amount', '0')),
                    company_pay_meals_per_day=_parse_positive_int(request.POST.get('company_pay_meals_per_day'), default=1, minimum=1, maximum=10),
                    subsidy_meals_per_day=_parse_positive_int(request.POST.get('subsidy_meals_per_day'), default=1, minimum=1, maximum=10),
                    cod_payment=request.POST.get('cod_payment') == 'on',
                    online_payment=request.POST.get('online_payment') == 'on',
                    monthly_payment=request.POST.get('monthly_payment') == 'on',
                    pos_cash_enabled=request.POST.get('pos_cash_enabled') == 'on',
                    pos_upi_enabled=request.POST.get('pos_upi_enabled') == 'on',
                    pos_card_enabled=request.POST.get('pos_card_enabled') == 'on',
                    pos_card_fee_percent=_parse_percent(request.POST.get('pos_card_fee_percent'), default='3.50'),
                    store_status=request.POST.get('store_status') == 'on',
                    order_from_time=parse_time((request.POST.get('order_from_time') or '').strip()) or None,
                    order_to_time=parse_time((request.POST.get('order_to_time') or '').strip()) or None,
                    order_open_days=selected_days,
                    require_customer_approval=request.POST.get('require_customer_approval') == 'on',
                    fulfillment_mode=request.POST.get('fulfillment_mode', 'pickup') or 'pickup',
                    royalty_enabled=request.POST.get('royalty_enabled') == 'on',
                    royalty_points_per_rupee=royalty_ppr,
                    royalty_min_redeem=int(request.POST.get('royalty_min_redeem','100') or '100'),
                    royalty_max_redeem_pct=int(request.POST.get('royalty_max_redeem_pct','50') or '50'),
                    royalty_reward_mode=request.POST.get('royalty_reward_mode','amount') or 'amount',
                    royalty_reward_period=request.POST.get('royalty_reward_period','monthly') or 'monthly',
                    royalty_rank1_points=int(request.POST.get('royalty_rank1_points','500') or '500'),
                    royalty_rank2_points=int(request.POST.get('royalty_rank2_points','250') or '250'),
                    royalty_rank3_points=int(request.POST.get('royalty_rank3_points','100') or '100'),
                    kiosk_theme_color=(request.POST.get('kiosk_theme_color') or '#1e3a5f').strip(),
                    kiosk_welcome_text=(request.POST.get('kiosk_welcome_text') or 'Touch to order').strip(),
                    web_order_prefix=(request.POST.get('web_order_prefix') or '').strip().upper()[:10],
                    kiosk_order_prefix=(request.POST.get('kiosk_order_prefix') or '').strip().upper()[:10],
                )
                if 'logo' in request.FILES:
                    co.logo = request.FILES['logo']
                    co.save(update_fields=['logo'])
                if 'kiosk_logo' in request.FILES:
                    co.kiosk_logo = request.FILES['kiosk_logo']
                    co.save(update_fields=['kiosk_logo'])
                messages.success(request, f'Company "{co.name}" created.')
                return redirect('dashboard:company_detail', pk=co.pk)
            except (ValidationError, IntegrityError, Exception) as e:
                messages.error(request, f'Error creating company: {e}')

    return render(request, 'dashboard/companies/form.html', {
        'page_title': 'Add Company',
        'action': 'Add',
        'order_day_choices': ORDER_DAY_CHOICES,
        'selected_order_days': selected_days,
    })


@staff_role_required('superadmin')
def company_edit(request, pk):
    company = get_object_or_404(Company, pk=pk)
    selected_days = list(company.enabled_order_days)

    if request.method == 'POST':
        selected_days = _selected_order_days(request.POST)

        if not selected_days:
            messages.error(request, 'Select at least one open day for ordering.')
        else:
            try:
                company.name = request.POST.get('name', company.name).strip()
                company.company_address = request.POST.get('company_address', '')
                company.company_gst = request.POST.get('company_gst', '')
                company.fssai_number = request.POST.get('fssai_number', '')
                company.phone = request.POST.get('phone', '')
                company.bill_company = int(request.POST.get('bill_company', 2) or 2)
                company.company_meal_amount = _parse_meal_amount(request.POST.get('company_meal_amount', '0'))
                company.company_pay_meals_per_day = _parse_positive_int(request.POST.get('company_pay_meals_per_day'), default=1, minimum=1, maximum=10)
                company.subsidy_meals_per_day = _parse_positive_int(request.POST.get('subsidy_meals_per_day'), default=1, minimum=1, maximum=10)
                company.cod_payment     = request.POST.get('cod_payment') == 'on'
                company.online_payment  = request.POST.get('online_payment') == 'on'
                company.monthly_payment = request.POST.get('monthly_payment') == 'on'
                company.pos_cash_enabled = request.POST.get('pos_cash_enabled') == 'on'
                company.pos_upi_enabled = request.POST.get('pos_upi_enabled') == 'on'
                company.pos_card_enabled = request.POST.get('pos_card_enabled') == 'on'
                company.pos_card_fee_percent = _parse_percent(request.POST.get('pos_card_fee_percent'), default='3.50')
                company.store_status    = request.POST.get('store_status') == 'on'
                company.order_from_time = parse_time((request.POST.get('order_from_time') or '').strip()) or None
                company.order_to_time   = parse_time((request.POST.get('order_to_time') or '').strip()) or None
                company.order_open_days = selected_days
                # Royalty config
                company.require_customer_approval = request.POST.get('require_customer_approval') == 'on'
                company.fulfillment_mode = request.POST.get('fulfillment_mode', 'pickup') or 'pickup'
                company.royalty_enabled = request.POST.get('royalty_enabled') == 'on'
                try:
                    company.royalty_points_per_rupee = Decimal(request.POST.get('royalty_points_per_rupee','1') or '1')
                    company.royalty_min_redeem = int(request.POST.get('royalty_min_redeem','100') or '100')
                    company.royalty_max_redeem_pct = int(request.POST.get('royalty_max_redeem_pct','50') or '50')
                    company.royalty_reward_mode   = request.POST.get('royalty_reward_mode','amount') or 'amount'
                    company.royalty_reward_period = request.POST.get('royalty_reward_period','monthly') or 'monthly'
                    company.royalty_rank1_points  = int(request.POST.get('royalty_rank1_points','500') or '500')
                    company.royalty_rank2_points  = int(request.POST.get('royalty_rank2_points','250') or '250')
                    company.royalty_rank3_points  = int(request.POST.get('royalty_rank3_points','100') or '100')
                except (ValueError, TypeError):
                    pass
                # Kiosk settings are managed from dedicated config pages.
                if 'kiosk_theme_color' in request.POST:
                    company.kiosk_theme_color = (request.POST.get('kiosk_theme_color') or '#1e3a5f').strip()
                if 'kiosk_welcome_text' in request.POST:
                    company.kiosk_welcome_text = (request.POST.get('kiosk_welcome_text') or 'Touch to order').strip()
                company.web_order_prefix     = (request.POST.get('web_order_prefix') or '').strip().upper()[:10]
                company.kiosk_order_prefix   = (request.POST.get('kiosk_order_prefix') or '').strip().upper()[:10]
                if 'logo' in request.FILES:
                    company.logo = request.FILES['logo']
                if 'kiosk_logo' in request.FILES:
                    company.kiosk_logo = request.FILES['kiosk_logo']
                company.save()

                messages.success(request, 'Company updated successfully.')
                return redirect('dashboard:company_detail', pk=pk)
            except (ValidationError, IntegrityError, Exception) as e:
                messages.error(request, f'Error saving company: {e}')

    return render(request, 'dashboard/companies/form.html', {
        'company': company,
        'page_title': f'Edit - {company.name}',
        'action': 'Save',
        'order_day_choices': ORDER_DAY_CHOICES,
        'selected_order_days': selected_days,
    })


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  BUILDINGS
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin')
def building_add(request):
    companies = _companies(request.user)
    locations = Location.objects.filter(is_active=True)
    states = State.objects.all().order_by('name')
    cities = City.objects.filter(is_active=True, is_deleted=False).select_related('state').order_by('name')
    if request.method == 'POST':
        co = Company.objects.filter(pk=request.POST.get('company')).first()
        name = request.POST.get('name','').strip()
        if not co or not name:
            messages.error(request, 'Company and name required.')
        else:
            loc = Location.objects.filter(pk=request.POST.get('location')).first()
            state = State.objects.filter(pk=request.POST.get('state')).first() if request.POST.get('state') else None
            city = City.objects.filter(pk=request.POST.get('city')).first() if request.POST.get('city') else None
            Building.objects.create(company=co, name=name, location=loc, state=state, city=city)
            messages.success(request, f'"{name}" added.')
            return redirect('dashboard:company_detail', pk=co.pk)
    return render(request, 'dashboard/companies/building_form.html', {
        'companies':companies,'locations':locations,'states':states,'cities':cities,'page_title':'Add Building'})


@staff_role_required('superadmin')
def building_list(request):
    qs = Building.objects.filter(is_deleted=False).select_related('company', 'state', 'city', 'location').order_by('company__name', 'name')
    company_id = request.GET.get('company', '').strip()
    if company_id:
        qs = qs.filter(company_id=company_id)
    return render(request, 'dashboard/companies/building_list.html', {
        'buildings': qs,
        'companies': _companies(request.user),
        'company_id': company_id,
        'page_title': 'Buildings',
    })


@staff_role_required('superadmin')
def building_edit(request, pk):
    building = get_object_or_404(Building, pk=pk, is_deleted=False)
    companies = _companies(request.user)
    locations = Location.objects.filter(is_active=True)
    states = State.objects.all().order_by('name')
    cities = City.objects.filter(is_active=True, is_deleted=False).select_related('state').order_by('name')
    if request.method == 'POST':
        co = Company.objects.filter(pk=request.POST.get('company')).first()
        name = request.POST.get('name','').strip()
        if not co or not name:
            messages.error(request, 'Company and name required.')
        else:
            building.company = co
            building.name = name
            building.location = Location.objects.filter(pk=request.POST.get('location')).first()
            building.state = State.objects.filter(pk=request.POST.get('state')).first() if request.POST.get('state') else None
            building.city = City.objects.filter(pk=request.POST.get('city')).first() if request.POST.get('city') else None
            building.save()
            messages.success(request, f'Building "{building.name}" updated.')
            return redirect('dashboard:building_list')
    return render(request, 'dashboard/companies/building_form.html', {
        'building': building, 'companies': companies, 'locations': locations, 'states': states, 'cities': cities, 'page_title': f'Edit Building - {building.name}'
    })


@require_POST
@staff_role_required('superadmin')
def building_toggle(request, pk):
    building = get_object_or_404(Building, pk=pk, is_deleted=False)
    building.is_active = not building.is_active
    building.save(update_fields=['is_active'])
    messages.success(request, f'Building "{building.name}" is now {"active" if building.is_active else "inactive"}.')
    return redirect(request.META.get('HTTP_REFERER') or 'dashboard:building_list')


@require_POST
@staff_role_required('superadmin')
def building_delete(request, pk):
    building = get_object_or_404(Building, pk=pk, is_deleted=False)
    building.is_deleted = True
    building.save(update_fields=['is_deleted'])
    messages.success(request, f'Building "{building.name}" deleted.')
    return redirect('dashboard:building_list')


@require_POST
@staff_role_required('superadmin')
def building_bulk_delete(request):
    ids = _bulk_post_ids(request)
    count = Building.objects.filter(pk__in=ids, is_deleted=False).update(is_deleted=True) if ids else 0
    _bulk_message(request, 'building', count)
    return redirect('dashboard:building_list')


# â”€â”€â”€ State CRUD â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin')
def state_list(request):
    states = State.objects.all().order_by('name')
    return render(request, 'dashboard/companies/state_list.html', {
        'states': states,
        'page_title': 'States',
    })


@staff_role_required('superadmin')
def state_add(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'State name is required.')
        elif State.objects.filter(name__iexact=name).exists():
            messages.error(request, f'A state named "{name}" already exists.')
        else:
            State.objects.create(name=name)
            messages.success(request, f'State "{name}" created.')
            return redirect('dashboard:state_list')
    return render(request, 'dashboard/companies/state_form.html', {'page_title': 'Add State'})


@staff_role_required('superadmin')
def state_edit(request, pk):
    state = get_object_or_404(State, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'State name is required.')
        elif State.objects.filter(name__iexact=name).exclude(pk=pk).exists():
            messages.error(request, f'A state named "{name}" already exists.')
        else:
            state.name = name
            state.save(update_fields=['name'])
            messages.success(request, f'State "{state.name}" updated.')
            return redirect('dashboard:state_list')
    return render(request, 'dashboard/companies/state_form.html', {
        'state': state,
        'page_title': f'Edit State - {state.name}',
    })


@require_POST
@staff_role_required('superadmin')
def state_delete(request, pk):
    state = get_object_or_404(State, pk=pk)
    name = state.name
    state.cities.all().update(state=None)
    state.buildings.all().update(state=None)
    state.delete()
    messages.success(request, f'State "{name}" deleted.')
    return redirect('dashboard:state_list')


@require_POST
@staff_role_required('superadmin')
def state_bulk_delete(request):
    ids = _bulk_post_ids(request)
    count = 0
    for state in State.objects.filter(pk__in=ids):
        state.cities.all().update(state=None)
        state.buildings.all().update(state=None)
        state.delete()
        count += 1
    _bulk_message(request, 'state', count)
    return redirect('dashboard:state_list')


@staff_role_required('superadmin')
def city_list(request):
    qs = City.objects.filter(is_deleted=False).select_related('state').order_by('state__name', 'name')
    state_id = request.GET.get('state', '').strip()
    if state_id:
        qs = qs.filter(state_id=state_id)
    return render(request, 'dashboard/companies/city_list.html', {
        'cities': qs,
        'states': State.objects.all().order_by('name'),
        'state_id': state_id,
        'page_title': 'Cities',
    })


@staff_role_required('superadmin')
def city_add(request):
    states = State.objects.all().order_by('name')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        state = State.objects.filter(pk=request.POST.get('state')).first() if request.POST.get('state') else None
        if not name:
            messages.error(request, 'City name is required.')
        else:
            City.objects.create(name=name, state=state, is_active=request.POST.get('is_active') == 'on')
            messages.success(request, f'City "{name}" created.')
            return redirect('dashboard:city_list')
    return render(request, 'dashboard/companies/city_form.html', {'states': states, 'page_title': 'Add City'})


@staff_role_required('superadmin')
def city_edit(request, pk):
    city = get_object_or_404(City, pk=pk, is_deleted=False)
    states = State.objects.all().order_by('name')
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'City name is required.')
        else:
            city.name = name
            city.state = State.objects.filter(pk=request.POST.get('state')).first() if request.POST.get('state') else None
            city.is_active = request.POST.get('is_active') == 'on'
            city.save()
            messages.success(request, f'City "{city.name}" updated.')
            return redirect('dashboard:city_list')
    return render(request, 'dashboard/companies/city_form.html', {'city': city, 'states': states, 'page_title': f'Edit City - {city.name}'})


@require_POST
@staff_role_required('superadmin')
def city_toggle(request, pk):
    city = get_object_or_404(City, pk=pk, is_deleted=False)
    city.is_active = not city.is_active
    city.save(update_fields=['is_active'])
    messages.success(request, f'City "{city.name}" is now {"active" if city.is_active else "inactive"}.')
    return redirect(request.META.get('HTTP_REFERER') or 'dashboard:city_list')


@require_POST
@staff_role_required('superadmin')
def city_delete(request, pk):
    city = get_object_or_404(City, pk=pk, is_deleted=False)
    city.is_deleted = True
    city.save(update_fields=['is_deleted'])
    messages.success(request, f'City "{city.name}" deleted.')
    return redirect('dashboard:city_list')


@require_POST
@staff_role_required('superadmin')
def city_bulk_delete(request):
    ids = _bulk_post_ids(request)
    count = City.objects.filter(pk__in=ids, is_deleted=False).update(is_deleted=True) if ids else 0
    _bulk_message(request, 'city', count)
    return redirect('dashboard:city_list')


@staff_role_required('superadmin')
def cafe_list(request):
    qs = Cafe.objects.filter(is_deleted=False).select_related('company', 'building').order_by('company__name', 'building__name', 'name')
    company_id = request.GET.get('company', '').strip()
    building_id = request.GET.get('building', '').strip()
    if company_id:
        qs = qs.filter(company_id=company_id)
    if building_id:
        qs = qs.filter(building_id=building_id)
    return render(request, 'dashboard/companies/cafe_list.html', {
        'cafes': qs,
        'companies': _companies(request.user),
        'buildings': Building.objects.filter(is_deleted=False).select_related('company').order_by('company__name', 'name'),
        'company_id': company_id,
        'building_id': building_id,
        'page_title': 'Cafeterias',
    })


@staff_role_required('superadmin')
def cafe_add(request):
    companies = _companies(request.user)
    if request.method == 'POST':
        company = Company.objects.filter(pk=request.POST.get('company')).first()
        name = request.POST.get('name', '').strip()
        if not company or not name:
            messages.error(request, 'Company and cafeteria name are required.')
        else:
            building = Building.objects.filter(pk=request.POST.get('building'), is_deleted=False).first()
            Cafe.objects.create(company=company, name=name, building=building,
                                is_active=request.POST.get('is_active') == 'on')
            messages.success(request, f'Cafeteria "{name}" created.')
            return redirect('dashboard:cafe_list')
    buildings = Building.objects.filter(is_deleted=False).select_related('company','location').order_by('company__name','name')
    return render(request, 'dashboard/companies/cafe_form.html', {
        'companies': companies, 'buildings': buildings, 'page_title': 'Add Cafeteria'
    })


@staff_role_required('superadmin')
def cafe_edit(request, pk):
    cafe = get_object_or_404(Cafe, pk=pk, is_deleted=False)
    companies = _companies(request.user)
    if request.method == 'POST':
        company = Company.objects.filter(pk=request.POST.get('company')).first()
        name = request.POST.get('name', '').strip()
        if not company or not name:
            messages.error(request, 'Company and cafeteria name are required.')
        else:
            cafe.company = company
            cafe.name = name
            cafe.building = Building.objects.filter(pk=request.POST.get('building'), is_deleted=False).first()
            cafe.is_active = request.POST.get('is_active') == 'on'
            cafe.save()
            messages.success(request, f'Cafeteria "{cafe.name}" updated.')
            return redirect('dashboard:cafe_list')
    buildings = Building.objects.filter(is_deleted=False).select_related('company','location').order_by('company__name','name')
    return render(request, 'dashboard/companies/cafe_form.html', {
        'cafe': cafe, 'companies': companies, 'buildings': buildings,
        'page_title': f'Edit Cafeteria - {cafe.name}'
    })


@require_POST
@staff_role_required('superadmin')
def cafe_toggle(request, pk):
    cafe = get_object_or_404(Cafe, pk=pk, is_deleted=False)
    cafe.is_active = not cafe.is_active
    cafe.save(update_fields=['is_active'])
    messages.success(request, f'Cafeteria "{cafe.name}" is now {"active" if cafe.is_active else "inactive"}.')
    return redirect(request.META.get('HTTP_REFERER') or 'dashboard:cafe_list')


@require_POST
@staff_role_required('superadmin')
def cafe_delete(request, pk):
    cafe = get_object_or_404(Cafe, pk=pk, is_deleted=False)
    cafe.is_deleted = True
    cafe.save(update_fields=['is_deleted'])
    messages.success(request, f'Cafeteria "{cafe.name}" deleted.')
    return redirect('dashboard:cafe_list')


@require_POST
@staff_role_required('superadmin')
def cafe_bulk_delete(request):
    ids = _bulk_post_ids(request)
    count = Cafe.objects.filter(pk__in=ids, is_deleted=False).update(is_deleted=True) if ids else 0
    _bulk_message(request, 'cafeteria', count)
    return redirect('dashboard:cafe_list')


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  CUSTOMERS
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin','admin')
def customer_list(request):
    user = request.user
    qs   = Customer.objects.filter(is_deleted=False).select_related('company','building')
    if not user.is_superadmin:
        qs = qs.filter(company__in=_companies(user))
    q              = request.GET.get('q','')
    company_id     = request.GET.get('company','')
    building_id    = request.GET.get('building','')
    subsidy_filter = request.GET.get('subsidy','')
    if q:
        qs = qs.filter(Q(name__icontains=q)|Q(email__icontains=q)|Q(phone__icontains=q))
    if company_id and (user.is_superadmin or _companies(user).filter(pk=company_id).exists()):
        qs = qs.filter(company_id=company_id)
    if building_id:
        qs = qs.filter(building_id=building_id)
    status_filter = request.GET.get('status', '')
    if subsidy_filter == '1':
        qs = qs.exclude(meal_benefit=Customer.MEAL_BENEFIT_NONE)
    elif subsidy_filter == '0':
        qs = qs.filter(meal_benefit=Customer.MEAL_BENEFIT_NONE)
    if status_filter == 'pending':
        qs = qs.filter(is_approved=False)
    elif status_filter == 'approved':
        qs = qs.filter(is_approved=True)
    total_count = qs.count()  # count AFTER all filters applied â€” used for headcount display
    page_obj = Paginator(qs.order_by('-id'), 50).get_page(request.GET.get('page'))
    return render(request, 'dashboard/customers/list.html', {
        'customers':page_obj,'page_obj':page_obj,'q':q,
        'company_id':company_id,'building_id':building_id,
        'subsidy_filter':subsidy_filter,
        'status_filter':status_filter,
        'companies':_companies(user),
        'buildings':Building.objects.filter(is_deleted=False).order_by('name'),
        'total_count':total_count,
        'page_title':'Customers',
        'lp': get_list_perms(user, 'perm_customers'),
    })


@staff_role_required('superadmin','admin')
def customer_add(request):
    if not user_can_action(request.user, 'perm_customers', 'add'):
        return _deny_dashboard_action(request)
    birth_day_value = ''
    birth_month_value = ''
    birth_date_error = ''
    initial_wallet_raw = ''
    initial_points_raw = ''
    initial_wallet = Decimal('0.00')
    initial_points = 0
    initial_balance_error = ''
    post_data = None
    if request.method == 'POST':
        birth_day_value = (request.POST.get('birth_day') or '').strip()
        birth_month_value = (request.POST.get('birth_month') or '').strip()
        initial_wallet_raw = (request.POST.get('initial_wallet_balance') or '').strip()
        initial_points_raw = (request.POST.get('initial_royalty_points') or '').strip()
        post_data = request.POST.copy()
        if not post_data.get('date_of_birth'):
            coerced_birth_date, birth_date_error = _coerce_customer_birth_date(birth_day_value, birth_month_value)
            if coerced_birth_date:
                post_data['date_of_birth'] = coerced_birth_date.isoformat()
        try:
            initial_wallet = Decimal(initial_wallet_raw or '0').quantize(Decimal('0.01'))
            if initial_wallet < 0:
                raise InvalidOperation()
        except (InvalidOperation, TypeError, ValueError):
            initial_wallet = Decimal('0.00')
            initial_balance_error = 'Initial wallet balance must be a valid non-negative amount.'
        try:
            initial_points = int(initial_points_raw or '0')
            if initial_points < 0:
                raise ValueError()
        except (TypeError, ValueError):
            initial_points = 0
            initial_balance_error = initial_balance_error or 'Initial royalty points must be a valid non-negative number.'
    form = CustomerRegisterForm(post_data or None)
    if not request.user.is_superadmin:
        form.fields['company'].queryset = _companies(request.user)
    if request.method == 'POST' and form.is_valid() and not birth_date_error and not initial_balance_error:
        # Admin-created customers are implicitly email-verified â€” they were
        # added directly by a staff member and do not go through the email
        # verification flow that self-registered customers use.
        customer = form.save(commit=False, email_verified=True)
        customer.meal_benefit = _normalize_customer_meal_benefit(
            customer,
            request.POST.get('meal_benefit', Customer.MEAL_BENEFIT_NONE),
        )
        customer.subsidy_amount_override = _parse_customer_subsidy_override(
            request.POST.get('subsidy_amount_override'),
            customer.meal_benefit,
        )
        customer.subsidy_eligible = customer.meal_benefit != Customer.MEAL_BENEFIT_NONE
        customer.wallet_balance = initial_wallet
        customer.royalty_points = initial_points
        customer.save()
        if initial_wallet > 0 or initial_points > 0:
            WalletTransaction.objects.create(
                customer=customer,
                txn_type=WalletTransaction.TYPE_ADJUSTMENT,
                wallet_delta=initial_wallet,
                points_delta=initial_points,
                balance_after=customer.wallet_balance,
                points_after=customer.royalty_points,
                note='Opening wallet/royalty balance set during customer creation.',
                created_by=request.user.email,
            )
        messages.success(request, 'Customer added.')
        return redirect('dashboard:customer_list')
    selected_company = None
    selected_company_id = ''
    if request.user.is_superadmin:
        selected_company_id = (post_data or request.POST).get('company') if request.method == 'POST' else ''
        if selected_company_id:
            selected_company = Company.objects.filter(pk=selected_company_id, is_deleted=False).first()
    else:
        selected_company_id = (post_data or request.POST).get('company') if request.method == 'POST' else ''
        selected_company = _companies(request.user).filter(pk=selected_company_id).first() if selected_company_id else get_primary_staff_company(request.user)
        selected_company_id = str(selected_company.pk) if selected_company else ''
    return render(request, 'dashboard/customers/form.html', {
        'form': form,
        'page_title': 'Add Customer',
        'birth_day_options': CUSTOMER_BIRTH_DAY_OPTIONS,
        'birth_month_options': CUSTOMER_BIRTH_MONTH_OPTIONS,
        'birth_day_value': birth_day_value,
        'birth_month_value': birth_month_value,
        'birth_date_error': birth_date_error,
        'selected_company': selected_company,
        'selected_company_id': selected_company_id,
        'company_benefit_data': form.fields['company'].queryset,
        'initial_wallet_raw': initial_wallet_raw,
        'initial_points_raw': initial_points_raw,
        'initial_balance_error': initial_balance_error,
    })


@staff_role_required('superadmin','admin')
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk, is_deleted=False)
    user = request.user
    _perm = check_module_permission(request, 'perm_customers')
    if _perm: return _perm
    if not user.is_superadmin and not user_can_access_company(user, customer.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:customer_list')
    buildings = Building.objects.filter(company=customer.company, is_deleted=False)
    if request.method == 'POST':
        if user.role != 'superadmin' and get_module_level(request.user, 'perm_customers') == 'full_edit':
            _diffs = {
                'name':        {'label': 'Name',     'before': customer.name,        'after': (request.POST.get('name') or customer.name).strip()},
                'phone':       {'label': 'Phone',    'before': customer.phone or '',  'after': (request.POST.get('phone') or '').strip()},
                'is_active':   {'label': 'Active',   'before': customer.is_active,   'after': request.POST.get('is_active') == 'on'},
                'is_approved': {'label': 'Approved', 'before': customer.is_approved, 'after': request.POST.get('is_approved') == 'on'},
            }
            _pc = create_pending_change(request, 'perm_customers', customer, _diffs)
            if _pc:
                messages.success(request, 'Your changes have been submitted for superadmin review.')
            else:
                messages.info(request, 'No changes detected.')
            return redirect('dashboard:customer_list')
        customer.name             = request.POST.get('name', customer.name).strip()
        customer.phone            = request.POST.get('phone', customer.phone).strip()
        customer.address          = request.POST.get('address','')
        dob_raw = (request.POST.get('date_of_birth') or '').strip()
        if dob_raw and not customer.date_of_birth:
            try:
                customer.date_of_birth = date.fromisoformat(dob_raw)
            except ValueError:
                messages.error(request, 'Invalid birth date.')
                return redirect('dashboard:customer_edit', pk=customer.pk)
        elif not customer.date_of_birth:
            coerced_birth_date, birth_date_error = _coerce_customer_birth_date(
                request.POST.get('birth_day'),
                request.POST.get('birth_month'),
            )
            if birth_date_error:
                messages.error(request, birth_date_error)
                return redirect('dashboard:customer_edit', pk=customer.pk)
            if coerced_birth_date:
                customer.date_of_birth = coerced_birth_date
        customer.cod_payment      = request.POST.get('cod_payment')=='on'
        customer.monthly_payment  = request.POST.get('monthly_payment')=='on'
        customer.meal_benefit     = _normalize_customer_meal_benefit(
            customer,
            request.POST.get('meal_benefit', Customer.MEAL_BENEFIT_NONE),
        )
        customer.subsidy_amount_override = _parse_customer_subsidy_override(
            request.POST.get('subsidy_amount_override'),
            customer.meal_benefit,
        )
        customer.subsidy_eligible = customer.meal_benefit != Customer.MEAL_BENEFIT_NONE
        customer.is_active    = request.POST.get('is_active')=='on'
        customer.is_approved  = request.POST.get('is_approved')=='on'
        bld = request.POST.get('building')
        customer.building = Building.objects.filter(
            pk=bld,
            company=customer.company,
            is_deleted=False,
        ).first() if bld else None
        pw = request.POST.get('new_password','').strip()
        if pw:
            customer.set_password(pw)
        customer.save()
        # Points adjustment
        pts_raw = (request.POST.get('royalty_adjust') or '').strip()
        if pts_raw:
            try:
                pts = int(pts_raw)
                if pts != 0:
                    from apps.accounts.models import WalletTransaction
                    customer.__class__.objects.filter(pk=customer.pk).update(
                        royalty_points=max(0, customer.royalty_points + pts)
                    )
                    customer.refresh_from_db(fields=['royalty_points'])
                    WalletTransaction.objects.create(
                        customer=customer, txn_type=WalletTransaction.TYPE_ADJUSTMENT,
                        points_delta=pts, balance_after=customer.wallet_balance,
                        points_after=customer.royalty_points,
                        note=f'Manual points adjustment by {request.user.email}',
                        created_by=request.user.email,
                    )
                    messages.success(request, f'Royalty points adjusted by {pts:+d}.')
            except (ValidationError, IntegrityError, ValueError) as e:
                messages.error(request, f'Points adjustment failed: {e}')
        messages.success(request, 'Customer updated.')
        return redirect('dashboard:customer_list')
    from apps.accounts.models import WalletTransaction as WT
    wallet_history = WT.objects.filter(customer=customer).order_by('-created_at')[:15]
    _lj, _pl = get_locked_html_names(request.user, 'perm_customers')
    return render(request, 'dashboard/customers/edit.html', {
        'customer': customer, 'buildings': buildings,
        'wallet_history': wallet_history,
        'birth_day_options': CUSTOMER_BIRTH_DAY_OPTIONS,
        'birth_month_options': CUSTOMER_BIRTH_MONTH_OPTIONS,
        'page_title': f'Edit - {customer.name}', 'perm_level': _pl or 'full_edit', 'locked_names_json': _lj})


@require_POST
@staff_role_required('superadmin','admin')
def customer_delete(request, pk):
    if not _has_full_module_edit(request.user, 'perm_customers'):
        return _deny_dashboard_action(request)
    customer = get_object_or_404(_customer_qs_for_user(request.user), pk=pk)
    customer.is_deleted = True
    customer.is_active = False
    customer.save(update_fields=['is_deleted', 'is_active'])
    messages.success(request, f'Customer "{customer.name}" deleted.')
    return redirect('dashboard:customer_list')


@require_POST
@staff_role_required('superadmin','admin')
def customer_bulk_delete(request):
    if not _has_full_module_edit(request.user, 'perm_customers'):
        return _deny_dashboard_action(request)
    ids = _bulk_post_ids(request)
    qs = _customer_qs_for_user(request.user).filter(pk__in=ids)
    count = qs.update(is_deleted=True, is_active=False) if ids else 0
    _bulk_message(request, 'customer', count)
    return redirect('dashboard:customer_list')


@require_POST
@staff_role_required('superadmin', 'admin')

def customer_toggle_subsidy(request, pk):
    """AJAX quick toggle: Off <-> Subsidy."""
    customer = get_object_or_404(Customer, pk=pk, is_deleted=False)
    user = request.user
    if not user_can_action(user, 'perm_customers', 'subsidy_toggle'):
        return JsonResponse({'error': 'Permission denied.'}, status=403)
    if not user.is_superadmin and not user_can_access_company(user, customer.company_id):
        return JsonResponse({'error': 'Access denied.'}, status=403)
    if customer.company.bill_company != 2:
        return JsonResponse({'error': 'Company is not in company-pay/subsidy mode.'}, status=400)
    customer.meal_benefit = Customer.MEAL_BENEFIT_NONE if customer.meal_benefit != Customer.MEAL_BENEFIT_NONE else Customer.MEAL_BENEFIT_SUBSIDY
    customer.subsidy_eligible = customer.meal_benefit != Customer.MEAL_BENEFIT_NONE
    customer.save(update_fields=['meal_benefit', 'subsidy_eligible'])
    return JsonResponse({
        'subsidy_eligible': customer.subsidy_eligible,
        'meal_benefit': customer.meal_benefit,
        'subsidy_source_label': customer.subsidy_source_label,
    })


@require_POST
@staff_role_required('superadmin', 'admin')
def customer_set_meal_benefit(request, pk):
    customer = get_object_or_404(Customer, pk=pk, is_deleted=False)
    user = request.user
    if not user_can_action(user, 'perm_customers', 'field_meal_benefit'):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard:no_access')
    if not user.is_superadmin and not user_can_access_company(user, customer.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:customer_list')
    benefit = request.POST.get('meal_benefit', Customer.MEAL_BENEFIT_NONE)
    if benefit not in {Customer.MEAL_BENEFIT_NONE, Customer.MEAL_BENEFIT_SUBSIDY, Customer.MEAL_BENEFIT_COMPANY_PAY}:
        benefit = Customer.MEAL_BENEFIT_NONE
    if customer.company.bill_company != 2 and benefit != Customer.MEAL_BENEFIT_NONE:
        messages.error(request, 'This company is not in company-paid meals mode.')
        return redirect(request.META.get('HTTP_REFERER', 'dashboard:customer_list'))
    customer.meal_benefit = benefit
    customer.subsidy_eligible = benefit != Customer.MEAL_BENEFIT_NONE
    if benefit != Customer.MEAL_BENEFIT_SUBSIDY:
        customer.subsidy_amount_override = None
    customer.save(update_fields=['meal_benefit', 'subsidy_eligible', 'subsidy_amount_override'])
    messages.success(request, f'Meal benefit updated for {customer.name}.')
    return redirect(request.META.get('HTTP_REFERER', 'dashboard:customer_list'))




def _promote_due_ready_orders(company=None):
    from apps.orders.models import CounterTicket

    now = timezone.now()
    qs = Order.objects.filter(
        order_status=OrderStatusChoices.CONFIRMED,
        is_deleted=False,
        auto_ready_at__isnull=False,
        auto_ready_at__lte=now,
    )
    if company is not None:
        qs = qs.filter(company=company)

    for order in qs:
        order.order_status = OrderStatusChoices.READY
        order.auto_ready_at = None
        order.save(update_fields=['order_status', 'auto_ready_at', 'updated_at'])

        order.counter_tickets.filter(
            status__in=[CounterTicket.STATUS_PENDING, CounterTicket.STATUS_PREPARING]
        ).update(status=CounterTicket.STATUS_READY, updated_at=now)

        OrderStatus.objects.create(
            order=order,
            status=OrderStatusChoices.READY,
            details='Auto-marked ready after configured preparation time.',
            created_at=now,
        )

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  ORDERS
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin','admin','pos','cafeman')
def order_list(request):
    _perm = check_module_permission(request, 'perm_orders')
    if _perm: return _perm
    user = request.user
    if user.is_superadmin:
        _promote_due_ready_orders(None)
    else:
        for company in _companies(user):
            _promote_due_ready_orders(company)
    qs   = Order.objects.filter(is_deleted=False).select_related(
        'customer','company').order_by('-created_at')
    if not user.is_superadmin:
        qs = qs.filter(company__in=_companies(user))
    status_filter  = request.GET.get('status','')
    date_filter    = request.GET.get('date','')
    company_filter = request.GET.get('company','')
    q              = request.GET.get('q','')
    if status_filter: qs = qs.filter(order_status=status_filter)
    if date_filter:   qs = qs.filter(created_at__date=date_filter)
    if company_filter and (user.is_superadmin or _companies(user).filter(pk=company_filter).exists()): qs = qs.filter(company_id=company_filter)
    if q:
        qs = qs.filter(
            Q(order_number__icontains=q) |
            Q(customer_name_snapshot__icontains=q) |
            Q(customer_phone_snapshot__icontains=q) |
            Q(customer__name__icontains=q) |
            Q(customer__phone__icontains=q)
        )
    # Subsidy filter (new feature: filter orders where company covered some amount)
    subsidy_filter = request.GET.get('subsidy', '')
    if subsidy_filter == '1':
        qs = qs.filter(bill_to_company__gt=0)
    page_obj = Paginator(qs, 50).get_page(request.GET.get('page'))
    for order in page_obj.object_list:
        order.display_order_number = _cashier_display_order_number(order.order_number)
    company_label = 'All Companies'
    if company_filter and user.is_superadmin:
        company_obj = Company.objects.filter(pk=company_filter, is_deleted=False).only('name').first()
        if company_obj:
            company_label = company_obj.name
    elif not user.is_superadmin:
        names = list(_companies(user).values_list('name', flat=True))
        company_label = ', '.join(names[:2]) + (' +' + str(len(names) - 2) if len(names) > 2 else '') if names else 'No Assigned Sites'

    report_date_label = date_filter or 'All Dates'

    if request.GET.get('export') == 'csv':
        return _export_orders_csv(qs)
    if request.GET.get('export') == 'excel':
        return _export_orders_excel(qs)
    if request.GET.get('export') == 'pdf':
        return _export_orders_pdf(qs, company_label=company_label, report_date_label=report_date_label)
    return render(request, 'dashboard/orders/list.html', {
        'orders':page_obj,'page_obj':page_obj,
        'status_filter':status_filter,'date_filter':date_filter,
        'company_filter':company_filter,'q':q,'subsidy_filter':subsidy_filter,
        'companies':_companies(user),'page_title':'Orders',
    })


def _export_orders_csv(qs):
    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = 'attachment; filename="orders.csv"'
    w = csv.writer(resp)
    w.writerow(['Order#','Customer','Phone','Company','Gross Subtotal','Offer Discount','Coupon Discount','Net Total','Customer Paid','Company Covered','Payment','Status','Date'])
    for o in qs[:5000]:
        w.writerow([
            o.order_number, o.display_customer_name, o.display_customer_phone,
            o.company.name, o.subtotal,
            o.offer_discount, o.coupon_discount,
            o.total_amount, o.my_pay, o.bill_to_company,
            o.get_payment_mode_display(), o.status_label,
            o.created_at.strftime('%d-%m-%Y %H:%M') if o.created_at else '',
        ])
    return resp


def _export_orders_excel(qs):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Orders'
        headers = ['Order#','Date','Customer','Phone','Company',
                   'Gross Subtotal','Offer Discount','Coupon Discount',
                   'Net Total','Customer Paid','Company Covered',
                   'Payment Mode','Status']
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        subsidy_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        for row_idx, o in enumerate(qs[:5000], 2):
            row_data = [
                o.order_number,
                o.created_at.strftime('%d-%m-%Y %H:%M') if o.created_at else '',
                o.display_customer_name, o.display_customer_phone, o.company.name,
                float(o.subtotal), float(o.offer_discount), float(o.coupon_discount),
                float(o.total_amount), float(o.my_pay), float(o.bill_to_company),
                o.get_payment_mode_display(), o.status_label,
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                if float(o.bill_to_company) > 0:
                    cell.fill = subsidy_fill
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = 'attachment; filename="orders_subsidy.xlsx"'
        return resp
    except ImportError:
        return _export_orders_csv(qs)




def _export_orders_pdf(qs, company_label='All Companies', report_date_label='All Dates'):
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    def fmt_money(value):
        try:
            return f"Rs. {Decimal(value or 0):.2f}"
        except Exception:
            return "Rs. 0.00"

    def clean(value, default=''):
        text = str(value or default).strip()
        return text or default

    orders = list(
        qs.select_related('company', 'customer')
          .prefetch_related('items__product__category')
          .order_by('created_at')[:1500]
    )

    effective_date_label = clean(report_date_label, '')
    if (not effective_date_label or effective_date_label == 'All Dates') and orders:
        date_values = [timezone.localtime(o.created_at).date() for o in orders if o.created_at]
        if date_values:
            min_date = min(date_values)
            max_date = max(date_values)
            if min_date == max_date:
                effective_date_label = min_date.strftime('%d-%m-%Y')
            else:
                effective_date_label = f"{min_date.strftime('%d-%m-%Y')} to {max_date.strftime('%d-%m-%Y')}"
    if not effective_date_label:
        effective_date_label = 'All Dates'

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=9 * mm,
        rightMargin=9 * mm,
        topMargin=9 * mm,
        bottomMargin=9 * mm,
    )

    styles = getSampleStyleSheet()
    header_box_style = ParagraphStyle(
        'HeaderBox',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        alignment=1,
        textColor=colors.white,
    )
    cell_style = ParagraphStyle(
        'Cell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11,
        textColor=colors.black,
    )
    cell_bold_style = ParagraphStyle(
        'CellBold',
        parent=cell_style,
        fontName='Helvetica-Bold',
        textColor=colors.white,
        alignment=1,
    )

    generated_on = timezone.localtime(timezone.now()).strftime('%d-%m-%Y %I:%M %p')

    header_html = (
        f'<font name="Helvetica-Bold" size="18">{clean(company_label, "All Companies")}</font><br/>'
        f'<font name="Helvetica" size="10">Powered by Neverno</font><br/>'
        f'<font name="Helvetica" size="10">Date: {effective_date_label}</font><br/>'
        f'<font name="Helvetica" size="9">Generated On: {generated_on}</font>'
    )

    header_table = Table([[Paragraph(header_html, header_box_style)]], colWidths=[doc.width])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1E3A5F')),
        ('BOX', (0, 0), (-1, -1), 0.6, colors.HexColor('#1E3A5F')),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))

    story = [header_table, Spacer(1, 8)]

    table_rows = [[
        Paragraph('Customer Name', cell_bold_style),
        Paragraph('Customer Number', cell_bold_style),
        Paragraph('Items', cell_bold_style),
        Paragraph('Category of Item', cell_bold_style),
        Paragraph('Amount', cell_bold_style),
        Paragraph('Date', cell_bold_style),
    ]]

    for order in orders:
        order_date = timezone.localtime(order.created_at).strftime('%d-%m-%Y') if order.created_at else ''
        customer_name = clean(getattr(order, 'display_customer_name', ''), 'Customer')
        customer_phone = clean(getattr(order, 'display_customer_phone', ''), '-')
        active_items = [item for item in order.items.all() if not getattr(item, 'is_deleted', False)]

        if not active_items:
            table_rows.append([
                Paragraph(customer_name, cell_style),
                Paragraph(customer_phone, cell_style),
                Paragraph('-', cell_style),
                Paragraph('-', cell_style),
                Paragraph(fmt_money(order.total_amount), cell_style),
                Paragraph(order_date, cell_style),
            ])
            continue

        for item in active_items:
            product = getattr(item, 'product', None)
            item_name = clean(getattr(product, 'name', ''), 'Item')
            qty = int(getattr(item, 'qty', 0) or 0)
            item_label = f'{item_name} x {qty}' if qty > 1 else item_name
            category_name = clean(getattr(getattr(product, 'category', None), 'name', ''), '-')
            amount = fmt_money(getattr(item, 'line_total', 0))
            table_rows.append([
                Paragraph(customer_name, cell_style),
                Paragraph(customer_phone, cell_style),
                Paragraph(item_label, cell_style),
                Paragraph(category_name, cell_style),
                Paragraph(amount, cell_style),
                Paragraph(order_date, cell_style),
            ])

    if len(table_rows) == 1:
        table_rows.append([
            Paragraph('No records found', cell_style),
            Paragraph('-', cell_style),
            Paragraph('-', cell_style),
            Paragraph('-', cell_style),
            Paragraph('-', cell_style),
            Paragraph('-', cell_style),
        ])

    table = Table(
        table_rows,
        repeatRows=1,
        colWidths=[52*mm, 36*mm, 85*mm, 52*mm, 30*mm, 28*mm],
    )
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('LEADING', (0, 0), (-1, -1), 11),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#B8C4CF')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F4F7FA')]),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
        ('ALIGN', (5, 1), (5, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
    ]))

    story.append(table)
    doc.build(story)

    pdf = buffer.getvalue()
    buffer.close()

    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = 'attachment; filename="orders_client_report.pdf"'
    return resp


@staff_role_required('superadmin','admin','pos','cafeman')
def order_detail(request, pk):
    user  = request.user
    if user.is_superadmin:
        _promote_due_ready_orders(None)
    else:
        for company in _companies(user):
            _promote_due_ready_orders(company)
    order = get_object_or_404(Order.objects.select_related('customer', 'company'), pk=pk, is_deleted=False)
    if not user.is_superadmin and not user_can_access_company(user, order.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:order_list')
    items    = order.items.select_related('product').all()
    statuses = order.status_history.order_by('created_at')
    return render(request, 'dashboard/orders/detail.html', {
        'order':order,'items':items,'statuses':statuses,
        'page_title':f'Order #{order.order_number}'})


@staff_role_required('superadmin','admin','pos','cafeman')
def order_update_status(request, pk):
    user  = request.user
    order = get_object_or_404(Order, pk=pk)
    if not user.is_superadmin and not user_can_access_company(user, order.company_id):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Access denied.'}, status=403)
        messages.error(request, 'Access denied.')
        return redirect('dashboard:order_list')
    if request.method == 'POST':
        try:
            new_status = int(request.POST.get('status', order.order_status))
        except (ValueError, TypeError):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Invalid status.'}, status=400)
            messages.error(request, 'Invalid status.')
            return redirect('dashboard:order_detail', pk=pk)
        # BUG 5 FIX: validate status BEFORE saving â€” reject invalid values
        valid_statuses = [s.value for s in OrderStatusChoices]
        if new_status not in valid_statuses:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': f'Invalid status value: {new_status}'}, status=400)
            messages.error(request, f'Invalid status: {new_status}')
            return redirect('dashboard:order_detail', pk=pk)
        details = request.POST.get('details', '')
        previous_status = order.order_status
        if previous_status in (OrderStatusChoices.DELIVERED, OrderStatusChoices.CANCELLED):
            err = 'This state is final and cannot be changed.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': err}, status=400)
            messages.error(request, err)
            return redirect('dashboard:order_detail', pk=pk)
        if new_status < previous_status:
            err = 'Order status cannot move backward.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': err}, status=400)
            messages.error(request, err)
            return redirect('dashboard:order_detail', pk=pk)
        if new_status == OrderStatusChoices.CANCELLED and user.role not in ('superadmin', 'admin'):
            err = 'Only admin can cancel orders.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': err}, status=403)
            messages.error(request, err)
            return redirect('dashboard:order_detail', pk=pk)
        order.order_status = new_status
        if new_status == OrderStatusChoices.CONFIRMED and previous_status != OrderStatusChoices.CONFIRMED:
            order.auto_ready_at = order.calculate_auto_ready_at(start_from=timezone.now())
            if order.payment_mode == PaymentModeChoices.CASH:
                order.payment_status = 'paid'
            elif order.payment_mode == PaymentModeChoices.MONTHLY and order.payment_status == 'pending':
                # Monthly orders confirmed by cashier should be marked approved, not left as pending
                order.payment_status = 'approved'
            if not details:
                details = 'Order confirmed by cashier/admin.'
        elif new_status in (OrderStatusChoices.READY, OrderStatusChoices.DELIVERED, OrderStatusChoices.CANCELLED):
            order.auto_ready_at = None
        order.save()
        if new_status == OrderStatusChoices.CANCELLED:
            for oi in order.items.filter(is_deleted=False).select_related('product'):
                if oi.product:
                    from apps.orders.views import _restock
                    _restock(oi.product, oi.qty, 'web', order.pk, order.company, f'Admin cancel {order.order_number}')
            from apps.accounts.models import WalletTransaction as _WT
            from django.db import transaction as _tx
            wallet_used = order.wallet_used or Decimal('0.00')
            points_redeemed = order.points_redeemed or 0
            if wallet_used > 0 or points_redeemed > 0:
                with _tx.atomic():
                    locked_customer = order.customer.__class__._default_manager.select_for_update().get(pk=order.customer_id)
                    if wallet_used > 0:
                        locked_customer.__class__.objects.filter(pk=locked_customer.pk).update(
                            wallet_balance=locked_customer.wallet_balance + wallet_used
                        )
                        locked_customer.refresh_from_db(fields=['wallet_balance', 'royalty_points'])
                        _WT.objects.create(
                            customer=locked_customer, txn_type=_WT.TYPE_REFUND,
                            wallet_delta=wallet_used, balance_after=locked_customer.wallet_balance,
                            points_after=locked_customer.royalty_points,
                            order_ref=order.order_number,
                            note=f'Wallet refunded on admin cancel of order {order.order_number}',
                            created_by=request.user.email,
                        )
                    if points_redeemed > 0:
                        locked_customer.__class__.objects.filter(pk=locked_customer.pk).update(
                            royalty_points=locked_customer.royalty_points + points_redeemed
                        )
                        locked_customer.refresh_from_db(fields=['wallet_balance', 'royalty_points'])
                        _WT.objects.create(
                            customer=locked_customer, txn_type=_WT.TYPE_ADJUSTMENT,
                            points_delta=points_redeemed, balance_after=locked_customer.wallet_balance,
                            points_after=locked_customer.royalty_points,
                            order_ref=order.order_number,
                            note=f'Royalty points restored on admin cancel of order {order.order_number}',
                            created_by=request.user.email,
                        )
        OrderStatus.objects.create(
            order=order, status=new_status, details=details,
            created_at=timezone.now()
        )
        # â”€â”€ Customer notification on status change â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
        try:
            from apps.accounts.views import _create_customer_notification
            from apps.core.models import Notification as _Notif
            _status_messages = {
                OrderStatusChoices.CONFIRMED: ('Order Confirmed', f'Your order #{order.order_number} has been confirmed and will be prepared shortly.'),
                OrderStatusChoices.PREPARING: ('Order Being Prepared', f'Your order #{order.order_number} is now being prepared.'),
                OrderStatusChoices.READY:     ('Order Ready', f'Your order #{order.order_number} is ready for pickup!'),
                OrderStatusChoices.DELIVERED: ('Order Delivered', f'Your order #{order.order_number} has been delivered. Enjoy your meal!'),
                OrderStatusChoices.CANCELLED: ('Order Cancelled', f'Your order #{order.order_number} has been cancelled.'),
            }
            if new_status in _status_messages and order.customer_id:
                _title, _message = _status_messages[new_status]
                _create_customer_notification(
                    order.customer,
                    notif_type=_Notif.TYPE_ORDER,
                    title=_title,
                    message=_message,
                    link=f'/orders/{order.pk}/',
                )
        except Exception:
            pass
        messages.success(request, 'Status updated.')
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success':True,'new_status':new_status,
                                 'label':OrderStatusChoices(new_status).label,
                                 'color':order.status_color})
    return redirect('dashboard:order_detail', pk=pk)


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  REVIEWS
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin','admin','reports')
def reviews_list(request):
    _perm = check_module_permission(request, 'perm_reviews')
    if _perm: return _perm
    from apps.reviews.models import Review
    user = request.user
    qs   = Review.objects.filter(is_deleted=False).select_related(
        'customer','customer__company','order').order_by('-created_at')
    if not user.is_superadmin:
        qs = qs.filter(customer__company__in=_companies(user))
    rating_filter = request.GET.get('rating','')
    if rating_filter: qs = qs.filter(rating=rating_filter)
    page_obj = Paginator(qs, 40).get_page(request.GET.get('page'))
    return render(request, 'dashboard/reviews/list.html', {
        'reviews':page_obj,'page_obj':page_obj,
        'rating_filter':rating_filter,'page_title':'Customer Reviews',
    })


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  STAFF
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin')
def staff_list(request):
    qs = StaffUser.objects.all().select_related('company').prefetch_related('site_access').order_by('role','name')
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            email = request.POST.get('email','').strip()
            if email and not StaffUser.objects.filter(email=email).exists():
                selected_sites = Company.objects.filter(
                    pk__in=request.POST.getlist('site_access'),
                    is_active=True,
                    is_deleted=False,
                ).order_by('name')
                co = Company.objects.filter(pk=request.POST.get('company')).first()
                if not co:
                    co = selected_sites.first()
                u  = StaffUser(
                    email=email, name=request.POST.get('name',''),
                    phone=request.POST.get('phone',''),
                    role=request.POST.get('role', StaffUser.ROLE_ADMIN),
                    company=co, is_staff=True, is_active=True)
                u.set_password(request.POST.get('password','admin123'))
                u.save()
                site_ids = set(selected_sites.values_list('pk', flat=True))
                if co:
                    site_ids.add(co.pk)
                u.site_access.set(Company.objects.filter(pk__in=site_ids))
                messages.success(request, f'{email} created. Set their permissions below.')
                return redirect('dashboard:permission_matrix', pk=u.pk)
            else:
                messages.error(request, 'Email exists or blank.')
        elif action == 'delete':
            StaffUser.objects.filter(pk=request.POST.get('uid')).exclude(role='superadmin').delete()
            messages.success(request, 'Deleted.')
        return redirect('dashboard:staff_list')
    staff_rows = list(qs)
    for staff_user in staff_rows:
        site_names = []
        seen = set()
        if staff_user.company_id and staff_user.company:
            site_names.append(staff_user.company.name)
            seen.add(staff_user.company_id)
        for site in staff_user.site_access.all():
            if site.pk not in seen:
                site_names.append(site.name)
                seen.add(site.pk)
        staff_user.site_names_display = ', '.join(site_names) or '--'
        staff_user.site_count = len(seen)
    return render(request, 'dashboard/staff/list.html', {
        'staff':staff_rows,'companies':_companies(request.user),
        'role_choices':StaffUser.ROLE_CHOICES,'page_title':'Staff Users'})


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  REPORTS
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin','admin','pos','cafeman','reports')
def reports(request):
    _perm = check_module_permission(request, 'perm_reports')
    if _perm: return _perm
    from datetime import date, timedelta, datetime
    from django.db.models import Sum, Count, FloatField, F as _F, ExpressionWrapper, Case, When, DecimalField

    user = request.user
    cf = _scope(user)

    today = date.today()
    from_date = request.GET.get('from_date', today.replace(day=1).isoformat())
    to_date = request.GET.get('to_date', today.isoformat())
    company_filter = request.GET.get('company', '')
    sale_particular = request.GET.get('sale_particular', '')

    try:
        fd = datetime.strptime(from_date, '%Y-%m-%d').date()
        td = datetime.strptime(to_date, '%Y-%m-%d').date()
    except ValueError:
        fd = today.replace(day=1)
        td = today
        from_date = fd.isoformat()
        to_date = td.isoformat()

    allowed_companies = _companies(user)

    # Manual company settlement entry
    if request.method == 'POST':
        if not (user.is_superadmin or user.role == StaffUser.ROLE_ADMIN):
            messages.error(request, 'You do not have permission to record company payments.')
            return redirect(_reports_redirect_url(
                request,
                request.POST.get('return_from_date', from_date),
                request.POST.get('return_to_date', to_date),
                request.POST.get('return_company', company_filter),
                request.POST.get('return_sale_particular', sale_particular),
            ))

        return_from_date = request.POST.get('return_from_date', from_date)
        return_to_date = request.POST.get('return_to_date', to_date)
        return_company = request.POST.get('return_company', company_filter)
        return_sale_particular = request.POST.get('return_sale_particular', sale_particular)

        settlement_company_id = request.POST.get('settlement_company', '').strip()
        payment_date_raw = request.POST.get('payment_date', '').strip()
        amount_raw = request.POST.get('amount_received', '').strip()
        reference_no = request.POST.get('reference_no', '').strip()
        notes = request.POST.get('notes', '').strip()

        settlement_company = allowed_companies.filter(pk=settlement_company_id).first() if settlement_company_id else None

        if not settlement_company:
            messages.error(request, 'Please select a valid company.')
            return redirect(_reports_redirect_url(request, return_from_date, return_to_date, return_company, return_sale_particular))

        try:
            payment_date = datetime.strptime(payment_date_raw, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'Please enter a valid payment date.')
            return redirect(_reports_redirect_url(request, return_from_date, return_to_date, return_company, return_sale_particular))

        try:
            amount_received = Decimal(amount_raw or '0')
        except (InvalidOperation, TypeError, ValueError):
            messages.error(request, 'Please enter a valid payment amount.')
            return redirect(_reports_redirect_url(request, return_from_date, return_to_date, return_company, return_sale_particular))

        if amount_received <= 0:
            messages.error(request, 'Payment amount must be greater than zero.')
            return redirect(_reports_redirect_url(request, return_from_date, return_to_date, return_company, return_sale_particular))

        current_outstanding = _company_outstanding_as_of(settlement_company, payment_date)
        if current_outstanding <= 0:
            messages.error(request, f'No outstanding amount exists for {settlement_company.name} as of {payment_date}.')
            return redirect(_reports_redirect_url(request, return_from_date, return_to_date, return_company, return_sale_particular))

        if amount_received > current_outstanding:
            messages.error(
                request,
                f'Entered amount exceeds outstanding balance. Outstanding for {settlement_company.name} is â‚¹{current_outstanding}.'
            )
            return redirect(_reports_redirect_url(request, return_from_date, return_to_date, return_company, return_sale_particular))

        CompanySettlement.objects.create(
            company=settlement_company,
            payment_date=payment_date,
            amount_received=amount_received,
            reference_no=reference_no,
            notes=notes,
            created_by=request.user,
        )
        messages.success(
            request,
            f'Payment of â‚¹{amount_received} recorded for {settlement_company.name}. Outstanding updated successfully.'
        )
        return redirect(_reports_redirect_url(request, return_from_date, return_to_date, return_company, return_sale_particular))

    order_qs = Order.objects.filter(
        is_deleted=False,
        created_at__date__gte=fd,
        created_at__date__lte=td,
        **cf
    )
    if company_filter and user.is_superadmin:
        order_qs = order_qs.filter(company_id=company_filter)
    wallet_sale_q = (
        Q(payment_mode=PaymentModeChoices.WALLET)
        | Q(wallet_used__gt=0)
        | Q(points_redeemed__gt=0)
    )
    if sale_particular == 'web':
        order_qs = order_qs.exclude(order_type__in=[ORDER_TYPE_KIOSK, ORDER_TYPE_WALLET_RECHARGE])
    elif sale_particular == 'kiosk':
        order_qs = order_qs.filter(order_type=ORDER_TYPE_KIOSK)
    elif sale_particular == 'pos':
        order_qs = order_qs.none()
    elif sale_particular == 'company':
        order_qs = order_qs.filter(bill_to_company__gt=0)
    elif sale_particular == 'monthly':
        order_qs = order_qs.filter(payment_mode=PaymentModeChoices.MONTHLY)
    elif sale_particular == 'wallet':
        order_qs = order_qs.exclude(order_type=ORDER_TYPE_WALLET_RECHARGE).filter(wallet_sale_q)
    elif sale_particular == 'wallet_recharge':
        order_qs = order_qs.filter(order_type=ORDER_TYPE_WALLET_RECHARGE)

    completed_web_qs = order_qs.filter(order_status__in=[2, 3, 4, 5])
    completed_recharge_qs = completed_web_qs.filter(order_type=ORDER_TYPE_WALLET_RECHARGE)
    completed_sales_qs = completed_web_qs.exclude(order_type=ORDER_TYPE_WALLET_RECHARGE)

    def _customer_paid_total(qs):
        row = qs.aggregate(my=Sum('my_pay'), wallet=Sum('wallet_used'), points=Sum('points_redeemed'))
        return (row['my'] or Decimal('0.00')) + (row['wallet'] or Decimal('0.00')) + Decimal(row['points'] or 0)

    def _wallet_value_total(qs):
        wallet_related = qs.filter(wallet_sale_q)
        customer_total = _customer_paid_total(wallet_related)
        non_wallet_cash = wallet_related.aggregate(t=Sum('my_pay'))['t'] or Decimal('0.00')
        return max(Decimal('0.00'), customer_total - non_wallet_cash)

    total_orders = order_qs.count()
    customer_web_cash_revenue = completed_sales_qs.aggregate(t=Sum('my_pay'))['t'] or Decimal('0.00')
    wallet_order_revenue = _wallet_value_total(completed_sales_qs)
    customer_web_revenue = customer_web_cash_revenue + wallet_order_revenue
    company_provisional = completed_sales_qs.aggregate(t=Sum('bill_to_company'))['t'] or Decimal('0.00')

    wallet_recharge_qs = WalletTransaction.objects.filter(
        txn_type=WalletTransaction.TYPE_TOPUP,
        created_at__date__gte=fd,
        created_at__date__lte=td,
    ).select_related('customer', 'customer__company')
    if not user.is_superadmin:
        wallet_recharge_qs = wallet_recharge_qs.filter(customer__company__in=allowed_companies)
    elif company_filter:
        wallet_recharge_qs = wallet_recharge_qs.filter(customer__company_id=company_filter)
    wallet_recharge_qs = wallet_recharge_qs.filter(
        order_ref__in=completed_recharge_qs.values_list('order_number', flat=True)
    )
    if sale_particular and sale_particular != 'wallet_recharge':
        wallet_recharge_qs = wallet_recharge_qs.none()
    wallet_recharge_revenue = completed_recharge_qs.aggregate(t=Sum('my_pay'))['t'] or Decimal('0.00')

    settlement_qs = CompanySettlement.objects.filter(
        is_deleted=False,
        payment_date__gte=fd,
        payment_date__lte=td,
        **cf
    )
    if company_filter and user.is_superadmin:
        settlement_qs = settlement_qs.filter(company_id=company_filter)

    company_received = settlement_qs.aggregate(t=Sum('amount_received'))['t'] or Decimal('0.00')

    # Snapshot outstanding by company as of selected end date
    snapshot_order_qs = Order.objects.filter(
        is_deleted=False,
        order_status__in=[2, 3, 4, 5],
        created_at__date__lte=td,
        **cf
    ).exclude(order_type=ORDER_TYPE_WALLET_RECHARGE)
    if company_filter and user.is_superadmin:
        snapshot_order_qs = snapshot_order_qs.filter(company_id=company_filter)

    snapshot_settlement_qs = CompanySettlement.objects.filter(
        is_deleted=False,
        payment_date__lte=td,
        **cf
    )
    if company_filter and user.is_superadmin:
        snapshot_settlement_qs = snapshot_settlement_qs.filter(company_id=company_filter)

    company_map = {}

    for row in snapshot_order_qs.values('company_id', 'company__name').annotate(
        cnt=Count('id'),
        customer_paid=Sum('my_pay'),
        provisional=Sum('bill_to_company'),
    ):
        company_map[row['company_id']] = {
            'company_id': row['company_id'],
            'company_name': row['company__name'] or '-',
            'cnt': row['cnt'] or 0,
            'customer_paid': row['customer_paid'] or Decimal('0.00'),
            'provisional': row['provisional'] or Decimal('0.00'),
            'received': Decimal('0.00'),
            'outstanding': Decimal('0.00'),
        }

    for row in snapshot_settlement_qs.values('company_id', 'company__name').annotate(
        received=Sum('amount_received')
    ):
        entry = company_map.setdefault(row['company_id'], {
            'company_id': row['company_id'],
            'company_name': row['company__name'] or '-',
            'cnt': 0,
            'customer_paid': Decimal('0.00'),
            'provisional': Decimal('0.00'),
            'received': Decimal('0.00'),
            'outstanding': Decimal('0.00'),
        })
        entry['received'] = row['received'] or Decimal('0.00')

    all_company_rows = []
    for entry in company_map.values():
        outstanding = entry['provisional'] - entry['received']
        entry['outstanding'] = outstanding if outstanding > 0 else Decimal('0.00')
        all_company_rows.append(entry)

    all_company_rows = sorted(
        all_company_rows,
        key=lambda x: (x['outstanding'], x['provisional'], x['company_name']),
        reverse=True
    )

    outstanding_total = sum((row['outstanding'] for row in all_company_rows), Decimal('0.00'))
    by_company = all_company_rows[:50]

    # Top web products
    top_products = OrderItem.objects.filter(
        order__in=completed_sales_qs
    ).annotate(
        line_total=ExpressionWrapper(_F('price') * _F('qty'), output_field=FloatField())
    ).values('product__name').annotate(
        qty=Sum('qty'),
        rev=Sum('line_total')
    ).order_by('-qty')[:20]

    # POS
    total_pos_orders = 0
    pos_revenue = 0
    pos_cash_orders = 0
    pos_upi_orders = 0
    pos_card_orders = 0
    pos_cash_amount = Decimal('0.00')
    pos_upi_amount = Decimal('0.00')
    pos_card_amount = Decimal('0.00')
    top_pos_products = []
    pos_qs = None
    try:
        from apps.pos.models import POSOrder, POSOrderItem
        pos_cf = {'company__in': allowed_companies} if not user.is_superadmin else {}
        pos_qs = POSOrder.objects.filter(
            is_deleted=False,
            **pos_cf
        ).exclude(created_at__isnull=True).filter(
            created_at__date__gte=fd,
            created_at__date__lte=td
        )
        if company_filter and (user.is_superadmin or allowed_companies.filter(pk=company_filter).exists()):
            pos_qs = pos_qs.filter(company_id=company_filter)
        if sale_particular and sale_particular != 'pos':
            pos_qs = pos_qs.none()

        total_pos_orders = pos_qs.count()
        pos_revenue = pos_qs.aggregate(t=Sum('total_amount'))['t'] or Decimal('0.00')
        pos_cash_orders = pos_qs.filter(payment_type=POSOrder.PAYMENT_CASH).count()
        pos_upi_orders = pos_qs.filter(payment_type=POSOrder.PAYMENT_UPI).count()
        pos_card_orders = pos_qs.filter(payment_type=POSOrder.PAYMENT_CARD).count()
        pos_cash_amount = pos_qs.filter(payment_type=POSOrder.PAYMENT_CASH).aggregate(t=Sum('total_amount'))['t'] or Decimal('0.00')
        pos_upi_amount = pos_qs.filter(payment_type=POSOrder.PAYMENT_UPI).aggregate(t=Sum('total_amount'))['t'] or Decimal('0.00')
        pos_card_amount = pos_qs.filter(payment_type=POSOrder.PAYMENT_CARD).aggregate(t=Sum('total_amount'))['t'] or Decimal('0.00')
        top_pos_products = POSOrderItem.objects.filter(order__in=pos_qs).values(
            'product_name'
        ).annotate(
            qty=Sum('qty'),
            rev=Sum('amount')
        ).order_by('-qty')[:20]
    except (ImportError, AttributeError, LookupError):
        # POS tables may be absent in some deployments
        pass

    platform_order_qs = order_qs.exclude(order_type=ORDER_TYPE_WALLET_RECHARGE)
    platform_web_orders = platform_order_qs.exclude(order_type=ORDER_TYPE_KIOSK).count()
    platform_kiosk_orders = platform_order_qs.filter(order_type=ORDER_TYPE_KIOSK).count()
    platform_pos_orders = total_pos_orders
    report_total_orders = platform_web_orders + platform_kiosk_orders + platform_pos_orders
    report_cash_orders = platform_order_qs.filter(payment_mode=PaymentModeChoices.CASH).count() + pos_cash_orders
    report_upi_orders = pos_upi_orders
    report_card_orders = pos_card_orders
    platform_web_amount = platform_order_qs.exclude(order_type=ORDER_TYPE_KIOSK).aggregate(t=Sum('total_amount'))['t'] or Decimal('0.00')
    platform_kiosk_amount = platform_order_qs.filter(order_type=ORDER_TYPE_KIOSK).aggregate(t=Sum('total_amount'))['t'] or Decimal('0.00')
    platform_pos_amount = pos_revenue
    report_total_amount = platform_web_amount + platform_kiosk_amount + platform_pos_amount
    report_cash_amount = (platform_order_qs.filter(payment_mode=PaymentModeChoices.CASH).aggregate(t=Sum('total_amount'))['t'] or Decimal('0.00')) + pos_cash_amount
    report_upi_amount = pos_upi_amount
    report_card_amount = pos_card_amount

    # Product-wise order summary (uses current report filters)
    order_summary_map = {}

    def _merge_order_summary(name, qty, order_count, rev):
        product_name = (name or 'Deleted / Unknown Product').strip() or 'Deleted / Unknown Product'
        entry = order_summary_map.setdefault(product_name, {
            'product_name': product_name,
            'qty': 0,
            'order_count': 0,
            'rev': Decimal('0.00'),
        })
        entry['qty'] += int(qty or 0)
        entry['order_count'] += int(order_count or 0)
        entry['rev'] += Decimal(str(rev or 0))

    web_summary_rows = OrderItem.objects.filter(
        is_deleted=False,
        order__in=completed_sales_qs
    ).annotate(
        line_total=ExpressionWrapper(
            _F('price') * _F('qty'),
            output_field=DecimalField(max_digits=14, decimal_places=2)
        )
    ).values('product__name').annotate(
        qty=Sum('qty'),
        order_count=Count('order_id', distinct=True),
        rev=Sum('line_total')
    ).order_by('-qty', 'product__name')

    for row in web_summary_rows:
        _merge_order_summary(
            row.get('product__name'),
            row.get('qty'),
            row.get('order_count'),
            row.get('rev'),
        )

    if pos_qs is not None:
        try:
            pos_summary_rows = POSOrderItem.objects.filter(
                order__in=pos_qs
            ).values('product_name').annotate(
                qty=Sum('qty'),
                order_count=Count('order_id', distinct=True),
                rev=Sum('amount')
            ).order_by('-qty', 'product_name')
            for row in pos_summary_rows:
                _merge_order_summary(
                    row.get('product_name'),
                    row.get('qty'),
                    row.get('order_count'),
                    row.get('rev'),
                )
        except Exception:
            pass

    order_summary_rows = sorted(
        order_summary_map.values(),
        key=lambda x: (-x['qty'], x['product_name'].lower())
    )
    order_summary_total_qty = sum((row['qty'] for row in order_summary_rows), 0)
    order_summary_total_orders = (
        completed_sales_qs.count() +
        (pos_qs.count() if pos_qs is not None else 0)
    )
    order_summary_total_rev = sum((row['rev'] for row in order_summary_rows), Decimal('0.00'))

    # Daily chart data
    daily_chart = []
    cur = fd
    while cur <= td:
        day_sales_qs = completed_sales_qs.filter(created_at__date=cur)
        web_day_qs = day_sales_qs.exclude(order_type=ORDER_TYPE_KIOSK)
        kiosk_day_qs = day_sales_qs.filter(order_type=ORDER_TYPE_KIOSK)
        customer_paid_day = _customer_paid_total(day_sales_qs)
        wallet_order_day = _wallet_value_total(day_sales_qs)
        web_sale_day = _customer_paid_total(web_day_qs)
        kiosk_sale_day = _customer_paid_total(kiosk_day_qs)
        company_pay_day = day_sales_qs.aggregate(t=Sum('bill_to_company'))['t'] or Decimal('0.00')
        monthly_day = day_sales_qs.filter(payment_mode=PaymentModeChoices.MONTHLY).aggregate(t=Sum('my_pay'))['t'] or Decimal('0.00')
        wallet_recharge_day = completed_recharge_qs.filter(created_at__date=cur).aggregate(t=Sum('my_pay'))['t'] or Decimal('0.00')

        company_received_day = settlement_qs.filter(
            payment_date=cur
        ).aggregate(t=Sum('amount_received'))['t'] or Decimal('0.00')

        try:
            pos_rev = Decimal('0.00')
            if pos_qs is not None:
                pos_rev = pos_qs.filter(
                    created_at__date=cur
                ).aggregate(t=Sum('total_amount'))['t'] or Decimal('0.00')
        except (ValueError, OverflowError, OSError):
            pos_rev = Decimal('0.00')

        daily_chart.append({
            'date': cur.strftime('%d %b'),
            'customer_paid': float(customer_paid_day),
            'company_received': float(company_received_day),
            'pos': float(pos_rev),
            'wallet_recharge': float(wallet_recharge_day),
            'web_sale': float(web_sale_day),
            'kiosk_sale': float(kiosk_sale_day),
            'company_pay': float(company_pay_day),
            'monthly': float(monthly_day),
            'wallet_order': float(wallet_order_day),
            'total_sales': float(web_sale_day + kiosk_sale_day + company_pay_day + pos_rev + wallet_recharge_day),
        })
        cur += timedelta(days=1)

    # Order Summary PDF export
    if request.GET.get('export') == 'order_summary_pdf':
        from io import BytesIO
        from django.http import HttpResponse
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
        styles = getSampleStyleSheet()

        sale_label_map = dict([
            ('', 'All Sales'),
            ('web', 'Web Sale'),
            ('kiosk', 'Kiosk Sale'),
            ('pos', 'POS Sale'),
            ('company', 'Company-paid'),
            ('monthly', 'Monthly Billing'),
            ('wallet', 'Wallet Orders'),
            ('wallet_recharge', 'Wallet Recharge'),
        ])
        sale_label = sale_label_map.get(sale_particular, 'All Sales')
        company_label = 'All Assigned Sites' if not user.is_superadmin else 'All Companies'
        if company_filter:
            selected = allowed_companies.filter(pk=company_filter).first()
            if selected:
                company_label = selected.name

        elements = [
            Paragraph("Order Summary", styles['Title']),
            Paragraph(f"Period: {from_date} to {to_date}", styles['Normal']),
            Paragraph(f"Company: {company_label}", styles['Normal']),
            Paragraph(f"Sale Particular: {sale_label}", styles['Normal']),
            Spacer(1, 10),
        ]

        table_data = [[
            'Product Name',
            'Qty Sold',
            'Orders',
            'Revenue (Rs.)',
        ]]

        for row in order_summary_rows[:5000]:
            table_data.append([
                row['product_name'],
                str(row['qty']),
                str(row['order_count']),
                f"{row['rev']:.2f}",
            ])

        table_data.append([
            'TOTAL',
            str(order_summary_total_qty),
            str(order_summary_total_orders),
            f"{order_summary_total_rev:.2f}",
        ])

        table = Table(table_data, repeatRows=1, colWidths=[320, 90, 90, 110])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EAF2FF')),
            ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#C7D3E3')),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F8FAFC')]),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)
        doc.build(elements)

        pdf = buffer.getvalue()
        buffer.close()

        resp = HttpResponse(pdf, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="order_summary_{from_date}_{to_date}.pdf"'
        return resp

    # Order Summary Excel export
    if request.GET.get('export') == 'order_summary_xlsx':
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Order Summary'

        sale_label_map = dict([
            ('', 'All Sales'),
            ('web', 'Web Sale'),
            ('kiosk', 'Kiosk Sale'),
            ('pos', 'POS Sale'),
            ('company', 'Company-paid'),
            ('monthly', 'Monthly Billing'),
            ('wallet', 'Wallet Orders'),
            ('wallet_recharge', 'Wallet Recharge'),
        ])
        sale_label = sale_label_map.get(sale_particular, 'All Sales')
        company_label = 'All Assigned Sites' if not user.is_superadmin else 'All Companies'
        if company_filter:
            selected = allowed_companies.filter(pk=company_filter).first()
            if selected:
                company_label = selected.name

        ws.append(['Order Summary'])
        ws.append([f'Period: {from_date} to {to_date}'])
        ws.append([f'Company: {company_label}'])
        ws.append([f'Sale Particular: {sale_label}'])
        ws.append([])
        ws.append(['Product Name', 'Qty Sold', 'Orders', 'Revenue'])

        header_fill = PatternFill(fill_type='solid', fgColor='1E3A5F')
        header_font = Font(color='FFFFFF', bold=True)
        bold_font = Font(bold=True)

        for cell in ws[6]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row in order_summary_rows[:5000]:
            ws.append([
                row['product_name'],
                row['qty'],
                row['order_count'],
                float(row['rev']),
            ])

        ws.append(['TOTAL', order_summary_total_qty, order_summary_total_orders, float(order_summary_total_rev)])
        for cell in ws[ws.max_row]:
            cell.font = bold_font

        ws.column_dimensions['A'].width = 42
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 16

        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="order_summary_{from_date}_{to_date}.xlsx"'
        wb.save(resp)
        return resp

    report_export_headers = [
        'Order#',
        'Date',
        'Customer',
        'Phone',
        'Company',
        'Gross Subtotal',
        'Offer Discount',
        'Coupon Discount',
        'Net Total',
        'Customer Paid',
        'Wallet Used',
        'Points Redeemed',
        'Company Covered',
        'Cash',
        'UPI',
        'Card',
        'Other Payment Mode',
        'Sale Particular',
        'Status',
    ]

    def _payment_bucket_amounts(payment_label, amount):
        mode = (payment_label or '').strip().lower()
        paid_amount = amount or Decimal('0.00')
        if 'cash' in mode:
            return [paid_amount, Decimal('0.00'), Decimal('0.00'), Decimal('0.00')]
        if 'upi' in mode:
            return [Decimal('0.00'), paid_amount, Decimal('0.00'), Decimal('0.00')]
        if 'card' in mode:
            return [Decimal('0.00'), Decimal('0.00'), paid_amount, Decimal('0.00')]
        return [Decimal('0.00'), Decimal('0.00'), Decimal('0.00'), paid_amount]

    def _report_export_rows():
        for o in order_qs.select_related('customer', 'company').order_by('-created_at')[:5000]:
            recharge_txn = o.recharge_transaction if getattr(o, 'is_wallet_recharge', False) else None
            payment_label = recharge_txn.get_payment_mode_display() if recharge_txn else o.get_payment_mode_display()
            paid_amount = o.my_pay or Decimal('0.00')
            is_wallet_related = (
                o.payment_mode == PaymentModeChoices.WALLET
                or (o.wallet_used or 0) > 0
                or (o.points_redeemed or 0) > 0
            )
            sale_label = (
                'Wallet Recharge' if o.order_type == ORDER_TYPE_WALLET_RECHARGE else
                'Kiosk Sale' if o.order_type == ORDER_TYPE_KIOSK else
                'Monthly Billing' if o.payment_mode == PaymentModeChoices.MONTHLY else
                'Wallet' if is_wallet_related else
                'Web Sale'
            )
            yield [
                o.order_number,
                o.created_at.strftime('%d-%m-%Y') if o.created_at else '',
                o.display_customer_name,
                o.display_customer_phone,
                o.company.name if o.company else '',
                o.subtotal or 0,
                o.offer_discount or 0,
                o.coupon_discount or 0,
                o.total_amount or 0,
                o.my_pay or 0,
                o.wallet_used or 0,
                o.points_redeemed or 0,
                o.bill_to_company or 0,
                *_payment_bucket_amounts(payment_label, paid_amount),
                sale_label,
                o.status_label,
            ]
        for p in (pos_qs.order_by('-created_at')[:5000] if pos_qs is not None else []):
            payment_label = p.get_payment_type_display()
            paid_amount = p.total_amount or Decimal('0.00')
            yield [
                p.order_number,
                p.created_at.strftime('%d-%m-%Y') if p.created_at else '',
                p.customer_name,
                p.customer_phone,
                p.company.name if p.company else '',
                p.base_amount or p.total_amount or 0,
                0,
                0,
                p.total_amount or 0,
                p.total_amount or 0,
                0,
                0,
                0,
                *_payment_bucket_amounts(payment_label, paid_amount),
                'POS Sale',
                'Completed',
            ]

    # Detailed Excel export
    if request.GET.get('export') == 'xlsx':
        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Report'
        ws.append(report_export_headers)

        header_fill = PatternFill(fill_type='solid', fgColor='1E3A5F')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for row in _report_export_rows():
            ws.append(row)

        for column_cells in ws.columns:
            header = column_cells[0]
            ws.column_dimensions[header.column_letter].width = min(max(len(str(header.value or '')) + 4, 14), 24)

        resp = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = f'attachment; filename="report_{from_date}_{to_date}.xlsx"'
        wb.save(resp)
        return resp

    # CSV export
    if request.GET.get('export') == 'csv':
        import csv
        from django.http import HttpResponse

        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = f'attachment; filename="report_{from_date}_{to_date}.csv"'
        w = csv.writer(resp)
        w.writerow(report_export_headers)
        for row in _report_export_rows():
            w.writerow(row)
        return resp

    orders = order_qs.select_related('customer', 'company').prefetch_related(
        'items'
    ).order_by('-created_at')[:500]

    recent_settlements = settlement_qs.select_related(
        'company', 'created_by'
    ).order_by('-payment_date', '-created_at')[:20]
    recent_recharges = wallet_recharge_qs.order_by('-created_at')[:50]

    return render(request, 'dashboard/reports.html', {
        'from_date': from_date,
        'to_date': to_date,
        'company_filter': company_filter,
        'sale_particular': sale_particular,
        'sale_particular_choices': [
            ('', 'All Sales'),
            ('web', 'Web Sale'),
            ('kiosk', 'Kiosk Sale'),
            ('pos', 'POS Sale'),
            ('company', 'Company-paid'),
            ('monthly', 'Monthly Billing'),
            ('wallet', 'Wallet Orders'),
            ('wallet_recharge', 'Wallet Recharge'),
        ],
        'companies': allowed_companies,
        'total_orders': total_orders,
        'report_total_orders': report_total_orders,
        'report_cash_orders': report_cash_orders,
        'report_upi_orders': report_upi_orders,
        'report_card_orders': report_card_orders,
        'platform_web_orders': platform_web_orders,
        'platform_pos_orders': platform_pos_orders,
        'platform_kiosk_orders': platform_kiosk_orders,
        'report_total_amount': report_total_amount,
        'report_cash_amount': report_cash_amount,
        'report_upi_amount': report_upi_amount,
        'report_card_amount': report_card_amount,
        'platform_web_amount': platform_web_amount,
        'platform_pos_amount': platform_pos_amount,
        'platform_kiosk_amount': platform_kiosk_amount,
        'customer_web_revenue': customer_web_revenue,
        'customer_web_cash_revenue': customer_web_cash_revenue,
        'wallet_order_revenue': wallet_order_revenue,
        'wallet_recharge_revenue': wallet_recharge_revenue,
        'company_provisional': company_provisional,
        'company_received': company_received,
        'outstanding_total': outstanding_total,
        'total_pos_orders': total_pos_orders,
        'pos_revenue': pos_revenue,
        'by_company': by_company,
        'top_products': top_products,
        'top_pos_products': top_pos_products,
        'order_summary_rows': order_summary_rows[:500],
        'order_summary_total_qty': order_summary_total_qty,
        'order_summary_total_orders': order_summary_total_orders,
        'order_summary_total_rev': order_summary_total_rev,
        'daily_chart': daily_chart,
        'orders': orders,
        'recent_settlements': recent_settlements,
        'recent_recharges': recent_recharges,
        'default_payment_date': today.isoformat(),
        'page_title': 'Reports',
    })


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  FLASH ORDER
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin','admin','pos')
def flash_order(request):
    from datetime import date
    from apps.menu.models import Product
    user     = request.user
    cf       = _scope(user)
    companies= _companies(user)

    products = Product.objects.filter(is_active=True, is_deleted=False, **cf).order_by('name')[:200]

    today = date.today().isoformat()

    recent_orders = Order.objects.filter(
        is_deleted=False, **cf,
        created_at__date=date.today()
    ).select_related('customer','company').prefetch_related(
        'items__product'
    ).order_by('-created_at')[:20]

    if request.method == 'POST':
        company_id    = request.POST.get('company')
        customer_name = request.POST.get('customer_name','').strip()
        customer_phone= request.POST.get('customer_phone','').strip()
        payment_mode  = (request.POST.get('payment_mode','online') or 'online').strip()
        delivery_date = request.POST.get('delivery_date', today)
        product_ids   = request.POST.getlist('product_id[]')
        qtys          = request.POST.getlist('qty[]')

        company = Company.objects.filter(pk=company_id, is_active=True, is_deleted=False).first()
        if not company:
            messages.error(request, 'Select a company.')
            return redirect('dashboard:flash_order')
        # Non-superadmin cannot place flash orders for other companies
        if not user.is_superadmin and not user_can_access_company(user, company.pk):
            messages.error(request, 'You can only place orders for assigned sites.')
            return redirect('dashboard:flash_order')

        if not customer_name:
            messages.error(request, 'Customer name is required.')
            return redirect('dashboard:flash_order')

        from apps.orders.models import PaymentModeChoices as _PMC
        _allowed_flash_modes = {_PMC.ONLINE, _PMC.CASH, _PMC.MONTHLY, _PMC.COMPANY}
        if payment_mode not in _allowed_flash_modes:
            messages.error(request, 'Invalid payment mode selected.')
            return redirect('dashboard:flash_order')

        # Find or create a walk-in customer for this company
        # Use phone as secondary lookup to avoid creating duplicates
        from apps.accounts.models import Customer
        import uuid as _uuid
        customer = Customer.objects.filter(
            company=company, name__iexact=customer_name
        ).first()
        if not customer and customer_phone:
            customer = Customer.objects.filter(
                company=company, phone=customer_phone, is_active=True
            ).first()
        if not customer:
            # Generate a unique placeholder email so the DB unique constraint
            # is never violated by multiple flash orders for different walk-in names
            placeholder_email = f'flash.{_uuid.uuid4().hex[:8]}@neverq.local'
            customer = Customer.objects.create(
                company=company, name=customer_name,
                phone=customer_phone, email=placeholder_email,
                password_hash='!', is_active=True,
            )

        subtotal    = 0
        cart_items  = []
        for pid, qty_str in zip(product_ids, qtys):
            try:
                # Scope product to the selected company â€” prevent cross-tenant product use
                product = Product.objects.get(pk=pid, is_active=True, is_deleted=False, company=company)
                qty     = max(1, int(qty_str or 1))
                line    = product.price * qty
                subtotal+= line
                cart_items.append({'product': product, 'qty': qty, 'price': product.price})
            except (Product.DoesNotExist, ValueError):
                pass

        if not cart_items:
            messages.error(request, 'Add at least one product.')
            return redirect('dashboard:flash_order')

        try:
            from django.utils import timezone as tz
            from datetime import datetime
            sched = tz.make_aware(datetime.strptime(delivery_date, '%Y-%m-%d'))
        except (ValueError, OverflowError, OSError):
            sched = None

        _flash_payment_status = (
            'approved' if payment_mode == _PMC.MONTHLY else
            'paid'     if payment_mode in (_PMC.CASH, _PMC.COMPANY) else
            'pending'
        )
        order = Order.objects.create(
            company=company, customer=customer,
            subtotal=subtotal, total_amount=subtotal,
            payment_mode=payment_mode,
            payment_status=_flash_payment_status,
            order_status=OrderStatusChoices.CONFIRMED,
            scheduled_date=sched,
        )
        for item in cart_items:
            from apps.orders.models import OrderItem
            OrderItem.objects.create(
                company=company, order=order,
                product=item['product'],
                price=item['price'], qty=item['qty'],
            )
        OrderStatus.objects.create(
            order=order, status=OrderStatusChoices.CONFIRMED,
            details='Flash order placed by staff.',
            created_at=timezone.now(),
        )
        messages.success(request, f'Flash Order #{order.order_number} placed! â‚¹{subtotal}')
        return redirect('dashboard:flash_order')

    return render(request, 'dashboard/flash_order.html', {
        'companies':     companies,
        'products':      products,
        'recent_orders': recent_orders,
        'today':         today,
        'page_title':    'Flash Order',
    })


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  CUSTOMER SEARCH AJAX  (for Flash Order autocomplete)
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin','admin','pos','cafeman')
def customer_search_ajax(request):
    q          = request.GET.get('q', '').strip()
    company_id = request.GET.get('company', '')
    user       = request.user
    qs         = Customer.objects.filter(is_active=True, is_deleted=False)
    if user.is_superadmin:
        # Superadmin: filter by provided company_id or return all
        if company_id:
            qs = qs.filter(company_id=company_id)
    elif _companies(user).exists():
        # Site-scoped staff: restrict to assigned sites and optional assigned company filter.
        qs = qs.filter(company__in=_companies(user))
        if company_id and _companies(user).filter(pk=company_id).exists():
            qs = qs.filter(company_id=company_id)
    else:
        # No company assigned and not superadmin: return nothing
        qs = qs.none()
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(phone__icontains=q))
    data = [{'id': c.pk, 'name': c.name, 'phone': c.phone} for c in qs[:15]]
    return JsonResponse({'customers': data})


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  COUPONS
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin','admin')
def coupon_list(request):
    from apps.core.models import Coupon
    user = request.user
    qs = Coupon.objects.all().select_related('company').order_by('-created_at')
    if not user.is_superadmin:
        qs = qs.filter(Q(company__in=_companies(user)) | Q(company__isnull=True))
    return render(request, 'dashboard/coupons/list.html', {
        'coupons': qs, 'page_title': 'Coupons',
        'lp': get_list_perms(user, 'perm_coupons'),
    })


@staff_role_required('superadmin','admin')
def coupon_add(request):
    from apps.core.models import Coupon
    user = request.user
    if not user_can_action(user, 'perm_coupons', 'add'):
        return _deny_dashboard_action(request)
    if request.method == 'POST':
        code = request.POST.get('code', '').strip().upper()
        if not code:
            messages.error(request, 'Coupon code is required.')
        elif Coupon.objects.filter(code=code).exists():
            messages.error(request, 'Coupon code already exists.')
        else:
            from decimal import Decimal
            company_id = request.POST.get('company')
            company = None
            if company_id:
                company = _companies(user).filter(pk=company_id).first()
            if not user.is_superadmin:
                company = company or get_primary_staff_company(user)

            Coupon.objects.create(
                company=company,
                code=code,
                description=request.POST.get('description', ''),
                discount_type=request.POST.get('discount_type', 'flat'),
                discount_value=Decimal(request.POST.get('discount_value', '0') or '0'),
                min_order=Decimal(request.POST.get('min_order', '0') or '0'),
                max_discount=Decimal(request.POST.get('max_discount', '0') or '0'),
                usage_limit=int(request.POST.get('usage_limit', '0') or '0'),
                valid_from=request.POST.get('valid_from') or None,
                valid_to=request.POST.get('valid_to') or None,
                is_active=request.POST.get('is_active') == 'on',
            )
            messages.success(request, f'Coupon "{code}" created.')
            return redirect('dashboard:coupon_list')
    return render(request, 'dashboard/coupons/form.html', {
        'companies': _companies(user),
        'page_title': 'Add Coupon', 'action': 'Add',
    })


@staff_role_required('superadmin','admin')
def coupon_edit(request, pk):
    from apps.core.models import Coupon
    from decimal import Decimal
    user = request.user
    _perm = check_module_permission(request, 'perm_coupons')
    if _perm: return _perm
    coupon = get_object_or_404(Coupon, pk=pk)
    if not user.is_superadmin and coupon.company and not user_can_access_company(user, coupon.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:coupon_list')
    if request.method == 'POST':
        if user.role != 'superadmin' and get_module_level(request.user, 'perm_coupons') == 'full_edit':
            _diffs = {
                'code':          {'label': 'Code',          'before': coupon.code or '',         'after': (request.POST.get('code') or coupon.code).strip().upper()},
                'description':   {'label': 'Description',   'before': coupon.description or '',  'after': request.POST.get('description', '')},
                'discount_type': {'label': 'Discount Type', 'before': coupon.discount_type,      'after': request.POST.get('discount_type', 'flat')},
                'is_active':     {'label': 'Active',        'before': coupon.is_active,          'after': request.POST.get('is_active') == 'on'},
            }
            _pc = create_pending_change(request, 'perm_coupons', coupon, _diffs)
            if _pc:
                messages.success(request, 'Your changes have been submitted for superadmin review.')
            else:
                messages.info(request, 'No changes detected.')
            return redirect('dashboard:coupon_list')
        coupon.code = request.POST.get('code', coupon.code).strip().upper()
        coupon.description = request.POST.get('description', '')
        coupon.discount_type = request.POST.get('discount_type', 'flat')
        coupon.discount_value = Decimal(request.POST.get('discount_value', '0') or '0')
        coupon.min_order = Decimal(request.POST.get('min_order', '0') or '0')
        coupon.max_discount = Decimal(request.POST.get('max_discount', '0') or '0')
        coupon.usage_limit = int(request.POST.get('usage_limit', '0') or '0')
        coupon.valid_from = request.POST.get('valid_from') or None
        coupon.valid_to = request.POST.get('valid_to') or None
        coupon.is_active = request.POST.get('is_active') == 'on'
        coupon.save()
        messages.success(request, 'Coupon updated.')
        return redirect('dashboard:coupon_list')
    _lj, _pl = get_locked_html_names(request.user, 'perm_coupons')
    return render(request, 'dashboard/coupons/form.html', {
        'coupon': coupon,
        'companies': _companies(user),
        'page_title': f'Edit Coupon - {coupon.code}', 'action': 'Save',
        'perm_level': _pl or 'full_edit', 'locked_names_json': _lj,
    })


@require_POST
@staff_role_required('superadmin','admin')
def coupon_delete(request, pk):
    if not _has_full_module_edit(request.user, 'perm_coupons'):
        return _deny_dashboard_action(request)
    coupon = get_object_or_404(Coupon, pk=pk)
    if not request.user.is_superadmin and coupon.company and not user_can_access_company(request.user, coupon.company_id):
        messages.error(request, 'Access denied.')
    else:
        coupon.delete()
        messages.success(request, 'Coupon deleted.')
    return redirect('dashboard:coupon_list')


@require_POST
@staff_role_required('superadmin','admin')
def coupon_bulk_delete(request):
    if not _has_full_module_edit(request.user, 'perm_coupons'):
        return _deny_dashboard_action(request)
    ids = _bulk_post_ids(request)
    qs = Coupon.objects.filter(pk__in=ids)
    if not request.user.is_superadmin:
        qs = qs.filter(Q(company__in=_companies(request.user)) | Q(company__isnull=True))
    count = qs.count()
    qs.delete()
    _bulk_message(request, 'coupon', count)
    return redirect('dashboard:coupon_list')


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  STATIC PAGES MANAGEMENT
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin')
def static_page_list(request):
    _perm = check_module_permission(request, 'perm_static_pages')
    if _perm: return _perm
    from apps.core.models import StaticPage
    pages = StaticPage.objects.all()
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add':
            slug = request.POST.get('slug', '').strip()
            title = request.POST.get('title', '').strip()
            content = request.POST.get('content', '')
            if slug and title:
                StaticPage.objects.update_or_create(
                    slug=slug, defaults={'title': title, 'content': content, 'is_active': True}
                )
                messages.success(request, f'Page "{title}" saved.')
            else:
                messages.error(request, 'Slug and title required.')
        elif action == 'delete':
            StaticPage.objects.filter(pk=request.POST.get('pk')).delete()
            messages.success(request, 'Page deleted.')
        return redirect('dashboard:static_page_list')
    return render(request, 'dashboard/static_pages/list.html', {
        'pages': pages, 'page_title': 'Static Pages',
        'slug_choices': [
            ('about-us', 'About Us'), ('terms-and-conditions', 'Terms & Conditions'),
            ('privacy-policy', 'Privacy Policy'), ('refund-policy', 'Refund Policy'),
            ('contact-us', 'Contact Us'),
        ],
    })


@staff_role_required('superadmin')
def static_page_edit(request, pk):
    from apps.core.models import StaticPage
    page = get_object_or_404(StaticPage, pk=pk)
    if request.method == 'POST':
        page.title = request.POST.get('title', page.title).strip()
        page.content = request.POST.get('content', '')
        page.is_active = request.POST.get('is_active') == 'on'
        page.save()
        messages.success(request, 'Page updated.')
        return redirect('dashboard:static_page_list')
    return render(request, 'dashboard/static_pages/edit.html', {
        'page_obj': page, 'page_title': f'Edit - {page.title}',
    })



# â”€â”€â”€ USER ACCESS CONTROL â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin')
def user_access(request):
    """List all staff users with their access status."""
    messages.info(request, 'Use Permission Matrix for staff access.')
    return redirect('dashboard:staff_list')

@staff_role_required('superadmin')
def user_access_edit(request, pk):
    """Configure access for a specific staff user."""
    messages.info(request, 'Use Permission Matrix for staff access.')
    return redirect('dashboard:permission_matrix', pk=pk)


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
#  PRODUCT QUANTITY MANAGEMENT (admin only)
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@require_POST
@staff_role_required('superadmin', 'admin', 'cafeman', 'pos')
def product_update_qty(request, pk):
    """Update web_qty and/or pos_qty for a product.
    When web_qty hits 0 the product is auto-disabled for web orders.
    """
    from apps.menu.models import Product
    user = request.user
    wants_web_qty = request.POST.get('web_qty') is not None
    wants_pos_qty = request.POST.get('pos_qty') is not None
    can_qty_update = user_can_action(user, 'perm_products', 'qty_update')
    can_web_qty_update = (
        can_qty_update
        or user_can_action(user, 'perm_products', 'field_web_qty')
        or (
            user.role == StaffUser.ROLE_POS
            and user_can_action(user, 'perm_products', 'cashier_edit')
            and not wants_pos_qty
        )
    )
    can_pos_qty_update = (
        can_qty_update
        or user_can_action(user, 'perm_products', 'field_pos_qty')
    )
    if (wants_web_qty and not can_web_qty_update) or (wants_pos_qty and not can_pos_qty_update):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    if not wants_web_qty and not wants_pos_qty:
        return JsonResponse({'success': False, 'error': 'No quantity value was provided.'}, status=400)
    try:
        product = Product.objects.get(pk=pk, is_deleted=False)
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product not found.'}, status=404)

    # Scope: non-superadmin can only edit their own company's products
    if not user.is_superadmin and not user_can_access_company(user, product.company_id):
        return JsonResponse({'success': False, 'error': 'Access denied.'}, status=403)

    update_fields = []
    web_qty_raw = request.POST.get('web_qty')
    pos_qty_raw = request.POST.get('pos_qty')

    if web_qty_raw is not None:
        try:
            web_qty = max(0, int(web_qty_raw))
            if product.web_qty != web_qty:
                product.web_qty = web_qty
                update_fields.append('web_qty')
            if web_qty == 0 and product.is_active:
                product.is_active = False
                update_fields.append('is_active')
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid web qty.'}, status=400)

    if pos_qty_raw is not None:
        try:
            pos_qty = max(0, int(pos_qty_raw))
            if product.pos_qty != pos_qty:
                product.pos_qty = pos_qty
                update_fields.append('pos_qty')
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid POS qty.'}, status=400)

    if update_fields:
        product.save(update_fields=update_fields)

    return JsonResponse({
        'success': True,
        'web_qty': product.web_qty,
        'pos_qty': product.pos_qty,
        'is_active': product.is_active,
        'auto_disabled': product.web_qty == 0,
    })


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  PRODUCT GALLERY
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@staff_role_required('superadmin', 'admin')
def product_gallery(request):
    from apps.menu.models import ProductGallery
    user = request.user
    qs = ProductGallery.objects.select_related('company', 'uploaded_by')
    if not user.is_superadmin:
        qs = qs.filter(Q(company__in=_companies(user)) | Q(company__isnull=True))
    qs = qs.order_by('-created_at')
    search = (request.GET.get('q') or '').strip()
    if search:
        qs = qs.filter(name__icontains=search)

    if request.method == 'POST':
        image_file = request.FILES.get('image')
        name = (request.POST.get('name') or '').strip() or Path(getattr(image_file, 'name', '') or 'Product Image').stem.replace('_', ' ').replace('-', ' ').strip() or 'Product Image'
        if not image_file:
            messages.error(request, 'Please select an image file.')
        else:
            company = None
            if not user.is_superadmin:
                company = _companies(user).filter(pk=request.POST.get('company')).first() or get_primary_staff_company(user)
            elif request.POST.get('company'):
                company = Company.objects.filter(pk=request.POST.get('company')).first()
            ProductGallery.objects.create(
                company=company, name=name,
                image=image_file, uploaded_by=user
            )
            messages.success(request, f'"{name}" added to product gallery.')
        return redirect('dashboard:product_gallery')

    companies = _companies(user)
    return render(request, 'dashboard/menu/product_gallery.html', {
        'gallery': qs,
        'search_query': search,
        'companies': companies,
        'can_upload': True,
        'page_title': 'Product Image Gallery',
    })


@require_POST
@staff_role_required('superadmin', 'admin')
def product_gallery_delete(request, pk):
    obj = get_object_or_404(ProductGallery, pk=pk)
    user = request.user
    if not user.is_superadmin and obj.company and not user_can_access_company(user, obj.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:product_gallery')
    obj.image.delete(save=False)
    obj.delete()
    messages.success(request, 'Image deleted.')
    return redirect('dashboard:product_gallery')


@require_POST
@staff_role_required('superadmin', 'admin')
def product_gallery_bulk_delete(request):
    ids = _bulk_post_ids(request)
    qs = ProductGallery.objects.filter(pk__in=ids)
    user = request.user
    if not user.is_superadmin:
        qs = qs.filter(Q(company__in=_companies(user)) | Q(company__isnull=True))
    count = 0
    for obj in qs:
        if obj.image:
            obj.image.delete(save=False)
        obj.delete()
        count += 1
    _bulk_message(request, 'image', count)
    return redirect('dashboard:product_gallery')


@staff_role_required('superadmin', 'admin')
def product_gallery_api(request):
    """AJAX endpoint â€” returns gallery images as JSON for product form picker."""
    from apps.menu.models import ProductGallery
    user = request.user
    qs = ProductGallery.objects.select_related('company')
    if not user.is_superadmin:
        qs = qs.filter(Q(company__in=_companies(user)) | Q(company__isnull=True))
    company_filter = request.GET.get('company')
    if company_filter and (user.is_superadmin or _companies(user).filter(pk=company_filter).exists()):
        qs = qs.filter(Q(company_id=company_filter) | Q(company__isnull=True))
    data = [
        {'id': g.pk, 'name': g.name, 'url': g.image.url}
        for g in qs.order_by('-created_at')[:200]
    ]
    return JsonResponse({'images': data})




@staff_role_required('superadmin', 'admin')
def product_gallery_rename(request, pk):
    from apps.menu.models import ProductGallery
    obj = get_object_or_404(ProductGallery, pk=pk)
    user = request.user
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    if not user.is_superadmin and obj.company and not user_can_access_company(user, obj.company_id):
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    name = (request.POST.get('name') or '').strip()
    if not name:
        return JsonResponse({'success': False, 'error': 'Name is required'}, status=400)
    obj.name = name
    obj.save(update_fields=['name'])
    return JsonResponse({'success': True, 'name': obj.name})


def _gallery_company_qs(model, user):
    qs = model.objects.select_related('company').order_by('-created_at')
    if not user.is_superadmin:
        qs = qs.filter(Q(company__in=_companies(user)) | Q(company__isnull=True))
    return qs


@staff_role_required('superadmin', 'admin')
def category_gallery(request):
    user = request.user
    qs = _gallery_company_qs(CategoryGallery, user)
    search = (request.GET.get('q') or '').strip()
    if search:
        qs = qs.filter(name__icontains=search)
    if request.method == 'POST':
        image_file = request.FILES.get('image')
        name = (request.POST.get('name') or '').strip() or Path(getattr(image_file, 'name', '') or 'Category Image').stem.replace('_', ' ').replace('-', ' ').strip() or 'Category Image'
        if not image_file:
            messages.error(request, 'Please select an image file.')
        else:
            company = (
                _companies(user).filter(pk=request.POST.get('company')).first() or get_primary_staff_company(user)
                if not user.is_superadmin
                else Company.objects.filter(pk=request.POST.get('company')).first() if request.POST.get('company') else None
            )
            CategoryGallery.objects.create(company=company, name=name, image=image_file)
            messages.success(request, f'"{name}" added to category gallery.')
            return redirect('dashboard:category_gallery')
    return render(request, 'dashboard/menu/category_gallery.html', {'gallery': qs, 'search_query': search, 'companies': _companies(user), 'page_title': 'Category Image Gallery'})


@require_POST
@staff_role_required('superadmin', 'admin')
def category_gallery_delete(request, pk):
    obj = get_object_or_404(CategoryGallery, pk=pk)
    user = request.user
    if not user.is_superadmin and obj.company and not user_can_access_company(user, obj.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:category_gallery')
    obj.image.delete(save=False)
    obj.delete()
    messages.success(request, 'Image deleted.')
    return redirect('dashboard:category_gallery')


@require_POST
@staff_role_required('superadmin', 'admin')
def category_gallery_bulk_delete(request):
    ids = _bulk_post_ids(request)
    qs = CategoryGallery.objects.filter(pk__in=ids)
    user = request.user
    if not user.is_superadmin:
        qs = qs.filter(Q(company__in=_companies(user)) | Q(company__isnull=True))
    count = 0
    for obj in qs:
        if obj.image:
            obj.image.delete(save=False)
        obj.delete()
        count += 1
    _bulk_message(request, 'image', count)
    return redirect('dashboard:category_gallery')


@staff_role_required('superadmin', 'admin')
def category_gallery_rename(request, pk):
    obj = get_object_or_404(CategoryGallery, pk=pk)
    user = request.user
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    if not user.is_superadmin and obj.company and not user_can_access_company(user, obj.company_id):
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    name = (request.POST.get('name') or '').strip()
    if not name:
        return JsonResponse({'success': False, 'error': 'Name is required'}, status=400)
    obj.name = name
    obj.save(update_fields=['name'])
    return JsonResponse({'success': True, 'name': obj.name})


@staff_role_required('superadmin', 'admin')
def category_gallery_api(request):
    user = request.user
    qs = _gallery_company_qs(CategoryGallery, user)
    company_filter = request.GET.get('company')
    if company_filter and (user.is_superadmin or _companies(user).filter(pk=company_filter).exists()):
        qs = qs.filter(Q(company_id=company_filter) | Q(company__isnull=True))
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(name__icontains=q)
    data = [{'id': g.pk, 'name': g.name, 'url': g.image.url, 'company': g.company.name if g.company else 'Global'} for g in qs[:200]]
    return JsonResponse({'images': data})


@staff_role_required('superadmin', 'admin')
def offering_gallery(request):
    user = request.user
    qs = _gallery_company_qs(OfferingGallery, user)
    search = (request.GET.get('q') or '').strip()
    if search:
        qs = qs.filter(name__icontains=search)
    if request.method == 'POST':
        image_file = request.FILES.get('image')
        name = (request.POST.get('name') or '').strip() or Path(getattr(image_file, 'name', '') or 'Offering Image').stem.replace('_', ' ').replace('-', ' ').strip() or 'Offering Image'
        if not image_file:
            messages.error(request, 'Please select an image file.')
        else:
            company = (
                _companies(user).filter(pk=request.POST.get('company')).first() or get_primary_staff_company(user)
                if not user.is_superadmin
                else Company.objects.filter(pk=request.POST.get('company')).first() if request.POST.get('company') else None
            )
            OfferingGallery.objects.create(company=company, name=name, image=image_file)
            messages.success(request, f'"{name}" added to offering gallery.')
            return redirect('dashboard:offering_gallery')
    return render(request, 'dashboard/menu/offering_gallery.html', {'gallery': qs, 'search_query': search, 'companies': _companies(user), 'page_title': 'Offering Image Gallery'})


@require_POST
@staff_role_required('superadmin', 'admin')
def offering_gallery_delete(request, pk):
    obj = get_object_or_404(OfferingGallery, pk=pk)
    user = request.user
    if not user.is_superadmin and obj.company and not user_can_access_company(user, obj.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:offering_gallery')
    obj.image.delete(save=False)
    obj.delete()
    messages.success(request, 'Image deleted.')
    return redirect('dashboard:offering_gallery')


@require_POST
@staff_role_required('superadmin', 'admin')
def offering_gallery_bulk_delete(request):
    ids = _bulk_post_ids(request)
    qs = OfferingGallery.objects.filter(pk__in=ids)
    user = request.user
    if not user.is_superadmin:
        qs = qs.filter(Q(company__in=_companies(user)) | Q(company__isnull=True))
    count = 0
    for obj in qs:
        if obj.image:
            obj.image.delete(save=False)
        obj.delete()
        count += 1
    _bulk_message(request, 'image', count)
    return redirect('dashboard:offering_gallery')


@staff_role_required('superadmin', 'admin')
def offering_gallery_rename(request, pk):
    obj = get_object_or_404(OfferingGallery, pk=pk)
    user = request.user
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    if not user.is_superadmin and obj.company and not user_can_access_company(user, obj.company_id):
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    name = (request.POST.get('name') or '').strip()
    if not name:
        return JsonResponse({'success': False, 'error': 'Name is required'}, status=400)
    obj.name = name
    obj.save(update_fields=['name'])
    return JsonResponse({'success': True, 'name': obj.name})


@staff_role_required('superadmin', 'admin')
def offering_gallery_api(request):
    user = request.user
    qs = _gallery_company_qs(OfferingGallery, user)
    company_filter = request.GET.get('company')
    if company_filter and (user.is_superadmin or _companies(user).filter(pk=company_filter).exists()):
        qs = qs.filter(Q(company_id=company_filter) | Q(company__isnull=True))
    q = (request.GET.get('q') or '').strip()
    if q:
        qs = qs.filter(name__icontains=q)
    data = [{'id': g.pk, 'name': g.name, 'url': g.image.url, 'company': g.company.name if g.company else 'Global'} for g in qs[:200]]
    return JsonResponse({'images': data})

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  CUSTOMER APPROVAL
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@require_POST
@staff_role_required('superadmin', 'admin')
def customer_approve(request, pk):
    """Quick AJAX approve/reject for new customer registrations."""
    customer = get_object_or_404(Customer, pk=pk, is_deleted=False)
    user = request.user
    if not user_can_action(user, 'perm_customers', 'approve'):
        return JsonResponse({'error': 'Permission denied.'}, status=403)
    if not user.is_superadmin and not user_can_access_company(user, customer.company_id):
        return JsonResponse({'error': 'Access denied.'}, status=403)
    action = request.POST.get('action', 'approve')
    customer.is_approved = (action == 'approve')
    customer.is_active = True
    customer.save(update_fields=['is_approved', 'is_active'])
    return JsonResponse({
        'success': True,
        'is_approved': customer.is_approved,
        'name': customer.name,
    })


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  WALLET MANAGEMENT (staff view per customer)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@staff_role_required('superadmin', 'admin')
def customer_wallet(request, pk):
    customer = get_object_or_404(Customer, pk=pk, is_deleted=False)
    user = request.user
    if not user_can_action(user, 'perm_customers', 'wallet_view'):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard:no_access')
    if not user.is_superadmin and not user_can_access_company(user, customer.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:customer_list')
    from apps.accounts.models import WalletTransaction
    history = WalletTransaction.objects.filter(customer=customer).order_by('-created_at')[:50]
    return render(request, 'dashboard/customers/wallet.html', {
        'customer': customer,
        'wallet_history': history,
        'page_title': f'Wallet - {customer.name}',
    })



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  LOCATION  (Company â†’ Location â†’ Building â†’ Cafeteria)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

def _company_wallet_recharge_modes(company):
    modes = []
    if getattr(company, 'pos_cash_enabled', True):
        modes.append((WalletTransaction.PAYMENT_CASH, 'Cash'))
    if getattr(company, 'pos_upi_enabled', True):
        modes.append((WalletTransaction.PAYMENT_UPI, 'UPI'))
    if getattr(company, 'pos_card_enabled', True):
        modes.append((WalletTransaction.PAYMENT_CARD, 'Card'))
    if getattr(company, 'online_payment', True):
        modes.append((WalletTransaction.PAYMENT_ONLINE, 'Online'))
    return modes


def _wallet_recharge_order_payment_mode(payment_mode):
    if payment_mode == WalletTransaction.PAYMENT_CASH:
        return PaymentModeChoices.CASH
    return PaymentModeChoices.ONLINE


@staff_role_required('superadmin', 'admin', 'pos')
def wallet_recharge(request):
    _perm = check_module_permission(request, 'perm_wallet_recharge')
    if _perm: return _perm
    user = request.user
    companies = _companies(user)
    company_filter = request.GET.get('company', '').strip()
    q = request.GET.get('q', '').strip()

    scoped_customers = Customer.objects.filter(is_deleted=False, is_active=True).select_related('company')
    if not user.is_superadmin:
        scoped_customers = scoped_customers.filter(company__in=companies)
        if company_filter and companies.filter(pk=company_filter).exists():
            scoped_customers = scoped_customers.filter(company_id=company_filter)
        elif company_filter:
            company_filter = ''
    elif company_filter:
        scoped_customers = scoped_customers.filter(company_id=company_filter)

    if q:
        scoped_customers = scoped_customers.filter(Q(name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q))

    if request.method == 'POST':
        customer_id = request.POST.get('customer', '').strip()
        customer_lookup = request.POST.get('customer_lookup', '').strip()
        amount_raw = request.POST.get('amount', '').strip()
        payment_mode = request.POST.get('payment_mode', WalletTransaction.PAYMENT_CASH)

        customer_qs = Customer.objects.filter(is_deleted=False, is_active=True).select_related('company')
        if not user.is_superadmin:
            customer_qs = customer_qs.filter(company__in=companies)
        customer = customer_qs.filter(pk=customer_id).first() if customer_id else None
        if not customer and customer_lookup:
            customer = customer_qs.filter(
                Q(phone__iexact=customer_lookup)
                | Q(email__iexact=customer_lookup)
            ).order_by('name').first()
            if not customer:
                name_matches = customer_qs.filter(name__icontains=customer_lookup).order_by('name')
                match_count = name_matches.count()
                if match_count == 1:
                    customer = name_matches.first()
                elif match_count > 1:
                    messages.error(request, 'Multiple customers match that name. Please select the exact customer from the list.')
                    return redirect('dashboard:wallet_recharge')

        if not customer:
            messages.error(request, 'Type or select a valid customer for recharge.')
            return redirect('dashboard:wallet_recharge')

        allowed_modes = [value for value, _ in _company_wallet_recharge_modes(customer.company)]
        if payment_mode not in allowed_modes:
            messages.error(request, 'This recharge payment mode is not enabled for the customer company.')
            return redirect('dashboard:wallet_recharge')

        try:
            amount = Decimal(amount_raw or '0').quantize(Decimal('0.01'))
        except (InvalidOperation, TypeError, ValueError):
            amount = Decimal('0.00')

        if amount < Decimal('100.00'):
            messages.error(request, 'Wallet recharge amount must be at least Rs.100.')
            return redirect('dashboard:wallet_recharge')
        if amount > Decimal('100000.00'):
            messages.error(request, 'Wallet recharge amount cannot exceed Rs.100000 in one entry.')
            return redirect('dashboard:wallet_recharge')

        fee_percent = Decimal(str(getattr(customer.company, 'pos_card_fee_percent', Decimal('3.50')) or 0))
        card_fee = (amount * fee_percent / Decimal('100')).quantize(Decimal('0.01')) if payment_mode == WalletTransaction.PAYMENT_CARD else Decimal('0.00')
        gross_amount = amount + card_fee

        with transaction.atomic():
            from apps.orders.views import _next_order_number

            locked_customer = Customer.objects.select_for_update().get(pk=customer.pk)
            locked_customer.wallet_balance = (locked_customer.wallet_balance or Decimal('0.00')) + amount
            locked_customer.save(update_fields=['wallet_balance'])
            recharge_order = Order.objects.create(
                company=locked_customer.company,
                customer=locked_customer,
                customer_name_snapshot=locked_customer.name,
                customer_phone_snapshot=locked_customer.phone,
                subtotal=amount,
                shipping_cost=Decimal('0.00'),
                bill_to_company=Decimal('0.00'),
                my_pay=gross_amount,
                total_amount=gross_amount,
                payment_mode=_wallet_recharge_order_payment_mode(payment_mode),
                payment_status='paid',
                transaction_id=f'wallet-recharge:{payment_mode}',
                order_type=ORDER_TYPE_WALLET_RECHARGE,
                order_status=OrderStatusChoices.DELIVERED,
                order_number=_next_order_number('WAL'),
            )
            OrderStatus.objects.create(
                order=recharge_order,
                status=OrderStatusChoices.DELIVERED,
                details=f'Wallet recharge completed via {dict(WalletTransaction.PAYMENT_CHOICES).get(payment_mode, payment_mode)}.',
                created_at=timezone.now(),
            )
            WalletTransaction.objects.create(
                customer=locked_customer,
                txn_type=WalletTransaction.TYPE_TOPUP,
                wallet_delta=amount,
                balance_after=locked_customer.wallet_balance,
                points_after=locked_customer.royalty_points,
                order_ref=recharge_order.order_number,
                payment_mode=payment_mode,
                card_fee_amount=card_fee,
                gross_amount=gross_amount,
                note=f'Wallet recharge via {payment_mode}',
                created_by=request.user.email,
            )
            # Customer notification on wallet recharge
            try:
                from apps.accounts.views import _create_customer_notification
                from apps.core.models import Notification as _Notif
                _create_customer_notification(
                    locked_customer,
                    notif_type=_Notif.TYPE_WALLET,
                    title='Wallet Topped Up',
                    message=f'Your wallet has been recharged with â‚¹{amount:.2f}. New balance: â‚¹{locked_customer.wallet_balance:.2f}.',
                    link='/auth/customer/wallet/',
                )
            except Exception:
                pass

        messages.success(request, f'Wallet recharged by Rs.{amount} for {customer.name}. Recharge order #{recharge_order.order_number} created.')
        return redirect('dashboard:wallet_recharge')

    history = WalletTransaction.objects.filter(txn_type=WalletTransaction.TYPE_TOPUP).select_related('customer', 'customer__company')
    if not user.is_superadmin:
        history = history.filter(customer__company__in=companies)
    elif company_filter:
        history = history.filter(customer__company_id=company_filter)

    form_company = get_primary_staff_company(user) if not user.is_superadmin else None
    if company_filter:
        form_company = companies.filter(pk=company_filter).first()
    payment_modes = _company_wallet_recharge_modes(form_company) if form_company else WalletTransaction.PAYMENT_CHOICES

    return render(request, 'dashboard/customers/wallet_recharge.html', {
        'customers': scoped_customers.order_by('name')[:250],
        'companies': companies,
        'company_filter': company_filter,
        'q': q,
        'payment_modes': payment_modes,
        'recent_recharges': history.order_by('-created_at')[:50],
        'page_title': 'Recharge Wallet',
    })


@staff_role_required('superadmin')
def location_list(request):
    from apps.core.models import Location
    qs = Location.objects.filter(is_deleted=False).order_by('name')
    return render(request, 'dashboard/companies/location_list.html', {
        'locations': qs,
        'page_title': 'Locations',
    })


@staff_role_required('superadmin')
def location_add(request):
    from apps.core.models import Location
    if request.method == 'POST':
        name = request.POST.get('name','').strip()
        if not name:
            messages.error(request, 'Location name is required.')
        else:
            Location.objects.create(name=name)
            messages.success(request, f'Location "{name}" created.')
            return redirect('dashboard:location_list')
    return render(request, 'dashboard/companies/location_form.html', {'page_title': 'Add Location'})


@staff_role_required('superadmin')
def location_edit(request, pk):
    from apps.core.models import Location
    loc = get_object_or_404(Location, pk=pk, is_deleted=False)
    if request.method == 'POST':
        name = request.POST.get('name','').strip()
        if not name:
            messages.error(request, 'Location name is required.')
        else:
            loc.name = name
            loc.is_active = request.POST.get('is_active') == 'on'
            loc.save()
            messages.success(request, f'Location "{loc.name}" updated.')
            return redirect('dashboard:location_list')
    return render(request, 'dashboard/companies/location_form.html', {
        'location': loc, 'page_title': f'Edit Location - {loc.name}'
    })


@require_POST
@staff_role_required('superadmin')
def location_delete(request, pk):
    loc = get_object_or_404(Location, pk=pk, is_deleted=False)
    loc.is_deleted = True
    loc.save(update_fields=['is_deleted'])
    messages.success(request, f'Location "{loc.name}" deleted.')
    return redirect('dashboard:location_list')


@require_POST
@staff_role_required('superadmin')
def location_bulk_delete(request):
    ids = _bulk_post_ids(request)
    count = Location.objects.filter(pk__in=ids, is_deleted=False).update(is_deleted=True) if ids else 0
    _bulk_message(request, 'location', count)
    return redirect('dashboard:location_list')



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  HIERARCHY OVERVIEW PAGE
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@staff_role_required('superadmin')
def hierarchy_overview(request):
    """
    Unified view: Company â†’ Location â†’ Building â†’ Cafeteria â†’ Counter
    Shows the full tree for superadmin, filtered to own company for admin.
    """
    from apps.core.models import Location
    from apps.menu.models import Cafe, Counter

    user = request.user
    companies = _companies(user)
    selected_company_id = request.GET.get('company','').strip()
    selected_company = companies.filter(pk=selected_company_id).first() if selected_company_id else None

    qs = Company.objects.filter(is_deleted=False, is_active=True).prefetch_related(
        'buildings__cafes__counters',
        'buildings__location',
    ).order_by('name')

    if not user.is_superadmin:
        qs = qs.filter(pk__in=companies.values_list('pk', flat=True))
    elif selected_company:
        qs = qs.filter(pk=selected_company.pk)

    return render(request, 'dashboard/companies/hierarchy.html', {
        'company_tree': qs,
        'companies': companies,
        'selected_company': selected_company,
        'page_title': 'Company Hierarchy',
    })



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  ROYALTY LEADERBOARD
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@staff_role_required('superadmin', 'admin')
def royalty_leaderboard(request):
    _perm = check_module_permission(request, 'perm_royalty_lb')
    if _perm: return _perm
    from apps.core.royalty_service import get_leaderboard, award_leaderboard_bonuses, _period_key
    from apps.core.models import RoyaltyAward

    user = request.user
    companies = _companies(user)
    company_id = request.GET.get('company', '')
    company = companies.filter(pk=company_id, is_active=True).first() if company_id else None
    if not user.is_superadmin and not company:
        company = get_primary_staff_company(user)

    period_key_param = request.GET.get('period', '')
    action = request.POST.get('action', '')

    msg = None
    if request.method == 'POST' and action == 'award' and company:
        pk = request.POST.get('period_key') or _period_key(company)
        results = award_leaderboard_bonuses(company, pk)
        awarded = [r for r in results if r['status'] == 'awarded']
        skipped = [r for r in results if r['status'] != 'awarded']
        parts = []
        if awarded:
            parts.append(f'Awarded: ' + ', '.join(f"#{r['rank']} {r['customer'].name} +{r['points']}pts" for r in awarded))
        if skipped:
            parts.append(f'Skipped (already awarded): {len(skipped)}')
        msg = ' | '.join(parts) if parts else 'No eligible customers.'
        messages.success(request, msg)

    leaderboard = []
    if company:
        pk = period_key_param or _period_key(company)
        leaderboard = get_leaderboard(company, pk)
        awards_this_period = RoyaltyAward.objects.filter(
            company=company, period_key=pk
        ).select_related('customer')
        awarded_ranks = {a.rank for a in awards_this_period}
    else:
        pk = ''
        awarded_ranks = set()

    return render(request, 'dashboard/royalty_leaderboard.html', {
        'companies': companies,
        'company': company,
        'leaderboard': leaderboard,
        'period_key': pk,
        'awarded_ranks': awarded_ranks,
        'page_title': 'Royalty Leaderboard',
    })


@staff_role_required('superadmin', 'admin', 'pos', 'cafeman')
def display_board_select(request):
    _perm = check_module_permission(request, 'perm_display_board')
    if _perm: return _perm

    companies = _companies(request.user)
    selected_company_id = (request.GET.get('company') or '').strip()
    selected_company = None

    if selected_company_id:
        selected_company = companies.filter(pk=selected_company_id).first()
    elif companies.count() == 1:
        selected_company = companies.first()

    board_companies = companies
    if selected_company:
        board_companies = companies.filter(pk=selected_company.pk)

    configs = (
        DisplayBoardConfig.objects
        .filter(company__in=board_companies, is_active=True)
        .select_related('company', 'building')
        .order_by('company__name', 'name')
    )

    return render(request, 'dashboard/orders/display_board_select.html', {
        'companies': companies,
        'selected_company': selected_company,
        'selected_company_id': selected_company_id,
        'configs': configs,
        'page_title': 'Display Board',
    })



# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  KIOSK CONFIG CRUD
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@staff_role_required('superadmin')
def kiosk_config_list(request):
    from apps.core.models import KioskConfig
    qs = KioskConfig.objects.select_related('company','building').order_by('company__name','name')
    if not request.user.is_superadmin and request.user.company:
        qs = qs.filter(company=request.user.company)
    return render(request, 'dashboard/companies/kiosk_config_list.html', {
        'configs': qs, 'page_title': 'Kiosk Configurations'
    })


@staff_role_required('superadmin')
def kiosk_config_add(request):
    from apps.core.models import KioskConfig
    companies = _companies(request.user)
    buildings = Building.objects.filter(is_deleted=False).select_related('company').order_by('company__name','name')
    if request.method == 'POST':
        co  = Company.objects.filter(pk=request.POST.get('company')).first()
        name = request.POST.get('name','').strip()
        if not co or not name:
            messages.error(request, 'Company and name are required.')
        else:
            bld = Building.objects.filter(pk=(request.POST.get('building') or '').strip(), is_deleted=False).first() if (request.POST.get('building') or '').strip() else None
            cfg = KioskConfig(
                company=co, building=bld, name=name,
                theme_color=request.POST.get('theme_color','').strip(),
                welcome_title=request.POST.get('welcome_title','').strip(),
                welcome_subtitle=request.POST.get('welcome_subtitle','').strip(),
                show_offerings=request.POST.get('show_offerings')=='on',
                show_categories=request.POST.get('show_categories')=='on',
                card_style=request.POST.get('card_style','standard'),
                is_active=request.POST.get('is_active')=='on',
            )
            if 'logo' in request.FILES: cfg.logo = request.FILES['logo']
            if 'hero_image' in request.FILES: cfg.hero_image = request.FILES['hero_image']
            cfg.save()
            messages.success(request, f'Kiosk config "{name}" created. URL param: ?kiosk={cfg.slug}')
            return redirect('dashboard:kiosk_config_list')
    return render(request, 'dashboard/companies/kiosk_config_form.html', {
        'companies': companies, 'buildings': buildings, 'page_title': 'Add Kiosk Config'
    })


@staff_role_required('superadmin')
def kiosk_config_edit(request, pk):
    from apps.core.models import KioskConfig
    cfg = get_object_or_404(KioskConfig, pk=pk)
    companies = _companies(request.user)
    buildings = Building.objects.filter(is_deleted=False).select_related('company').order_by('company__name','name')
    if request.method == 'POST':
        co  = Company.objects.filter(pk=request.POST.get('company')).first()
        name = request.POST.get('name','').strip()
        if not co or not name:
            messages.error(request, 'Company and name are required.')
        else:
            cfg.company=co
            cfg.building=Building.objects.filter(pk=request.POST.get('building'), is_deleted=False).first()
            cfg.name=name
            cfg.theme_color=request.POST.get('theme_color','').strip()
            cfg.welcome_title=request.POST.get('welcome_title','').strip()
            cfg.welcome_subtitle=request.POST.get('welcome_subtitle','').strip()
            cfg.show_offerings=request.POST.get('show_offerings')=='on'
            cfg.show_categories=request.POST.get('show_categories')=='on'
            cfg.card_style=request.POST.get('card_style','standard')
            cfg.is_active=request.POST.get('is_active')=='on'
            if 'logo' in request.FILES: cfg.logo = request.FILES['logo']
            if 'hero_image' in request.FILES: cfg.hero_image = request.FILES['hero_image']
            cfg.save()
            messages.success(request, f'Kiosk config "{cfg.name}" updated.')
            return redirect('dashboard:kiosk_config_list')
    return render(request, 'dashboard/companies/kiosk_config_form.html', {
        'cfg': cfg, 'companies': companies, 'buildings': buildings, 'page_title': f'Edit - {cfg.name}'
    })


@require_POST
@staff_role_required('superadmin')
def kiosk_config_delete(request, pk):
    cfg = get_object_or_404(KioskConfig, pk=pk)
    name = cfg.name
    cfg.delete()
    messages.success(request, f'Kiosk config "{name}" deleted.')
    return redirect('dashboard:kiosk_config_list')


@require_POST
@staff_role_required('superadmin')
def kiosk_config_bulk_delete(request):
    ids = _bulk_post_ids(request)
    qs = KioskConfig.objects.filter(pk__in=ids)
    count = qs.count()
    qs.delete()
    _bulk_message(request, 'kiosk config', count)
    return redirect('dashboard:kiosk_config_list')



@staff_role_required('superadmin')
def web_config_list(request):
    from apps.core.models import WebViewConfig
    qs = WebViewConfig.objects.select_related('company','building').order_by('company__name','name')
    if not request.user.is_superadmin and request.user.company:
        qs = qs.filter(company=request.user.company)
    return render(request, 'dashboard/companies/web_config_list.html', {
        'configs': qs, 'page_title': 'Web View Configurations'
    })


@staff_role_required('superadmin')
def web_config_add(request):
    from apps.core.models import WebViewConfig
    companies = _companies(request.user)
    buildings = Building.objects.filter(is_deleted=False).select_related('company').order_by('company__name','name')
    if request.method == 'POST':
        co = Company.objects.filter(pk=request.POST.get('company')).first()
        name = request.POST.get('name','').strip()
        if not co or not name:
            messages.error(request, 'Company and name are required.')
        else:
            bld = Building.objects.filter(pk=(request.POST.get('building') or '').strip(), is_deleted=False).first() if (request.POST.get('building') or '').strip() else None
            cfg = WebViewConfig(
                company=co, building=bld, name=name,
                theme_color=request.POST.get('theme_color','').strip(),
                navbar_color=request.POST.get('navbar_color','').strip(),
                welcome_title=request.POST.get('welcome_title','').strip(),
                welcome_subtitle=request.POST.get('welcome_subtitle','').strip(),
                show_offerings=request.POST.get('show_offerings')=='on',
                show_categories=request.POST.get('show_categories')=='on',
                card_style=request.POST.get('card_style','standard'),
                is_active=request.POST.get('is_active')=='on',
            )
            if 'logo' in request.FILES: cfg.logo = request.FILES['logo']
            if 'hero_image' in request.FILES: cfg.hero_image = request.FILES['hero_image']
            cfg.save()
            messages.success(request, f'Web view config "{name}" created. Preview param: ?web={cfg.slug}')
            return redirect('dashboard:web_config_list')
    return render(request, 'dashboard/companies/web_config_form.html', {
        'companies': companies, 'buildings': buildings, 'page_title': 'Add Web View Config'
    })


@staff_role_required('superadmin')
def web_config_edit(request, pk):
    from apps.core.models import WebViewConfig
    cfg = get_object_or_404(WebViewConfig, pk=pk)
    companies = _companies(request.user)
    buildings = Building.objects.filter(is_deleted=False).select_related('company').order_by('company__name','name')
    if request.method == 'POST':
        co = Company.objects.filter(pk=request.POST.get('company')).first()
        name = request.POST.get('name','').strip()
        if not co or not name:
            messages.error(request, 'Company and name are required.')
        else:
            cfg.company = co
            cfg.building = Building.objects.filter(pk=(request.POST.get('building') or '').strip(), is_deleted=False).first() if (request.POST.get('building') or '').strip() else None
            cfg.name = name
            cfg.theme_color = request.POST.get('theme_color','').strip()
            cfg.navbar_color = request.POST.get('navbar_color','').strip()
            cfg.welcome_title = request.POST.get('welcome_title','').strip()
            cfg.welcome_subtitle = request.POST.get('welcome_subtitle','').strip()
            cfg.show_offerings = request.POST.get('show_offerings')=='on'
            cfg.show_categories = request.POST.get('show_categories')=='on'
            cfg.card_style = request.POST.get('card_style','standard')
            cfg.is_active = request.POST.get('is_active')=='on'
            if 'logo' in request.FILES: cfg.logo = request.FILES['logo']
            if 'hero_image' in request.FILES: cfg.hero_image = request.FILES['hero_image']
            cfg.save()
            messages.success(request, f'Web view config "{cfg.name}" updated.')
            return redirect('dashboard:web_config_list')
    return render(request, 'dashboard/companies/web_config_form.html', {
        'cfg': cfg, 'companies': companies, 'buildings': buildings, 'page_title': f'Edit - {cfg.name}'
    })


@require_POST
@staff_role_required('superadmin')
def web_config_delete(request, pk):
    cfg = get_object_or_404(WebViewConfig, pk=pk)
    name = cfg.name
    cfg.delete()
    messages.success(request, f'Web view config "{name}" deleted.')
    return redirect('dashboard:web_config_list')


@require_POST
@staff_role_required('superadmin')
def web_config_bulk_delete(request):
    ids = _bulk_post_ids(request)
    qs = WebViewConfig.objects.filter(pk__in=ids)
    count = qs.count()
    qs.delete()
    _bulk_message(request, 'web config', count)
    return redirect('dashboard:web_config_list')



@staff_role_required('superadmin')
def display_config_list(request):
    from apps.core.models import DisplayBoardConfig
    qs = DisplayBoardConfig.objects.select_related('company','building').order_by('company__name','name')
    return render(request, 'dashboard/companies/display_config_list.html', {
        'configs': qs, 'page_title': 'Display Board Configurations'
    })


@staff_role_required('superadmin')
def display_config_add(request):
    from apps.core.models import DisplayBoardConfig
    companies = _companies(request.user)
    selected_company_id = (request.POST.get('company') if request.method == 'POST' else request.GET.get('company', '')).strip()
    selected_company = Company.objects.filter(pk=selected_company_id).first() if selected_company_id else None
    buildings = Building.objects.filter(is_deleted=False)
    if selected_company:
        buildings = buildings.filter(company=selected_company)
    else:
        buildings = buildings.none()
    buildings = buildings.select_related('company').order_by('name')
    if request.method == 'POST':
        co = selected_company
        name = request.POST.get('name','').strip()
        if not co or not name:
            messages.error(request, 'Company and name are required.')
        else:
            building_id = (request.POST.get('building') or '').strip()
            bld = None
            if building_id:
                bld = Building.objects.filter(pk=building_id, company=co, is_deleted=False).first()
                if not bld:
                    messages.error(request, 'Selected building does not belong to the selected company.')
                    return render(request, 'dashboard/companies/display_config_form.html', {
                        'companies': companies, 'buildings': buildings, 'page_title': 'Add Display Board Config',
                        'selected_company_id': selected_company_id,
                    })
            cfg = DisplayBoardConfig(
                company=co, building=bld, name=name,
                theme_color=request.POST.get('theme_color','').strip(),
                heading_text=request.POST.get('heading_text','').strip(),
                side_text=request.POST.get('side_text','').strip(),
                waiting_text=request.POST.get('waiting_text','').strip(),
                promo_embed_url=request.POST.get('promo_embed_url','').strip(),
                footer_text=request.POST.get('footer_text','').strip(),
                pending_label=request.POST.get('pending_label','Pending').strip() or 'Pending',
                confirmed_label=request.POST.get('confirmed_label','Order Placed').strip() or 'Order Placed',
                preparing_label=request.POST.get('preparing_label','Preparing').strip() or 'Preparing',
                ready_label=request.POST.get('ready_label','Food Ready').strip() or 'Food Ready',
                show_clock=request.POST.get('show_clock')=='on',
                show_company_filter=request.POST.get('show_company_filter')=='on',
                show_status_legend=request.POST.get('show_status_legend')=='on',
                sound_enabled=request.POST.get('sound_enabled')=='on',
                voice_enabled=request.POST.get('voice_enabled')=='on',
                is_active=request.POST.get('is_active')=='on',
            )
            if 'logo' in request.FILES: cfg.logo = request.FILES['logo']
            if 'footer_logo' in request.FILES: cfg.footer_logo = request.FILES['footer_logo']
            if 'background_image' in request.FILES: cfg.background_image = request.FILES['background_image']
            cfg.save()
            messages.success(request, f'Display board config "{name}" created.')
            return redirect('dashboard:display_config_list')
    return render(request, 'dashboard/companies/display_config_form.html', {
        'companies': companies, 'buildings': buildings, 'page_title': 'Add Display Board Config',
        'selected_company_id': selected_company_id,
    })


@staff_role_required('superadmin')
def display_config_edit(request, pk):
    from apps.core.models import DisplayBoardConfig
    cfg = get_object_or_404(DisplayBoardConfig, pk=pk)
    companies = _companies(request.user)
    selected_company_id = (request.POST.get('company') if request.method == 'POST' else request.GET.get('company', str(cfg.company_id or ''))).strip()
    selected_company = Company.objects.filter(pk=selected_company_id).first() if selected_company_id else cfg.company
    buildings = Building.objects.filter(is_deleted=False)
    if selected_company:
        buildings = buildings.filter(company=selected_company)
    else:
        buildings = buildings.none()
    buildings = buildings.select_related('company').order_by('name')
    if request.method == 'POST':
        co = selected_company
        name = request.POST.get('name','').strip()
        if not co or not name:
            messages.error(request, 'Company and name are required.')
        else:
            building_id = (request.POST.get('building') or '').strip()
            bld = None
            if building_id:
                bld = Building.objects.filter(pk=building_id, company=co, is_deleted=False).first()
                if not bld:
                    messages.error(request, 'Selected building does not belong to the selected company.')
                    return render(request, 'dashboard/companies/display_config_form.html', {
                        'cfg': cfg, 'companies': companies, 'buildings': buildings, 'page_title': f'Edit - {cfg.name}',
                        'selected_company_id': selected_company_id,
                    })
            cfg.company = co
            cfg.building = bld
            cfg.name = name
            cfg.theme_color = request.POST.get('theme_color','').strip()
            cfg.heading_text = request.POST.get('heading_text','').strip()
            cfg.side_text = request.POST.get('side_text','').strip()
            cfg.waiting_text = request.POST.get('waiting_text','').strip()
            cfg.promo_embed_url = request.POST.get('promo_embed_url','').strip()
            cfg.footer_text = request.POST.get('footer_text','').strip()
            cfg.pending_label = request.POST.get('pending_label','Pending').strip() or 'Pending'
            cfg.confirmed_label = request.POST.get('confirmed_label','Order Placed').strip() or 'Order Placed'
            cfg.preparing_label = request.POST.get('preparing_label','Preparing').strip() or 'Preparing'
            cfg.ready_label = request.POST.get('ready_label','Food Ready').strip() or 'Food Ready'
            cfg.show_clock = request.POST.get('show_clock')=='on'
            cfg.show_company_filter = request.POST.get('show_company_filter')=='on'
            cfg.show_status_legend = request.POST.get('show_status_legend')=='on'
            cfg.sound_enabled = request.POST.get('sound_enabled')=='on'
            cfg.voice_enabled = request.POST.get('voice_enabled')=='on'
            cfg.is_active = request.POST.get('is_active')=='on'
            if 'logo' in request.FILES: cfg.logo = request.FILES['logo']
            if 'footer_logo' in request.FILES: cfg.footer_logo = request.FILES['footer_logo']
            if 'background_image' in request.FILES: cfg.background_image = request.FILES['background_image']
            cfg.save()
            messages.success(request, f'Display board config "{cfg.name}" updated.')
            return redirect('dashboard:display_config_list')
    return render(request, 'dashboard/companies/display_config_form.html', {
        'cfg': cfg, 'companies': companies, 'buildings': buildings, 'page_title': f'Edit - {cfg.name}',
        'selected_company_id': selected_company_id,
    })


@require_POST
@staff_role_required('superadmin')
def display_config_delete(request, pk):
    cfg = get_object_or_404(DisplayBoardConfig, pk=pk)
    name = cfg.name
    cfg.delete()
    messages.success(request, f'Display board config "{name}" deleted.')
    return redirect('dashboard:display_config_list')


@require_POST
@staff_role_required('superadmin')
def display_config_bulk_delete(request):
    ids = _bulk_post_ids(request)
    qs = DisplayBoardConfig.objects.filter(pk__in=ids)
    count = qs.count()
    qs.delete()
    _bulk_message(request, 'display config', count)
    return redirect('dashboard:display_config_list')


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
#  BROADCAST NOTIFICATIONS (admin â†’ customers)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

@staff_role_required('superadmin', 'admin')
def broadcast_notification_list(request):
    """List broadcast notifications grouped by broadcast batch."""
    _perm = check_module_permission(request, 'perm_broadcast_notifications')
    if _perm: return _perm
    from apps.core.models import Notification
    from django.core.paginator import Paginator
    from collections import defaultdict
    user = request.user
    qs = Notification.objects.filter(
        notif_type__in=['broadcast', 'system', 'promo']
    ).select_related('company', 'customer').order_by('-created_at')
    if not user.is_superadmin:
        qs = qs.filter(company__in=_companies(user))
    # Company filter (superadmin only)
    filter_company_id = request.GET.get('company', '').strip()
    if filter_company_id and (user.is_superadmin or _companies(user).filter(pk=filter_company_id).exists()):
        qs = qs.filter(company__pk=filter_company_id)
    elif filter_company_id:
        filter_company_id = ''
    # Group into broadcast batches (same title+message+company+type+minute)
    all_notifs = list(qs[:2000])
    groups = defaultdict(list)
    for n in all_notifs:
        minute = n.created_at.replace(second=0, microsecond=0)
        key = (n.title, n.message, n.company_id, n.notif_type, minute)
        groups[key].append(n)
    broadcast_groups = []
    for (title, message, company_id, notif_type, minute), notifs in groups.items():
        first = notifs[0]
        broadcast_groups.append({
            'title': title,
            'message': message,
            'notif_type': notif_type,
            'notif_type_display': first.get_notif_type_display(),
            'company': first.company,
            'sent_at': minute,
            'count': len(notifs),
            'recipients': notifs,
        })
    broadcast_groups.sort(key=lambda x: x['sent_at'], reverse=True)
    paginator = Paginator(broadcast_groups, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    return render(request, 'dashboard/notifications/broadcast_list.html', {
        'broadcast_groups': page_obj,
        'page_obj': page_obj,
        'companies': _companies(user),
        'filter_company_id': filter_company_id,
        'page_title': 'Broadcast Notifications',
    })


@staff_role_required('superadmin', 'admin')
def broadcast_notification_send(request):
    """Send a broadcast notification to all or specific customers."""
    _perm = check_module_permission(request, 'perm_broadcast_notifications')
    if _perm: return _perm
    from apps.core.models import Notification
    from apps.accounts.models import Customer
    user = request.user
    companies = _companies(user)

    if request.method == 'POST':
        if not user_can_action(user, 'perm_broadcast_notifications', 'send'):
            return _deny_dashboard_action(request)
        title = (request.POST.get('title') or '').strip()
        message = (request.POST.get('message') or '').strip()
        notif_type = request.POST.get('notif_type', 'broadcast')
        company_id = request.POST.get('company', '').strip()
        customer_id = request.POST.get('customer', '').strip()

        if notif_type not in ('broadcast', 'system', 'promo'):
            notif_type = 'broadcast'

        if not title:
            messages.error(request, 'Title is required.')
            return redirect('dashboard:broadcast_notification_send')

        # Determine target company
        target_company = None
        if company_id:
            target_company = companies.filter(pk=company_id).first()
        elif not user.is_superadmin:
            target_company = get_primary_staff_company(user)
        if not user.is_superadmin and target_company and not user_can_access_company(user, target_company.pk):
            messages.error(request, 'Access denied.')
            return redirect('dashboard:broadcast_notification_send')

        # Determine target customers
        if customer_id:
            # Specific customer
            qs = Customer.objects.filter(pk=customer_id, is_deleted=False, is_active=True)
            if target_company:
                qs = qs.filter(company=target_company)
        elif target_company:
            # All customers of this company
            qs = Customer.objects.filter(
                company=target_company, is_deleted=False, is_active=True
            )
        else:
            # Superadmin â€” all customers
            qs = Customer.objects.filter(is_deleted=False, is_active=True)

        image_file = request.FILES.get('image')
        count = 0
        for cust in qs.iterator():
            n = Notification.objects.create(
                company=cust.company,
                customer=cust,
                notif_type=notif_type,
                title=title,
                message=message,
            )
            if image_file:
                image_file.seek(0)
                n.image.save(image_file.name, image_file, save=True)
            count += 1

        messages.success(request, f'Notification sent to {count} customer{"s" if count != 1 else ""}.')
        return redirect('dashboard:broadcast_notification_list')

    # GET â€” show form
    # Customer search for specific targeting
    selected_company = None
    if user.is_superadmin:
        company_id = request.GET.get('company', '')
        if company_id:
            selected_company = Company.objects.filter(pk=company_id, is_deleted=False).first()
    else:
        company_id = request.GET.get('company', '')
        selected_company = companies.filter(pk=company_id).first() if company_id else get_primary_staff_company(user)

    customers = []
    if selected_company:
        customers = Customer.objects.filter(
            company=selected_company, is_deleted=False, is_active=True
        ).order_by('name')[:200]

    return render(request, 'dashboard/notifications/broadcast_send.html', {
        'companies': companies,
        'customers': customers,
        'selected_company': selected_company,
        'page_title': 'Send Broadcast Notification',

    })

@staff_role_required('superadmin', 'admin')
def broadcast_notification_delete(request, pk):
    _perm = check_module_permission(request, 'perm_broadcast_notifications')
    if _perm: return _perm
    from apps.core.models import Notification
    notif = get_object_or_404(Notification, pk=pk, notif_type__in=['broadcast','wallet'])
    if not _has_full_module_edit(request.user, 'perm_broadcast_notifications'):
        return _deny_dashboard_action(request)
    if not request.user.is_superadmin and notif.company_id and not user_can_access_company(request.user, notif.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:broadcast_notification_list')
    if request.method == 'POST':
        notif.delete()
        messages.success(request, 'Notification deleted.')
    return redirect('dashboard:broadcast_notification_list')

@staff_role_required('superadmin', 'admin')
def broadcast_notification_bulk_delete(request):
    _perm = check_module_permission(request, 'perm_broadcast_notifications')
    if _perm: return _perm
    from apps.core.models import Notification
    if request.method == 'POST':
        if not _has_full_module_edit(request.user, 'perm_broadcast_notifications'):
            return _deny_dashboard_action(request)
        pks = request.POST.getlist('pks')
        if pks:
            qs = Notification.objects.filter(
                pk__in=pks,
                notif_type__in=['broadcast', 'system', 'promo']
            )
            if not request.user.is_superadmin:
                qs = qs.filter(company__in=_companies(request.user))
            deleted, _ = qs.delete()
            messages.success(request, f'Deleted broadcast ({deleted} entries).')
        else:
            messages.error(request, 'Nothing to delete.')
    return redirect('dashboard:broadcast_notification_list')


# â”€â”€â”€ GRANULAR PERMISSION MATRIX VIEWS â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

@staff_role_required('superadmin')
def permission_matrix(request, pk):
    """Granular per-field permission matrix for a staff member."""
    from apps.accounts.models import StaffModulePermission
    from apps.core.access import ACTION_REGISTRY, PERM_SECTION_ORDER

    target_user = get_object_or_404(StaffUser, pk=pk)
    if target_user.role == StaffUser.ROLE_SUPERADMIN:
        messages.error(request, 'Superadmin permissions cannot be configured.')
        return redirect('dashboard:staff_list')

    existing = {
        p.module_key: p
        for p in StaffModulePermission.objects.filter(staff_user=target_user)
    }

    if request.method == 'POST':
        if request.POST.get('site_access_submitted') == '1':
            selected_sites = Company.objects.filter(
                pk__in=request.POST.getlist('site_access'),
                is_active=True,
                is_deleted=False,
            ).order_by('name')
            selected_site_ids = list(selected_sites.values_list('pk', flat=True))
            target_user.site_access.set(selected_sites)
            if target_user.company_id not in selected_site_ids:
                target_user.company = selected_sites.first()
                target_user.save(update_fields=['company'])

        StaffModulePermission.objects.filter(staff_user=target_user).delete()
        created = 0
        for module_key, meta in ACTION_REGISTRY.items():
            level = request.POST.get(f'level__{module_key}', '').strip()
            if not level:
                continue
            allowed_actions = []
            if level == 'part_edit':
                allowed_actions = [
                    ak for ak in meta['actions'].keys()
                    if request.POST.get(f'action__{module_key}__{ak}') == 'on'
                ]
            StaffModulePermission.objects.create(
                staff_user=target_user,
                module_key=module_key,
                level=level,
                allowed_actions=allowed_actions,
            )
            created += 1
        messages.success(
            request,
            f'Permissions and site access saved for {target_user.name} - {created} module(s) configured.'
        )
        return redirect('dashboard:staff_list')

    # Build display matrix grouped by section
    sections = {}
    for module_key, meta in ACTION_REGISTRY.items():
        sec = meta['section']
        if sec not in sections:
            sections[sec] = []
        perm = existing.get(module_key)
        current_level   = perm.level if perm else ''
        current_actions = set(perm.allowed_actions or []) if perm else set()
        sections[sec].append({
            'module_key':   module_key,
            'label':        meta['label'],
            'current_level': current_level,
            'has_actions':  bool(meta['actions']),
            'actions': [
                {'key': ak, 'label': al, 'checked': ak in current_actions}
                for ak, al in meta['actions'].items()
            ],
        })

    ordered_sections = []
    for sec_name in PERM_SECTION_ORDER:
        if sec_name in sections:
            ordered_sections.append((sec_name, sections[sec_name]))
    for sec_name, items in sections.items():
        if sec_name not in PERM_SECTION_ORDER:
            ordered_sections.append((sec_name, items))

    return render(request, 'dashboard/permission_matrix.html', {
        'target_user':      target_user,
        'companies':        Company.objects.filter(is_active=True, is_deleted=False).order_by('name'),
        'selected_site_ids': set(get_staff_site_companies(target_user).values_list('pk', flat=True)),
        'ordered_sections': ordered_sections,
        'page_title':       f'Permission Matrix - {target_user.name}',
    })


@staff_role_required('superadmin')
def pending_changes_list(request):
    """List of pending changes awaiting superadmin approval."""
    from apps.accounts.models import PendingChange

    pending = PendingChange.objects.filter(
        status='pending'
    ).select_related('staff_user').order_by('-created_at')

    recent = PendingChange.objects.exclude(
        status='pending'
    ).select_related('staff_user', 'reviewed_by').order_by('-reviewed_at')[:30]

    return render(request, 'dashboard/pending_changes.html', {
        'pending':       pending,
        'recent':        recent,
        'page_title':    'Pending Changes',
        'pending_count': pending.count(),
    })


def _apply_pending_change(change):
    """
    Apply field_diffs from a PendingChange to the real DB record.
    Returns (True, 'OK') or (False, error_message).
    """
    MODEL_MAP = {
        'perm_products':    ('apps.menu.models',     'Product'),
        'perm_categories':  ('apps.menu.models',     'Category'),
        'perm_offerings':   ('apps.menu.models',     'Offering'),
        'perm_counters':    ('apps.menu.models',     'Counter'),
        'perm_offers':      ('apps.menu.models',     'Offer'),
        'perm_customers':   ('apps.accounts.models', 'Customer'),
        'perm_coupons':     ('apps.core.models',     'Coupon'),
    }
    mapping = MODEL_MAP.get(change.module_key)
    if not mapping:
        return False, f'No model mapping for module: {change.module_key}'
    try:
        import importlib
        mod        = importlib.import_module(mapping[0])
        ModelClass = getattr(mod, mapping[1])
        obj        = ModelClass.objects.get(pk=change.object_id)
        update_fields = []
        for field_name, diff in change.field_diffs.items():
            setattr(obj, field_name, diff.get('after'))
            update_fields.append(field_name)
        if update_fields:
            obj.save(update_fields=update_fields)
        return True, 'OK'
    except Exception as e:
        return False, str(e)


@staff_role_required('superadmin')
def pending_change_review(request, pk):
    """Approve or reject a pending change."""
    from apps.accounts.models import PendingChange
    from django.utils import timezone

    if request.method != 'POST':
        return redirect('dashboard:pending_changes')

    change = get_object_or_404(PendingChange, pk=pk, status=PendingChange.STATUS_PENDING)
    action = request.POST.get('action', '')
    note   = request.POST.get('review_note', '').strip()

    if action == 'approve':
        ok, msg = _apply_pending_change(change)
        if ok:
            change.status      = PendingChange.STATUS_APPROVED
            change.reviewed_by = request.user
            change.review_note = note
            change.reviewed_at = timezone.now()
            change.save()
            messages.success(request, f'Change approved and applied: {change.object_label}')
        else:
            messages.error(request, f'Could not apply change: {msg}')
    elif action == 'reject':
        change.status      = PendingChange.STATUS_REJECTED
        change.reviewed_by = request.user
        change.review_note = note
        change.reviewed_at = timezone.now()
        change.save()
        messages.success(request, f'Change rejected: {change.object_label}')
    else:
        messages.error(request, 'Unknown action.')

    return redirect('dashboard:pending_changes')
