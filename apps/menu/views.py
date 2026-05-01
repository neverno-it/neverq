import csv
import io
import json
import re
import uuid
from pathlib import Path
from decimal import Decimal, InvalidOperation
from django.core.cache import cache
from django.core.files.base import ContentFile
from django.core.exceptions import ValidationError
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, Http404
from django.views.decorators.http import require_POST
from django.views.decorators.cache import never_cache
from django.db import IntegrityError
from django.db.models import Q, Count
from django.utils import timezone
from django.utils.text import slugify
from django.utils.dateparse import parse_date, parse_datetime, parse_time
from django.utils.http import url_has_allowed_host_and_scheme
from apps.accounts.decorators import customer_login_required, staff_role_required
from apps.core.access import (
    check_module_permission,
    create_pending_change,
    get_list_perms,
    get_locked_html_names,
    get_module_level,
    get_primary_staff_company,
    get_staff_site_companies,
    has_any_granular_perms,
    user_can_access_company,
    user_can_action,
)
from apps.core.models import Company, Building
from .models import Category, CategoryCompanyStatus, Product, Advertise, MediaAsset, HolidaySchedule, Cafe, FoodType, Schedule, Offering, Counter, ProductCounter, Offer, ProductCompanyPrice, OfferingGallery, ProductGallery, PORTAL_BANNER_WIDTH, PORTAL_BANNER_HEIGHT, PORTAL_BANNER_LABEL
from .pricing import PRICING_MODE_STAFF, get_effective_price
from apps.orders.views import _promote_due_ready_orders


# ════════════════════════════════════════════════════════════════
#  CUSTOMER-FACING  (menu browsing + cart)
#  ALL product lookups scoped to customer.company
# ════════════════════════════════════════════════════════════════

def _money(value):
    """Return a formatted 2-decimal string — used ONLY in JSON responses and template context.
    Do NOT use this for arithmetic; use Decimal(str(value)) directly instead."""
    try:
        return f"{Decimal(str(value)):.2f}"
    except (InvalidOperation, TypeError, ValueError):
        return "0.00"


def _dashboard_return_or_default(request, fallback_name, **kwargs):
    return_to = (request.POST.get('return_to') or request.GET.get('return_to') or '').strip()
    return_anchor = (request.POST.get('return_anchor') or request.GET.get('return_anchor') or '').strip()
    if return_to and url_has_allowed_host_and_scheme(
        return_to,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        target = return_to.split('#', 1)[0]
        if return_anchor and re.fullmatch(r'[-A-Za-z0-9_:.]+', return_anchor):
            target = f'{target}#{return_anchor}'
        return redirect(target)
    return redirect(fallback_name, **kwargs)


def _to_decimal(value):
    """Safe Decimal conversion for arithmetic."""
    try:
        return Decimal(str(value or 0))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal('0.00')


def _cart_quantity_map(cart):
    return {str(pid): int(item.get('qty', 0)) for pid, item in (cart or {}).items()}


def _get_order_min_qty(product):
    try:
        return max(1, int(getattr(product, 'min_qty', 1) or 1))
    except (TypeError, ValueError):
        return 1


def _get_order_max_qty(product):
    try:
        return max(1, int(getattr(product, 'max_qty', 999999) or 999999))
    except (TypeError, ValueError):
        return 999999


def _get_web_max_qty(product):
    """
    Effective maximum units allowed on a web/kiosk order line.

    max_qty is the per-order cap. web_qty is only stock: -1 means unlimited
    stock, 0 means sold out, positive numbers cap the order by remaining stock.
    """
    min_qty = _get_order_min_qty(product)
    cap = _get_order_max_qty(product)
    web_qty = getattr(product, 'web_qty', -1)
    try:
        web_qty = int(web_qty)
    except (TypeError, ValueError):
        web_qty = -1

    if web_qty >= 0:
        cap = min(cap, web_qty)

    return cap if cap >= min_qty else 0



def _ordering_closed_payload(company):
    return {
        'success': False,
        'error': company.ordering_status_message or 'Ordering is currently closed for this store.',
    }


def _fresh_company(customer):
    company = Company.objects.get(pk=customer.company_id)
    customer.company = company
    return company



def _product_is_visible_for_customer(product, company, food_pref=''):
    if not company.is_store_open:
        return False
    if not getattr(product, 'is_active', False) or getattr(product, 'is_deleted', False):
        return False
    if getattr(product, 'company_id', None) != getattr(company, 'id', None):
        return False
    if not product.is_available_now():
        return False
    if _get_web_max_qty(product) <= 0:
        return False

    # schedule_bypass is already enforced inside product.is_available_now().
    # Do not re-hide the product here by re-checking its parent category/offering
    # schedules, otherwise the bypass toggle appears to save but has no visible effect
    # on customer-facing menu listings.
    if getattr(product, 'schedule_bypass', False):
        return _product_matches_food_pref(product, food_pref)

    if product.category_id and getattr(product, 'category', None) and not product.category.is_active_now(company):
        return False
    if product.offering_id and getattr(product, 'offering', None) and not product.offering.is_active_now():
        return False
    return _product_matches_food_pref(product, food_pref)

def _product_matches_food_pref(product, food_pref):
    """
    Returns True if the product matches the requested food preference.

    Priority order:
      1. Category icon_type  (ICON_VEG=1, ICON_NONVEG=2) — explicit admin setting, most reliable.
      2. Per-product FoodType names — fallback for products tagged individually.
      3. If names list is empty (no food-type tagged) and category is unclassified → exclude
         from filtered views so customers never see mis-classified items.
    """
    food_pref = (food_pref or '').strip().lower()
    if food_pref not in {'veg', 'nonveg'}:
        return True

    # 1. Category icon_type is the authoritative flag
    cat_icon = getattr(product.category, 'icon_type', 0)
    if food_pref == 'veg' and cat_icon == 1:    # Category.ICON_VEG
        return True
    if food_pref == 'nonveg' and cat_icon == 2:  # Category.ICON_NONVEG
        return True
    # If category is explicitly the *opposite* type, short-circuit False
    if food_pref == 'veg' and cat_icon == 2:
        return False
    if food_pref == 'nonveg' and cat_icon == 1:
        return False

    # 2. Fall back to per-product FoodType name matching
    names = [ft.name.lower() for ft in product.food_type.all()]
    if not names:
        # No classification at all — hide from filtered results to avoid misleading customers
        return False
    if food_pref == 'veg':
        return any('veg' in n and 'non' not in n for n in names)
    return any('non' in n for n in names)


def _resolve_site_price(product, company, building=None, cafe=None):
    return get_effective_price(
        product,
        company,
        building=building,
        cafe=cafe,
        pricing_mode=PRICING_MODE_STAFF,
    )


def _attach_display_prices(products, company, building=None, cafe=None, used_offer_ids=None):
    """
    Attach display_price, live_offer, and display_discounted_price to each product.

    live_offer is ALWAYS set (badge is informational — shows even if already used today).
    display_discounted_price is only set if the offer has NOT been used today,
    so the strike-through price only appears when the discount is actually applicable.
    """
    _used = used_offer_ids or set()
    for product in products:
        product.display_price = _resolve_site_price(product, company, building=building, cafe=cafe)
        product.live_offer = _get_live_offer_for_product_menu(company, product, cafe=cafe)
        product.display_discounted_price = None
        product.effective_web_qty = _get_web_max_qty(product)
        # Only show struck-through/discounted price if offer has not been used today
        if product.live_offer and product.live_offer.pk not in _used:
            if product.live_offer.offer_type == Offer.TYPE_PERCENT:
                eff, saving = _apply_offer_to_line_menu(product.live_offer, product.display_price, 1)
                if saving > 0:
                    product.display_discounted_price = eff
    return products


def _normalize_bulk_header(value):
    raw = str(value or '').strip().lower().replace(' ', '_').replace('-', '_')
    normalized = re.sub(r'[^a-z0-9]+', '_', raw).strip('_')
    aliases = {
        'staff_base_price': 'base_price',
        'staff_base_price_rs': 'base_price',
        'staff_base_price_inr': 'base_price',
        'staff_price': 'base_price',
        'staff_price_rs': 'base_price',
        'price_rs': 'base_price',
        'base_price_rs': 'base_price',
        'visitor_price': 'company_price',
        'visitor_price_rs': 'company_price',
        'visitor_price_inr': 'company_price',
        'pos_visitor_price': 'company_price',
        'room_service_extra': 'room_service_extra_percent',
        'room_service_extra_percent': 'room_service_extra_percent',
        'room_service_extra_percentage': 'room_service_extra_percent',
        'room_service_extra_pct': 'room_service_extra_percent',
        'packing': 'packing_price',
        'packing_rs': 'packing_price',
        'packing_price_rs': 'packing_price',
        'packing_price_inr': 'packing_price',
        'min_qty_per_order': 'min_qty',
        'minimum_qty_per_order': 'min_qty',
        'minimum_quantity_per_order': 'min_qty',
        'max_qty_per_order': 'max_qty',
        'maximum_qty_per_order': 'max_qty',
        'maximum_quantity_per_order': 'max_qty',
        'web_stock': 'web_qty',
        'web_stock_web_qty': 'web_qty',
        'web_quantity': 'web_qty',
        'web_ordering_stock': 'web_qty',
        'pos_stock': 'pos_qty',
        'pos_stock_pos_qty': 'pos_qty',
        'pos_quantity': 'pos_qty',
        'pos_counter_stock': 'pos_qty',
        'prep_time': 'preparation_time_minutes',
        'prep_time_minutes': 'preparation_time_minutes',
        'preparation_time': 'preparation_time_minutes',
        'preparation_time_minutes': 'preparation_time_minutes',
        'calories_kcal': 'calories',
        'kcal': 'calories',
    }
    return aliases.get(normalized, normalized)


def _sheet_key(name):
    return _normalize_bulk_header(name).replace('__', '_')


def _load_bulk_upload_payload(upload):
    """Return {'mode': 'xlsx'|'csv', 'sheets': {'products': [...], ...}}."""
    name = (getattr(upload, 'name', '') or '').lower()
    if name.endswith('.xlsx') or name.endswith('.xlsm'):
        from openpyxl import load_workbook
        wb = load_workbook(upload, read_only=True, data_only=True)
        sheets = {}
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [_normalize_bulk_header(h) for h in rows[0]]
            sheet_rows = []
            for row in rows[1:]:
                if not any(cell not in (None, '') for cell in row):
                    continue
                row_dict = {}
                for i, header in enumerate(headers):
                    if not header:
                        continue
                    row_dict[header] = row[i] if i < len(row) else None
                sheet_rows.append(row_dict)
            sheets[_sheet_key(ws.title)] = sheet_rows
        return {'mode': 'xlsx', 'sheets': sheets}

    decoded = upload.read().decode('utf-8-sig', errors='replace')
    rows = []
    for row in csv.DictReader(io.StringIO(decoded)):
        rows.append({_normalize_bulk_header(k): v for k, v in row.items()})
    return {'mode': 'csv', 'sheets': {'products': rows}}


def _parse_bool(value, default=False):
    if value is None or value == '':
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {'1', 'true', 'yes', 'y', 'on'}:
        return True
    if text in {'0', 'false', 'no', 'n', 'off'}:
        return False
    return default


def _parse_int(value, default=0, minimum=None):
    if value in (None, ''):
        result = default
    else:
        try:
            result = int(str(value).strip())
        except (TypeError, ValueError):
            raise ValueError(f'Invalid integer: {value}')
    if minimum is not None and result < minimum:
        result = minimum
    return result


def _parse_decimal(value, default='0'):
    raw = default if value in (None, '') else value
    try:
        return Decimal(str(raw).strip())
    except (InvalidOperation, AttributeError, TypeError):
        raise ValueError(f'Invalid decimal: {value}')


def _clamp_estimated_calories(value, upper_limit=1800):
    rounded = int(round(float(value) / 10) * 10)
    return max(5, min(int(upper_limit), rounded))


def _estimate_product_calories(name, description=''):
    text = re.sub(r'\s+', ' ', f'{name or ""} {description or ""}'.lower()).strip()
    if not text:
        return None

    explicit = re.search(r'\b(\d{2,4})\s*(kcal|calories|cal)\b', text)
    if explicit:
        return max(0, min(5000, int(explicit.group(1))))

    counted_rules = [
        (r'\b([1-9]\d?)\s*(idli|idlis)\b', 70),
        (r'\b([1-9]\d?)\s*(egg|eggs)\b', 80),
        (r'\b([1-9]\d?)\s*(roti|rotis|chapati|chapatis|phulka|phulkas)\b', 110),
        (r'\b([1-9]\d?)\s*(poori|pooris|puri|puris)\b', 120),
        (r'\b([1-9]\d?)\s*(momo|momos)\b', 45),
    ]
    estimate = None
    for pattern, kcal in counted_rules:
        match = re.search(pattern, text)
        if match:
            estimate = int(match.group(1)) * kcal
            break

    rules = [
        (r'\b(mutton|lamb)\s+biryani\b', 850),
        (r'\bchicken\s+biryani\b', 760),
        (r'\bveg(etable)?\s+biryani\b', 650),
        (r'\bbiryani\b', 720),
        (r'\bfried\s+rice\b', 620),
        (r'\b(noodles|chowmein|chow mein)\b', 590),
        (r'\bpasta\b', 620),
        (r'\b(rice\s+bowl|bowl)\b', 550),
        (r'\b(thali|full\s+meal)\b', 850),
        (r'\b(combo|platter)\b', 680),
        (r'\bpizza\b', 700),
        (r'\bburger\b', 560),
        (r'\b(sandwich|toastie)\b', 330),
        (r'\b(wrap|roll|kathi)\b', 430),
        (r'\bdosa\b', 190),
        (r'\b(uttapam|uthappam)\b', 300),
        (r'\b(upma|poha)\b', 280),
        (r'\b(paratha|parotta)\b', 330),
        (r'\b(roti|rotis|chapati|chapatis|phulka|phulkas)\b', 90),
        (r'\bsamosa\b', 260),
        (r'\b(vada|vadai)\b', 180),
        (r'\b(pakora|pakoda|bhaji)\b', 350),
        (r'\bpaneer\b', 450),
        (r'\bchicken\b', 460),
        (r'\bfish\b', 410),
        (r'\b(dal|daal)\b', 240),
        (r'\b(chole|chana|channa|rajma)\b', 360),
        (r'\b(dahi|curd|yogurt|yoghurt|raita)\b', 120),
        (r'\b(jeera\s+rice|plain\s+rice|steamed\s+rice|basmati\s+rice|rice)\b', 260),
        (r'\b(soup|shorba)\b', 140),
        (r'\bsalad\b', 180),
        (r'\b(cake|pastry|brownie|muffin)\b', 450),
        (r'\b(ice\s*cream|kulfi)\b', 260),
        (r'\b(gulab\s+jamun|rasgulla|sweet|dessert)\b', 320),
        (r'\b(lassi|smoothie|shake)\b', 240),
        (r'\bjuice\b', 140),
        (r'\b(tea|chai|coffee)\b', 90),
    ]
    for pattern, kcal in rules:
        if re.search(pattern, text):
            estimate = max(estimate, kcal) if estimate else kcal
            break

    if not estimate:
        return None

    if re.search(r'\b(fried|crispy|cheesy|cheese|butter|cream|malai|mayo)\b', text):
        estimate += 100
    if re.search(r'\b(with\s+rice|rice\s+and|fries|extra|loaded)\b', text):
        estimate += 120
    if re.search(r'\b(gravy|masala|curry)\b', text):
        estimate += 60
    if re.search(r'\b(grilled|steamed|roasted|baked|tandoori)\b', text):
        estimate -= 50
    if re.search(r'\b(small|mini|half)\b', text):
        estimate -= 80
    if re.search(r'\b(large|jumbo|double|full)\b', text):
        estimate += 130

    return _clamp_estimated_calories(estimate)


def _parse_calories_value(value, name='', description=''):
    raw = '' if value is None else str(value).strip()
    if not raw:
        return _estimate_product_calories(name, description)
    if raw.isdigit():
        return max(0, int(raw))
    try:
        decimal_value = Decimal(raw)
        if decimal_value >= 0:
            return int(decimal_value)
    except (InvalidOperation, ValueError):
        pass
    explicit = re.search(r'\b(\d{1,4})\s*(kcal|calories|cal)\b', raw, flags=re.IGNORECASE)
    if explicit:
        return max(0, min(5000, int(explicit.group(1))))
    raise ValueError(f'Invalid calories: {value}')


def _parse_date_value(value):
    from datetime import date, datetime
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    parsed = parse_date(str(value).strip())
    if parsed:
        return parsed
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(f'Invalid date: {value}')


def _parse_time_value(value):
    from datetime import datetime, time
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    if isinstance(value, time):
        return value.replace(microsecond=0)
    parsed = parse_time(str(value).strip())
    if parsed:
        return parsed
    for fmt in ('%I:%M %p', '%I %p', '%H:%M'):
        try:
            return datetime.strptime(str(value).strip(), fmt).time()
        except ValueError:
            continue
    raise ValueError(f'Invalid time: {value}')


def _company_name_map(companies):
    return {c.name.strip().lower(): c for c in companies}


def _resolve_company_from_row(user, selected_company, row, allowed_companies_by_name):
    if selected_company:
        return selected_company
    if not user.is_superadmin:
        return get_primary_staff_company(user)

    company_value = row.get('company') or row.get('company_name') or row.get('company_id')
    if company_value in (None, ''):
        return None
    raw = str(company_value).strip()
    if raw.isdigit():
        return Company.objects.filter(pk=int(raw), is_active=True, is_deleted=False).first()
    return allowed_companies_by_name.get(raw.lower()) or Company.objects.filter(name__iexact=raw, is_active=True, is_deleted=False).first()


def _resolve_category_for_company(company, row, allow_create=False, default_active=True):
    category_id = str(row.get('category_id') or '').strip()
    category_slug = str(row.get('category_slug') or '').strip()
    category_name = str(row.get('category') or row.get('category_name') or '').strip()

    qs = Category.objects.filter(is_deleted=False)
    category = None
    if category_id.isdigit():
        category = qs.filter(pk=int(category_id)).first()
    if not category and category_slug:
        category = qs.filter(slug__iexact=slugify(category_slug)).first()
    if not category and category_name:
        category = qs.filter(name__iexact=category_name).filter(Q(companies=company) | Q(companies__isnull=True)).distinct().first()
        if not category:
            category = qs.filter(name__iexact=category_name).first()

    created = False
    if not category and allow_create and category_name:
        category = Category.objects.create(
            name=category_name,
            slug=slugify(category_name),
            is_active=_parse_bool(row.get('category_active'), default_active),
            is_deleted=False,
        )
        created = True

    if category and company:
        if not category.companies.filter(pk=company.pk).exists():
            category.companies.add(company)
        CategoryCompanyStatus.objects.get_or_create(
            category=category,
            company=company,
            defaults={'is_active': _parse_bool(row.get('category_active'), category.is_active)},
        )

    return category, created


def _resolve_offering_for_company(company, row, allow_create=False, default_active=True):
    offering_id = str(row.get('offering_id') or '').strip()
    offering_name = str(row.get('offering') or row.get('offering_name') or '').strip()
    offering = None
    if offering_id.isdigit():
        offering = Offering.objects.filter(pk=int(offering_id), company=company, is_deleted=False).first()
    if not offering and offering_name:
        offering = Offering.objects.filter(company=company, name__iexact=offering_name, is_deleted=False).first()
    created = False
    if not offering and allow_create and offering_name:
        offering = Offering.objects.create(
            company=company,
            name=offering_name,
            slug=slugify(offering_name),
            is_active=_parse_bool(row.get('offering_active'), default_active),
            is_deleted=False,
        )
        created = True
    return offering, created


def _resolve_product_for_import(company, row):
    product_code = str(row.get('product_code') or row.get('code') or '').strip()
    product_name = str(row.get('product_name') or row.get('name') or '').strip()
    qs = Product.objects.filter(company=company)
    if product_code:
        product = qs.filter(code__iexact=product_code).first()
        if product:
            return product
    product_slug = _product_slug_from_import_row(row, name=product_name)
    if product_slug:
        product = qs.filter(slug__iexact=product_slug).order_by('id').first()
        if product:
            return product
    if product_name:
        return qs.filter(name__iexact=product_name).order_by('id').first()
    return None


def _resolve_building_for_company(company, row):
    building_value = str(row.get('building') or row.get('building_name') or '').strip()
    if not building_value:
        return None
    qs = Building.objects.filter(company=company, is_deleted=False)
    state_value = str(row.get('state') or '').strip()
    city_value = str(row.get('city') or '').strip()
    if state_value:
        qs = qs.filter(state__name__iexact=state_value)
    if city_value:
        qs = qs.filter(city__name__iexact=city_value)
    building = qs.filter(name__iexact=building_value).first()
    if not building:
        building = Building.objects.filter(company=company, name__iexact=building_value, is_deleted=False).first()
    return building


def _resolve_cafe_for_company(company, row, building=None):
    cafe_value = str(row.get('cafe') or row.get('cafe_name') or '').strip()
    if not cafe_value:
        return None
    qs = Cafe.objects.filter(company=company, is_deleted=False)
    if building:
        qs = qs.filter(Q(building=building) | Q(building__isnull=True))
    cafe = qs.filter(name__iexact=cafe_value).first()
    if not cafe:
        cafe = Cafe.objects.filter(company=company, name__iexact=cafe_value, is_deleted=False).first()
    return cafe


def _resolve_counter_for_company(company, row, cafe=None):
    counter_value = str(row.get('counter') or row.get('counter_name') or '').strip()
    if not counter_value:
        return None
    qs = Counter.objects.filter(company=company, is_deleted=False)
    if cafe:
        qs = qs.filter(Q(cafe=cafe) | Q(cafe__isnull=True))
    counter = qs.filter(name__iexact=counter_value).order_by('position_order', 'id').first()
    if not counter:
        counter = Counter.objects.filter(company=company, name__iexact=counter_value, is_deleted=False).order_by('position_order', 'id').first()
    return counter


def _validate_site_chain(company, building=None, cafe=None, counter=None, state_name='', city_name=''):
    if building and building.company_id != company.id:
        raise ValueError('Selected building does not belong to the selected company.')
    if cafe and cafe.company_id != company.id:
        raise ValueError('Selected cafe does not belong to the selected company.')
    if cafe and building and cafe.building_id and cafe.building_id != building.id:
        raise ValueError('Selected cafe does not belong to the selected building.')
    if counter and counter.company_id != company.id:
        raise ValueError('Selected counter does not belong to the selected company.')
    if counter and cafe and counter.cafe_id and counter.cafe_id != cafe.id:
        raise ValueError('Selected counter does not belong to the selected cafe.')
    if state_name and building and building.state_id and building.state.name.lower() != state_name.lower():
        raise ValueError('State does not match the selected building.')
    if city_name and building and building.city_id and building.city.name.lower() != city_name.lower():
        raise ValueError('City does not match the selected building.')


def _set_product_food_types(product, veg_type_value):
    food_type_ids = _canonical_food_type_ids_from_value(veg_type_value)
    if not food_type_ids:
        return
    product.food_type.set(FoodType.objects.filter(pk__in=food_type_ids))


def _canonical_food_type_ids_from_value(value):
    raw = str(value or '').strip()
    if not raw:
        return []
    parts = [part.strip() for part in raw.replace('|', ',').replace('/', ',').split(',') if part.strip()]
    kinds = set()
    for part in parts:
        normalized = part.lower().replace('-', '').replace('_', '').replace(' ', '')
        if normalized in {'veg', 'vegetarian'}:
            kinds.add('veg')
        elif normalized in {'nonveg', 'nonvegetarian', 'nonvegitarian', 'nonvegeterian'} or 'nonveg' in normalized:
            kinds.add('nonveg')
        else:
            raise ValueError('Food type must be Veg or Non-Veg.')
    if len(kinds) > 1:
        raise ValueError('Choose only one food type: Veg or Non-Veg.')
    wanted = next(iter(kinds), '')
    for option in _canonical_food_type_options():
        if option['kind'] == wanted:
            return [option['pk']]
    return []


def _row_has_nonblank(row, *keys):
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        if str(value).strip():
            return True
    return False


def _coalesce_row_value(row, *keys):
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        if str(value).strip():
            return value
    return ''


def _build_sample_product_workbook_bytes():
    """
    Build the sample bulk-upload Excel workbook.

    Hierarchy: Offering → Category → Product
    No subcategory field — it has been removed from the active flow.

    Sheets:
      Products       — master product records (required)
      CounterMapping — product→counter mappings (optional)
      ReadMe         — field reference guide
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # ── Sheet 1: Products ─────────────────────────────────────────
    ws = wb.active
    ws.title = 'Products'

    # Columns aligned as closely as possible to the live product form.
    # Legacy aliases such as veg_type / is_featured remain supported by parser.
    product_headers = [
        # Identity
        'company',           # required — must match an existing company name
        'product_code',      # optional — unique code/SKU; used to find existing product
        'product_name',      # required
        'description',       # optional

        # Hierarchy — Offering → Category → Product (NO subcategory)
        'offering',          # optional — name of offering; auto-created if missing
        'offering_active',   # Yes/No — only used when auto-creating
        'category',          # required — name of category; auto-created if missing
        'category_active',   # Yes/No — only used when auto-creating

        # Food type
        'food_types',        # Veg / Non-Veg — matches the product form
        'counters',          # optional — comma/pipe-separated pickup counter names; first is default

        # Pricing
        'Staff/Base Price (Rs)',     # required — staff/base selling price
        'Visitor Price (Rs)',        # optional — POS visitor rate
        'Room Service Extra (%)',    # optional — % added to visitor price for room service
        'Packing Price (Rs)',        # optional — added at checkout

        # Stock
        'min_qty',           # default 1
        'max_qty',           # default 10
        'web_qty',           # -1 = unlimited; 0 = out of stock; N = capped
        'pos_qty',           # POS stock (default 0)

        # Display
        'position_order',    # sort order (default 0)
        'is_active',         # Yes/No (default Yes)
        'is_kiosk_active',   # Yes/No (default Yes) — kiosk visibility
        'featured_in_web',   # Yes/No (default No) — pin on web home
        'featured_in_kiosk_extra',  # Yes/No (default No) — kiosk featured section
        'is_free_meal_product',     # Yes/No (default No)

        # Schedule (all optional — leave blank for always-available)
        'menu_date',         # YYYY-MM-DD — restrict to this single date
        'schedule_enabled',  # Yes/No — set Yes to enable date/time window
        'start_date',        # YYYY-MM-DD
        'end_date',          # YYYY-MM-DD
        'start_time',        # HH:MM  (24h)
        'end_time',          # HH:MM  (24h)

        # Other
        'preparation_time_minutes',   # default 10
        'calories',                   # optional kcal count
        'schedule_bypass',            # Yes/No — superadmin import only
    ]
    ws.append(product_headers)

    # Sample row 1 — scheduled item
    ws.append([
        'Acme Foods', 'VT-001', 'Veg Thali', 'Daily lunch combo',
        'Lunch', 'Yes', 'Meals', 'Yes',
        'Veg', 'Counter 1',
        '70.00', '85.00', '10.00', '5.00',
        '1', '25', '-1', '20',
        '1', 'Yes',
        'Yes', 'Yes', 'No', 'No',
        '', 'Yes', '2026-01-01', '2026-12-31', '11:30', '15:30',
        '12', '450', 'No',
    ])
    # Sample row 2 — always-available item
    ws.append([
        'Acme Foods', 'CR-001', 'Chicken Roll', 'Popular snack',
        'Snacks', 'Yes', 'Rolls', 'Yes',
        'Non-Veg', 'Snacks Counter',
        '65.00', '0.00', '0.00', '0.00',
        '1', '30', '50', '20',
        '2', 'Yes',
        'Yes', 'No', 'No', 'No',
        '', 'No', '', '', '', '',
        '8', '320', 'No',
    ])

    # Style header row
    header_fill = PatternFill('solid', fgColor='0D2137')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # ── Sheet 2: CounterMapping ───────────────────────────────────
    ws3 = wb.create_sheet('CounterMapping')
    counter_headers = [
        'company', 'product_code',
        'state', 'city', 'building', 'cafe',
        'counter', 'is_primary', 'is_active', 'position_order',
    ]
    ws3.append(counter_headers)
    ws3.append(['Acme Foods', 'VT-001', 'West Bengal', 'Kolkata', 'HQ Tower', 'Main Cafe', 'Counter 1', 'Yes', 'Yes', '0'])
    ws3.append(['Acme Foods', 'CR-001', 'West Bengal', 'Kolkata', 'HQ Tower', 'Main Cafe', 'Snacks Counter', 'Yes', 'Yes', '0'])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # ── Sheet 3: ReadMe ───────────────────────────────────────────
    ws4 = wb.create_sheet('ReadMe')
    readme_rows = [
        ['FIELD', 'SHEET', 'REQUIRED?', 'ACCEPTED VALUES / FORMAT', 'NOTES'],
        # Products
        ['company',          'Products', 'Yes', 'Exact company name', 'Must already exist in NeverQ'],
        ['product_code',     'Products', 'No',  'Any text (e.g. VT-001)', 'Used to find/update existing product'],
        ['product_name',     'Products', 'Yes', 'Any text', ''],
        ['description',      'Products', 'No',  'Any text', ''],
        ['offering',         'Products', 'No',  'Offering name', 'Auto-created if Enable Offering Creation is on'],
        ['offering_active',  'Products', 'No',  'Yes / No', 'Only used when auto-creating'],
        ['category',         'Products', 'Yes', 'Category name', 'Auto-created if Enable Category Creation is on'],
        ['category_active',  'Products', 'No',  'Yes / No', 'Only used when auto-creating'],
        ['food_types',       'Products', 'No',  'Veg / Non-Veg', 'Matches the two food type choices on the product form'],
        ['counters',         'Products', 'No',  'Counter names separated by comma or |', 'Pickup counters from the product form; ignored if CounterMapping sheet is present'],
        ['Staff/Base Price (Rs)', 'Products', 'Yes', 'Number (e.g. 70.00)', 'Shown on web, kiosk, and POS staff sales'],
        ['Visitor Price (Rs)',    'Products', 'No',  'Number', 'Used for POS visitor sales; leave 0 to hide Visitor in POS'],
        ['Room Service Extra (%)', 'Products', 'No', 'Number', 'Percentage added to room service orders; leave 0 to hide Room Svc in POS'],
        ['Packing Price (Rs)',    'Products', 'No',  'Number', 'Added on top at checkout'],
        ['min_qty',          'Products', 'No',  'Integer ≥ 1', 'Default: 1'],
        ['max_qty',          'Products', 'No',  'Integer ≥ 1', 'Default: 10'],
        ['web_qty',          'Products', 'No',  '-1 = unlimited; 0 = out of stock; N = cap', 'Default: -1'],
        ['pos_qty',          'Products', 'No',  'Integer ≥ 0', 'Default: 0'],
        ['position_order',   'Products', 'No',  'Integer', 'Controls display order; default 0'],
        ['is_active',        'Products', 'No',  'Yes / No', 'Default: Yes'],
        ['is_kiosk_active',  'Products', 'No',  'Yes / No', 'Default: Yes — controls kiosk visibility independently'],
        ['featured_in_web',  'Products', 'No',  'Yes / No', 'Default: No — pin on customer web portal home'],
        ['featured_in_kiosk_extra', 'Products', 'No', 'Yes / No', 'Default: No — show in kiosk Featured section (max 10)'],
        ['is_free_meal_product', 'Products', 'No', 'Yes / No', 'Maps to the free meal toggle on the product form'],
        ['menu_date',        'Products', 'No',  'YYYY-MM-DD', 'Restrict product to this exact date only'],
        ['schedule_enabled', 'Products', 'No',  'Yes / No', 'Set Yes to restrict by date/time'],
        ['start_date',       'Products', 'No',  'YYYY-MM-DD', 'Only used when schedule_enabled=Yes'],
        ['end_date',         'Products', 'No',  'YYYY-MM-DD', 'Only used when schedule_enabled=Yes'],
        ['start_time',       'Products', 'No',  'HH:MM (24h)', 'Only used when schedule_enabled=Yes'],
        ['end_time',         'Products', 'No',  'HH:MM (24h)', 'Only used when schedule_enabled=Yes'],
        ['preparation_time_minutes', 'Products', 'No', 'Integer', 'Default: 10'],
        ['calories',         'Products', 'No',  'Integer (kcal)', 'Leave blank to auto-estimate from product name + description'],
        ['schedule_bypass',  'Products', 'No',  'Yes / No', 'Superadmin import only — ignored for non-superadmin uploads'],
        # Hierarchy note
        ['IMPORTANT', '', '', '', 'Hierarchy is Offering → Category → Product. There is NO subcategory in this system.'],
        ['LEGACY', '', '', '', 'Legacy aliases such as base_price, company_price, room_service_extra_percent, packing_price, veg_type, and is_featured are still accepted.'],
    ]
    for row in readme_rows:
        ws4.append(row)
    for cell in ws4[1]:
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.fill = header_fill

    # Auto-width all sheets
    for worksheet in wb.worksheets:
        worksheet.freeze_panes = 'A2'
        for col_cells in worksheet.columns:
            max_len = max((len(str(c.value or '')) for c in col_cells), default=10)
            worksheet.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 32)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


@customer_login_required
def menu(request):
    """Entry point for customer menu — delegates to customer_menu."""
    return customer_menu(request)


def _get_own_product(product_id, customer):
    """Fetch product ONLY if it belongs to the customer's company. Returns None if foreign."""
    try:
        return Product.objects.select_related('category', 'company').prefetch_related('food_type').get(
            pk=int(product_id),
            is_deleted=False,
            company=customer.company,
        )
    except (Product.DoesNotExist, TypeError, ValueError):
        return None


def _get_live_offer_for_product_menu(company, product, cafe=None):
    """
    Return the best live Offer for a product (menu-side helper, kept in sync with orders.views version).
    Priority: cafe-scoped single-product → cafe-scoped multi-product (M2M) →
              global single-product → global multi-product (M2M) →
              cafe-scoped site-wide (no products at all) → global site-wide (no products at all).

    Offers with M2M products set are product-specific and must ONLY match the
    products explicitly listed in them — they are never treated as site-wide.
    """
    from apps.menu.models import Offer

    # 1. cafe-scoped single-product offer
    if cafe:
        for offer in Offer.objects.filter(company=company, cafe=cafe, is_active=True, is_deleted=False, product=product).order_by('-created_at'):
            if offer.is_live:
                return offer

    # 2. cafe-scoped multi-product (M2M) offer
    if cafe:
        for offer in Offer.objects.filter(
            company=company, cafe=cafe, is_active=True, is_deleted=False,
            product__isnull=True, products=product
        ).order_by('-created_at'):
            if offer.is_live:
                return offer

    # 3. global single-product offer
    for offer in Offer.objects.filter(company=company, is_active=True, is_deleted=False, product=product, cafe__isnull=True).order_by('-created_at'):
        if offer.is_live:
            return offer

    # 4. global multi-product (M2M) offer
    for offer in Offer.objects.filter(
        company=company, is_active=True, is_deleted=False,
        product__isnull=True, products=product, cafe__isnull=True
    ).order_by('-created_at'):
        if offer.is_live:
            return offer

    # 5. cafe-scoped site-wide offer (no single-product FK, no M2M products at all)
    if cafe:
        for offer in Offer.objects.filter(
            company=company, cafe=cafe, is_active=True, is_deleted=False,
            product__isnull=True, products__isnull=True
        ).order_by('-created_at'):
            if offer.is_live:
                return offer

    # 6. global site-wide offer (no single-product FK, no M2M products at all)
    for offer in Offer.objects.filter(
        company=company, is_active=True, is_deleted=False,
        product__isnull=True, cafe__isnull=True, products__isnull=True
    ).order_by('-created_at'):
        if offer.is_live:
            return offer

    return None


def _preload_cart_products_and_offers_menu(customer, cart, cafe=None):
    from apps.menu.models import Offer

    product_ids = []
    for raw_product_id in (cart or {}).keys():
        try:
            product_ids.append(int(raw_product_id))
        except (TypeError, ValueError):
            continue

    if not product_ids:
        return {}, {}

    company = customer.company
    products = list(
        Product.objects.select_related('category', 'company')
        .prefetch_related('food_type')
        .filter(pk__in=product_ids, is_deleted=False, company=company)
    )
    product_map = {product.pk: product for product in products}
    if not product_map:
        return {}, {}

    offer_qs = Offer.objects.filter(
        company=company,
        is_active=True,
        is_deleted=False,
    )
    if cafe:
        offer_qs = offer_qs.filter(Q(cafe=cafe) | Q(cafe__isnull=True))
    else:
        offer_qs = offer_qs.filter(cafe__isnull=True)

    offers = list(
        offer_qs.select_related('product', 'cafe')
        .prefetch_related('products')
        .order_by('-created_at')
    )

    cafe_id = getattr(cafe, 'pk', None)
    all_product_ids = list(product_map.keys())
    best_offer_by_product = {}
    best_priority_by_product = {}

    for offer in offers:
        if not offer.is_live:
            continue

        m2m_product_ids = [p.pk for p in offer.products.all()]

        if cafe_id and offer.cafe_id == cafe_id:
            if offer.product_id is not None:
                priority = 1
                target_product_ids = [offer.product_id]
            elif m2m_product_ids:
                priority = 2
                target_product_ids = m2m_product_ids
            else:
                priority = 5
                target_product_ids = all_product_ids
        elif offer.cafe_id is None:
            if offer.product_id is not None:
                priority = 3
                target_product_ids = [offer.product_id]
            elif m2m_product_ids:
                priority = 4
                target_product_ids = m2m_product_ids
            else:
                priority = 6
                target_product_ids = all_product_ids
        else:
            continue

        for target_product_id in target_product_ids:
            if target_product_id not in product_map:
                continue
            current_priority = best_priority_by_product.get(target_product_id)
            if current_priority is None or priority < current_priority:
                best_priority_by_product[target_product_id] = priority
                best_offer_by_product[target_product_id] = offer

    return product_map, best_offer_by_product


def _get_live_cart_offer(company, cafe=None):
    """Return a live cart-level (TYPE_CART) offer for the company, or None."""
    qs = Offer.objects.filter(
        company=company, is_active=True, is_deleted=False, offer_type=Offer.TYPE_CART
    ).order_by('-created_at')
    if cafe:
        for offer in qs.filter(cafe=cafe):
            if offer.is_live:
                return offer
    for offer in qs.filter(cafe__isnull=True):
        if offer.is_live:
            return offer
    return None


def _resolve_customer_cafe(customer, company, request=None):
    """
    Resolve the cafe for a web customer in this order of priority:
      1. explicit customer.cafe
      2. session-selected web cafe (web_cafe_id), if valid
      3. first active cafe for the customer's building
      4. None
    """
    cafe = getattr(customer, 'cafe', None)
    if cafe is not None:
        return cafe

    building = getattr(customer, 'building', None)

    selected_cafe_id = None
    if request is not None:
        selected_cafe_id = request.session.get('web_cafe_id')

    if selected_cafe_id:
        selected = Cafe.objects.filter(
            pk=selected_cafe_id,
            company=company,
            is_active=True,
            is_deleted=False,
        ).first()
        if selected:
            if building is None or selected.building_id == getattr(building, 'pk', None):
                return selected
        else:
            request.session.pop('web_cafe_id', None)

    if building is not None:
        return Cafe.objects.filter(
            building=building,
            company=company,
            is_active=True,
            is_deleted=False,
        ).first()

    return None


def _apply_offer_to_line_menu(offer, unit_price, qty, cart_subtotal=None):
    """
    Return (effective_line_total, offer_saving).
    Kept in sync with orders.views._apply_offer_to_line.
    cart_subtotal is the gross cart value at the time of evaluation, used to check
    min_order_value thresholds. Pass it so FLAT/PERCENT offers with minimums
    are correctly applied or withheld.
    """
    if offer is None:
        return unit_price * qty, Decimal('0.00')
    gross = unit_price * qty
    # Min order threshold check
    if offer.min_order_value and cart_subtotal is not None:
        if cart_subtotal < offer.min_order_value:
            return gross, Decimal('0.00')
    if offer.offer_type == offer.TYPE_FREE:
        return Decimal('0.00'), gross
    if offer.offer_type == offer.TYPE_BOGO:
        free_units = qty // 2
        saving = unit_price * free_units
        return max(Decimal('0.00'), gross - saving), saving
    if offer.offer_type == offer.TYPE_PERCENT:
        rate = min(Decimal('100'), max(Decimal('0'), offer.value))
        saving = (gross * rate / Decimal('100')).quantize(Decimal('0.01'))
        if offer.max_discount:
            saving = min(saving, offer.max_discount)
        return max(Decimal('0.00'), gross - saving), saving
    return gross, Decimal('0.00')


def _build_cart_summary(customer, cart, benefit_date=None, request=None):
    """
    Build cart totals with correct site-wise pricing (cafe-aware) and offer discounts.
    Keep this aligned with orders.views._build_cart_summary.
    """
    company = _fresh_company(customer)
    building = getattr(customer, 'building', None)
    cafe = _resolve_customer_cafe(customer, company, request=request)
    cart = cart or {}
    items = []
    subtotal = Decimal('0.00')
    offer_discount = Decimal('0.00')
    eligible_product_ids = set(company.free_meal_products.filter(is_deleted=False).values_list('pk', flat=True))
    eligible_subtotal = Decimal('0.00')

    from .models import OfferUsage as _OfferUsage
    _today = timezone.localdate()
    _used_offer_ids = set(
        _OfferUsage.objects.filter(customer=customer, used_on=_today)
        .values_list('offer_id', flat=True)
    )
    ONE_USE_TYPES = {Offer.TYPE_BOGO, Offer.TYPE_FREE, Offer.TYPE_PERCENT, Offer.TYPE_FLAT, Offer.TYPE_CART}
    _applied_one_use_offer_ids = set()

    _cart_products, _live_offers = _preload_cart_products_and_offers_menu(customer, cart, cafe=cafe)

    for pid, item in list(cart.items()):
        try:
            product_pk = int(pid)
        except (TypeError, ValueError):
            cart.pop(pid, None)
            continue

        product = _cart_products.get(product_pk)
        if not product:
            cart.pop(pid, None)
            continue

        try:
            qty = max(0, int(item.get('qty', 0)))
        except (TypeError, ValueError):
            qty = 0

        if qty <= 0:
            cart.pop(pid, None)
            continue

        min_qty = _get_order_min_qty(product)
        max_qty = _get_web_max_qty(product)
        if max_qty <= 0:
            cart.pop(pid, None)
            continue
        qty = max(min_qty, min(qty, max_qty))
        item['qty'] = qty

        site_price = _resolve_site_price(product, company, building=building, cafe=cafe)
        live_offer = _live_offers.get(product.pk)

        if live_offer and live_offer.pk in _used_offer_ids:
            live_offer = None
        if live_offer and live_offer.offer_type in ONE_USE_TYPES and live_offer.pk in _applied_one_use_offer_ids:
            live_offer = None

        line_total, line_saving = _apply_offer_to_line_menu(
            live_offer,
            site_price,
            qty,
            cart_subtotal=subtotal + site_price * qty,
        )

        product.display_price = site_price
        product.effective_web_qty = _get_web_max_qty(product)
        is_free_meal_eligible = not eligible_product_ids or product.pk in eligible_product_ids

        subtotal += site_price * qty
        offer_discount += line_saving

        if is_free_meal_eligible:
            eligible_subtotal += line_total

        if live_offer and line_saving > 0 and live_offer.offer_type in ONE_USE_TYPES:
            _applied_one_use_offer_ids.add(live_offer.pk)

        items.append({
            'product': product,
            'qty': qty,
            'site_price': site_price,
            'line_total': line_total,
            'gross_line_total': site_price * qty,
            'line_saving': line_saving,
            'offer': live_offer,
            'is_available_now': product.is_available_now(),
            'is_free_meal_eligible': is_free_meal_eligible,
        })

    packing = sum((_to_decimal(item['product'].packing_price) * item['qty']) for item in items)

    # Cart-level FLAT / CART offers
    cart_level_offer = None
    cart_offer_saving = Decimal('0.00')
    for offer_type in (Offer.TYPE_FLAT, Offer.TYPE_CART):
        _qs = Offer.objects.filter(
            company=company, is_active=True, is_deleted=False, offer_type=offer_type
        ).order_by('-created_at')

        if cafe:
            for o in _qs.filter(cafe=cafe):
                if o.is_live and o.pk not in _used_offer_ids:
                    cart_level_offer = o
                    break

        if not cart_level_offer:
            for o in _qs.filter(cafe__isnull=True):
                if o.is_live and o.pk not in _used_offer_ids:
                    cart_level_offer = o
                    break

        if cart_level_offer:
            break

    if cart_level_offer:
        min_order = cart_level_offer.min_order_value or Decimal('0')
        if subtotal >= min_order:
            if cart_level_offer.offer_type == Offer.TYPE_FLAT:
                cart_offer_saving = min(cart_level_offer.value, subtotal)
            elif cart_level_offer.offer_type == Offer.TYPE_CART:
                rate = min(Decimal('100'), max(Decimal('0'), cart_level_offer.value))
                cart_offer_saving = (subtotal * rate / Decimal('100')).quantize(Decimal('0.01'))
                if cart_level_offer.max_discount:
                    cart_offer_saving = min(cart_offer_saving, cart_level_offer.max_discount)
        else:
            cart_level_offer = None

    offer_discount += cart_offer_saving
    after_offer_subtotal = subtotal - offer_discount
    gross_total = after_offer_subtotal + packing

    if not eligible_product_ids:
        eligible_subtotal = after_offer_subtotal

    benefit_date = benefit_date or timezone.localdate()
    subsidy = customer.company_cover_for_amount(eligible_subtotal, benefit_date)
    my_pay = max(Decimal('0.00'), gross_total - subsidy)
    bill_to_company = subsidy
    company_cover_label = (
        'Company-paid meal'
        if customer.meal_benefit == customer.MEAL_BENEFIT_COMPANY_PAY and subsidy > 0
        else 'Company subsidy'
    )
    benefit_used_today = (
        customer.meal_benefit in (customer.MEAL_BENEFIT_COMPANY_PAY, customer.MEAL_BENEFIT_SUBSIDY) and
        customer.benefit_limit_for_date(benefit_date) > 0 and
        customer.benefit_used_on(benefit_date)
    )

    return {
    'cart': cart,
    'items': items,
    'cart_map': _cart_quantity_map(cart),
    'subtotal': subtotal,
    'offer_discount': offer_discount,
    'after_offer_subtotal': after_offer_subtotal,
    'cart_level_offer': cart_level_offer,
    'cart_offer_saving': cart_offer_saving,
    'eligible_subtotal': eligible_subtotal,
    'packing': packing,
    'subsidy': subsidy,
    'my_pay': my_pay,
    'bill_to_company': bill_to_company,
    'company_cover_label': company_cover_label,
    'benefit_used_today': benefit_used_today,
    'total': gross_total,
    'cart_count': sum(item['qty'] for item in items),
}


def _get_cart_unlock_nudge(company, subtotal):
    subtotal = _to_decimal(subtotal)
    for offer_type in (Offer.TYPE_FLAT, Offer.TYPE_CART):
        for offer in Offer.objects.filter(
            company=company, is_active=True, is_deleted=False, offer_type=offer_type
        ).order_by('min_order_value', '-created_at'):
            if not offer.is_live or not offer.min_order_value:
                continue
            remaining_value = max(Decimal('0.00'), _to_decimal(offer.min_order_value) - subtotal)
            return {
                'offer': offer,
                'title': offer.title,
                'min_order_value': _to_decimal(offer.min_order_value),
                'remaining_value': remaining_value,
            }
    return None


def _mark_free_meal_products(products, company):
    eligible_ids = set(company.free_meal_products.filter(is_deleted=False).values_list('pk', flat=True)) if company else set()
    for product in products:
        product.is_free_meal_eligible = product.pk in eligible_ids
    return products


@never_cache
@customer_login_required
def customer_menu(request):
    customer = request.current_customer
    company = _fresh_company(customer)
    building = getattr(customer, 'building', None)
    cart = request.session.get('cart', {})
    cart_summary = _build_cart_summary(customer, cart, request=request)
    request.session['cart'] = cart_summary['cart']
    request.session.modified = True

    q = request.GET.get('q', '').strip()
    food_pref = (request.GET.get('food') or '').strip().lower()
    offering_filter = (request.GET.get('offering') or '').strip()
    calorie_max_raw = (request.GET.get('calorie_max') or '').strip()
    calorie_max = None
    try:
        calorie_max = int(calorie_max_raw) if calorie_max_raw else None
    except (ValueError, TypeError):
        calorie_max = None
    categories = [c for c in Category.objects.filter(
        companies=company,
        is_deleted=False,
        parent__isnull=True,
    ).prefetch_related('schedules', 'company_statuses').order_by('position_order', 'name') if c.is_active_now(company)]

    ad_qs = Advertise.objects.filter(
        is_active=True,
        status=Advertise.STATUS_APPROVED,
    ).filter(companies=company).distinct().prefetch_related('holiday_schedules').order_by('position_order')
    adverts = [ad for ad in ad_qs if ad.is_live]

    live_offering_qs = Offering.objects.filter(company=company, is_deleted=False, is_active=True).order_by('position_order', 'name')
    offerings = [off for off in live_offering_qs if off.is_active_now()]
    selected_offering = None
    if offering_filter:
        selected_offering = next((off for off in offerings if str(off.pk) == offering_filter or off.slug == offering_filter), None)

    offer_qs = Offer.objects.filter(company=company, is_deleted=False).select_related('product', 'cafe').prefetch_related('products').order_by('-created_at')
    _all_live = [offer for offer in offer_qs if offer.is_live]
    # Hide one-use offers that this customer has already used today.
    # Keep the menu display aligned with checkout / OfferUsage enforcement.
    from .models import OfferUsage
    _ONE_USE = {
        Offer.TYPE_BOGO,
        Offer.TYPE_FREE,
        Offer.TYPE_PERCENT,
        Offer.TYPE_FLAT,
        Offer.TYPE_CART,
    }
    _used_ids = set(
        OfferUsage.objects.filter(customer=customer, offer__in=_all_live, used_on=timezone.localdate())
        .values_list('offer_id', flat=True)
    )
    live_offers = [
        o for o in _all_live
        if o.offer_type not in _ONE_USE or o.pk not in _used_ids
    ]

    featured_qs = Product.objects.filter(
        company=company,
        is_active=True,
        is_deleted=False,
    ).select_related('category', 'offering').prefetch_related('food_type', 'category__company_statuses').annotate(
        order_count=Count('orderitem')
    ).order_by('-rating', '-order_count', 'category__position_order', 'category__name', 'position_order', 'name')
    if q:
        featured_qs = featured_qs.filter(
            Q(name__icontains=q)
            | Q(description__icontains=q)
            | Q(code__icontains=q)
            | Q(category__name__icontains=q)
            | Q(offering__name__icontains=q)
        )
    if selected_offering:
        featured_qs = featured_qs.filter(offering=selected_offering)
    pinned_featured_products = []
    fallback_featured_products = []
    featured_products = []
    visible_products = []
    visible_qs = Product.objects.filter(
        company=company,
        is_active=True,
        is_deleted=False,
    ).select_related('category', 'offering').prefetch_related('food_type', 'category__company_statuses').order_by(
        'category__position_order', 'category__name', 'position_order', 'name'
    )
    if q:
        visible_qs = visible_qs.filter(
            Q(name__icontains=q)
            | Q(description__icontains=q)
            | Q(code__icontains=q)
            | Q(category__name__icontains=q)
            | Q(offering__name__icontains=q)
        )
    if selected_offering:
        visible_qs = visible_qs.filter(offering=selected_offering)

    for product in featured_qs:
        if len(pinned_featured_products) >= 8 and len(fallback_featured_products) >= 8:
            break
        if _product_is_visible_for_customer(product, company, food_pref):
            if calorie_max is None or (product.calories is not None and product.calories <= calorie_max):
                if product.featured_in_web:
                    pinned_featured_products.append(product)
                if len(fallback_featured_products) < 8:
                    fallback_featured_products.append(product)

    featured_products = pinned_featured_products[:8] if pinned_featured_products else fallback_featured_products

    for product in visible_qs:
        if _product_is_visible_for_customer(product, company, food_pref):
            if calorie_max is None or (product.calories is not None and product.calories <= calorie_max):
                visible_products.append(product)
    # Build used-offer set for display (badge always shown, but discounted price only if not used)
    from .models import OfferUsage as _OUDisplay
    _display_used_ids = set(
        _OUDisplay.objects.filter(customer=customer, used_on=timezone.localdate())
        .values_list('offer_id', flat=True)
    )
    _selected_cafe = _resolve_customer_cafe(customer, company, request=request)
    _attach_display_prices(featured_products, company, building=building, cafe=_selected_cafe, used_offer_ids=_display_used_ids)
    _attach_display_prices(visible_products, company, building=building, cafe=_selected_cafe, used_offer_ids=_display_used_ids)
    _mark_free_meal_products(featured_products, company)
    _mark_free_meal_products(visible_products, company)
    # live_offer is already set by _attach_display_prices — no need to overwrite

    offering_sections = []
    if not selected_offering:
        for offering in offerings:
            section_products = []
            section_qs = Product.objects.filter(
                company=company,
                is_active=True,
                is_deleted=False,
                offering=offering,
            )
            if q:
                section_qs = section_qs.filter(
                    Q(name__icontains=q)
                    | Q(description__icontains=q)
                    | Q(code__icontains=q)
                    | Q(category__name__icontains=q)
                    | Q(offering__name__icontains=q)
                )
            for product in section_qs.select_related('category', 'offering').prefetch_related('food_type', 'category__company_statuses').order_by(
                'category__position_order', 'category__name', 'position_order', 'name'
            ):
                if _product_is_visible_for_customer(product, company, food_pref):
                    if calorie_max is None or (product.calories is not None and product.calories <= calorie_max):  # fixed: removed clause that leaked unclassified products
                        section_products.append(product)
                if len(section_products) >= 4:
                    break
            if section_products:
                _attach_display_prices(section_products, company, building=building, cafe=_selected_cafe, used_offer_ids=_display_used_ids)
                _mark_free_meal_products(section_products, company)
                offering_sections.append((offering, section_products))

    recent_orders = customer.orders.filter(is_deleted=False).prefetch_related(
        'items__product'
    ).order_by('-created_at')[:3]
    live_products_count = Product.objects.filter(company=company, is_active=True, is_deleted=False).count()
    live_offers_count = len(live_offers)
    from apps.menu.models import Cafe as _Cafe
    company_cafes_list = list(_Cafe.objects.filter(company=company, is_active=True, is_deleted=False).order_by('name'))
    company_cafes_count = len(company_cafes_list)
    _home_selected_cafe_id = request.session.get('web_cafe_id')
    # Validate selected cafe still belongs to this company
    if _home_selected_cafe_id and not any(c.pk == _home_selected_cafe_id for c in company_cafes_list):
        request.session.pop('web_cafe_id', None)
        _home_selected_cafe_id = None
    subsidy_reminder = ''
    try:
        benefit_remaining = customer.benefit_remaining_on(timezone.localdate())
    except Exception:
        benefit_remaining = 0
    if benefit_remaining > 0 and getattr(company, 'bill_company', None) == 2 and customer.meal_benefit != customer.MEAL_BENEFIT_NONE:
        benefit_name = 'company-paid meal' if customer.meal_benefit == customer.MEAL_BENEFIT_COMPANY_PAY else 'subsidy benefit'
        subsidy_reminder = (
            f'You still have {benefit_remaining} {benefit_name}{"s" if benefit_remaining != 1 else ""} remaining today. '
            'Please use your available benefit before the day ends.'
        )

    return render(request, 'menu/home.html', {
        'categories': categories,
        'adverts': adverts,
        'live_offers': live_offers,
        'featured_products': featured_products,
        'home_products_preview': visible_products,
        'home_products_preview_initial': 3,
        'offering_sections': offering_sections,
        'offerings': offerings,
        'selected_offering': selected_offering,
        'recent_orders': recent_orders,
        'live_products_count': live_products_count,
        'live_offers_count': live_offers_count,
        'company_cafes': company_cafes_list,
        'company_cafes_count': company_cafes_count,
        'selected_cafe_id': _home_selected_cafe_id,
        'cart_map': cart_summary['cart_map'],
        'cart_count': cart_summary['cart_count'],
        'company': company,
        'customer': customer,
        'search_query': q,
        'food_pref': food_pref,
        'offering_filter': offering_filter,
        'calorie_max': calorie_max,
        'calorie_presets': [200, 400, 600, 800],
        'subsidy_reminder': subsidy_reminder,
        'page_title': 'Menu',
    })


@never_cache
@customer_login_required
def category_detail(request, slug):
    customer = request.current_customer
    company = _fresh_company(customer)
    building = getattr(customer, 'building', None)
    if not company.is_store_open:
        messages.error(request, company.ordering_status_message or 'Ordering is currently closed.')
        return redirect('menu:menu')

    category = get_object_or_404(
        Category,
        slug=slug,
        is_deleted=False,
        companies=company,
    )
    if not category.is_active_now(company):
        messages.error(request, 'This category is not available right now.')
        return redirect('menu:menu')

    q = request.GET.get('q', '').strip()
    food_pref = (request.GET.get('food') or '').strip().lower()
    offering_filter = (request.GET.get('offering') or '').strip()
    calorie_max_raw = (request.GET.get('calorie_max') or '').strip()
    calorie_max = None
    try:
        calorie_max = int(calorie_max_raw) if calorie_max_raw else None
    except (TypeError, ValueError):
        calorie_max = None
    selected_offering = None
    if offering_filter:
        selected_offering = Offering.objects.filter(company=company, is_deleted=False).filter(Q(pk=offering_filter) | Q(slug=offering_filter)).first()
        if selected_offering and not selected_offering.is_active_now():
            selected_offering = None
    product_filter = Q(is_active=True, is_deleted=False, company=company, category=category)
    if selected_offering:
        product_filter &= Q(offering=selected_offering)
    if q:
        product_filter &= Q(name__icontains=q) | Q(description__icontains=q) | Q(code__icontains=q)

    products = [
        p for p in Product.objects.filter(product_filter).select_related('category', 'offering').prefetch_related('food_type', 'category__company_statuses').order_by(
            'category__position_order', 'category__name', 'position_order', 'name'
        )
        if _product_is_visible_for_customer(p, company, food_pref)
        and (calorie_max is None or (p.calories is not None and p.calories <= calorie_max))
    ]
    _attach_display_prices(products, company, building=building, cafe=_resolve_customer_cafe(customer, company, request=request))
    _mark_free_meal_products(products, company)

    cart_summary = _build_cart_summary(customer, request.session.get('cart', {}), request=request)
    request.session['cart'] = cart_summary['cart']
    request.session.modified = True

    visible_count = len(products)

    return render(request, 'menu/category_detail.html', {
        'category': category,
        'sub_categories': [],
        'sub_sections': [],
        'products': products,
        'company': company,
        'customer': customer,
        'cart_count': cart_summary['cart_count'],
        'cart_map': cart_summary['cart_map'],
        'search_query': q,
        'food_pref': food_pref,
        'offering_filter': offering_filter,
        'calorie_max': calorie_max,
        'calorie_presets': [200, 400, 600, 800],
        'selected_offering': selected_offering,
        'visible_count': visible_count,
        'page_title': category.name,
    })


@never_cache
@customer_login_required
def offering_detail(request, slug):
    customer = request.current_customer
    company = _fresh_company(customer)
    building = getattr(customer, 'building', None)
    if not company.is_store_open:
        messages.error(request, company.ordering_status_message or 'Ordering is currently closed.')
        return redirect('menu:menu')

    offering = Offering.objects.filter(
        company=company,
        is_active=True,
        is_deleted=False,
    ).filter(Q(slug=slug) | Q(pk=slug if str(slug).isdigit() else None)).order_by('position_order', 'name').first()
    if not offering:
        raise Http404('Offering not found.')
    if not offering.is_active_now():
        messages.error(request, 'This offering is not available right now.')
        return redirect('menu:menu')

    q = request.GET.get('q', '').strip()
    food_pref = (request.GET.get('food') or '').strip().lower()
    calorie_max_raw = (request.GET.get('calorie_max') or '').strip()
    calorie_max = None
    try:
        calorie_max = int(calorie_max_raw) if calorie_max_raw else None
    except (TypeError, ValueError):
        calorie_max = None

    product_filter = Q(is_active=True, is_deleted=False, company=company, offering=offering)
    if q:
        product_filter &= Q(name__icontains=q) | Q(description__icontains=q) | Q(code__icontains=q) | Q(category__name__icontains=q)

    products = [
        p for p in Product.objects.filter(product_filter).select_related('category', 'offering').prefetch_related('food_type', 'category__company_statuses').order_by(
            'category__position_order', 'category__name', 'position_order', 'name'
        )
        if _product_is_visible_for_customer(p, company, food_pref)
        and (calorie_max is None or (p.calories is not None and p.calories <= calorie_max))
    ]
    _attach_display_prices(products, company, building=building, cafe=_resolve_customer_cafe(customer, company, request=request))
    _mark_free_meal_products(products, company)

    cart_summary = _build_cart_summary(customer, request.session.get('cart', {}), request=request)
    request.session['cart'] = cart_summary['cart']
    request.session.modified = True

    visible_count = len(products)
    category_count = len({p.category_id for p in products if p.category_id})

    return render(request, 'menu/offering_detail.html', {
        'offering': offering,
        'products': products,
        'company': company,
        'customer': customer,
        'cart_count': cart_summary['cart_count'],
        'cart_map': cart_summary['cart_map'],
        'search_query': q,
        'food_pref': food_pref,
        'calorie_max': calorie_max,
        'calorie_presets': [200, 400, 600, 800],
        'visible_count': visible_count,
        'category_count': category_count,
        'page_title': offering.name,
    })


@never_cache
@customer_login_required
def product_detail(request, pk):
    customer = request.current_customer
    _fresh_company(customer)
    company = customer.company
    # Try to find the product scoped to the customer's company.
    # If it doesn't exist or is inactive/deleted/wrong-company, redirect to menu
    # with a friendly message instead of a hard 404 — this prevents broken links
    # from offer popups where the admin may have deactivated a product after the
    # offer was created.
    product = Product.objects.select_related(
        'category', 'offering'
    ).prefetch_related('food_type', 'category__company_statuses', 'counter_mappings__counter').filter(
        pk=pk,
        is_active=True,
        is_deleted=False,
        company=customer.company,
    ).first()
    if product is None:
        messages.warning(request, 'That product is no longer available.')
        return redirect('menu:menu')
    if _get_web_max_qty(product) <= 0:
        messages.warning(request, 'That product is sold out right now.')
        return redirect('menu:menu')
    if not _product_is_visible_for_customer(product, company):
        messages.warning(request, 'That product is no longer available.')
        return redirect('menu:menu')
    cart_summary = _build_cart_summary(customer, request.session.get('cart', {}), request=request)
    current_qty = cart_summary['cart_map'].get(str(product.pk), 0)
    _cafe = _resolve_customer_cafe(customer, company, request=request)
    _building = getattr(customer, 'building', None)
    site_price = _resolve_site_price(product, customer.company, building=_building, cafe=_cafe)
    product.display_price = site_price
    product.effective_web_qty = _get_web_max_qty(product)
    _mark_free_meal_products([product], customer.company)

    # ── Compute active offer for this product so template can show discounted price ──
    product_offer = _get_live_offer_for_product_menu(customer.company, product, cafe=_cafe)
    discounted_price = None
    offer_saving_preview = Decimal('0.00')
    if product_offer and product_offer.offer_type not in (Offer.TYPE_FLAT, Offer.TYPE_CART):
        # Per-product offer: compute effective price for qty=1 preview
        eff, saving = _apply_offer_to_line_menu(product_offer, site_price, 1)
        if saving > 0:
            discounted_price = eff
            offer_saving_preview = saving

    # ── Also get any live cart-level offer for threshold display ──
    cart_level_offer = cart_summary.get('cart_level_offer')

    similar_products = [
        item for item in Product.objects.filter(
            company=customer.company,
            is_active=True,
            is_deleted=False,
            category=product.category,
        ).exclude(pk=product.pk).select_related('category').prefetch_related('food_type', 'category__company_statuses').order_by('-rating', 'position_order', 'name')
        if _product_is_visible_for_customer(item, company)
    ][:4]
    _attach_display_prices(similar_products, customer.company, building=_building, cafe=_cafe)
    _mark_free_meal_products(similar_products, customer.company)

    return render(request, 'menu/product_detail.html', {
        'product': product,
        'customer': customer,
        'cart_count': cart_summary['cart_count'],
        'cart_map': cart_summary['cart_map'],
        'current_qty': current_qty,
        'in_cart': current_qty > 0,
        'similar_products': similar_products,
        # Offer info
        'product_offer': product_offer,
        'discounted_price': discounted_price,
        'offer_saving_preview': offer_saving_preview,
        'cart_level_offer': cart_level_offer,
        'cart_offer_saving': cart_summary.get('cart_offer_saving', Decimal('0.00')),
        'subtotal': cart_summary.get('subtotal', Decimal('0.00')),
        'page_title': product.name,
    })


# ─── Cart ───────────────────────────────────────────────────

@never_cache
@customer_login_required
def cart_view(request):
    customer = request.current_customer
    _fresh_company(customer)
    company = customer.company
    summary = _build_cart_summary(customer, request.session.get('cart', {}), request=request)
    request.session['cart'] = summary['cart']
    request.session.modified = True
    # Collect all live cart-level offers (for unlock nudge display)
    cafe = _resolve_customer_cafe(customer, company, request=request)
    _live_cart_offers = []
    for _ot in (Offer.TYPE_FLAT, Offer.TYPE_CART):
        _qs = Offer.objects.filter(company=company, is_active=True, is_deleted=False,
                                   offer_type=_ot).order_by('min_order_value')
        for _o in _qs:
            if _o.is_live:
                _live_cart_offers.append(_o)
    # ── BOGO / FREE nudges: one per eligible product with odd qty ──────────────
    # For each cart item that has a BOGO or FREE offer and the customer has an
    # odd quantity (1, 3, 5…), show a nudge: "Add one more X to get Buy 1 Get 1".
    bogo_nudges = []
    _already_used = set(
        __import__('apps.menu.models', fromlist=['OfferUsage']).OfferUsage.objects.filter(
            customer=customer, used_on=timezone.localdate()
        ).values_list('offer_id', flat=True)
    )
    for _item in summary.get('items', []):
        _p = _item['product']
        _qty = _item['qty']
        _offer = _get_live_offer_for_product_menu(company, _p, cafe=cafe)
        if not _offer:
            continue
        if _offer.pk in _already_used:
            continue
        if _offer.offer_type == Offer.TYPE_BOGO and _qty % 2 == 1:
            bogo_nudges.append({
                'product': _p,
                'offer': _offer,
                'message': f'Add 1 more “{_p.name}” to get Buy 1 Get 1 Free!',
                'type': 'bogo',
            })
        elif _offer.offer_type == Offer.TYPE_FREE and _qty == 0:
            # FREE: show nudge only if product isn't in cart at all — already handled elsewhere
            pass

    # Build cafe list for picker (only shown when company has multiple cafes)
    from apps.menu.models import Cafe as _CafeMenu
    _company_cafes = list(
        _CafeMenu.objects.filter(company=company, is_active=True, is_deleted=False)
        .order_by('name')
    )
    _selected_cafe_id = request.session.get('web_cafe_id')
    # Ensure selected cafe still belongs to this company
    if _selected_cafe_id and not any(c.pk == _selected_cafe_id for c in _company_cafes):
        request.session.pop('web_cafe_id', None)
        _selected_cafe_id = None

    cart_unlock_nudge = _get_cart_unlock_nudge(company, summary['subtotal'])

    return render(request, 'menu/cart.html', {
        **summary,
        'company': company,
        'customer': customer,
        'live_cart_offers': _live_cart_offers,
        'cart_unlock_nudge': cart_unlock_nudge,
        'bogo_nudges': bogo_nudges,
        'company_cafes': _company_cafes,
        'selected_cafe_id': _selected_cafe_id,
        'ordering_closed_message': company.ordering_status_message if not company.is_store_open else '',
        'page_title': 'Your Cart',
    })


@require_POST
@customer_login_required
def cart_add(request, product_id):
    customer = request.current_customer
    _fresh_company(customer)
    product = _get_own_product(product_id, customer)
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if not product:
        payload = {'success': False, 'error': 'Product not available for your account.'}
        if is_ajax:
            return JsonResponse(payload, status=403)
        messages.error(request, payload['error'])
        return redirect('menu:menu')

    if not customer.company.is_store_open:
        payload = _ordering_closed_payload(customer.company)
        if is_ajax:
            return JsonResponse(payload, status=400)
        messages.error(request, payload['error'])
        return redirect('menu:menu')

    if not product.is_active or not product.is_available_now():
        payload = {'success': False, 'error': 'This item is not available right now.'}
        if is_ajax:
            return JsonResponse(payload, status=400)
        messages.error(request, payload['error'])
        return redirect('menu:menu')

    max_qty = _get_web_max_qty(product)
    if max_qty <= 0:
        payload = {'success': False, 'error': 'This item is sold out for now.'}
        if is_ajax:
            return JsonResponse(payload, status=400)
        messages.error(request, payload['error'])
        return redirect('menu:menu')

    cart = request.session.get('cart', {})
    key = str(product.pk)
    try:
        qty = max(1, int(request.POST.get('qty', 1)))
    except (TypeError, ValueError):
        qty = 1

    current = int(cart.get(key, {}).get('qty', 0) or 0)
    min_qty = _get_order_min_qty(product)
    requested_qty = max(qty, min_qty) if current <= 0 else current + qty
    new_qty = min(requested_qty, max_qty)
    if new_qty < min_qty:
        payload = {'success': False, 'error': f'Minimum order quantity for {product.name} is {min_qty}.'}
        if is_ajax:
            return JsonResponse(payload, status=400)
        messages.error(request, payload['error'])
        return redirect('menu:menu')

    cart[key] = {
        'qty': new_qty,
        'price': str(product.price),
        'name': product.name,
    }
    request.session['cart'] = cart
    request.session.modified = True

    summary = _build_cart_summary(customer, cart, request=request)
    request.session['cart'] = summary['cart']
    request.session.modified = True

    payload = {
        'success': True,
        'cart_count': summary['cart_count'],
        'qty': new_qty,
        'item_name': product.name,
        'line_total': _money(next(
            (i['line_total'] for i in summary.get('items', []) if str(i['product'].pk) == str(product.pk)),
            _resolve_site_price(
                product,
                customer.company,
                building=getattr(customer, 'building', None),
                cafe=_resolve_customer_cafe(customer, customer.company, request=request),
            ) * new_qty
        )),
        'subtotal': _money(summary['subtotal']),
        'packing': _money(summary['packing']),
        'my_pay': _money(summary['my_pay']),
        'offer_discount': _money(summary.get('offer_discount', 0)),
        'cart_offer_saving': _money(summary.get('cart_offer_saving', 0)),
        'cart_level_offer_title': summary['cart_level_offer'].title if summary.get('cart_level_offer') else '',
        'cart_offer_min_order': str((_get_cart_unlock_nudge(customer.company, summary['subtotal']) or {}).get('min_order_value') or 0),
        'cart_offer_remaining': _money(((_get_cart_unlock_nudge(customer.company, summary['subtotal']) or {}).get('remaining_value')) or 0),
        'cart_offer_nudge_title': ((_get_cart_unlock_nudge(customer.company, summary['subtotal']) or {}).get('title')) or '',
        'bill_to_company': _money(summary.get('bill_to_company', 0)),
        'total': _money(summary.get('total', 0)),
    }
    if is_ajax:
        return JsonResponse(payload)

    messages.success(request, f'“{product.name}” added to cart.')
    return redirect('menu:cart')


@require_POST
@customer_login_required
def cart_remove(request, product_id):
    customer = request.current_customer
    cart = request.session.get('cart', {})
    cart.pop(str(product_id), None)
    request.session['cart'] = cart
    request.session.modified = True

    summary = _build_cart_summary(customer, cart, request=request)
    request.session['cart'] = summary['cart']
    request.session.modified = True

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'cart_count': summary['cart_count'],
            'subtotal': _money(summary['subtotal']),
            'packing': _money(summary['packing']),
            'my_pay': _money(summary['my_pay']),
            'bill_to_company': _money(summary['bill_to_company']),
            'total': _money(summary['total']),
            'offer_discount': _money(summary.get('offer_discount', 0)),
            'cart_offer_saving': _money(summary.get('cart_offer_saving', 0)),
            'cart_level_offer_title': summary['cart_level_offer'].title if summary.get('cart_level_offer') else '',
            'cart_offer_min_order': str((_get_cart_unlock_nudge(customer.company, summary['subtotal']) or {}).get('min_order_value') or 0),
            'cart_offer_remaining': _money(((_get_cart_unlock_nudge(customer.company, summary['subtotal']) or {}).get('remaining_value')) or 0),
            'cart_offer_nudge_title': ((_get_cart_unlock_nudge(customer.company, summary['subtotal']) or {}).get('title')) or '',
        })

    messages.success(request, 'Item removed from your cart.')
    return redirect('menu:cart')


@require_POST
@customer_login_required
def cart_update_qty(request, product_id):
    customer = request.current_customer
    _fresh_company(customer)
    product = _get_own_product(product_id, customer)
    if not product:
        return JsonResponse({'success': False, 'error': 'Product not available.'}, status=403)

    cart = request.session.get('cart', {})
    key = str(product_id)
    current_qty = int(cart.get(key, {}).get('qty', 0) or 0)

    try:
        qty = int(request.POST.get('qty', 1))
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'error': 'Invalid quantity.'}, status=400)

    if qty > current_qty and not customer.company.is_store_open:
        return JsonResponse(_ordering_closed_payload(customer.company), status=400)

    if qty > current_qty and (not product.is_active or not product.is_available_now()):
        return JsonResponse({'success': False, 'error': 'This item is not available right now.'}, status=400)

    if qty <= 0:
        cart.pop(key, None)
        qty = 0
    else:
        max_qty = _get_web_max_qty(product)
        if max_qty <= 0:
            return JsonResponse({'success': False, 'error': 'This item is sold out for now.'}, status=400)

        qty = max(int(product.min_qty or 1), min(qty, max_qty))
        cart[key] = {
            'qty': qty,
            'price': str(product.price),
            'name': product.name,
        }

    request.session['cart'] = cart
    request.session.modified = True

    summary = _build_cart_summary(customer, cart, request=request)
    request.session['cart'] = summary['cart']
    request.session.modified = True

    return JsonResponse({
        'success': True,
        'cart_count': summary['cart_count'],
        'qty': qty,
        'line_total': _money(next(
            (i['line_total'] for i in summary.get('items', []) if str(i['product'].pk) == str(product_id)),
            _resolve_site_price(
                product,
                customer.company,
                building=getattr(customer, 'building', None),
                cafe=_resolve_customer_cafe(customer, customer.company, request=request),
            ) * qty
        )) if qty > 0 else '0.00',
        'subtotal': _money(summary['subtotal']),
        'packing': _money(summary['packing']),
        'my_pay': _money(summary['my_pay']),
        'bill_to_company': _money(summary['bill_to_company']),
        'total': _money(summary['total']),
        'offer_discount': _money(summary.get('offer_discount', 0)),
        'cart_offer_saving': _money(summary.get('cart_offer_saving', 0)),
        'cart_level_offer_title': summary['cart_level_offer'].title if summary.get('cart_level_offer') else '',
        'cart_offer_min_order': str((_get_cart_unlock_nudge(customer.company, summary['subtotal']) or {}).get('min_order_value') or 0),
        'cart_offer_remaining': _money(((_get_cart_unlock_nudge(customer.company, summary['subtotal']) or {}).get('remaining_value')) or 0),
        'cart_offer_nudge_title': ((_get_cart_unlock_nudge(customer.company, summary['subtotal']) or {}).get('title')) or '',
    })


# ════════════════════════════════════════════════════════════════
#  DASHBOARD  —  Products
# ════════════════════════════════════════════════════════════════

def _user_companies(user):
    return get_staff_site_companies(user)


def _user_categories(user):
    if user.is_superadmin:
        return Category.objects.filter(is_deleted=False).prefetch_related('companies').order_by('position_order', 'name')
    companies = _user_companies(user)
    return Category.objects.filter(companies__in=companies, is_deleted=False).prefetch_related('companies').distinct().order_by('position_order', 'name')


def _user_offerings(user):
    qs = Offering.objects.filter(is_deleted=False).select_related('company').order_by('company__name', 'position_order', 'name')
    if user.is_superadmin:
        return qs
    return qs.filter(company__in=_user_companies(user))


def _user_counters(user):
    qs = Counter.objects.filter(is_active=True, is_deleted=False).select_related('company', 'cafe').order_by('company__name', 'position_order', 'name')
    if user.is_superadmin:
        return qs
    return qs.filter(company__in=_user_companies(user))


def _optional_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _default_product_form_company_id(request, product, companies):
    posted = _optional_int(request.POST.get('company')) if request.method == 'POST' else None
    if posted:
        return posted
    if product and product.company_id:
        return product.company_id
    company_list = list(companies[:2]) if hasattr(companies, '__getitem__') else list(companies)
    return company_list[0].pk if len(company_list) == 1 else None


def _food_type_kind(food_type):
    name = str(getattr(food_type, 'name', '') or '').strip().lower().replace('-', '').replace(' ', '')
    if 'non' in name:
        return 'nonveg'
    if 'veg' in name:
        return 'veg'
    return ''


def _canonical_food_type_options():
    existing = list(FoodType.objects.filter(is_active=True, name__icontains='veg').order_by('id'))

    def pick(kind, preferred):
        for ft in existing:
            if ft.name.strip().lower() == preferred.lower():
                return ft
        for ft in existing:
            if _food_type_kind(ft) == kind:
                return ft
        obj, _ = FoodType.objects.get_or_create(name=preferred, defaults={'is_active': True})
        if not obj.is_active:
            obj.is_active = True
            obj.save(update_fields=['is_active'])
        return obj

    veg = pick('veg', 'Veg')
    nonveg = pick('nonveg', 'Non-Veg')
    return [
        {'pk': veg.pk, 'name': 'Veg', 'kind': 'veg'},
        {'pk': nonveg.pk, 'name': 'Non-Veg', 'kind': 'nonveg'},
    ]


def _selected_food_type_ids_from_request(request):
    raw_values = request.POST.getlist('food_types') or request.POST.getlist('food_type')
    raw = str(raw_values[0] if raw_values else '').strip()
    if not raw:
        return []
    if raw.isdigit():
        return [int(raw)]
    wanted = 'nonveg' if 'non' in raw.lower() else 'veg'
    for option in _canonical_food_type_options():
        if option['kind'] == wanted:
            return [option['pk']]
    return []


def _selected_food_type_ids_for_product(product):
    selected = list(product.food_type.all())
    if any(_food_type_kind(ft) == 'nonveg' for ft in selected):
        wanted = 'nonveg'
    elif any(_food_type_kind(ft) == 'veg' for ft in selected):
        wanted = 'veg'
    else:
        return []
    for option in _canonical_food_type_options():
        if option['kind'] == wanted:
            return [option['pk']]
    return []


def _selected_company_for_user(user, company_id):
    if user.is_superadmin:
        return Company.objects.filter(pk=company_id, is_active=True, is_deleted=False).first() if company_id else None
    companies = _user_companies(user)
    if company_id:
        return companies.filter(pk=company_id).first()
    return get_primary_staff_company(user)


def _scoped_building_for_company(company, pk):
    if not company or not pk:
        return None
    return Building.objects.filter(pk=pk, company=company, is_deleted=False).first()


def _scoped_cafe_for_company(company, pk):
    if not company or not pk:
        return None
    return Cafe.objects.select_related('building').filter(pk=pk, company=company, is_deleted=False).first()


def _scoped_category_for_company(company, pk):
    if not company or not pk:
        return None
    return Category.objects.filter(pk=pk, companies=company, is_deleted=False).first()


def _scoped_offering_for_company(company, pk):
    if not company or not pk:
        return None
    return Offering.objects.filter(pk=pk, company=company, is_deleted=False).first()


def _bulk_post_ids(request):
    ids = []
    for raw in request.POST.getlist('ids'):
        try:
            ids.append(int(raw))
        except (TypeError, ValueError):
            continue
    return ids


def _bulk_message(request, label, count, skipped=0):
    if count:
        msg = f'{count} {label}{"s" if count != 1 else ""} deleted.'
        if skipped:
            msg += f' {skipped} skipped.'
        messages.success(request, msg)
    elif skipped:
        messages.warning(request, f'No items deleted. {skipped} skipped.')
    else:
        messages.warning(request, 'No items selected.')


def _image_upload_label(image_file, fallback='Image'):
    original_name = getattr(image_file, 'name', '') or ''
    label = Path(original_name).stem.replace('_', ' ').replace('-', ' ').strip()
    return label or (fallback or 'Image')


def _copy_upload_to_gallery(model, *, image_file, company=None, user=None, name=''):
    if not image_file:
        return None
    original_name = getattr(image_file, 'name', '') or 'image-upload'
    try:
        image_file.seek(0)
    except Exception:
        pass
    payload = image_file.read()
    try:
        image_file.seek(0)
    except Exception:
        pass
    if not payload:
        return None
    obj = model(company=company, name=(name or '').strip() or _image_upload_label(image_file))
    if hasattr(obj, 'uploaded_by_id'):
        obj.uploaded_by = user
    obj.image.save(original_name, ContentFile(payload), save=True)
    return obj

def _is_product_free_meal_enabled(product):
    if not product or not product.pk or not product.company_id:
        return False
    return product.company.free_meal_products.filter(pk=product.pk).exists()

def _store_banner_upload_in_media_library(*, image_file, company, user, ad_name=''):
    """Create a reusable MediaAsset copy for a banner upload without changing banner logic.

    Reads the upload bytes once and resets the file pointer so the caller
    can still assign the same file object to ad.image afterwards.
    """
    if not image_file or not company:
        return None

    original_name = getattr(image_file, 'name', '') or 'banner-upload'
    asset_name = (ad_name or '').strip() or _image_upload_label(image_file, 'Banner Upload')

    # Read bytes first
    try:
        image_file.seek(0)
    except Exception:
        pass
    payload = image_file.read()
    if not payload:
        return None

    # Always reset so caller's ad.image assignment still works
    try:
        image_file.seek(0)
    except Exception:
        pass

    asset = MediaAsset(company=company, name=asset_name, uploaded_by=user)
    # Use a fresh ContentFile so we don't share the pointer with the caller
    asset.image.save(original_name, ContentFile(payload), save=False)
    try:
        asset.full_clean()
    except ValidationError:
        raise
    asset.save()
    asset.companies.set(Company.objects.filter(is_active=True, is_deleted=False).exclude(pk=company.pk))

    return asset


def _own_product_or_404(pk, user):
    """Get product scoped to user's company, or 404."""
    qs = Product.objects.filter(pk=pk, is_deleted=False)
    if not user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(user))
    return get_object_or_404(qs)


def _deny_dashboard_action(request, message='Permission denied.'):
    messages.error(request, message)
    return redirect('dashboard:no_access')


def _can_product_edit(user):
    if user.is_superadmin:
        return True
    return get_module_level(user, 'perm_products') in ('part_edit', 'full_edit')


def _can_product_delete(user):
    if user.is_superadmin:
        return True
    return get_module_level(user, 'perm_products') == 'full_edit'


def _own_category_or_404(pk, user):
    """Get category scoped to user's company, or 404."""
    qs = Category.objects.filter(pk=pk, is_deleted=False)
    if not user.is_superadmin:
        qs = qs.filter(companies__in=_user_companies(user)).distinct()
    return get_object_or_404(qs)


def _own_advert_or_404(pk, user):
    """Get advert scoped to user's company, or 404."""
    qs = Advertise.objects.filter(pk=pk)
    if not user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(user))
    return get_object_or_404(qs)


@staff_role_required('superadmin','admin','pos')
def dashboard_product_list(request):
    user = request.user
    lp = get_list_perms(user, 'perm_products')
    q = request.GET.get('q', '').strip()
    cat_id = request.GET.get('cat', '').strip()
    company_id = request.GET.get('company', '').strip()

    qs = Product.objects.filter(is_deleted=False).select_related(
        'category', 'company'
    ).prefetch_related('free_meal_companies').order_by(
        'company__name', 'category__position_order', 'category__name', 'position_order', 'name'
    )

    # Company scope
    if user.is_superadmin:
        if company_id:
            qs = qs.filter(company_id=company_id)
            category_qs = Category.objects.filter(
                companies__pk=company_id,
                is_deleted=False
            ).distinct().order_by('position_order', 'name')
        else:
            category_qs = Category.objects.filter(is_deleted=False).order_by('position_order', 'name')
    else:
        companies = _user_companies(user)
        if companies.exists():
            qs = qs.filter(company__in=companies)
            if company_id and companies.filter(pk=company_id).exists():
                qs = qs.filter(company_id=company_id)
                category_qs = Category.objects.filter(
                    companies__pk=company_id,
                    is_deleted=False
                ).distinct().order_by('position_order', 'name')
            else:
                category_qs = Category.objects.filter(
                    companies__in=companies,
                    is_deleted=False
                ).distinct().order_by('position_order', 'name')
        else:
            qs = qs.none()
            category_qs = Category.objects.none()

    # Other filters
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
    if cat_id:
        qs = qs.filter(category_id=cat_id)

    products = list(qs)

    # Free-meal badge visibility for all staff roles
    for p in products:
        mapped_company_ids = {co.pk for co in p.free_meal_companies.all()}
        p.is_free_meal_eligible = p.company_id in mapped_company_ids

    # Phase fix: pass offerings so the Bulk Copy modal's override-offering dropdown works
    if user.is_superadmin:
        offering_qs = Offering.objects.filter(is_deleted=False).select_related('company').order_by('company__name', 'name')
    else:
        offering_qs = Offering.objects.filter(company__in=_user_companies(user), is_deleted=False).order_by('company__name', 'name')
    return render(request, 'dashboard/menu/product_list.html', {
        'products': products,
        'categories': category_qs,
        'offerings': offering_qs,
        'companies': _user_companies(user),
        'can_manage_products': user.role == 'superadmin' or any(
            lp.get(action) for action in ('delete', 'reorder', 'copy', 'bulk_copy')
        ),
        'can_limited_product_edit': user.role == 'pos' and lp.get('cashier_edit'),
        'lp': lp,
        'q': q,
        'cat_id': cat_id,
        'company_id': company_id,
        'page_title': 'Products',
    })


def _dashboard_product_export_qs(request):
    user = request.user
    q = request.GET.get('q', '').strip()
    cat_id = request.GET.get('cat', '').strip()
    company_id = request.GET.get('company', '').strip()

    qs = Product.objects.filter(is_deleted=False).select_related(
        'category', 'company', 'offering'
    ).prefetch_related('food_type').order_by(
        'company__name', 'category__position_order', 'category__name', 'position_order', 'name'
    )

    if user.is_superadmin:
        if company_id:
            qs = qs.filter(company_id=company_id)
    else:
        companies = _user_companies(user)
        if companies.exists():
            qs = qs.filter(company__in=companies)
            if company_id and companies.filter(pk=company_id).exists():
                qs = qs.filter(company_id=company_id)
        else:
            qs = qs.none()

    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q))
    if cat_id:
        qs = qs.filter(category_id=cat_id)

    raw_ids = request.GET.get('ids', '').strip()
    if raw_ids:
        ids = [pk for pk in re.split(r'[\s,]+', raw_ids) if pk.isdigit()]
        if ids:
            qs = qs.filter(pk__in=ids)
    return qs


@staff_role_required('superadmin', 'admin')
def product_excel_download(request):
    if not user_can_action(request.user, 'perm_products', 'export'):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard:no_access')
    products = list(_dashboard_product_export_qs(request))
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="products-export.csv"'
        writer = csv.writer(resp)
        writer.writerow(['Product', 'Code', 'Company', 'Category', 'Offering', 'Staff/Base Price', 'Visitor Price', 'Room Service Extra %', 'Packing', 'Web Qty', 'POS Qty', 'Web Active', 'Kiosk Active', 'Calories', 'Image URL', 'Description'])
        for p in products:
            writer.writerow([
                p.name, p.code, p.company.name if p.company else '', p.category.name if p.category else '',
                p.offering.name if p.offering else '', p.price, p.company_price, p.room_service_extra_percent, p.packing_price, p.web_qty, p.pos_qty,
                'Yes' if p.is_active else 'No', 'Yes' if p.is_kiosk_active else 'No', p.calories or '',
                request.build_absolute_uri(p.image.url) if p.image else '', p.description,
            ])
        return resp

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'
    headers = [
        'Photo', 'Product', 'Code', 'Company', 'Category', 'Offering', 'Staff/Base Price', 'Visitor Price', 'Room Service Extra %', 'Packing',
        'Web Qty', 'POS Qty', 'Web Active', 'Kiosk Active', 'Featured Web', 'Featured Kiosk',
        'Bypass Schedule', 'Calories', 'Food Type', 'Image URL', 'Description',
    ]
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = 'A2'
    widths = [14, 28, 16, 22, 22, 22, 12, 12, 18, 12, 10, 10, 12, 12, 12, 13, 14, 10, 20, 42, 48]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row_idx, p in enumerate(products, 2):
        image_url = request.build_absolute_uri(p.image.url) if p.image else ''
        food_types = ', '.join(ft.name for ft in p.food_type.all())
        row_data = [
            '', p.name, p.code, p.company.name if p.company else '', p.category.name if p.category else '',
            p.offering.name if p.offering else '', float(p.price or 0), float(p.company_price or 0), float(p.room_service_extra_percent or 0), float(p.packing_price or 0),
            p.web_qty, p.pos_qty, 'Yes' if p.is_active else 'No', 'Yes' if p.is_kiosk_active else 'No',
            'Yes' if p.featured_in_web else 'No', 'Yes' if p.featured_in_kiosk_extra else 'No',
            'Yes' if p.schedule_bypass else 'No', p.calories or '', food_types, image_url, p.description,
        ]
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 48
        if p.image:
            try:
                suffix = Path(getattr(p.image, 'name', '')).suffix.lower()
                if suffix in {'.jpg', '.jpeg', '.png'}:
                    xl_img = XLImage(p.image.path)
                else:
                    from PIL import Image as PILImage
                    image_buffer = io.BytesIO()
                    with PILImage.open(p.image.path) as source:
                        source.convert('RGBA').save(image_buffer, format='PNG')
                    image_buffer.seek(0)
                    xl_img = XLImage(image_buffer)
                xl_img.width = 64
                xl_img.height = 42
                ws.add_image(xl_img, f'A{row_idx}')
            except Exception:
                pass
        if image_url:
            ws.cell(row=row_idx, column=18).hyperlink = image_url
            ws.cell(row=row_idx, column=18).style = 'Hyperlink'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    resp = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="products-export.xlsx"'
    return resp


@staff_role_required('superadmin','admin')
def dashboard_product_add(request):
    user = request.user
    if not user_can_action(user, 'perm_products', 'add'):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard:no_access')
    if request.method == 'POST':
        err = _save_product(request, None, user)
        if err is None:
            messages.success(request, 'Product added successfully.')
            return redirect('dashboard:product_list')
        for e in err:
            messages.error(request, e)
    # Preserve selected_types on POST-failure re-render so the user's food-type
    # picks are not silently dropped when validation fails.
    selected_types_on_post = _selected_food_type_ids_from_request(request) if request.method == 'POST' else []
    form_companies = _user_companies(user)
    selected_company_id = _default_product_form_company_id(request, None, form_companies)
    return render(request, 'dashboard/menu/product_form.html', {
        'companies': form_companies,
        'categories': _user_categories(user),
        'selected_company_id': selected_company_id,
        'selected_category_id': _optional_int(request.POST.get('category')) if request.method == 'POST' else None,
        'selected_offering_id': _optional_int(request.POST.get('offering')) if request.method == 'POST' else None,
        'food_type_options': _canonical_food_type_options(),
        'selected_types': selected_types_on_post,
        'offerings': _user_offerings(user),
        'counters': _user_counters(user),
        'selected_counter_ids': [int(x) for x in request.POST.getlist('counter_ids') if str(x).isdigit()] if request.method == 'POST' else [],
        'is_free_meal_product': request.POST.get('is_free_meal_product') == 'on',
        'page_title': 'Add Product',
        'action': 'Add',
    })


@staff_role_required('superadmin','admin')
def dashboard_product_edit(request, pk):
    user = request.user
    if not _can_product_edit(user):
        return _deny_dashboard_action(request)
    _perm = check_module_permission(request, 'perm_products')
    if _perm: return _perm
    product = _own_product_or_404(pk, user)

    if request.method == 'POST':
        if user.role != 'superadmin' and get_module_level(request.user, 'perm_products') == 'full_edit':
            _diffs = {
                'name':          {'label': 'Name',          'before': product.name,                       'after': (request.POST.get('name') or '').strip()},
                'code':          {'label': 'SKU/Code',      'before': product.code or '',                 'after': (request.POST.get('code') or '').strip()},
                'price':         {'label': 'Staff/Base Price', 'before': str(product.price),              'after': request.POST.get('price', '')},
                'company_price': {'label': 'Visitor Price', 'before': str(product.company_price or 0),    'after': request.POST.get('company_price', '')},
                'room_service_extra_percent': {'label': 'Room Service Extra %', 'before': str(product.room_service_extra_percent or 0), 'after': request.POST.get('room_service_extra_percent', '')},
                'packing_price': {'label': 'Packing Price', 'before': str(product.packing_price or 0),    'after': request.POST.get('packing_price', '')},
                'description':   {'label': 'Description',  'before': product.description or '',           'after': (request.POST.get('description') or '').strip()},
                'is_active':     {'label': 'Active',        'before': product.is_active,                  'after': request.POST.get('is_active') == 'on'},
                'position_order':{'label': 'Sort Order',    'before': product.position_order,             'after': int(request.POST.get('position_order') or 0)},
                'min_qty':       {'label': 'Min Qty',       'before': product.min_qty,                    'after': max(1, int(request.POST.get('min_qty') or 1))},
                'max_qty':       {'label': 'Max Qty',       'before': product.max_qty,                    'after': max(1, int(request.POST.get('max_qty') or 1))},
                'preparation_time_minutes': {'label': 'Prep Time', 'before': product.preparation_time_minutes, 'after': max(0, int(request.POST.get('preparation_time_minutes') or 0))},
            }
            _pc = create_pending_change(request, 'perm_products', product, _diffs)
            if _pc:
                messages.success(request, 'Your changes have been submitted for superadmin review.')
            else:
                messages.info(request, 'No changes detected.')
            return _dashboard_return_or_default(request, 'dashboard:product_list')
        err = _save_product(request, product, user)
        if err is None:
            messages.success(request, 'Product updated.')
            return _dashboard_return_or_default(request, 'dashboard:product_list')
        for e in err:
            messages.error(request, e)
        is_free_meal_product = request.POST.get('is_free_meal_product') == 'on'
    else:
        is_free_meal_product = _is_product_free_meal_enabled(product)
    _lj, _pl = get_locked_html_names(request.user, 'perm_products')
    form_companies = _user_companies(user)
    selected_company_id = _default_product_form_company_id(request, product, form_companies)
    return render(request, 'dashboard/menu/product_form.html', {
        'product': product,
        'companies': form_companies,
        'categories': _user_categories(user),
        'selected_company_id': selected_company_id,
        'selected_category_id': _optional_int(request.POST.get('category')) if request.method == 'POST' else product.category_id,
        'selected_offering_id': _optional_int(request.POST.get('offering')) if request.method == 'POST' else product.offering_id,
        'food_type_options': _canonical_food_type_options(),
        'offerings': _user_offerings(user),
        'counters': _user_counters(user),
        'selected_counter_ids': list(product.counter_mappings.values_list('counter_id', flat=True)),
        'selected_types': _selected_food_type_ids_from_request(request) if request.method == 'POST' else _selected_food_type_ids_for_product(product),
        'is_free_meal_product': is_free_meal_product,
        'page_title': f'Edit - {product.name}',
        'action': 'Save',
        'perm_level': _pl or 'full_edit', 'locked_names_json': _lj,
    })


@require_POST
@staff_role_required('superadmin', 'admin', 'pos', 'cafeman')
def dashboard_product_toggle(request, pk):
    if has_any_granular_perms(request.user) and not user_can_action(request.user, 'perm_products', 'toggle'):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    p = _own_product_or_404(pk, request.user)
    p.is_active = not p.is_active
    p.save(update_fields=['is_active'])
    return JsonResponse({'success': True, 'is_active': p.is_active})



@require_POST
@staff_role_required('superadmin')
def dashboard_product_schedule_bypass_toggle(request, pk):
    """Superadmin-only: toggle schedule_bypass on a product."""
    p = _own_product_or_404(pk, request.user)
    p.schedule_bypass = not p.schedule_bypass
    p.save(update_fields=['schedule_bypass'])
    return JsonResponse({'success': True, 'schedule_bypass': p.schedule_bypass})


@require_POST
@staff_role_required('superadmin', 'admin')
def dashboard_product_featured_toggle(request, pk):
    """Toggle featured_in_kiosk_extra — visible in kiosk Featured section."""
    if has_any_granular_perms(request.user) and not user_can_action(request.user, 'perm_products', 'featured_toggle'):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    p = _own_product_or_404(pk, request.user)
    p.featured_in_kiosk_extra = not p.featured_in_kiosk_extra
    p.save(update_fields=['featured_in_kiosk_extra'])
    return JsonResponse({'success': True, 'featured': p.featured_in_kiosk_extra})


@require_POST
@staff_role_required('superadmin', 'admin')
def dashboard_product_web_featured_toggle(request, pk):
    """Toggle featured_in_web — visible in customer web portal Featured section."""
    if has_any_granular_perms(request.user) and not user_can_action(request.user, 'perm_products', 'web_featured_toggle'):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    p = _own_product_or_404(pk, request.user)
    p.featured_in_web = not p.featured_in_web
    p.save(update_fields=['featured_in_web'])
    return JsonResponse({'success': True, 'featured_web': p.featured_in_web})


@require_POST
@staff_role_required('superadmin', 'admin', 'cafeman')
def dashboard_product_kiosk_toggle(request, pk):
    """Toggle kiosk-specific active flag — independent of the web status."""
    if has_any_granular_perms(request.user) and not user_can_action(request.user, 'perm_products', 'kiosk_toggle'):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    p = _own_product_or_404(pk, request.user)
    p.is_kiosk_active = not p.is_kiosk_active
    p.save(update_fields=['is_kiosk_active'])
    return JsonResponse({'success': True, 'is_kiosk_active': p.is_kiosk_active})


@require_POST
@staff_role_required('superadmin', 'admin', 'pos', 'cafeman')
def dashboard_product_pos_toggle(request, pk):
    """Toggle POS-specific active flag — independent of the web and kiosk status."""
    if has_any_granular_perms(request.user) and not (
        user_can_action(request.user, 'perm_products', 'pos_toggle')
        or user_can_action(request.user, 'perm_products', 'field_is_pos_active')
    ):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    p = _own_product_or_404(pk, request.user)
    p.is_pos_active = not p.is_pos_active
    p.save(update_fields=['is_pos_active'])
    return JsonResponse({'success': True, 'is_pos_active': p.is_pos_active})


@require_POST
@staff_role_required('superadmin','admin')
def dashboard_product_delete(request, pk):
    if not _can_product_delete(request.user):
        return _deny_dashboard_action(request)
    p = _own_product_or_404(pk, request.user)
    p.is_deleted = True
    p.is_active = False
    p.save(update_fields=['is_deleted', 'is_active'])
    messages.success(request, f'"{p.name}" deleted.')
    return redirect('dashboard:product_list')


@require_POST
@staff_role_required('superadmin','admin')
def dashboard_product_bulk_delete(request):
    if not _can_product_delete(request.user):
        return _deny_dashboard_action(request)
    ids = _bulk_post_ids(request)
    qs = Product.objects.filter(pk__in=ids, is_deleted=False)
    if not request.user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(request.user))
    count = qs.update(is_deleted=True, is_active=False) if ids else 0
    _bulk_message(request, 'product', count)
    return redirect('dashboard:product_list')


def _save_product(request, instance, user):
    errors = []
    original_name = instance.name if instance is not None and instance.pk else ''
    locked_names = set()
    if instance is not None and not user.is_superadmin:
        locked_json, level = get_locked_html_names(user, 'perm_products')
        if level == 'part_edit':
            try:
                locked_names = set(json.loads(locked_json))
            except (TypeError, ValueError):
                locked_names = set()

    def is_locked(html_name):
        return instance is not None and html_name in locked_names

    name = (instance.name if is_locked('name') else request.POST.get('name', '')).strip()
    description = (
        instance.description or ''
        if is_locked('description')
        else request.POST.get('description', '').strip()
    )
    if not name:
        errors.append('Product name is required.')
        return errors
    try:
        price = instance.price if is_locked('price') else Decimal(request.POST.get('price', '0') or '0')
    except InvalidOperation:
        errors.append('Invalid price.')
        return errors

    try:
        company_price = instance.company_price if is_locked('company_price') else Decimal(request.POST.get('company_price', '0') or '0')
    except InvalidOperation:
        errors.append('Invalid visitor price.')
        return errors

    try:
        room_service_extra_percent = (
            instance.room_service_extra_percent
            if is_locked('room_service_extra_percent')
            else Decimal(request.POST.get('room_service_extra_percent', '0') or '0')
        )
        if room_service_extra_percent < 0:
            room_service_extra_percent = Decimal('0')
    except InvalidOperation:
        errors.append('Invalid room service extra percentage.')
        return errors

    try:
        packing_price = instance.packing_price if is_locked('packing_price') else Decimal(request.POST.get('packing_price', '0') or '0')
    except InvalidOperation:
        errors.append('Invalid packing price.')
        return errors

    try:
        min_qty = instance.min_qty if is_locked('min_qty') else max(1, int(request.POST.get('min_qty', '1') or '1'))
        max_qty = instance.max_qty if is_locked('max_qty') else max(1, int(request.POST.get('max_qty', '10') or '10'))
        web_qty = instance.web_qty if is_locked('web_qty') else max(-1, int(request.POST.get('web_qty', '-1') or '-1'))
        pos_qty = instance.pos_qty if is_locked('pos_qty') else max(0, int(request.POST.get('pos_qty', '0') or '0'))
        position_order = instance.position_order if is_locked('position_order') else int(request.POST.get('position_order', '0') or '0')
        preparation_time_minutes = (
            instance.preparation_time_minutes
            if is_locked('preparation_time_minutes')
            else max(0, int(request.POST.get('preparation_time_minutes', '10') or '10'))
        )
        calories = (
            instance.calories
            if is_locked('calories')
            else _parse_calories_value(request.POST.get('calories', '').strip(), name, description)
        )
    except (TypeError, ValueError):
        errors.append('Min qty, max qty, web qty, sort order, POS qty, preparation time, and calories must be valid numbers.')
        return errors
    if max_qty < min_qty:
        errors.append('Max Qty per Order cannot be less than Min Qty per Order.')
        return errors

    cat_id = request.POST.get('category')
    company_id = request.POST.get('company')

    if is_locked('company'):
        company = instance.company
    else:
        company = _selected_company_for_user(user, company_id)

    if not company:
        errors.append('Company is required.')
        return errors

    if is_locked('category'):
        category = instance.category
    else:
        category = _scoped_category_for_company(company, cat_id)

    if not category:
        errors.append('Category is required or not available for your company.')
        return errors
    if not category.companies.filter(pk=company.pk).exists():
        errors.append('Selected category is not designated for this company.')
        return errors

    if instance is None:
        instance = Product()

    if not user.is_superadmin and instance.pk and not user_can_access_company(user, instance.company_id):
        errors.append('You do not have permission to edit this product.')
        return errors

    duplicate_name_qs = Product.objects.filter(company=company, is_deleted=False, name__iexact=name)
    if instance.pk:
        duplicate_name_qs = duplicate_name_qs.exclude(pk=instance.pk)
    if duplicate_name_qs.exists():
        errors.append('Duplicate value: a product with this name already exists in this company.')
        return errors

    code_value = (instance.code or '') if is_locked('code') else request.POST.get('code', '').strip()
    if code_value:
        duplicate_code_qs = Product.objects.filter(company=company, is_deleted=False, code__iexact=code_value)
        if instance.pk:
            duplicate_code_qs = duplicate_code_qs.exclude(pk=instance.pk)
        if duplicate_code_qs.exists():
            errors.append('Duplicate value: a product with this SKU / code already exists in this company.')
            return errors

    instance.name = name
    if not instance.pk:
        instance.slug = ''
    elif not is_locked('name') and name != original_name:
        instance.slug = ''
    instance.code = code_value
    instance.category = category

    instance.sub_category = None   # deprecated field — kept dormant
    instance.sub_list     = None   # deprecated field — kept dormant
    instance.company = company
    offering_id = request.POST.get('offering')
    if not is_locked('offering'):
        offering = _scoped_offering_for_company(company, offering_id)
        if offering_id and not offering:
            errors.append('Selected offering is not available for this company.')
            return errors
        instance.offering = offering
    elif instance.offering_id and instance.offering.company_id != company.pk:
        errors.append('Selected offering is not available for this company.')
        return errors
    if not is_locked('menu_date'):
        instance.menu_date = request.POST.get('menu_date') or None
    instance.price = price
    instance.company_price = company_price
    instance.room_service_extra_percent = room_service_extra_percent
    instance.packing_price = packing_price
    instance.min_qty = min_qty
    instance.max_qty = max_qty
    instance.web_qty = web_qty
    instance.pos_qty = pos_qty
    instance.preparation_time_minutes = preparation_time_minutes
    instance.calories = calories
    instance.description = description
    instance.position_order = position_order
    if not is_locked('is_active'):
        instance.is_active = request.POST.get('is_active') == 'on'
    if not is_locked('is_kiosk_active'):
        instance.is_kiosk_active = request.POST.get('is_kiosk_active') == 'on'
    if not is_locked('is_pos_active'):
        instance.is_pos_active = request.POST.get('is_pos_active') == 'on'
    if not is_locked('featured_in_web'):
        instance.featured_in_web = request.POST.get('featured_in_web') == 'on'
    if not is_locked('featured_in_kiosk_extra'):
        instance.featured_in_kiosk_extra = request.POST.get('featured_in_kiosk_extra') == 'on'
    # schedule_bypass: only superadmin can enable; non-superadmin always resets to False
    if request.user.role == 'superadmin':
        instance.schedule_bypass = request.POST.get('schedule_bypass') == 'on'
    elif instance.pk is None:
        instance.schedule_bypass = False

    if not is_locked('available_from'):
        ft_from = parse_time(request.POST.get('available_from', '') or '')
        if ft_from:
            instance.available_from = ft_from
    if not is_locked('available_to'):
        ft_to = parse_time(request.POST.get('available_to', '') or '')
        if ft_to:
            instance.available_to = ft_to

    if not is_locked('image') and 'image' in request.FILES:
        uploaded_image = request.FILES['image']
        _copy_upload_to_gallery(
            ProductGallery,
            image_file=uploaded_image,
            company=company,
            user=request.user,
            name=name,
        )
        instance.image = uploaded_image
    elif not is_locked('gallery_image_url') and request.POST.get('gallery_image_url'):
        # Selected from gallery — copy the file reference
        gallery_url = request.POST.get('gallery_image_url', '').strip()
        if gallery_url:
            # Extract relative path from URL for ImageField
            try:
                from urllib.parse import urlparse
                path = urlparse(gallery_url).path
                # Find gallery entry by URL match
                from django.conf import settings
                media_prefix = settings.MEDIA_URL
                if path.startswith(media_prefix):
                    relative = path[len(media_prefix):]
                    # Set the image field to the same file path
                    instance.image = relative
            except (AttributeError, ValueError):
                pass

    instance.save()

    if not is_locked('counter_ids'):
        counter_ids = request.POST.getlist('counter_ids') or request.POST.getlist('counters')
        ProductCounter.objects.filter(product=instance).delete()
        for idx, counter in enumerate(Counter.objects.filter(pk__in=counter_ids, company=company, is_deleted=False, is_active=True)):
            ProductCounter.objects.create(product=instance, counter=counter, position_order=idx)

    if not is_locked('food_types'):
        food_type_ids = _selected_food_type_ids_from_request(request)
        instance.food_type.set(FoodType.objects.filter(pk__in=food_type_ids) if food_type_ids else [])

    if is_locked('is_free_meal_product'):
        return None

    is_free_meal_product = request.POST.get('is_free_meal_product') == 'on'
    instance.free_meal_companies.clear()

    if is_free_meal_product and instance.company_id:
        instance.company.free_meal_products.add(instance)
    elif instance.company_id:
        instance.company.free_meal_products.remove(instance)

    return None




@staff_role_required('superadmin','admin')
def dashboard_product_copy(request, pk):
    import uuid as _uuid
    if has_any_granular_perms(request.user) and not user_can_action(request.user, 'perm_products', 'copy'):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard:no_access')
    source = _own_product_or_404(pk, request.user)
    new_product = Product.objects.get(pk=source.pk)
    new_product.pk = None
    # Use UUID suffix to avoid UniqueConstraint(company, slug) conflict on repeated copies
    new_product.slug = f"{slugify(source.name)}-{_uuid.uuid4().hex[:6]}"
    new_product.code = ''
    new_product.name = f"{source.name} Copy"
    new_product.save()
    new_product.food_type.set(source.food_type.all())
    for idx, mapping in enumerate(source.counter_mappings.select_related('counter').all()):
        ProductCounter.objects.create(product=new_product, counter=mapping.counter, position_order=idx)
    messages.success(request, f'Copied "{source.name}" into a new product draft.')
    return redirect('dashboard:product_edit', pk=new_product.pk)


@staff_role_required('superadmin','admin')
def product_bulk_sample_download(request):
    data = _build_sample_product_workbook_bytes()
    resp = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="neverq_product_bulk_sample.xlsx"'
    return resp


@staff_role_required('superadmin', 'admin')
def dashboard_offering_list(request):
    qs = Offering.objects.filter(is_deleted=False).select_related('company').prefetch_related('schedules').order_by('company__name','position_order','name')
    if not request.user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(request.user))
    return render(request, 'dashboard/menu/offering_list.html', {'offerings': qs, 'page_title': 'Offerings', 'lp': get_list_perms(request.user, 'perm_offerings')})


@staff_role_required('superadmin', 'admin')
def dashboard_offering_add(request):
    if not user_can_action(request.user, 'perm_offerings', 'add'):
        return _deny_dashboard_action(request)
    companies = _user_companies(request.user)
    if request.method == 'POST':
        company = _selected_company_for_user(request.user, request.POST.get('company'))
        name = (request.POST.get('name') or '').strip()
        if company and name:
            valid_days = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
            off = Offering.objects.create(
                company=company,
                name=name,
                available_from=parse_time((request.POST.get('available_from') or '').strip()) or None,
                available_to=parse_time((request.POST.get('available_to') or '').strip()) or None,
                prep_start_time=parse_time((request.POST.get('prep_start_time') or '').strip()) or None,
                position_order=int(request.POST.get('position_order') or 0),
                open_days=[d for d in request.POST.getlist('open_days') if d in valid_days],
                is_active=request.POST.get('is_active') == 'on',
            )
            if 'image' in request.FILES:
                uploaded_image = request.FILES['image']
                _copy_upload_to_gallery(
                    OfferingGallery,
                    image_file=uploaded_image,
                    company=company,
                    user=request.user,
                    name=name,
                )
                off.image = uploaded_image
                off.save(update_fields=['image'])
            elif request.POST.get('gallery_image_url'):
                from urllib.parse import urlparse
                from django.conf import settings
                path = urlparse(request.POST.get('gallery_image_url')).path
                if path.startswith(settings.MEDIA_URL):
                    off.image = path[len(settings.MEDIA_URL):]
                    off.save(update_fields=['image'])
            # Phase 2: save multi-window schedules
            _save_offering_schedules(request.POST, off)
            messages.success(request, 'Offering added successfully.')
            return redirect('dashboard:offering_list')
        messages.error(request, 'Company and offering name are required.')
    return render(request, 'dashboard/menu/offering_form.html', {
        'companies': companies, 'page_title': 'Add Offering', 'action': 'Add',
        'weekday_choices': CATEGORY_WEEKDAY_CHOICES,
        'offering_open_days': request.POST.getlist('open_days') if request.method == 'POST' else [],
    })


@staff_role_required('superadmin', 'admin')
def dashboard_offering_edit(request, pk):
    offering = get_object_or_404(Offering, pk=pk, is_deleted=False)
    _perm = check_module_permission(request, 'perm_offerings')
    if _perm: return _perm
    if not request.user.is_superadmin and not user_can_access_company(request.user, offering.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:offering_list')
    if request.method == 'POST':
        if request.user.role != 'superadmin' and get_module_level(request.user, 'perm_offerings') == 'full_edit':
            _diffs = {
                'name':           {'label': 'Name',       'before': offering.name,           'after': (request.POST.get('name') or offering.name).strip()},
                'is_active':      {'label': 'Active',     'before': offering.is_active,      'after': request.POST.get('is_active') == 'on'},
                'position_order': {'label': 'Sort Order', 'before': offering.position_order, 'after': int(request.POST.get('position_order') or 0)},
            }
            _pc = create_pending_change(request, 'perm_offerings', offering, _diffs)
            if _pc:
                messages.success(request, 'Your changes have been submitted for superadmin review.')
            else:
                messages.info(request, 'No changes detected.')
            return _dashboard_return_or_default(request, 'dashboard:offering_list')
        offering.name = (request.POST.get('name') or offering.name).strip()
        offering.available_from = parse_time((request.POST.get('available_from') or '').strip()) or None
        offering.available_to = parse_time((request.POST.get('available_to') or '').strip()) or None
        offering.prep_start_time = parse_time((request.POST.get('prep_start_time') or '').strip()) or None
        offering.position_order = int(request.POST.get('position_order') or 0)
        valid_days = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']
        offering.open_days = [d for d in request.POST.getlist('open_days') if d in valid_days]
        offering.is_active = request.POST.get('is_active') == 'on'
        if 'image' in request.FILES:
            uploaded_image = request.FILES['image']
            _copy_upload_to_gallery(
                OfferingGallery,
                image_file=uploaded_image,
                company=offering.company,
                user=request.user,
                name=offering.name,
            )
            offering.image = uploaded_image
        elif request.POST.get('gallery_image_url'):
            from urllib.parse import urlparse
            from django.conf import settings
            path = urlparse(request.POST.get('gallery_image_url')).path
            if path.startswith(settings.MEDIA_URL):
                offering.image = path[len(settings.MEDIA_URL):]
        offering.save()
        # Phase 2: save multi-window schedules
        _save_offering_schedules(request.POST, offering)
        messages.success(request, 'Offering updated.')
        return _dashboard_return_or_default(request, 'dashboard:offering_list')
    # Pass existing schedules to template
    existing_schedules = [
        {'day': s.display_day, 'start': s.start_time.strftime('%H:%M'), 'end': s.end_time.strftime('%H:%M')}
        for s in offering.schedules.all()
    ]
    _lj, _pl = get_locked_html_names(request.user, 'perm_offerings')
    return render(request, 'dashboard/menu/offering_form.html', {
        'offering': offering, 'companies': [offering.company],
        'page_title': f'Edit - {offering.name}', 'action': 'Save',
        'weekday_choices': CATEGORY_WEEKDAY_CHOICES,
        'offering_open_days': request.POST.getlist('open_days') if request.method == 'POST' else (offering.open_days or []),
        'existing_schedules': existing_schedules,
        'perm_level': _pl or 'full_edit', 'locked_names_json': _lj,
    })


@staff_role_required('superadmin')
def dashboard_counter_list(request):
    qs = Counter.objects.filter(is_deleted=False).select_related('company','cafe').order_by('company__name','position_order','name')
    if not request.user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(request.user))
    cafes = Cafe.objects.filter(is_deleted=False, is_active=True)
    if not request.user.is_superadmin:
        cafes = cafes.filter(company__in=_user_companies(request.user))
    return render(request, 'dashboard/menu/counter_list.html', {'counters': qs, 'cafes': cafes, 'page_title': 'Counters', 'lp': get_list_perms(request.user, 'perm_counters')})


@staff_role_required('superadmin')
def dashboard_counter_add(request):
    if not user_can_action(request.user, 'perm_counters', 'add'):
        return _deny_dashboard_action(request)
    companies = _user_companies(request.user)
    cafes = Cafe.objects.filter(is_deleted=False, is_active=True)
    if not request.user.is_superadmin:
        cafes = cafes.filter(company__in=companies)
    if request.method == 'POST':
        company = _selected_company_for_user(request.user, request.POST.get('company'))
        name = (request.POST.get('name') or '').strip()
        if company and name:
            cafe = Cafe.objects.filter(pk=request.POST.get('cafe'), company=company).first() if request.POST.get('cafe') else None
            Counter.objects.create(
                company=company,
                cafe=cafe,
                name=name,
                code=(request.POST.get('code') or '').strip(),
                printer_label=(request.POST.get('printer_label') or '').strip(),
                auto_print_on_ready=request.POST.get('auto_print_on_ready') == 'on',
                auto_print_on_scan=request.POST.get('auto_print_on_scan') != 'off',
                position_order=int(request.POST.get('position_order') or 0),
                is_active=request.POST.get('is_active') == 'on'
            )
            messages.success(request, 'Counter added successfully.')
            return redirect('dashboard:counter_list')
        messages.error(request, 'Company and counter name are required.')
    return render(request, 'dashboard/menu/counter_form.html', {'companies': companies, 'cafes': cafes, 'page_title': 'Add Counter', 'action': 'Add'})


@staff_role_required('superadmin')
def dashboard_counter_edit(request, pk):
    counter = get_object_or_404(Counter, pk=pk, is_deleted=False)
    _perm = check_module_permission(request, 'perm_counters')
    if _perm: return _perm
    if not request.user.is_superadmin and not user_can_access_company(request.user, counter.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:counter_list')
    cafes = Cafe.objects.filter(is_deleted=False, is_active=True, company=counter.company)
    if request.method == 'POST':
        if request.user.role != 'superadmin' and get_module_level(request.user, 'perm_counters') == 'full_edit':
            _diffs = {
                'name':           {'label': 'Name',          'before': counter.name,              'after': (request.POST.get('name') or counter.name).strip()},
                'code':           {'label': 'Code',          'before': counter.code or '',         'after': (request.POST.get('code') or '').strip()},
                'is_active':      {'label': 'Active',        'before': counter.is_active,          'after': request.POST.get('is_active') == 'on'},
                'position_order': {'label': 'Sort Order',    'before': counter.position_order,     'after': int(request.POST.get('position_order') or 0)},
                'printer_label':  {'label': 'Printer Label', 'before': counter.printer_label or '', 'after': (request.POST.get('printer_label') or '').strip()},
            }
            _pc = create_pending_change(request, 'perm_counters', counter, _diffs)
            if _pc:
                messages.success(request, 'Your changes have been submitted for superadmin review.')
            else:
                messages.info(request, 'No changes detected.')
            return _dashboard_return_or_default(request, 'dashboard:counter_list')
        counter.name = (request.POST.get('name') or counter.name).strip()
        counter.code = (request.POST.get('code') or '').strip()
        counter.cafe = Cafe.objects.filter(pk=request.POST.get('cafe'), company=counter.company).first() if request.POST.get('cafe') else None
        counter.printer_label = (request.POST.get('printer_label') or '').strip()
        counter.auto_print_on_ready = request.POST.get('auto_print_on_ready') == 'on'
        counter.auto_print_on_scan = request.POST.get('auto_print_on_scan') != 'off'
        counter.position_order = int(request.POST.get('position_order') or 0)
        counter.is_active = request.POST.get('is_active') == 'on'
        counter.save()
        messages.success(request, 'Counter updated.')
        return _dashboard_return_or_default(request, 'dashboard:counter_list')
    _lj, _pl = get_locked_html_names(request.user, 'perm_counters')
    return render(request, 'dashboard/menu/counter_form.html', {'counter': counter, 'companies': [counter.company], 'cafes': cafes, 'page_title': f'Edit - {counter.name}', 'action': 'Save', 'perm_level': _pl or 'full_edit', 'locked_names_json': _lj})


@staff_role_required('superadmin','admin','cafeman','pos','reports')
def dashboard_offer_list(request):
    qs = Offer.objects.filter(is_deleted=False).select_related('company','product','cafe').order_by('-created_at')
    if not request.user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(request.user))
    return render(request, 'dashboard/menu/offer_list.html', {'offers': qs, 'page_title': 'Offers', 'lp': get_list_perms(request.user, 'perm_offers')})


@staff_role_required('superadmin', 'admin')
def offer_cafe_options(request):
    """AJAX: return cafes for a given company_id as JSON (used by offer form cascade)."""
    from django.http import JsonResponse
    company_id = request.GET.get('company_id', '').strip()
    if not company_id:
        return JsonResponse({'cafes': []})
    # Superadmin can query any company; admin can only query their own.
    if request.user.is_superadmin:
        qs = Cafe.objects.filter(company_id=company_id, is_active=True, is_deleted=False).order_by('name')
    else:
        if not user_can_access_company(request.user, company_id):
            return JsonResponse({'cafes': []})
        qs = Cafe.objects.filter(company_id=company_id, is_active=True, is_deleted=False).order_by('name')
    return JsonResponse({'cafes': [{'pk': c.pk, 'name': c.name} for c in qs]})


@staff_role_required('superadmin','admin')
def dashboard_offer_add(request):
    if not user_can_action(request.user, 'perm_offers', 'add'):
        return _deny_dashboard_action(request)
    companies = _user_companies(request.user)
    company = get_primary_staff_company(request.user) if not request.user.is_superadmin else None
    products = Product.objects.filter(company__in=companies, is_deleted=False) if not request.user.is_superadmin else Product.objects.filter(is_deleted=False)
    cafes = Cafe.objects.filter(company__in=companies, is_deleted=False) if not request.user.is_superadmin else Cafe.objects.filter(is_deleted=False)
    if request.method == 'POST':
        selected_company = _selected_company_for_user(request.user, request.POST.get('company'))
        title = (request.POST.get('title') or '').strip()
        if selected_company and title:
            from django.utils.dateparse import parse_datetime as _pdt
            _sd = request.POST.get('start_datetime','').strip()
            _ed = request.POST.get('end_datetime','').strip()
            from django.utils.timezone import make_aware
            def _parse_aware(s):
                if not s: return None
                try: return make_aware(_pdt(s.replace('T',' ') + ':00'))
                except (ValueError, OverflowError, OSError): return None
            scope = request.POST.get('product_scope', 'none')
            single_product = None
            if scope == 'single':
                single_product = Product.objects.filter(pk=request.POST.get('product'), company=selected_company).first() if request.POST.get('product') else None
            def _safe_decimal(val):
                try: return Decimal(str(val)) if val else None
                except Exception: return None
            offer = Offer(
                company=selected_company,
                cafe=Cafe.objects.filter(pk=request.POST.get('cafe'), company=selected_company).first() if request.POST.get('cafe') else None,
                product=single_product,
                title=title,
                offer_type=request.POST.get('offer_type') or Offer.TYPE_PERCENT,
                value=Decimal(str(request.POST.get('value') or '0')),
                min_order_value=_safe_decimal(request.POST.get('min_order_value', '').strip()),
                max_discount=_safe_decimal(request.POST.get('max_discount', '').strip()),
                start_datetime=_parse_aware(_sd),
                end_datetime=_parse_aware(_ed),
                is_popup_enabled=request.POST.get('is_popup_enabled') == 'on',
                is_active=request.POST.get('is_active') == 'on',
            )
            if 'popup_image' in request.FILES:
                offer.popup_image = request.FILES['popup_image']
            offer.save()
            # Multi-product M2M
            if scope == 'multi':
                product_ids = request.POST.getlist('product_ids')
                offer.products.set(Product.objects.filter(pk__in=product_ids, company=selected_company))
            else:
                offer.products.clear()
            messages.success(request, 'Offer created successfully.')
            return redirect('dashboard:offer_list')
        messages.error(request, 'Company and offer title are required.')
    return render(request, 'dashboard/menu/offer_form.html', {'companies': companies, 'products': products, 'cafes': cafes, 'offer_type_choices': Offer.OFFER_TYPE_CHOICES, 'page_title': 'Add Offer', 'action': 'Add'})


@staff_role_required('superadmin','admin')
def dashboard_offer_edit(request, pk):
    offer = get_object_or_404(Offer, pk=pk, is_deleted=False)
    _perm = check_module_permission(request, 'perm_offers')
    if _perm: return _perm
    if not request.user.is_superadmin and not user_can_access_company(request.user, offer.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:offer_list')
    products = Product.objects.filter(company=offer.company, is_deleted=False)
    cafes = Cafe.objects.filter(company=offer.company, is_deleted=False)
    if request.method == 'POST':
        if request.user.role != 'superadmin' and get_module_level(request.user, 'perm_offers') == 'full_edit':
            _diffs = {
                'title':      {'label': 'Title',      'before': offer.title,      'after': (request.POST.get('title') or offer.title).strip()},
                'is_active':  {'label': 'Active',     'before': offer.is_active,  'after': request.POST.get('is_active') == 'on'},
                'offer_type': {'label': 'Offer Type', 'before': offer.offer_type, 'after': request.POST.get('offer_type') or offer.offer_type},
                'value':      {'label': 'Value',      'before': str(offer.value), 'after': request.POST.get('value', '')},
            }
            _pc = create_pending_change(request, 'perm_offers', offer, _diffs)
            if _pc:
                messages.success(request, 'Your changes have been submitted for superadmin review.')
            else:
                messages.info(request, 'No changes detected.')
            return redirect('dashboard:offer_list')
        from django.utils.dateparse import parse_datetime as _pdt
        _sd = request.POST.get('start_datetime','').strip()
        _ed = request.POST.get('end_datetime','').strip()
        from django.utils.timezone import make_aware as _make_aware
        def _parse_aware_edit(s):
            if not s: return None
            try: return _make_aware(_pdt(s.replace('T',' ') + ':00'))
            except Exception: return None  # datetime parse helper — safe fallback
        scope = request.POST.get('product_scope', 'none')
        offer.title = (request.POST.get('title') or offer.title).strip()
        offer.cafe = Cafe.objects.filter(pk=request.POST.get('cafe'), company=offer.company).first() if request.POST.get('cafe') else None
        offer.product = Product.objects.filter(pk=request.POST.get('product'), company=offer.company).first() if scope == 'single' and request.POST.get('product') else None
        def _safe_dec_edit(val):
            try: return Decimal(str(val)) if val else None
            except Exception: return None
        offer.offer_type       = request.POST.get('offer_type') or offer.offer_type
        offer.value            = Decimal(str(request.POST.get('value') or offer.value))
        offer.min_order_value  = _safe_dec_edit(request.POST.get('min_order_value', '').strip())
        offer.max_discount     = _safe_dec_edit(request.POST.get('max_discount', '').strip())
        offer.start_datetime   = _parse_aware_edit(_sd)
        offer.end_datetime = _parse_aware_edit(_ed)
        offer.is_popup_enabled = request.POST.get('is_popup_enabled') == 'on'
        offer.is_active = request.POST.get('is_active') == 'on'
        if 'popup_image' in request.FILES:
            offer.popup_image = request.FILES['popup_image']
        offer.save()
        # Multi-product M2M
        if scope == 'multi':
            product_ids = request.POST.getlist('product_ids')
            offer.products.set(Product.objects.filter(pk__in=product_ids, company=offer.company))
        else:
            offer.products.clear()
        messages.success(request, 'Offer updated.')
        return redirect('dashboard:offer_list')
    _lj, _pl = get_locked_html_names(request.user, 'perm_offers')
    return render(request, 'dashboard/menu/offer_form.html', {'offer': offer, 'companies': [offer.company], 'products': products, 'cafes': cafes, 'offer_type_choices': Offer.OFFER_TYPE_CHOICES, 'page_title': f'Edit - {offer.title}', 'action': 'Save', 'perm_level': _pl or 'full_edit', 'locked_names_json': _lj})




@require_POST
@staff_role_required('superadmin')
def dashboard_offer_reset_usage(request, pk):
    """
    Superadmin tool: wipes ALL OfferUsage records for the given offer.
    This resets the one-use-per-customer limit so every customer can
    redeem the offer again — useful after testing or if offers were
    accidentally consumed during a buggy order flow.
    """
    offer = get_object_or_404(Offer, pk=pk, is_deleted=False)
    from apps.menu.models import OfferUsage
    deleted_count, _ = OfferUsage.objects.filter(offer=offer).delete()
    messages.success(
        request,
        f'Usage reset for "{offer.title}": {deleted_count} redemption record(s) cleared. '
        f'All customers can now use this offer again.'
    )
    return redirect('dashboard:offer_edit', pk=pk)


@require_POST
@staff_role_required('superadmin', 'admin')
def dashboard_offering_toggle(request, pk):
    if has_any_granular_perms(request.user) and not user_can_action(request.user, 'perm_offerings', 'toggle'):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard:no_access')
    offering = get_object_or_404(Offering, pk=pk, is_deleted=False)
    if not request.user.is_superadmin and not user_can_access_company(request.user, offering.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:offering_list')
    offering.is_active = not offering.is_active
    offering.save(update_fields=['is_active'])
    messages.success(request, f'Offering {"activated" if offering.is_active else "deactivated"}.')
    return redirect('dashboard:offering_list')


@require_POST
@staff_role_required('superadmin', 'admin')
def dashboard_offering_delete(request, pk):
    if get_module_level(request.user, 'perm_offerings') != 'full_edit':
        return _deny_dashboard_action(request)
    offering = get_object_or_404(Offering, pk=pk, is_deleted=False)
    if not request.user.is_superadmin and not user_can_access_company(request.user, offering.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:offering_list')
    offering.is_deleted = True
    offering.is_active = False
    offering.save(update_fields=['is_deleted', 'is_active'])
    messages.success(request, f'Offering "{offering.name}" deleted.')
    return redirect('dashboard:offering_list')


@require_POST
@staff_role_required('superadmin', 'admin')
def dashboard_offering_bulk_delete(request):
    if get_module_level(request.user, 'perm_offerings') != 'full_edit':
        return _deny_dashboard_action(request)
    ids = _bulk_post_ids(request)
    qs = Offering.objects.filter(pk__in=ids, is_deleted=False)
    if not request.user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(request.user))
    count = qs.update(is_deleted=True, is_active=False) if ids else 0
    _bulk_message(request, 'offering', count)
    return redirect('dashboard:offering_list')


@require_POST
@staff_role_required('superadmin')
def dashboard_counter_toggle(request, pk):
    if has_any_granular_perms(request.user) and not user_can_action(request.user, 'perm_counters', 'toggle'):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard:no_access')
    counter = get_object_or_404(Counter, pk=pk, is_deleted=False)
    if not request.user.is_superadmin and not user_can_access_company(request.user, counter.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:counter_list')
    counter.is_active = not counter.is_active
    counter.save(update_fields=['is_active'])
    messages.success(request, f'Counter {"activated" if counter.is_active else "deactivated"}.')
    return redirect('dashboard:counter_list')


@require_POST
@staff_role_required('superadmin')
def dashboard_counter_delete(request, pk):
    if get_module_level(request.user, 'perm_counters') != 'full_edit':
        return _deny_dashboard_action(request)
    counter = get_object_or_404(Counter, pk=pk, is_deleted=False)
    if not request.user.is_superadmin and not user_can_access_company(request.user, counter.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:counter_list')
    counter.is_deleted = True
    counter.is_active = False
    counter.save(update_fields=['is_deleted', 'is_active'])
    messages.success(request, f'Counter "{counter.name}" deleted.')
    return redirect('dashboard:counter_list')


@require_POST
@staff_role_required('superadmin')
def dashboard_counter_bulk_delete(request):
    if get_module_level(request.user, 'perm_counters') != 'full_edit':
        return _deny_dashboard_action(request)
    ids = _bulk_post_ids(request)
    qs = Counter.objects.filter(pk__in=ids, is_deleted=False)
    if not request.user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(request.user))
    count = qs.update(is_deleted=True, is_active=False) if ids else 0
    _bulk_message(request, 'counter', count)
    return redirect('dashboard:counter_list')


@require_POST
@staff_role_required('superadmin','admin')
def dashboard_offer_toggle(request, pk):
    if has_any_granular_perms(request.user) and not user_can_action(request.user, 'perm_offers', 'toggle'):
        messages.error(request, 'Permission denied.')
        return redirect('dashboard:no_access')
    offer = get_object_or_404(Offer, pk=pk, is_deleted=False)
    if not request.user.is_superadmin and not user_can_access_company(request.user, offer.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:offer_list')
    offer.is_active = not offer.is_active
    offer.save(update_fields=['is_active'])
    messages.success(request, f'Offer {"activated" if offer.is_active else "deactivated"}.')
    return redirect('dashboard:offer_list')


@require_POST
@staff_role_required('superadmin','admin')
def dashboard_offer_delete(request, pk):
    if get_module_level(request.user, 'perm_offers') != 'full_edit':
        return _deny_dashboard_action(request)
    offer = get_object_or_404(Offer, pk=pk, is_deleted=False)
    if not request.user.is_superadmin and not user_can_access_company(request.user, offer.company_id):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:offer_list')
    offer.is_deleted = True
    offer.is_active = False
    offer.save(update_fields=['is_deleted', 'is_active'])
    messages.success(request, f'Offer "{offer.title}" deleted.')
    return redirect('dashboard:offer_list')


@require_POST
@staff_role_required('superadmin','admin')
def dashboard_offer_bulk_delete(request):
    if get_module_level(request.user, 'perm_offers') != 'full_edit':
        return _deny_dashboard_action(request)
    ids = _bulk_post_ids(request)
    qs = Offer.objects.filter(pk__in=ids, is_deleted=False)
    if not request.user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(request.user))
    count = qs.update(is_deleted=True, is_active=False) if ids else 0
    _bulk_message(request, 'offer', count)
    return redirect('dashboard:offer_list')



# ════════════════════════════════════════════════════════════════
#  COUNTER TICKETS  —  per-counter fulfilment board
# ════════════════════════════════════════════════════════════════

@staff_role_required('superadmin','admin','cafeman','pos')
def counter_ticket_board(request):
    from apps.orders.models import CounterTicket, OrderStatusChoices as _OSC
    user = request.user

    # Keep auto-ready order status and counter ticket board in sync
    if user.is_superadmin:
        _promote_due_ready_orders()
    else:
        for company in _user_companies(user):
            _promote_due_ready_orders(company=company)

    qs = CounterTicket.objects.filter(
        status__in=['pending', 'preparing', 'ready']
    ).exclude(
        order__order_status__in=[_OSC.DELIVERED, _OSC.CANCELLED]
    ).select_related(
        'order', 'order__customer', 'counter', 'company'
    ).prefetch_related('order__items__product')

    if not user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(user))

    counter_filter = request.GET.get('counter', '')
    if counter_filter:
        qs = qs.filter(counter_id=counter_filter)

    counters = Counter.objects.filter(is_deleted=False, is_active=True)
    if not user.is_superadmin:
        counters = counters.filter(company__in=_user_companies(user))

    return render(request, 'dashboard/orders/counter_ticket_board.html', {
        'tickets': qs,
        'counters': counters,
        'counter_filter': counter_filter,
        'page_title': 'Counter Ticket Board',
    })


@require_POST
@staff_role_required('superadmin','admin','cafeman','pos')
def counter_ticket_update(request, pk):
    from apps.orders.models import CounterTicket, OrderStatus, OrderStatusChoices
    from django.utils import timezone
    ticket = get_object_or_404(CounterTicket, pk=pk)
    if not request.user.is_superadmin and not user_can_access_company(request.user, ticket.company_id):
        return JsonResponse({'success': False, 'error': 'Access denied'}, status=403)
    new_status = request.POST.get('status')
    allowed = [s for s, _ in CounterTicket.STATUS_CHOICES]
    auto_print = False
    if new_status in allowed:
        order = ticket.order
        if order.order_status in (OrderStatusChoices.DELIVERED, OrderStatusChoices.CANCELLED):
            error_msg = 'This order is already in a final state and ticket status cannot change it.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': error_msg}, status=400)
            messages.error(request, error_msg)
            return redirect('dashboard:counter_ticket_board')
        if new_status == CounterTicket.STATUS_COLLECTED:
            now = timezone.now()
            for item in order.items.filter(counter=ticket.counter, is_deleted=False, picked_up_at__isnull=True):
                item.picked_up_at = now
                item.save(update_fields=['picked_up_at'])
            ticket.status = CounterTicket.STATUS_COLLECTED
            ticket.collected_at = now
            ticket.save(update_fields=['status', 'collected_at', 'updated_at'])
            remaining = order.items.filter(is_deleted=False, picked_up_at__isnull=True).count()
            if remaining == 0:
                if order.order_status < OrderStatusChoices.DELIVERED:
                    order.order_status = OrderStatusChoices.DELIVERED
                    order.auto_ready_at = None
                    order.save(update_fields=['order_status', 'auto_ready_at', 'updated_at'])
                    OrderStatus.objects.create(order=order, status=OrderStatusChoices.DELIVERED, details=f'{ticket.ticket_number} collected and order completed.', created_at=now)
            elif order.order_status < OrderStatusChoices.READY:
                order.order_status = OrderStatusChoices.READY
                order.save(update_fields=['order_status', 'updated_at'])
                OrderStatus.objects.create(order=order, status=OrderStatusChoices.READY, details=f'{ticket.ticket_number} collected.', created_at=now)
        else:
            ticket.status = new_status
            ticket.save(update_fields=['status', 'updated_at'])
            if new_status == CounterTicket.STATUS_READY and ticket.counter and ticket.counter.auto_print_on_ready:
                auto_print = True
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if is_ajax:
        return JsonResponse({'success': True, 'status': ticket.status, 'kot': ticket.kot_data(), 'auto_print': auto_print})
    return redirect('dashboard:counter_ticket_board')


@staff_role_required('superadmin','admin','cafeman','pos')
def counter_ticket_kot(request, pk):
    from apps.orders.models import CounterTicket
    ticket = get_object_or_404(CounterTicket, pk=pk)
    if not request.user.is_superadmin and not user_can_access_company(request.user, ticket.company_id):
        return JsonResponse({'error': 'Access denied'}, status=403)
    return JsonResponse(ticket.kot_data())


# ════════════════════════════════════════════════════════════════
#  SITE-WISE PRICING  —  per-company product price overrides
# ════════════════════════════════════════════════════════════════

@staff_role_required('superadmin', 'admin')
def site_price_list(request):
    _perm = check_module_permission(request, 'perm_site_prices')
    if _perm: return _perm
    from apps.menu.models import ProductCompanyPrice

    user = request.user
    scoped_company = user.company if not user.is_superadmin else None

    # ── Filter params ───────────────────────────────────────────
    q               = request.GET.get('q', '').strip()
    filter_company  = request.GET.get('company', '').strip()
    filter_building = request.GET.get('building', '').strip()
    filter_cafe     = request.GET.get('cafe', '').strip()
    filter_category = request.GET.get('category', '').strip()
    filter_offering = request.GET.get('offering', '').strip()
    filter_status   = request.GET.get('status', '').strip()

    product_qs = Product.objects.filter(is_deleted=False).select_related(
        'category', 'offering', 'company'
    ).prefetch_related('food_type')
    if scoped_company:
        product_qs = product_qs.filter(company=scoped_company)

    if q:
        product_qs = product_qs.filter(
            Q(name__icontains=q) | Q(code__icontains=q) |
            Q(category__name__icontains=q) | Q(offering__name__icontains=q)
        )

    # Company filter (superadmin only — non-superadmin is already scoped)
    if filter_company and user.is_superadmin:
        product_qs = product_qs.filter(company_id=filter_company)

    # Building filter: show all products whose company owns this building.
    # Previously this only showed products that already had a site price override
    # for that building — so it always returned 0 on a fresh setup.
    if filter_building:
        company_ids = Building.objects.filter(pk=filter_building).values_list('company_id', flat=True)
        product_qs = product_qs.filter(company_id__in=company_ids)

    # Cafe filter: show all products whose company owns this cafe.
    if filter_cafe:
        from apps.menu.models import Cafe as _Cafe
        company_ids = _Cafe.objects.filter(pk=filter_cafe).values_list('company_id', flat=True)
        product_qs = product_qs.filter(company_id__in=company_ids)

    if filter_category:
        product_qs = product_qs.filter(category_id=filter_category)
    if filter_offering:
        product_qs = product_qs.filter(offering_id=filter_offering)
    if filter_status == 'active':
        product_qs = product_qs.filter(is_active=True)
    elif filter_status == 'inactive':
        product_qs = product_qs.filter(is_active=False)

    product_qs = product_qs.order_by('company__name', 'category__name', 'position_order', 'name')

    existing_qs = ProductCompanyPrice.objects.select_related(
        'product', 'product__category', 'product__offering', 'company', 'building', 'cafe',
    ).order_by('company__name', 'building__name', 'cafe__name', 'product__name')
    if scoped_company:
        existing_qs = existing_qs.filter(company=scoped_company)
    overridden_pks = set(existing_qs.values_list('product_id', flat=True))

    companies   = _user_companies(user)
    building_qs = Building.objects.filter(is_deleted=False)
    cafe_qs     = Cafe.objects.filter(is_deleted=False).select_related('building')
    category_qs = Category.objects.filter(is_deleted=False)
    offering_qs = Offering.objects.filter(is_deleted=False)

    # For non-superadmin, always scope to their company
    if scoped_company:
        building_qs = building_qs.filter(company=scoped_company)
        cafe_qs     = cafe_qs.filter(company=scoped_company)
        category_qs = category_qs.filter(companies=scoped_company)
        offering_qs = offering_qs.filter(company=scoped_company)
    elif filter_company:
        # Superadmin selected a specific company — scope dropdowns to that company
        building_qs = building_qs.filter(company_id=filter_company)
        cafe_qs     = cafe_qs.filter(company_id=filter_company)
        category_qs = category_qs.filter(companies__pk=filter_company)
        offering_qs = offering_qs.filter(company_id=filter_company)
    elif filter_building:
        # Scope cafes to the selected building
        cafe_qs = cafe_qs.filter(building_id=filter_building)

    is_filtered = bool(q or filter_company or filter_building or filter_cafe or filter_category or filter_offering or filter_status)

    return render(request, 'dashboard/menu/site_price_list.html', {
        'products':       product_qs,
        'overridden_pks': overridden_pks,
        'companies':      companies,
        'buildings':      building_qs.order_by('name'),
        'cafes':          cafe_qs.order_by('building__name', 'name'),
        'categories':     category_qs.order_by('name'),
        'offerings':      offering_qs.order_by('name'),
        'f_q':            q,
        'f_company':      filter_company,
        'f_building':     filter_building,
        'f_cafe':         filter_cafe,
        'f_category':     filter_category,
        'f_offering':     filter_offering,
        'f_status':       filter_status,
        'total_count':    product_qs.count(),
        'is_filtered':    is_filtered,
        'page_title':     'Site Pricing',
    })


@staff_role_required('superadmin','admin')
def site_price_set(request):
    """Create or update a site-wise price override at company/building/cafe level."""
    from apps.menu.models import ProductCompanyPrice
    companies = _user_companies(request.user)
    selected_company = request.user.company if not request.user.is_superadmin else None
    edit_obj = None
    if request.GET.get('edit'):
        edit_obj = get_object_or_404(ProductCompanyPrice.objects.select_related('company','building','cafe','product'), pk=request.GET.get('edit'))
        if not request.user.is_superadmin and request.user.company_id != edit_obj.company_id:
            messages.error(request, 'Access denied.')
            return redirect('dashboard:site_price_list')
        selected_company = edit_obj.company
    if request.method == 'POST':
        edit_id = request.POST.get('edit_id')
        selected_company = _selected_company_for_user(request.user, request.POST.get('company')) or selected_company
        product_id = request.POST.get('product')
        building_id = request.POST.get('building')
        cafe_id = request.POST.get('cafe')
        building = _scoped_building_for_company(selected_company, building_id)
        cafe = _scoped_cafe_for_company(selected_company, cafe_id)
        price_val  = request.POST.get('price', '').strip()
        is_active  = request.POST.get('is_active') == 'on'
        if selected_company and product_id and price_val:
            try:
                product = Product.objects.get(pk=product_id, company=selected_company, is_deleted=False)
                if building_id and not building:
                    raise ValueError('Selected building does not belong to the selected company.')
                if cafe_id and not cafe:
                    raise ValueError('Selected cafe does not belong to the selected company.')
                if cafe and cafe.building_id:
                    if building and building.pk != cafe.building_id:
                        raise ValueError('Selected cafe belongs to a different building.')
                    building = cafe.building
                price = Decimal(price_val)
                if edit_id:
                    obj = get_object_or_404(ProductCompanyPrice, pk=edit_id)
                    obj.product = product
                    obj.company = selected_company
                    obj.building = building
                    obj.cafe = cafe
                    obj.price = price
                    obj.is_active = is_active
                    obj.full_clean()
                    obj.save()
                    created = False
                else:
                    lookup = {'product': product}
                    if cafe:
                        lookup['cafe'] = cafe
                    elif building:
                        lookup['building'] = building
                    else:
                        lookup['company'] = selected_company
                        lookup['building'] = None
                        lookup['cafe'] = None
                    defaults = {'company': selected_company, 'building': building if not cafe else None, 'cafe': cafe, 'price': price, 'is_active': is_active}
                    obj, created = ProductCompanyPrice.objects.update_or_create(defaults=defaults, **lookup)
                messages.success(request, f'Price {"set" if created else "updated"} for {product.name} @ {obj.scope_label}.')
                return redirect('dashboard:site_price_list')
            except (ValidationError, IntegrityError, ValueError, TypeError) as e:
                messages.error(request, f'Error: {e}')
        else:
            messages.error(request, 'Company, product, and price are required.')
    scoped_company = selected_company or (edit_obj.company if edit_obj else None)
    products = Product.objects.filter(company=scoped_company, is_deleted=False) if scoped_company else Product.objects.filter(is_deleted=False)
    buildings = Building.objects.filter(company=scoped_company, is_deleted=False).order_by('name') if scoped_company else Building.objects.filter(is_deleted=False).select_related('company').order_by('company__name','name')
    cafes = Cafe.objects.filter(company=scoped_company, is_deleted=False).select_related('building').order_by('building__name','name') if scoped_company else Cafe.objects.filter(is_deleted=False).select_related('company','building').order_by('company__name','building__name','name')
    return render(request, 'dashboard/menu/site_price_form.html', {
        'companies': companies,
        'products': products,
        'buildings': buildings,
        'cafes': cafes,
        'edit_obj': edit_obj,
        'selected_company': scoped_company,
        'page_title': 'Set Site Price' if not edit_obj else f'Edit Site Price — {edit_obj.product.name}',
    })


@staff_role_required('superadmin', 'admin')
def site_price_copy_product(request, pk):
    """
    Copy a product to a target site (company/building/cafe).

    GET  — returns modal context JSON for AJAX modal population.
    POST — creates (or updates) a ProductCompanyPrice override.
           If copying cross-company, also creates a new Product record
           under the target company so the product actually appears in
           that company's customer menu.
    """
    from apps.menu.models import ProductCompanyPrice
    from django.core.exceptions import ValidationError

    product = get_object_or_404(Product, pk=pk, is_deleted=False)
    user = request.user
    if not user.is_superadmin and user.company_id != product.company_id:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'error': 'Access denied.'}, status=403)
        messages.error(request, 'Access denied.')
        return redirect('dashboard:site_price_list')

    companies  = _user_companies(user)
    all_buildings = Building.objects.filter(is_deleted=False).select_related('company').order_by('company__name', 'name')
    all_cafes     = Cafe.objects.filter(is_deleted=False).select_related('building', 'company').order_by('company__name', 'building__name', 'name')
    all_categories = Category.objects.filter(is_deleted=False).order_by('name')
    all_offerings  = Offering.objects.filter(is_deleted=False).order_by('name')

    if not user.is_superadmin and user.company:
        all_buildings  = all_buildings.filter(company=user.company)
        all_cafes      = all_cafes.filter(company=user.company)
        all_categories = all_categories.filter(companies=user.company)
        all_offerings  = all_offerings.filter(company=user.company)

    if request.method == 'GET':
        # Return data for modal population
        return JsonResponse({
            'product': {
                'pk': product.pk,
                'name': product.name,
                'price': str(product.price),
                'position_order': product.position_order,
                'category_id': product.category_id,
                'offering_id': product.offering_id,
                'company_id': product.company_id,
            },
            'companies':  list(companies.values('pk', 'name')),
            'buildings':  list(all_buildings.values('pk', 'name', 'company_id')),
            'cafes':      list(all_cafes.values('pk', 'name', 'building_id', 'company_id')),
            'categories': list(all_categories.values('pk', 'name')),
            'offerings':  list(all_offerings.values('pk', 'name')),
        })

    # ── POST: perform the copy ─────────────────────────────────────
    target_company_id = request.POST.get('company') or product.company_id
    target_company    = _selected_company_for_user(user, target_company_id) or product.company
    building_id = request.POST.get('building')
    cafe_id = request.POST.get('cafe')
    building = _scoped_building_for_company(target_company, building_id)
    cafe = _scoped_cafe_for_company(target_company, cafe_id)
    if building_id and not building:
        messages.error(request, 'Selected building does not belong to the target company.')
        return redirect('dashboard:site_price_list')
    if cafe_id and not cafe:
        messages.error(request, 'Selected cafe does not belong to the target company.')
        return redirect('dashboard:site_price_list')
    if cafe and cafe.building_id and not building:
        building = cafe.building

    price_raw = (request.POST.get('price') or '').strip()
    try:
        price = Decimal(price_raw) if price_raw else product.price
    except (InvalidOperation, ValueError):
        price = product.price

    category_id  = request.POST.get('category') or product.category_id
    offering_id  = request.POST.get('offering') or ''
    category = _scoped_category_for_company(target_company, category_id)
    if category_id and not category:
        messages.error(request, 'Selected category does not belong to the target company.')
        return redirect('dashboard:site_price_list')
    offering = None
    if offering_id and offering_id != '__clear__':
        offering = _scoped_offering_for_company(target_company, offering_id)
        if not offering:
            messages.error(request, 'Selected offering does not belong to the target company.')
            return redirect('dashboard:site_price_list')
    position_raw = (request.POST.get('position_order') or '').strip()
    is_active    = request.POST.get('is_active') == 'on'

    # ── Cross-company copy: duplicate product into target company ──
    if target_company.pk != product.company_id:
        from django.utils.text import slugify as _slugify
        import uuid as _uuid
        new_product = Product(
            company      = target_company,
            category     = category,
            name         = product.name,
            slug         = f'{_slugify(product.name)}-{_uuid.uuid4().hex[:6]}',
            code         = product.code,
            price        = price,
            company_price= product.company_price,
            room_service_extra_percent=product.room_service_extra_percent,
            packing_price= product.packing_price,
            min_qty      = product.min_qty,
            max_qty      = product.max_qty,
            web_qty      = product.web_qty,
            pos_qty      = product.pos_qty,
            description  = product.description,
            is_active    = is_active,
            position_order = int(position_raw) if position_raw else product.position_order,
            preparation_time_minutes = product.preparation_time_minutes,
        )
        if offering_id and offering_id != '__clear__':
            new_product.offering = offering
        new_product.save()
        # Copy image reference if exists
        if product.image:
            new_product.image = product.image
            new_product.save(update_fields=['image'])
        messages.success(request, f'"{product.name}" copied to {target_company.name} as a new product. It now appears in that company\'s customer menu.')
        return redirect('dashboard:site_price_list')

    # ── Same-company copy: create/update a site price override ────
    # Apply field overrides on the original product
    changed = []
    if position_raw:
        try:
            product.position_order = int(position_raw)
            changed.append('position_order')
        except (ValueError, TypeError):
            pass
    if category_id:
        product.category = category
        changed.append('category')
    if offering_id == '__clear__':
        product.offering = None
        changed.append('offering')
    elif offering_id:
        product.offering = offering
        changed.append('offering')
    if changed:
        product.save(update_fields=changed)

    lookup = {'product': product}
    defaults = {'company': target_company, 'price': price, 'is_active': is_active}
    if cafe:
        lookup['cafe'] = cafe
        defaults['building'] = cafe.building
        defaults['cafe']     = cafe
    elif building:
        lookup['building'] = building
        defaults['building'] = building
        defaults['cafe']     = None
    else:
        lookup['company']  = target_company
        lookup['building'] = None
        lookup['cafe']     = None
        defaults['building'] = None
        defaults['cafe']     = None

    try:
        obj, created = ProductCompanyPrice.objects.update_or_create(defaults=defaults, **lookup)
        messages.success(request, f'{"Created" if created else "Updated"}: {product.name} → {obj.scope_label} @ ₹{price}')
    except (ValidationError, IntegrityError, ValueError) as e:
        messages.error(request, f'Copy failed: {e}')
    return redirect('dashboard:site_price_list')


@require_POST
@staff_role_required('superadmin','admin')
def site_price_delete(request, pk):
    from apps.menu.models import ProductCompanyPrice
    obj = get_object_or_404(ProductCompanyPrice, pk=pk)
    if not request.user.is_superadmin and request.user.company_id != obj.company_id:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:site_price_list')
    obj.delete()
    messages.success(request, 'Price override removed.')
    return redirect('dashboard:site_price_list')


@staff_role_required('superadmin', 'admin')
def site_price_copy(request, pk):
    """Copy a single site-price entry to a new site, with optional field overrides."""
    from apps.menu.models import ProductCompanyPrice
    source = get_object_or_404(ProductCompanyPrice.objects.select_related('product', 'company', 'building', 'cafe'), pk=pk)
    if not request.user.is_superadmin and request.user.company_id != source.company_id:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:site_price_list')

    companies = _user_companies(request.user)

    if request.method == 'POST':
        target_company = _selected_company_for_user(request.user, request.POST.get('company'))
        if not target_company:
            target_company = source.company
        building_id = request.POST.get('building')
        cafe_id = request.POST.get('cafe')
        building = _scoped_building_for_company(target_company, building_id)
        cafe = _scoped_cafe_for_company(target_company, cafe_id)
        if building_id and not building:
            messages.error(request, 'Selected building does not belong to the target company.')
            return redirect('dashboard:site_price_list')
        if cafe_id and not cafe:
            messages.error(request, 'Selected cafe does not belong to the target company.')
            return redirect('dashboard:site_price_list')
        if cafe and cafe.building_id and not building:
            building = cafe.building

        price_val = request.POST.get('price', '').strip() or str(source.price)
        try:
            price = Decimal(price_val)
        except (InvalidOperation, ValueError):
            price = source.price

        # Allow changing product fields via copy
        product_id = request.POST.get('product')
        product = source.product
        if product_id:
            try:
                product = Product.objects.get(pk=product_id, company=target_company, is_deleted=False)
            except Product.DoesNotExist:
                messages.error(request, 'Selected product does not belong to the target company.')
                return redirect('dashboard:site_price_list')
        elif product.company_id != target_company.pk:
            messages.error(request, 'Selected product does not belong to the target company.')
            return redirect('dashboard:site_price_list')

        # position_order (slno) — stored on the product itself for ordering
        slno_val = request.POST.get('position_order', '').strip()
        if slno_val:
            try:
                product.position_order = int(slno_val)
                product.save(update_fields=['position_order'])
            except (ValueError, TypeError):
                pass

        # category
        category_id = request.POST.get('category')
        if category_id:
            cat = _scoped_category_for_company(target_company, category_id)
            if not cat:
                messages.error(request, 'Selected category does not belong to the target company.')
                return redirect('dashboard:site_price_list')
            product.category = cat
            product.save(update_fields=['category'])

        # offering
        offering_id = request.POST.get('offering')
        if offering_id == '__clear__':
            product.offering = None
            product.save(update_fields=['offering'])
        elif offering_id:
            off = _scoped_offering_for_company(target_company, offering_id)
            if not off:
                messages.error(request, 'Selected offering does not belong to the target company.')
                return redirect('dashboard:site_price_list')
            product.offering = off
            product.save(update_fields=['offering'])

        is_active = request.POST.get('is_active') == 'on'

        # Build lookup for upsert
        lookup = {'product': product}
        if cafe:
            lookup['cafe'] = cafe
        elif building:
            lookup['building'] = building
        else:
            lookup['company'] = target_company
            lookup['building'] = None
            lookup['cafe'] = None
        defaults = {
            'company': target_company,
            'building': building if not cafe else None,
            'cafe': cafe,
            'price': price,
            'is_active': is_active,
        }
        try:
            obj, created = ProductCompanyPrice.objects.update_or_create(defaults=defaults, **lookup)
            messages.success(request, f'{"Copied" if created else "Updated"}: {product.name} → {obj.scope_label} @ ₹{price}')
        except (ValidationError, IntegrityError, ValueError) as e:
            messages.error(request, f'Copy failed: {e}')
        return redirect('dashboard:site_price_list')

    # GET — return JSON context for the modal (AJAX) or redirect
    import json as _json
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        scoped_company = source.company
        products = list(Product.objects.filter(company=scoped_company, is_deleted=False).values('pk', 'name', 'price', 'position_order', 'category_id'))
        buildings = list(Building.objects.filter(company=scoped_company, is_deleted=False).order_by('name').values('pk', 'name'))
        cafes = list(Cafe.objects.filter(company=scoped_company, is_deleted=False).select_related('building').order_by('building__name', 'name').values('pk', 'name', 'building_id'))
        categories = list(Category.objects.filter(companies=scoped_company, is_deleted=False).values('pk', 'name'))
        offerings = list(Offering.objects.filter(company=scoped_company, is_deleted=False).values('pk', 'name'))
        companies_data = list(companies.values('pk', 'name'))
        return JsonResponse({
            'source': {
                'pk': source.pk,
                'product_id': source.product_id,
                'product_name': source.product.name,
                'company_id': source.company_id,
                'company_name': source.company.name,
                'building_id': source.building_id,
                'cafe_id': source.cafe_id,
                'price': str(source.price),
                'position_order': source.product.position_order,
                'category_id': source.product.category_id,
                'offering_id': source.product.offering_id,
                'is_active': source.is_active,
            },
            'companies': companies_data,
            'products': products,
            'buildings': buildings,
            'cafes': cafes,
            'categories': categories,
            'offerings': offerings,
        })
    return redirect('dashboard:site_price_list')


@require_POST
@staff_role_required('superadmin', 'admin')
def site_price_bulk_copy(request):
    """Bulk copy selected products as site-price overrides to a target site."""
    from apps.menu.models import ProductCompanyPrice

    ids_raw = request.POST.get('ids', '')
    try:
        pks = [int(x) for x in ids_raw.split(',') if x.strip().isdigit()]
    except (ValueError, AttributeError):
        pks = []

    if not pks:
        messages.error(request, 'No items selected for bulk copy.')
        return redirect('dashboard:site_price_list')

    target_company = _selected_company_for_user(request.user, request.POST.get('company'))
    building_id = request.POST.get('building')
    cafe_id = request.POST.get('cafe')
    if (building_id or cafe_id) and not target_company:
        messages.error(request, 'Select a target company before selecting a building or cafe.')
        return redirect('dashboard:site_price_list')
    building = _scoped_building_for_company(target_company, building_id)
    cafe = _scoped_cafe_for_company(target_company, cafe_id)
    if building_id and not building:
        messages.error(request, 'Selected building does not belong to the target company.')
        return redirect('dashboard:site_price_list')
    if cafe_id and not cafe:
        messages.error(request, 'Selected cafe does not belong to the target company.')
        return redirect('dashboard:site_price_list')
    if cafe and cafe.building_id and not building:
        building = cafe.building

    price_override = request.POST.get('price_override', '').strip()
    is_active = request.POST.get('is_active') == 'on'

    # PKs are now Product PKs (not ProductCompanyPrice PKs)
    products_qs = Product.objects.filter(pk__in=pks, is_deleted=False)
    if not request.user.is_superadmin and request.user.company:
        products_qs = products_qs.filter(company=request.user.company)

    copied = 0
    errors = 0
    for product in products_qs:
        effective_company = target_company or product.company
        if not effective_company:
            errors += 1
            continue

        price = Decimal(price_override) if price_override else product.price

        lookup = {'product': product}
        if cafe:
            lookup['cafe'] = cafe
        elif building:
            lookup['building'] = building
        else:
            lookup['company'] = effective_company
            lookup['building'] = None
            lookup['cafe'] = None
        defaults = {
            'company': effective_company,
            'building': building if not cafe else None,
            'cafe': cafe,
            'price': price,
            'is_active': is_active,
        }
        try:
            ProductCompanyPrice.objects.update_or_create(defaults=defaults, **lookup)
            copied += 1
        except (ValidationError, IntegrityError):
            errors += 1

    if copied:
        messages.success(request, f'Bulk copy complete: {copied} product{"s" if copied != 1 else ""} copied to site.')
    if errors:
        messages.warning(request, f'{errors} product(s) could not be copied due to conflicts.')
    return redirect('dashboard:site_price_list')


# ════════════════════════════════════════════════════════════════
#  CATEGORY BULK COPY
#  Copy all products in a category (optionally under an offering)
#  to a target company, preserving Offering → Category → Product.
# ════════════════════════════════════════════════════════════════

@staff_role_required('superadmin', 'admin')
def category_bulk_copy(request):
    """
    POST: Copy all products from a source category (and optionally a source
    offering) to a target company.

    Hierarchy enforced: Offering → Category → Product.
    Sub-category logic is completely absent from this flow.

    If the target company already has a category with the same name it is
    reused; otherwise a new one is created and linked to the target company.
    Same logic for offerings.
    """
    if request.method == 'GET':
        user = request.user
        companies   = _user_companies(user)
        source_co   = user.company if not user.is_superadmin else None
        category_qs = Category.objects.filter(is_deleted=False)
        offering_qs = Offering.objects.filter(is_deleted=False)
        if source_co:
            category_qs = category_qs.filter(companies=source_co)
            offering_qs = offering_qs.filter(company=source_co)
        return render(request, 'dashboard/menu/category_bulk_copy.html', {
            'companies':   companies,
            'categories':  category_qs.order_by('name'),
            'offerings':   offering_qs.order_by('name'),
            'page_title':  'Category Bulk Copy',
        })

    # ── POST ──────────────────────────────────────────────────────
    user = request.user
    src_category_id  = request.POST.get('source_category')
    src_offering_id  = request.POST.get('source_offering') or None
    target_company_id = request.POST.get('target_company')

    target_company = _selected_company_for_user(user, target_company_id)
    if not target_company:
        messages.error(request, 'Target company is required.')
        return redirect('dashboard:category_bulk_copy')

    src_category = Category.objects.filter(pk=src_category_id, is_deleted=False).first()
    if not src_category:
        messages.error(request, 'Source category not found.')
        return redirect('dashboard:category_bulk_copy')

    # Non-superadmin may only copy from their own company's categories
    if not user.is_superadmin and user.company:
        if not src_category.companies.filter(pk=user.company.pk).exists():
            messages.error(request, 'Access denied: category does not belong to your company.')
            return redirect('dashboard:category_bulk_copy')

    # Resolve / create target category
    tgt_category, _ = Category.objects.get_or_create(
        slug=src_category.slug,
        defaults={
            'name': src_category.name,
            'is_active': src_category.is_active,
            'position_order': src_category.position_order,
            'icon_type': src_category.icon_type,
        }
    )
    # Ensure this category is linked to the target company
    if not tgt_category.companies.filter(pk=target_company.pk).exists():
        tgt_category.companies.add(target_company)
    CategoryCompanyStatus.objects.get_or_create(
        category=tgt_category,
        company=target_company,
        defaults={'is_active': src_category.is_active_for_company(getattr(user, 'company', None)) if not user.is_superadmin else src_category.is_active},
    )

    # Resolve source offering (optional filter)
    src_offering = Offering.objects.filter(pk=src_offering_id, is_deleted=False).first() if src_offering_id else None

    # Build source product queryset — flat, no subcategory
    product_qs = Product.objects.filter(
        category=src_category,
        is_deleted=False,
    ).select_related('category', 'offering').prefetch_related('food_type', 'counter_mappings__counter')
    if src_offering:
        product_qs = product_qs.filter(offering=src_offering)

    if not product_qs.exists():
        messages.warning(request, 'No products found in the selected category/offering to copy.')
        return redirect('dashboard:category_bulk_copy')

    copied = 0
    skipped = 0
    import uuid as _uuid
    from django.utils.text import slugify as _slugify

    for src_product in product_qs:
        # Resolve / create target offering (same name, different company)
        tgt_offering = None
        if src_product.offering:
            tgt_offering, _ = Offering.objects.get_or_create(
                company=target_company,
                name=src_product.offering.name,
                defaults={
                    'slug': src_product.offering.slug,
                    'is_active': src_product.offering.is_active,
                    'position_order': src_product.offering.position_order,
                    'open_days': src_product.offering.open_days,
                    'available_from': src_product.offering.available_from,
                    'available_to': src_product.offering.available_to,
                }
            )

        # Check if target company already has a product with same name in this category
        existing = Product.objects.filter(
            company=target_company,
            category=tgt_category,
            name=src_product.name,
            is_deleted=False,
        ).first()

        if existing:
            skipped += 1
            continue

        # Create new product under target company — flat hierarchy only
        new_p = Product(
            company       = target_company,
            category      = tgt_category,
            offering      = tgt_offering,
            sub_category  = None,   # deprecated — always None
            sub_list      = None,   # deprecated — always None
            name          = src_product.name,
            slug          = f'{_slugify(src_product.name)}-{_uuid.uuid4().hex[:6]}',
            code          = src_product.code,
            price         = src_product.price,
            company_price = src_product.company_price,
            room_service_extra_percent = src_product.room_service_extra_percent,
            packing_price = src_product.packing_price,
            min_qty       = src_product.min_qty,
            max_qty       = src_product.max_qty,
            web_qty       = src_product.web_qty,
            pos_qty       = src_product.pos_qty,
            description   = src_product.description,
            position_order= src_product.position_order,
            is_active     = src_product.is_active,
            is_kiosk_active = src_product.is_kiosk_active,
            preparation_time_minutes = src_product.preparation_time_minutes,
            calories      = src_product.calories,
        )
        if src_product.image:
            new_p.image = src_product.image
        new_p.save()

        # Copy food types
        new_p.food_type.set(src_product.food_type.all())
        copied += 1

    if copied:
        messages.success(request, f'Bulk copy complete: {copied} product{"s" if copied != 1 else ""} copied to {target_company.name} → {tgt_category.name}.')
    if skipped:
        messages.info(request, f'{skipped} product{"s" if skipped != 1 else ""} skipped (already exist in target).')
    return redirect('dashboard:category_bulk_copy')


# ════════════════════════════════════════════════════════════════
#  STOCK MANAGEMENT  —  web_qty / pos_qty / ledger
# ════════════════════════════════════════════════════════════════

@staff_role_required('superadmin','admin')
def stock_management(request):
    from apps.menu.models import StockLedger
    user    = request.user
    company = user.company if not user.is_superadmin else None
    qs = Product.objects.filter(is_deleted=False)
    if company:
        qs = qs.filter(company=company)

    if request.method == 'POST':
        pk        = request.POST.get('product_id')
        web_qty   = request.POST.get('web_qty')
        pos_qty   = request.POST.get('pos_qty')
        note      = request.POST.get('note', 'Manual adjustment').strip()
        try:
            product = qs.get(pk=pk)
            target_company = product.company or company
            if web_qty is not None and web_qty != '':
                old = product.web_qty
                product.web_qty = int(web_qty)
                diff = product.web_qty - old
                if diff != 0:
                    StockLedger.objects.create(
                        product=product, company=target_company,
                        source='manual', ref_id=0, qty=diff, note=note,
                    )
            if pos_qty is not None and pos_qty != '':
                old = product.pos_qty
                product.pos_qty = int(pos_qty)
                diff = product.pos_qty - old
                if diff != 0:
                    StockLedger.objects.create(
                        product=product, company=target_company,
                        source='manual', ref_id=0, qty=diff, note=note,
                    )
            product.save(update_fields=['web_qty','pos_qty'])
            messages.success(request, f'{product.name} stock updated.')
        except (Product.DoesNotExist, ValueError) as e:
            messages.error(request, f'Error: {e}')
        return redirect('dashboard:stock_management')

    products_qs = qs.select_related('company').order_by('name')
    recent_ledger = StockLedger.objects.select_related('product','company').order_by('-created_at')
    if company:
        recent_ledger = recent_ledger.filter(company=company)
    stock_stats = {
        'total_products': products_qs.count(),
        'low_web': products_qs.filter(web_qty__gte=0, web_qty__lte=5).count(),
        'low_pos': products_qs.filter(pos_qty__gte=0, pos_qty__lte=5).count(),
        'unlimited_web': products_qs.filter(web_qty=-1).count(),
    }
    return render(request, 'dashboard/menu/stock_management.html', {
        'products': products_qs,
        'recent_ledger': recent_ledger[:100],
        'stock_stats': stock_stats,
        'page_title': 'Stock Management',
    })

# ════════════════════════════════════════════════════════════════
#  DASHBOARD  —  Categories
# ════════════════════════════════════════════════════════════════

CATEGORY_WEEKDAY_CHOICES = [
    ('Mon', 'Monday'), ('Tue', 'Tuesday'), ('Wed', 'Wednesday'),
    ('Thu', 'Thursday'), ('Fri', 'Friday'), ('Sat', 'Saturday'), ('Sun', 'Sunday'),
]
CATEGORY_VALID_DAYS = [day for day, _label in CATEGORY_WEEKDAY_CHOICES]


def _schedule_context_rows(schedules):
    return [
        {
            'day': s.display_day,
            'start': s.start_time.strftime('%H:%M'),
            'end': s.end_time.strftime('%H:%M'),
        }
        for s in schedules
    ]


def _category_site_availability_context(category, user):
    if not category or not category.pk:
        return []
    companies = _user_companies(user).filter(pk__in=category.companies.values_list('pk', flat=True))
    companies = list(companies.order_by('name'))
    if not companies:
        return []

    company_ids = [company.pk for company in companies]
    statuses = {
        status.company_id: status
        for status in CategoryCompanyStatus.objects.filter(
            category=category,
            company_id__in=company_ids,
        )
    }
    schedules_by_company = {company_id: [] for company_id in company_ids}
    for schedule in category.schedules.filter(company_id__in=company_ids).order_by('display_day', 'start_time'):
        schedules_by_company.setdefault(schedule.company_id, []).append(schedule)

    site_rows = []
    for company in companies:
        status = statuses.get(company.pk)
        use_custom = bool(status and status.use_custom_availability)
        site_rows.append({
            'company': company,
            'use_custom': use_custom,
            'open_days': list(status.open_days or []) if use_custom else [],
            'schedules': _schedule_context_rows(schedules_by_company.get(company.pk, [])),
        })
    return site_rows


def _save_category_site_availability(request, category, selected_companies):
    posted_ids = {
        int(company_id)
        for company_id in request.POST.getlist('site_schedule_company_ids')
        if str(company_id).isdigit()
    }
    if not posted_ids:
        return

    selected_by_id = {company.pk: company for company in selected_companies}
    for company_id in sorted(posted_ids):
        company = selected_by_id.get(company_id)
        if not company:
            continue
        status, _created = CategoryCompanyStatus.objects.get_or_create(
            category=category,
            company=company,
            defaults={'is_active': category.is_active},
        )
        enabled = request.POST.get(f'site_schedule_enabled_{company_id}') == 'on'
        status.use_custom_availability = enabled
        status.open_days = [
            day for day in request.POST.getlist(f'site_open_days_{company_id}')
            if day in CATEGORY_VALID_DAYS
        ] if enabled else []
        status.save(update_fields=['use_custom_availability', 'open_days'])

        category.schedules.filter(company=company).delete()
        if not enabled:
            continue

        from datetime import time as _time
        idx = 0
        while idx <= 20:
            day = (request.POST.get(f'site_window_day_{company_id}_{idx}') or '').strip()
            start = (request.POST.get(f'site_window_start_{company_id}_{idx}') or '').strip()
            end = (request.POST.get(f'site_window_end_{company_id}_{idx}') or '').strip()
            if not (day and start and end):
                break
            try:
                Schedule.objects.create(
                    category=category,
                    company=company,
                    display_day=day,
                    start_time=_time.fromisoformat(start),
                    end_time=_time.fromisoformat(end),
                )
            except ValueError:
                pass
            idx += 1


def _attach_category_site_statuses(categories, user):
    categories = list(categories)
    category_ids = [cat.pk for cat in categories]
    status_map = {}
    if category_ids:
        for status in CategoryCompanyStatus.objects.filter(category_id__in=category_ids).select_related('company'):
            status_map[(status.category_id, status.company_id)] = status

    for cat in categories:
        companies = list(cat.companies.all())
        cat._company_status_obj_map = {
            company.pk: status_map.get((cat.pk, company.pk))
            for company in companies
        }
        cat._company_status_map = {
            company.pk: getattr(status_map.get((cat.pk, company.pk)), 'is_active', cat.is_active)
            for company in companies
        }
        if not user.is_superadmin:
            allowed_ids = set(_user_companies(user).values_list('pk', flat=True))
            companies = [company for company in companies if company.pk in allowed_ids]
        cat.site_statuses = [
            {
                'company': company,
                'is_active': cat._company_status_map.get(company.pk, cat.is_active),
                'use_custom_availability': bool(getattr(cat._company_status_obj_map.get(company.pk), 'use_custom_availability', False)),
            }
            for company in companies
        ]
        cat.global_schedules = [
            schedule for schedule in list(cat.schedules.all())
            if schedule.company_id is None
        ]
        primary_company = get_primary_staff_company(user)
        if primary_company:
            cat.display_is_active = cat.is_active_for_company(primary_company)
        else:
            cat.display_is_active = any(site['is_active'] for site in cat.site_statuses) if cat.site_statuses else cat.is_active
    return categories


@staff_role_required('superadmin', 'admin', 'pos')
def dashboard_category_list(request):
    user = request.user
    qs = Category.objects.filter(
        parent__isnull=True, is_deleted=False
    ).prefetch_related('children__children', 'schedules', 'companies').order_by('position_order', 'name')
    if not user.is_superadmin:
        qs = qs.filter(companies__in=_user_companies(user)).distinct()
    categories = _attach_category_site_statuses(qs, user)
    lp = get_list_perms(request.user, 'perm_categories')
    can_manage_categories = any(lp.get(k) for k in ('add', 'edit', 'delete', 'toggle', 'reorder'))
    return render(request, 'dashboard/menu/category_list.html', {
        'categories': categories, 'page_title': 'Categories',
        'can_drag_sort': lp.get('reorder') and len(categories) > 1,
        'can_manage_categories': can_manage_categories,
        'lp': lp,
    })


@staff_role_required('superadmin', 'admin')
def dashboard_category_add(request):
    user    = request.user
    if not user_can_action(user, 'perm_categories', 'add'):
        return _deny_dashboard_action(request)
    parents = Category.objects.filter(parent__isnull=True, is_deleted=False).order_by('name')
    if not user.is_superadmin:
        parents = parents.filter(companies__in=_user_companies(user)).distinct()
    if request.method == 'POST':
        err = _save_category(request, None, user)
        if err is None:
            messages.success(request, 'Category added.')
            return redirect('dashboard:category_list')
        for e in err:
            messages.error(request, e)
    return render(request, 'dashboard/menu/category_form.html', {
        'companies': _user_companies(user),
        'parents': parents,
        'weekday_choices': [('Mon','Monday'),('Tue','Tuesday'),('Wed','Wednesday'),
                            ('Thu','Thursday'),('Fri','Friday'),('Sat','Saturday'),('Sun','Sunday')],
        'category_open_days': request.POST.getlist('open_days') if request.method == 'POST' else [],
        'schedule_start_time': request.POST.get('schedule_start_time','') if request.method == 'POST' else '',
        'schedule_end_time': request.POST.get('schedule_end_time','') if request.method == 'POST' else '',
        'page_title': 'Add Category', 'action': 'Add',
    })


@staff_role_required('superadmin', 'admin')
def dashboard_category_edit(request, pk):
    user     = request.user
    _perm = check_module_permission(request, 'perm_categories')
    if _perm: return _perm
    category = _own_category_or_404(pk, user)   # scoped
    parents  = Category.objects.filter(
        parent__isnull=True, is_deleted=False
    ).exclude(pk=pk).order_by('name')
    if not user.is_superadmin:
        parents = parents.filter(companies__in=_user_companies(user)).distinct()
    if request.method == 'POST':
        if user.role != 'superadmin' and get_module_level(request.user, 'perm_categories') == 'full_edit':
            _diffs = {
                'name':           {'label': 'Name',       'before': category.name,           'after': (request.POST.get('name') or '').strip()},
                'tagline':        {'label': 'Tagline',    'before': category.tagline or '',   'after': (request.POST.get('tagline') or '').strip()},
                'is_active':      {'label': 'Active',     'before': category.is_active,      'after': request.POST.get('is_active') == 'on'},
                'position_order': {'label': 'Sort Order', 'before': category.position_order, 'after': int(request.POST.get('position_order') or 0)},
                'preparation_time_minutes': {'label': 'Prep Time', 'before': category.preparation_time_minutes, 'after': max(0, int(request.POST.get('preparation_time_minutes') or 0))},
            }
            _pc = create_pending_change(request, 'perm_categories', category, _diffs)
            if _pc:
                messages.success(request, 'Your changes have been submitted for superadmin review.')
            else:
                messages.info(request, 'No changes detected.')
            return _dashboard_return_or_default(request, 'dashboard:category_list')
        err = _save_category(request, category, user)
        if err is None:
            messages.success(request, 'Category updated.')
            return _dashboard_return_or_default(request, 'dashboard:category_list')
        for e in err:
            messages.error(request, e)
    # Pass existing schedules to template for multi-window display
    existing_schedules = _schedule_context_rows(
        category.schedules.filter(company__isnull=True)
    ) if request.method != 'POST' else []
    _lj, _pl = get_locked_html_names(request.user, 'perm_categories')
    return render(request, 'dashboard/menu/category_form.html', {
        'category': category,
        'companies': _user_companies(user),
        'parents': parents,
        'selected_companies': list(category.companies.values_list('id', flat=True)),
        'weekday_choices': CATEGORY_WEEKDAY_CHOICES,
        'category_open_days': request.POST.getlist('open_days') if request.method == 'POST' else (category.open_days or []),
        'existing_schedules': existing_schedules,
        'category_site_availability': _category_site_availability_context(category, user),
        'page_title': f'Edit - {category.name}', 'action': 'Save',
        'perm_level': _pl or 'full_edit', 'locked_names_json': _lj,
    })


@require_POST
@staff_role_required('superadmin', 'admin', 'pos')
def dashboard_category_toggle(request, pk):
    if has_any_granular_perms(request.user) and not user_can_action(request.user, 'perm_categories', 'toggle'):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    cat = _own_category_or_404(pk, request.user)
    company = None
    if request.user.is_superadmin:
        company_id = request.POST.get('company_id')
        if company_id:
            company = Company.objects.filter(pk=company_id, is_active=True, is_deleted=False).first()
        else:
            linked_ids = list(cat.companies.values_list('pk', flat=True))
            if len(linked_ids) == 1:
                company = Company.objects.filter(pk=linked_ids[0], is_active=True, is_deleted=False).first()
        if not company:
            return JsonResponse({'success': False, 'error': 'Choose a site to update this shared category.'}, status=400)
    else:
        company_id = request.POST.get('company_id')
        company = _user_companies(request.user).filter(pk=company_id).first() if company_id else get_primary_staff_company(request.user)

    if not company or not cat.companies.filter(pk=company.pk).exists():
        return JsonResponse({'success': False, 'error': 'Selected site is not linked to this category.'}, status=400)

    status, _ = CategoryCompanyStatus.objects.get_or_create(
        category=cat,
        company=company,
        defaults={'is_active': cat.is_active},
    )
    status.is_active = not status.is_active
    status.save(update_fields=['is_active'])
    return JsonResponse({'success': True, 'is_active': status.is_active, 'company_id': company.pk})


@require_POST
@staff_role_required('superadmin', 'admin')
def dashboard_category_delete(request, pk):
    if get_module_level(request.user, 'perm_categories') != 'full_edit':
        return _deny_dashboard_action(request)
    cat = _own_category_or_404(pk, request.user)
    cat.is_deleted = True
    cat.is_active = False
    cat.save(update_fields=['is_deleted', 'is_active'])
    messages.success(request, f'"{cat.name}" deleted.')
    return redirect('dashboard:category_list')


@require_POST
@staff_role_required('superadmin', 'admin')
def dashboard_category_bulk_delete(request):
    if get_module_level(request.user, 'perm_categories') != 'full_edit':
        return _deny_dashboard_action(request)
    ids = _bulk_post_ids(request)
    qs = Category.objects.filter(pk__in=ids, is_deleted=False)
    if not request.user.is_superadmin:
        qs = qs.filter(companies__in=_user_companies(request.user)).distinct()
    count = qs.update(is_deleted=True, is_active=False) if ids else 0
    _bulk_message(request, 'category', count)
    return redirect('dashboard:category_list')


def _selected_category_companies(request, user):
    if user.is_superadmin:
        company_ids = [cid for cid in request.POST.getlist('companies') if str(cid).strip()]
        if not company_ids:
            return []
        return list(Company.objects.filter(pk__in=company_ids))
    allowed = _user_companies(user)
    company_ids = [cid for cid in request.POST.getlist('companies') if str(cid).strip()]
    if company_ids:
        return list(allowed.filter(pk__in=company_ids))
    primary_company = get_primary_staff_company(user)
    return [primary_company] if primary_company else []


def _category_slug_seed(name):
    return slugify(name) or f'category-{uuid.uuid4().hex[:8]}'


def _unique_category_slug(name, instance_pk=None):
    base = _category_slug_seed(name)
    slug = base
    n = 1
    qs = Category.objects.all()
    if instance_pk:
        qs = qs.exclude(pk=instance_pk)
    while qs.filter(slug=slug).exists():
        slug = f'{base}-{n}'
        n += 1
    return slug


def _save_category(request, instance, user):
    name = request.POST.get('name', '').strip()
    if not name:
        return ['Category name is required.']
    is_create = instance is None
    selected_companies = _selected_category_companies(request, user)
    slug_seed = _category_slug_seed(name)

    duplicate_qs = Category.objects.filter(is_deleted=False).filter(
        Q(slug=slug_seed) | Q(name__iexact=name)
    )
    if is_create:
        existing = duplicate_qs.order_by('position_order', 'name', 'id').first()
        if existing:
            if selected_companies:
                existing.companies.add(*selected_companies)
                requested_active = request.POST.get('is_active') == 'on'
                for company in selected_companies:
                    CategoryCompanyStatus.objects.get_or_create(
                        category=existing,
                        company=company,
                        defaults={'is_active': requested_active},
                    )
            return None
    elif duplicate_qs.exclude(pk=instance.pk).exists():
        return ['A category with this name already exists. Please add this company to the existing category instead.']

    if instance is None:
        instance = Category()
    instance.name          = name
    instance.slug          = _unique_category_slug(name, instance.pk)
    parent_id = request.POST.get('parent_id')
    if parent_id:
        parent_qs = Category.objects.filter(pk=parent_id, is_deleted=False)
        if not user.is_superadmin:
            parent_qs = parent_qs.filter(companies__in=_user_companies(user)).distinct()
        instance.parent = parent_qs.first()
        if parent_id and not instance.parent:
            return ['Selected parent category is not available for your company.']
    else:
        instance.parent = None
    instance.icon_type     = int(request.POST.get('icon_type', '0') or '0')
    instance.tagline       = request.POST.get('tagline', '').strip()
    instance.position_order= int(request.POST.get('position_order', '0') or '0')
    instance.cat_type      = 1 if not instance.parent else 2
    instance.is_active     = request.POST.get('is_active') == 'on'
    try:
        instance.preparation_time_minutes = max(0, int(request.POST.get('preparation_time_minutes', '0') or '0'))
    except (TypeError, ValueError):
        instance.preparation_time_minutes = 0
    valid_days = CATEGORY_VALID_DAYS
    instance.open_days = [d for d in request.POST.getlist('open_days') if d in valid_days]
    try:
        instance.save()
    except IntegrityError:
        return ['A category with this name already exists. Please add this company to the existing category instead.']
    # Phase 2: multi-window schedule support
    # Accepts indexed POST fields: window_day_0, window_start_0, window_end_0, window_day_1 …
    if hasattr(instance, 'schedules'):
        instance.schedules.filter(company__isnull=True).delete()
        from datetime import time as _time
        idx = 0
        while True:
            day   = (request.POST.get(f'window_day_{idx}') or '').strip()
            start = (request.POST.get(f'window_start_{idx}') or '').strip()
            end   = (request.POST.get(f'window_end_{idx}') or '').strip()
            if not (day and start and end):
                break
            try:
                Schedule.objects.create(
                    category=instance,
                    display_day=day,
                    start_time=_time.fromisoformat(start),
                    end_time=_time.fromisoformat(end),
                )
            except ValueError:
                pass
            idx += 1
            if idx > 20:  # safety ceiling
                break
        # Backward compat: if old single-field names sent, create one window
        if idx == 0:
            s_start = (request.POST.get('schedule_start_time') or '').strip()
            s_end   = (request.POST.get('schedule_end_time') or '').strip()
            if s_start and s_end:
                try:
                    schedule_days = instance.open_days or ['All']
                    for day in schedule_days:
                        Schedule.objects.create(
                            category=instance,
                            display_day=day,
                            start_time=_time.fromisoformat(s_start),
                            end_time=_time.fromisoformat(s_end),
                        )
                except ValueError:
                    pass
    if not is_create:
        instance.companies.clear()
    if selected_companies:
        instance.companies.add(*selected_companies)
        for company in selected_companies:
            CategoryCompanyStatus.objects.get_or_create(
                category=instance,
                company=company,
                defaults={'is_active': instance.is_active},
            )
        _save_category_site_availability(request, instance, selected_companies)
    return None



# ════════════════════════════════════════════════════════════════
#  DASHBOARD  —  Banners / Adverts  (scheduling + approval + media library)
# ════════════════════════════════════════════════════════════════

@staff_role_required('superadmin', 'admin', 'pos')
def dashboard_advertise_list(request):
    """
    superadmin : full card view — all ads, can approve/reject.
    admin/pos  : dropdown view — approved live banners + their own submissions.
    """
    user = request.user

    if user.is_superadmin:
        qs = Advertise.objects.select_related(
            'company', 'created_by', 'reviewed_by', 'media_asset'
        ).prefetch_related(
            'companies', 'holiday_schedules'
        ).order_by('-created_at')
        status_filter = request.GET.get('status', '')
        if status_filter:
            qs = qs.filter(status=status_filter)
        pending_count = Advertise.objects.filter(status=Advertise.STATUS_PENDING).count()
        from django.utils import timezone as _tz
        return render(request, 'dashboard/advertise/list.html', {
            'adverts':       qs,
            'status_filter': status_filter,
            'pending_count': pending_count,
            'today':         _tz.localdate(),
            'page_title':    'Advertisement Banners',
        })

    # admin / pos — dropdown view
    if not user.company:
        messages.error(request, 'No company assigned to your account.')
        return redirect('dashboard:home')

    from django.utils import timezone as _tz

    # Fetch all approved+active ads for this company (own OR targeted via M2M)
    # then apply is_live in Python so holiday-only logic runs correctly
    from django.db.models import Q, Count
    _all_approved = Advertise.objects.filter(
        status=Advertise.STATUS_APPROVED,
        is_active=True,
    ).filter(companies=user.company).distinct().select_related('media_asset').prefetch_related(
        'companies', 'holiday_schedules'
    ).order_by('position_order')

    live_banners      = [ad for ad in _all_approved if ad.is_live]
    scheduled_banners = [ad for ad in _all_approved if not ad.is_live]

    my_submissions = Advertise.objects.filter(
        company=user.company, created_by=user
    ).select_related('media_asset').prefetch_related(
        'companies', 'holiday_schedules'
    ).order_by('-created_at')[:20]

    return render(request, 'dashboard/advertise/manager_view.html', {
        'live_banners':      live_banners,
        'scheduled_banners': scheduled_banners,
        'my_submissions':    my_submissions,
        'page_title':        'Banners',
    })


@staff_role_required('superadmin', 'admin', 'pos')
def dashboard_advertise_add(request):
    """
    superadmin : full form — image upload, any company, multi-site, holidays, auto-approved.
    admin/pos  : simplified — pick image from library, select sites, add holidays, status=pending.
    """
    user = request.user
    if not user.is_superadmin and not user.company:
        messages.error(request, 'No company assigned to your account.')
        return redirect('dashboard:home')

    all_companies = _user_companies(user)
    if user.is_superadmin:
        assets = MediaAsset.objects.all().order_by('-created_at')
    else:
        from django.db.models import Q as _Q
        assets = MediaAsset.objects.filter(
            _Q(company=user.company) | _Q(companies=user.company)
        ).order_by('-created_at')
    holidays = HolidaySchedule.objects.filter(is_active=True).order_by('month', 'day')

    if request.method == 'POST':
        # Derive owning company from the first selected site checkbox (no separate dropdown)
        target_pks = request.POST.getlist('companies')
        if user.is_superadmin:
            company = Company.objects.filter(pk=target_pks[0]).first() if target_pks else None
        else:
            company = user.company
        if not company:
            messages.error(request, 'Please select at least one site.')
        else:
            status = (Advertise.STATUS_APPROVED if user.is_superadmin
                      else Advertise.STATUS_PENDING)

            asset_obj = None
            asset_id  = request.POST.get('media_asset') or None
            if asset_id:
                from django.db.models import Q as _Q2
                aq = MediaAsset.objects.filter(pk=asset_id)
                if not user.is_superadmin:
                    aq = aq.filter(_Q2(company=company) | _Q2(companies=company))
                asset_obj = aq.first()

            from django.utils import timezone as _tz
            uploaded_image = request.FILES.get('image') if user.is_superadmin else None
            ad_name = request.POST.get('name', '').strip()
            if uploaded_image:
                asset_obj = None
            if user.is_superadmin and uploaded_image:
                try:
                    asset_obj = _store_banner_upload_in_media_library(
                        image_file=uploaded_image,
                        company=company,
                        user=user,
                        ad_name=ad_name,
                    )
                except ValidationError as exc:
                    messages.error(request, '; '.join(exc.messages) if getattr(exc, 'messages', None) else 'Banner image is invalid.')
                    return render(request, 'dashboard/advertise/form.html', {
                        'all_companies': all_companies,
                        'assets': assets,
                        'holidays': holidays,
                        'page_title': 'Add Banner',
                        'action': 'Submit',
                        'banner_width': PORTAL_BANNER_WIDTH,
                        'banner_height': PORTAL_BANNER_HEIGHT,
                        'banner_size_label': PORTAL_BANNER_LABEL,
                    })

            ad = Advertise(
                company        = company,
                name           = ad_name,
                position_order = int(request.POST.get('position_order', '0') or '0'),
                is_active      = request.POST.get('is_active') == 'on',
                start_date     = request.POST.get('start_date') or None,
                end_date       = request.POST.get('end_date')   or None,
                status         = status,
                created_by     = user,
                media_asset    = asset_obj,
                created_at     = _tz.now(),
            )
            if user.is_superadmin and uploaded_image:
                ad.image = uploaded_image
            ad.save()

            # Save companies M2M (target_pks already fetched above to derive owner)
            if target_pks:
                ad.companies.set(Company.objects.filter(pk__in=target_pks, is_active=True, is_deleted=False))
            else:
                ad.companies.clear()

            # Save holiday schedules M2M
            holiday_pks = request.POST.getlist('holiday_schedules')
            ad.holiday_schedules.set(HolidaySchedule.objects.filter(pk__in=holiday_pks))

            if status == Advertise.STATUS_APPROVED:
                messages.success(request, f'Banner "{ad.name or f"#{ad.pk}"}" created and is live.')
            else:
                messages.success(request, f'Banner "{ad.name or f"#{ad.pk}"}" submitted — awaiting superadmin approval.')
            return redirect('dashboard:advertise_list')

    return render(request, 'dashboard/advertise/form.html', {
        'all_companies': all_companies,
        'assets':        assets,
        'holidays':      holidays,
        'page_title':    'Add Banner',
        'action':        'Submit',
        'banner_width': PORTAL_BANNER_WIDTH,
        'banner_height': PORTAL_BANNER_HEIGHT,
        'banner_size_label': PORTAL_BANNER_LABEL,
    })
@staff_role_required('superadmin', 'admin', 'pos')
def dashboard_advertise_edit(request, pk):
    ad   = _own_advert_or_404(pk, request.user)
    user = request.user

    if not user.is_superadmin:
        if ad.created_by != user:
            messages.error(request, 'You can only edit banners you submitted.')
            return redirect('dashboard:advertise_list')
        if ad.status == Advertise.STATUS_APPROVED:
            messages.warning(request, 'Approved banners cannot be edited. Submit a new one.')
            return redirect('dashboard:advertise_list')

    if user.is_superadmin:
        assets = MediaAsset.objects.all().order_by('-created_at')
    else:
        from django.db.models import Q as _Q
        assets = MediaAsset.objects.filter(
            _Q(company=ad.company) | _Q(companies=ad.company)
        ).order_by('-created_at')
    all_companies = _user_companies(user)
    holidays      = HolidaySchedule.objects.filter(is_active=True).order_by('month', 'day')

    if request.method == 'POST':
        ad.name           = request.POST.get('name', '').strip()
        ad.position_order = int(request.POST.get('position_order', '0') or '0')
        ad.is_active      = request.POST.get('is_active') == 'on'
        ad.start_date     = request.POST.get('start_date') or None
        ad.end_date       = request.POST.get('end_date')   or None

        asset_id = request.POST.get('media_asset') or None
        if asset_id:
            from django.db.models import Q as _Q2
            aq = MediaAsset.objects.filter(pk=asset_id)
            if not user.is_superadmin:
                aq = aq.filter(_Q2(company=ad.company) | _Q2(companies=ad.company))
            ad.media_asset = aq.first()
        elif request.POST.get('clear_asset') == '1':
            ad.media_asset = None

        if user.is_superadmin and 'image' in request.FILES:
            uploaded_image = request.FILES['image']
            try:
                ad.media_asset = _store_banner_upload_in_media_library(
                    image_file=uploaded_image,
                    company=ad.company,
                    user=user,
                    ad_name=ad.name,
                )
            except ValidationError as exc:
                messages.error(request, '; '.join(exc.messages) if getattr(exc, 'messages', None) else 'Banner image is invalid.')
                return render(request, 'dashboard/advertise/form.html', {
                    'ad':            ad,
                    'assets':        assets,
                    'all_companies': all_companies,
                    'holidays':      holidays,
                    'page_title':    f'Edit Banner #{ad.pk}',
                    'action':        'Save',
                    'banner_width': PORTAL_BANNER_WIDTH,
                    'banner_height': PORTAL_BANNER_HEIGHT,
                    'banner_size_label': PORTAL_BANNER_LABEL,
                })
            ad.image = uploaded_image
        if not user.is_superadmin:
            ad.status = Advertise.STATUS_PENDING

        ad.save()

        # Save companies M2M
        target_pks = request.POST.getlist('companies')
        if target_pks:
            ad.companies.set(Company.objects.filter(pk__in=target_pks, is_active=True, is_deleted=False))
        else:
            ad.companies.clear()

        # Save holiday schedules M2M
        holiday_pks = request.POST.getlist('holiday_schedules')
        ad.holiday_schedules.set(HolidaySchedule.objects.filter(pk__in=holiday_pks))

        msg = 'Banner updated.' if user.is_superadmin else 'Banner updated — awaiting superadmin approval.'
        messages.success(request, msg)
        return _dashboard_return_or_default(request, 'dashboard:advertise_list')

    return render(request, 'dashboard/advertise/form.html', {
        'ad':            ad,
        'assets':        assets,
        'all_companies': all_companies,
        'holidays':      holidays,
        'page_title':    f'Edit Banner #{ad.pk}',
        'action':        'Save',
        'banner_width': PORTAL_BANNER_WIDTH,
        'banner_height': PORTAL_BANNER_HEIGHT,
        'banner_size_label': PORTAL_BANNER_LABEL,
    })
@staff_role_required('superadmin')
def dashboard_advertise_approve(request, pk):
    """Only superadmin can approve or reject banners."""
    ad     = get_object_or_404(Advertise, pk=pk)
    action = request.POST.get('action')
    note   = request.POST.get('review_note', '').strip()
    if action == 'approve':
        ad.status = Advertise.STATUS_APPROVED
        ad.reviewed_by = request.user
        ad.review_note = note
        ad.save()
        messages.success(request, f'Banner "{ad.name or f"#{ad.pk}"}" approved and is now live.')
    elif action == 'reject':
        ad.status = Advertise.STATUS_REJECTED
        ad.reviewed_by = request.user
        ad.review_note = note
        ad.save()
        messages.warning(request, f'Banner "{ad.name or f"#{ad.pk}"}" rejected.')
    else:
        messages.error(request, 'Invalid action.')
    return redirect('dashboard:advertise_list')


@require_POST
@staff_role_required('superadmin', 'admin', 'pos')
def dashboard_advertise_delete(request, pk):
    ad   = _own_advert_or_404(pk, request.user)
    user = request.user
    if not user.is_superadmin:
        if ad.created_by != user:
            messages.error(request, 'You can only delete banners you submitted.')
            return redirect('dashboard:advertise_list')
        if ad.status == Advertise.STATUS_APPROVED:
            messages.error(request, 'Approved banners cannot be deleted. Contact superadmin.')
            return redirect('dashboard:advertise_list')
    ad.delete()
    messages.success(request, 'Banner deleted.')
    return redirect('dashboard:advertise_list')


@require_POST
@staff_role_required('superadmin', 'admin', 'pos')
def dashboard_advertise_bulk_delete(request):
    ids = _bulk_post_ids(request)
    user = request.user
    count = 0
    skipped = 0
    for ad in Advertise.objects.filter(pk__in=ids):
        if not user.is_superadmin and (ad.created_by != user or ad.status == Advertise.STATUS_APPROVED):
            skipped += 1
            continue
        ad.delete()
        count += 1
    _bulk_message(request, 'banner', count, skipped)
    return redirect('dashboard:advertise_list')


# ────────────────────────────────────────────────────────────────
#  MEDIA LIBRARY — superadmin manages; admin/pos browse only
# ────────────────────────────────────────────────────────────────

@staff_role_required('superadmin', 'admin', 'pos')
def media_library(request):
    user   = request.user
    assets = MediaAsset.objects.select_related('company', 'uploaded_by')
    if not user.is_superadmin and user.company:
        assets = assets.filter(Q(company=user.company) | Q(companies=user.company))
    assets = assets.order_by('-created_at')

    if request.method == 'POST':
        if not user.is_superadmin:
            messages.error(request, 'Only superadmin can upload images to the library.')
            return redirect('dashboard:media_library')
        active_companies = Company.objects.filter(is_active=True, is_deleted=False).order_by('name')
        company = user.company if getattr(user, 'company_id', None) else active_companies.first()
        if not company:
            messages.error(request, 'Create an active company before uploading banner media.')
            return redirect('dashboard:media_library')
        image_file = request.FILES.get('image')
        if not image_file:
            messages.error(request, 'Please select an image file.')
            return redirect('dashboard:media_library')
        name = _image_upload_label(image_file, 'Banner Image')
        asset = MediaAsset(company=company, name=name, image=image_file, uploaded_by=user)
        try:
            asset.full_clean()
        except ValidationError as exc:
            messages.error(request, '; '.join(exc.messages) if getattr(exc, 'messages', None) else 'Banner image is invalid.')
            return redirect('dashboard:media_library')
        asset.save()
        asset.companies.set(active_companies.exclude(pk=company.pk))
        messages.success(request, f'"{name}" added to media library.')
        return redirect('dashboard:media_library')

    return render(request, 'dashboard/advertise/media_library.html', {
        'assets':     assets,
        'companies':  Company.objects.filter(is_active=True, is_deleted=False).order_by('name'),
        'can_upload': user.is_superadmin,
        'page_title': 'Media Library',
        'banner_width': PORTAL_BANNER_WIDTH,
        'banner_height': PORTAL_BANNER_HEIGHT,
        'banner_size_label': PORTAL_BANNER_LABEL,
    })


@require_POST
@staff_role_required('superadmin')
def media_library_rename(request, pk):
    asset = get_object_or_404(MediaAsset, pk=pk)
    name = (request.POST.get('name') or '').strip()
    if not name:
        return JsonResponse({'success': False, 'error': 'Name is required.'}, status=400)
    asset.name = name
    asset.save(update_fields=['name'])
    return JsonResponse({'success': True, 'name': asset.name})


@require_POST
@staff_role_required('superadmin')
def media_library_delete(request, pk):
    asset = get_object_or_404(MediaAsset, pk=pk)
    if asset.adverts.exists():
        messages.error(request, 'Cannot delete — this image is used by one or more banners.')
        return redirect('dashboard:media_library')
    asset.delete()
    messages.success(request, 'Image deleted from media library.')
    return redirect('dashboard:media_library')


@require_POST
@staff_role_required('superadmin')
def media_library_bulk_delete(request):
    ids = _bulk_post_ids(request)
    count = 0
    skipped = 0
    for asset in MediaAsset.objects.filter(pk__in=ids):
        if asset.adverts.exists():
            skipped += 1
            continue
        asset.delete()
        count += 1
    _bulk_message(request, 'image', count, skipped)
    return redirect('dashboard:media_library')


# ════════════════════════════════════════════════════════════════
#  HOLIDAY SCHEDULES  (superadmin only — recurring annual dates)
# ════════════════════════════════════════════════════════════════

@staff_role_required('superadmin')
def holiday_list(request):
    """List all holiday schedules. Handle add/delete via same page."""
    from django.utils import timezone as _tz
    today = _tz.localdate()
    holidays = HolidaySchedule.objects.all().prefetch_related('adverts').order_by('month', 'day')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add':
            name  = request.POST.get('name', '').strip()
            month = int(request.POST.get('month', 0) or 0)
            day   = int(request.POST.get('day', 0) or 0)
            desc  = request.POST.get('description', '').strip()
            if not name or not (1 <= month <= 12) or not (1 <= day <= 31):
                messages.error(request, 'Name, valid month (1–12), and day (1–31) are required.')
            elif HolidaySchedule.objects.filter(month=month, day=day).exists():
                messages.error(request,
                    f'A holiday already exists for {day:02d}/{month:02d}. Edit it instead.')
            else:
                HolidaySchedule.objects.create(
                    name=name, month=month, day=day,
                    description=desc, created_by=request.user
                )
                messages.success(request, f'"{name}" holiday added.')
            return redirect('dashboard:holiday_list')

        elif action == 'toggle':
            pk = request.POST.get('pk')
            h  = get_object_or_404(HolidaySchedule, pk=pk)
            h.is_active = not h.is_active
            h.save()
            return redirect('dashboard:holiday_list')

        elif action == 'delete':
            pk = request.POST.get('pk')
            h  = get_object_or_404(HolidaySchedule, pk=pk)
            h.delete()
            messages.success(request, 'Holiday deleted.')
            return redirect('dashboard:holiday_list')

    return render(request, 'dashboard/advertise/holiday_list.html', {
        'holidays':   holidays,
        'today':      today,
        'months': [
            (1,'January'),(2,'February'),(3,'March'),(4,'April'),
            (5,'May'),(6,'June'),(7,'July'),(8,'August'),
            (9,'September'),(10,'October'),(11,'November'),(12,'December'),
        ],
        'page_title': 'Holiday Schedules',
    })


@staff_role_required('superadmin')
def holiday_edit(request, pk):
    """Edit an existing holiday schedule."""
    h = get_object_or_404(HolidaySchedule, pk=pk)
    if request.method == 'POST':
        h.name        = request.POST.get('name', h.name).strip()
        h.description = request.POST.get('description', '').strip()
        h.is_active   = request.POST.get('is_active') == 'on'
        month = int(request.POST.get('month', h.month) or h.month)
        day   = int(request.POST.get('day',   h.day)   or h.day)
        if 1 <= month <= 12 and 1 <= day <= 31:
            # Check uniqueness only if date changed
            if (month != h.month or day != h.day):
                if HolidaySchedule.objects.filter(month=month, day=day).exclude(pk=pk).exists():
                    messages.error(request, f'Another holiday already on {day:02d}/{month:02d}.')
                    return redirect('dashboard:holiday_edit', pk=pk)
            h.month = month
            h.day   = day
        h.save()
        messages.success(request, f'"{h.name}" updated.')
        return redirect('dashboard:holiday_list')

    return render(request, 'dashboard/advertise/holiday_edit.html', {
        'holiday':    h,
        'page_title': f'Edit Holiday — {h.name}',
    })


# ════════════════════════════════════════════════════════════════
#  PRODUCT BULK UPLOAD
# ════════════════════════════════════════════════════════════════

def _bulk_upload_summary(payload=None, *, is_preview=False):
    product_rows = payload['sheets'].get('products', []) if payload else []
    counter_rows = (payload['sheets'].get('countermapping') or payload['sheets'].get('counter_mapping') or []) if payload else []
    return {
        'mode': payload['mode'] if payload else '',
        'is_preview': is_preview,
        'products_sheet_rows': len(product_rows),
        'counter_mapping_sheet_rows': len(counter_rows),
        'products_created': 0,
        'products_updated': 0,
        'counter_mappings_created': 0,
        'counter_mappings_updated': 0,
        'offerings_created': 0,
        'categories_created': 0,
        'skipped': 0,
        'errors': [],
        'warnings': [],
        'preview_products': [],
        'preview_counter_mappings': [],
        'preview_products_ready': 0,
        'preview_products_blocked': 0,
        'preview_counter_ready': 0,
        'preview_counter_blocked': 0,
        'calories_estimated': 0,
        'preview_token': '',
        'can_import': False,
    }


def _bulk_add_issue(summary, sheet, line_no, message, *, product='', level='error'):
    target = summary['warnings'] if level == 'warning' else summary['errors']
    target.append({
        'sheet': sheet,
        'line': line_no,
        'message': str(message),
        'product': product,
        'level': level,
    })


def _bulk_price_from_row(row):
    value = row.get('base_price') if row.get('base_price') not in (None, '') else row.get('price')
    if value in (None, ''):
        raise ValueError('Staff/Base Price is required.')
    price = _parse_decimal(value, '0')
    if price < 0:
        raise ValueError('Staff/Base Price cannot be negative.')
    return price


def _bulk_product_lookup_keys(company, row, *, name=''):
    keys = []
    if not company:
        return keys
    code = str(row.get('product_code') or row.get('code') or '').strip().lower()
    product_name = (name or str(row.get('product_name') or row.get('name') or '')).strip().lower()
    product_slug = _product_slug_from_import_row(row, name=name or product_name)
    if code:
        keys.append(('code', company.pk, code))
    if product_slug:
        keys.append(('slug', company.pk, product_slug))
    if product_name:
        keys.append(('name', company.pk, product_name))
    return keys


def _product_slug_from_import_row(row, *, name=''):
    raw_slug = str(row.get('slug') or '').strip()
    if raw_slug:
        return slugify(raw_slug)
    product_name = (name or str(row.get('product_name') or row.get('name') or '')).strip()
    return slugify(product_name) if product_name else ''


def _counter_names_from_product_row(row):
    raw = _coalesce_row_value(row, 'counters', 'counter_names', 'pickup_counters', 'counter_ids')
    if not raw:
        return []
    return [part.strip() for part in str(raw).replace('|', ',').split(',') if part and part.strip()]


def _resolve_counter_from_product_value(company, value):
    raw = str(value or '').strip()
    if not raw:
        return None
    qs = Counter.objects.filter(company=company, is_deleted=False, is_active=True)
    if raw.isdigit():
        return qs.filter(pk=int(raw)).first()
    return qs.filter(name__iexact=raw).order_by('position_order', 'id').first()


def _build_product_bulk_preview(user, payload, selected_company, import_options):
    company_lookup = _company_name_map(_user_companies(user))
    product_rows = payload['sheets'].get('products', [])
    counter_rows = payload['sheets'].get('countermapping') or payload['sheets'].get('counter_mapping') or []
    summary = _bulk_upload_summary(payload, is_preview=True)

    seen_product_keys = set()
    planned_product_keys = set()
    has_dedicated_counter_sheet = bool(counter_rows)

    for line_no, row in enumerate(product_rows, start=2):
        row_errors = []
        row_notes = []
        name = str(row.get('product_name') or row.get('name') or '').strip()
        code = str(row.get('product_code') or row.get('code') or '').strip()
        description = str(row.get('description') or '').strip()
        company = None
        category_label = ''
        offering_label = ''
        calories = None
        calories_source = 'blank'
        action = 'Create'

        try:
            company = _resolve_company_from_row(user, selected_company, row, company_lookup)
            if not company:
                row_errors.append('Company missing or invalid.')
            if not name:
                row_errors.append('Product name is required.')

            for key in _bulk_product_lookup_keys(company, row, name=name):
                if key in seen_product_keys:
                    row_errors.append('Duplicate product in this workbook. Product names that make the same slug are also duplicates.')
                    break
                seen_product_keys.add(key)

            if company:
                category, _ = _resolve_category_for_company(company, row, allow_create=False)
                category_name = str(row.get('category') or row.get('category_name') or '').strip()
                if category:
                    category_label = category.name
                elif category_name and import_options['auto_create_categories']:
                    category_label = f'{category_name} (will be created)'
                    row_notes.append('Category will be created.')
                else:
                    row_errors.append('Category not found.')

                offering, _ = _resolve_offering_for_company(company, row, allow_create=False)
                offering_name = str(row.get('offering') or row.get('offering_name') or '').strip()
                if offering:
                    offering_label = offering.name
                elif offering_name and import_options['auto_create_offerings']:
                    offering_label = f'{offering_name} (will be created)'
                    row_notes.append('Offering will be created.')
                elif offering_name:
                    row_errors.append('Offering not found.')

                existing = _resolve_product_for_import(company, row)
                if existing:
                    action = 'Update'
                    if not import_options['replace_product_data']:
                        row_errors.append('Product already exists and Replace Product Data is off.')
            _bulk_price_from_row(row)
            min_qty = _parse_int(row.get('min_qty'), 1, minimum=1)
            max_qty = _parse_int(row.get('max_qty'), 10, minimum=1)
            if max_qty < min_qty:
                row_errors.append('Max Qty per Order cannot be less than Min Qty per Order.')
            _parse_int(row.get('web_qty'), -1)
            _parse_int(row.get('pos_qty'), 0, minimum=0)
            _parse_int(row.get('position_order'), 0)
            _parse_int(row.get('preparation_time_minutes'), 10, minimum=0)
            _parse_decimal(row.get('company_price'), '0')
            _parse_decimal(row.get('room_service_extra_percent'), '0')
            _parse_decimal(row.get('packing_price'), '0')
            _canonical_food_type_ids_from_value(_coalesce_row_value(row, 'food_types', 'food_type', 'veg_type'))
            calories_cell = row.get('calories')
            calories_raw = '' if calories_cell is None else str(calories_cell).strip()
            calories = _parse_calories_value(calories_raw, name, description)
            if calories_raw:
                calories_source = 'Excel'
            elif calories is not None:
                calories_source = 'Auto'
                row_notes.append(f'Calories will auto-fill as {calories} kcal.')
                summary['calories_estimated'] += 1
            if not has_dedicated_counter_sheet:
                for counter_value in _counter_names_from_product_row(row):
                    if company and not _resolve_counter_from_product_value(company, counter_value):
                        row_errors.append(f'Counter not found: {counter_value}')
        except Exception as exc:
            row_errors.append(str(exc))

        ready = not row_errors
        if ready:
            summary['preview_products_ready'] += 1
            for key in _bulk_product_lookup_keys(company, row, name=name):
                planned_product_keys.add(key)
        else:
            summary['preview_products_blocked'] += 1
            summary['skipped'] += 1
            for err in row_errors:
                _bulk_add_issue(summary, 'Products', line_no, err, product=name or code)

        summary['preview_products'].append({
            'line': line_no,
            'status': 'Ready' if ready else 'Blocked',
            'action': action,
            'product': name or '(missing name)',
            'code': code,
            'company': company.name if company else '',
            'category': category_label,
            'offering': offering_label,
            'price': str(row.get('base_price') if row.get('base_price') not in (None, '') else row.get('price') or ''),
            'calories': calories if calories is not None else '',
            'calories_source': calories_source,
            'notes': row_notes,
            'errors': row_errors,
        })

    for line_no, row in enumerate(counter_rows, start=2):
        row_errors = []
        product_name = str(row.get('product_name') or row.get('name') or row.get('product_code') or row.get('code') or '').strip()
        counter_name = str(row.get('counter') or row.get('counter_name') or '').strip()
        company = None
        try:
            company = _resolve_company_from_row(user, selected_company, row, company_lookup)
            if not company:
                row_errors.append('Company missing or invalid.')
            product = _resolve_product_for_import(company, row) if company else None
            planned = any(key in planned_product_keys for key in _bulk_product_lookup_keys(company, row)) if company else False
            if not product and not planned:
                row_errors.append('Product not found. Use product_code or include the product in the Products sheet.')
            building = _resolve_building_for_company(company, row) if company else None
            cafe = _resolve_cafe_for_company(company, row, building=building) if company else None
            counter = _resolve_counter_for_company(company, row, cafe=cafe) if company else None
            if not counter:
                row_errors.append('Counter not found.')
            if company:
                _validate_site_chain(
                    company,
                    building=building,
                    cafe=cafe,
                    counter=counter,
                    state_name=str(row.get('state') or '').strip(),
                    city_name=str(row.get('city') or '').strip(),
                )
        except Exception as exc:
            row_errors.append(str(exc))
        ready = not row_errors
        if ready:
            summary['preview_counter_ready'] += 1
        else:
            summary['preview_counter_blocked'] += 1
            summary['skipped'] += 1
            for err in row_errors:
                _bulk_add_issue(summary, 'CounterMapping', line_no, err, product=product_name)
        summary['preview_counter_mappings'].append({
            'line': line_no,
            'status': 'Ready' if ready else 'Blocked',
            'product': product_name or '(missing product)',
            'counter': counter_name or '(missing counter)',
            'errors': row_errors,
        })

    summary['can_import'] = bool(
        summary['preview_products_ready']
        or summary['preview_counter_ready']
    )
    return summary


@staff_role_required('superadmin','admin')
def dashboard_product_bulk_upload(request):
    user = request.user
    companies = _user_companies(user)
    categories = _user_categories(user)
    summary = None
    selected_company_id = request.POST.get('company') if request.method == 'POST' else request.GET.get('company')
    selected_company = _selected_company_for_user(user, selected_company_id)
    company_lookup = _company_name_map(companies)

    import_options = {
        'replace_product_data': _parse_bool(request.POST.get('replace_product_data'), True) if request.method == 'POST' else True,
        'replace_counter_mappings': _parse_bool(request.POST.get('replace_counter_mappings'), False) if request.method == 'POST' else False,
        'auto_create_categories': _parse_bool(request.POST.get('auto_create_categories'), True) if request.method == 'POST' else True,
        'auto_create_offerings': _parse_bool(request.POST.get('auto_create_offerings'), True) if request.method == 'POST' else True,
    }

    if request.method == 'POST':
        action = request.POST.get('action') or 'preview'
        preview_token = (request.POST.get('preview_token') or '').strip()
        cached_import = None

        if action == 'import_verified':
            cached_import = cache.get(f'product-bulk-preview:{preview_token}') if preview_token else None
            if not cached_import:
                summary = _bulk_upload_summary(None, is_preview=True)
                _bulk_add_issue(summary, 'Upload', '', 'The verified upload expired. Please verify the workbook again.')
                messages.error(request, 'The verified upload expired. Please verify the workbook again.')
                return render(request, 'dashboard/menu/product_bulk_upload.html', {
                    'companies': companies,
                    'categories': categories,
                    'summary': summary,
                    'page_title': 'Bulk Upload Products',
                    'selected_company_id': str(selected_company.id) if selected_company else '',
                    'import_options': import_options,
                    'sheet_names': ['Products', 'CounterMapping'],
                })
            if str(cached_import.get('user_id')) != str(user.pk):
                summary = _bulk_upload_summary(None, is_preview=True)
                _bulk_add_issue(summary, 'Upload', '', 'This verified upload belongs to another user. Please verify the workbook again.')
                messages.error(request, 'Please verify the workbook again before uploading.')
                return render(request, 'dashboard/menu/product_bulk_upload.html', {
                    'companies': companies,
                    'categories': categories,
                    'summary': summary,
                    'page_title': 'Bulk Upload Products',
                    'selected_company_id': str(selected_company.id) if selected_company else '',
                    'import_options': import_options,
                    'sheet_names': ['Products', 'CounterMapping'],
                })
            payload = cached_import['payload']
            import_options = cached_import.get('import_options') or import_options
            selected_company = _selected_company_for_user(user, cached_import.get('selected_company_id'))
            selected_company_id = str(selected_company.id) if selected_company else ''
        else:
            upload = request.FILES.get('csv_file')
            if not upload:
                summary = _bulk_upload_summary(None, is_preview=True)
                _bulk_add_issue(summary, 'Upload', '', 'Please choose a CSV or XLSX file first.')
                return render(request, 'dashboard/menu/product_bulk_upload.html', {
                    'companies': companies,
                    'categories': categories,
                    'summary': summary,
                    'page_title': 'Bulk Upload Products',
                    'selected_company_id': str(selected_company.id) if selected_company else '',
                    'import_options': import_options,
                    'sheet_names': ['Products', 'CounterMapping'],
                })

            try:
                payload = _load_bulk_upload_payload(upload)
            except (ValueError, KeyError, OSError, TypeError):
                summary = _bulk_upload_summary(None, is_preview=True)
                _bulk_add_issue(summary, 'Upload', '', 'Could not read the uploaded file. Use UTF-8 CSV or XLSX.')
                return render(request, 'dashboard/menu/product_bulk_upload.html', {
                    'companies': companies,
                    'categories': categories,
                    'summary': summary,
                    'page_title': 'Bulk Upload Products',
                    'selected_company_id': str(selected_company.id) if selected_company else '',
                    'import_options': import_options,
                    'sheet_names': ['Products', 'CounterMapping'],
                })

            summary = _build_product_bulk_preview(user, payload, selected_company, import_options)
            preview_token = uuid.uuid4().hex
            cache.set(
                f'product-bulk-preview:{preview_token}',
                {
                    'payload': payload,
                    'selected_company_id': str(selected_company.id) if selected_company else '',
                    'import_options': import_options,
                    'user_id': user.pk,
                },
                timeout=60 * 30,
            )
            summary['preview_token'] = preview_token
            return render(request, 'dashboard/menu/product_bulk_upload.html', {
                'companies': companies,
                'categories': categories,
                'summary': summary,
                'page_title': 'Bulk Upload Products',
                'selected_company_id': str(selected_company.id) if selected_company else '',
                'import_options': import_options,
                'sheet_names': ['Products', 'CounterMapping'],
            })

        product_rows = payload['sheets'].get('products', [])
        counter_rows = payload['sheets'].get('countermapping') or payload['sheets'].get('counter_mapping') or []

        summary = _bulk_upload_summary(payload, is_preview=False)
        errors = summary['errors']
        cleared_counter_products = set()
        counter_position_tracker = {}
        has_dedicated_counter_sheet = bool(counter_rows)
        seen_import_product_keys = set()

        for line_no, row in enumerate(product_rows, start=2):
            try:
                company = _resolve_company_from_row(user, selected_company, row, company_lookup)
                if not company:
                    raise ValueError('Company missing or invalid.')

                name = str(row.get('product_name') or row.get('name') or '').strip()
                if not name:
                    raise ValueError('Product name is required.')

                duplicate_in_file = False
                for key in _bulk_product_lookup_keys(company, row, name=name):
                    if key in seen_import_product_keys:
                        duplicate_in_file = True
                        break
                    seen_import_product_keys.add(key)
                if duplicate_in_file:
                    summary['skipped'] += 1
                    _bulk_add_issue(summary, 'Products', line_no, 'Duplicate product in this workbook. Product names that make the same slug are also duplicates.', product=name)
                    continue

                existing = _resolve_product_for_import(company, row)
                if existing and not import_options['replace_product_data']:
                    summary['skipped'] += 1
                    _bulk_add_issue(summary, 'Products', line_no, 'Product already exists and Replace Product Data is off.', product=name)
                    continue

                category, category_created = _resolve_category_for_company(
                    company,
                    row,
                    allow_create=import_options['auto_create_categories'],
                )
                if not category:
                    raise ValueError('Category not found.')
                if category_created:
                    summary['categories_created'] += 1

                offering, offering_created = _resolve_offering_for_company(
                    company,
                    row,
                    allow_create=import_options['auto_create_offerings'],
                )
                if offering_created:
                    summary['offerings_created'] += 1

                schedule_enabled = _parse_bool(
                    row.get('schedule_enabled'),
                    default=bool(row.get('start_date') or row.get('end_date') or row.get('start_time') or row.get('end_time')),
                )
                start_date = _parse_date_value(row.get('start_date'))
                end_date = _parse_date_value(row.get('end_date'))
                start_time = _parse_time_value(row.get('start_time'))
                end_time = _parse_time_value(row.get('end_time'))
                explicit_menu_date = _parse_date_value(row.get('menu_date'))
                if start_date and end_date and start_date > end_date:
                    raise ValueError('Schedule end date cannot be before start date.')

                description = str(row.get('description') or '').strip()
                min_qty = _parse_int(row.get('min_qty'), 1, minimum=1)
                max_qty = _parse_int(row.get('max_qty'), 10, minimum=1)
                if max_qty < min_qty:
                    raise ValueError('Max Qty per Order cannot be less than Min Qty per Order.')
                food_type_ids = _canonical_food_type_ids_from_value(_coalesce_row_value(row, 'food_types', 'food_type', 'veg_type'))
                calories_cell = row.get('calories')
                calories_raw = '' if calories_cell is None else str(calories_cell).strip()
                calories = _parse_calories_value(calories_raw, name, description)
                if not calories_raw and calories is not None:
                    summary['calories_estimated'] += 1
                product_counter_values = _counter_names_from_product_row(row) if not has_dedicated_counter_sheet else []
                for counter_value in product_counter_values:
                    if not _resolve_counter_from_product_value(company, counter_value):
                        raise ValueError(f'Counter not found: {counter_value}')

                raw_slug = str(row.get('slug') or '').strip()
                defaults = {
                    'category': category,
                    'offering': offering,
                    'company': company,
                    'slug': raw_slug or (existing.slug if existing else slugify(name)),
                    'name': name,
                    'code': str(row.get('product_code') or row.get('code') or '').strip(),
                    'price': _bulk_price_from_row(row),
                    'company_price': _parse_decimal(row.get('company_price'), '0'),
                    'room_service_extra_percent': _parse_decimal(row.get('room_service_extra_percent'), '0'),
                    'packing_price': _parse_decimal(row.get('packing_price'), '0'),
                    'min_qty': min_qty,
                    'max_qty': max_qty,
                    'web_qty': _parse_int(row.get('web_qty'), -1),
                    'pos_qty': _parse_int(row.get('pos_qty'), 0, minimum=0),
                    'description': description,
                    'position_order': _parse_int(row.get('position_order'), 0),
                    'is_active': _parse_bool(row.get('is_active'), True),
                    'is_kiosk_active': _parse_bool(row.get('is_kiosk_active', row.get('kiosk_active')), True),
                    'featured_in_web': _parse_bool(row.get('featured_in_web'), False),
                    'featured_in_kiosk_extra': _parse_bool(row.get('featured_in_kiosk_extra', row.get('is_featured')), False),
                    'is_deleted': False,
                    'preparation_time_minutes': _parse_int(row.get('preparation_time_minutes'), 10, minimum=0),
                    'calories': calories,
                    'schedule_bypass': _parse_bool(row.get('schedule_bypass'), False) if user.role == 'superadmin' else False,
                }

                if schedule_enabled:
                    defaults['available_from'] = start_time
                    defaults['available_to'] = end_time
                    defaults['menu_date'] = explicit_menu_date or (start_date if start_date and end_date and start_date == end_date else None)
                    defaults['start_datetime'] = timezone.make_aware(parse_datetime(f'{start_date.isoformat()} 00:00:00')) if start_date else None
                    defaults['end_datetime'] = timezone.make_aware(parse_datetime(f'{end_date.isoformat()} 23:59:59')) if end_date else None
                else:
                    defaults['menu_date'] = explicit_menu_date
                    defaults['available_from'] = None
                    defaults['available_to'] = None
                    defaults['start_datetime'] = None
                    defaults['end_datetime'] = None

                if existing:
                    try:
                        for key, value in defaults.items():
                            setattr(existing, key, value)
                        existing.save()
                    except IntegrityError:
                        summary['skipped'] += 1
                        _bulk_add_issue(summary, 'Products', line_no, 'A product with this name/slug already exists for this company.', product=name)
                        continue
                    product_obj = existing
                    summary['products_updated'] += 1
                else:
                    try:
                        product_obj = Product.objects.create(**defaults)
                    except IntegrityError:
                        summary['skipped'] += 1
                        _bulk_add_issue(summary, 'Products', line_no, 'A product with this name/slug already exists for this company.', product=name)
                        continue
                    summary['products_created'] += 1

                product_obj.food_type.set(FoodType.objects.filter(pk__in=food_type_ids) if food_type_ids else [])

                if _row_has_nonblank(row, 'is_free_meal_product') or not existing:
                    product_obj.free_meal_companies.clear()
                    if _parse_bool(row.get('is_free_meal_product'), False) and product_obj.company_id:
                        product_obj.company.free_meal_products.add(product_obj)

                # Allow pickup counters on Products sheet when CounterMapping sheet is absent.
                if product_counter_values and not has_dedicated_counter_sheet:
                    if import_options['replace_counter_mappings']:
                        ProductCounter.objects.filter(product=product_obj).delete()
                    for idx, counter_value in enumerate(product_counter_values):
                        counter = _resolve_counter_from_product_value(company, counter_value)
                        mapping, created = ProductCounter.objects.get_or_create(
                            product=product_obj,
                            counter=counter,
                            defaults={'position_order': idx, 'is_active': True},
                        )
                        if created:
                            summary['counter_mappings_created'] += 1
                        else:
                            mapping.position_order = idx
                            mapping.is_active = True
                            mapping.save(update_fields=['position_order', 'is_active'])
                            summary['counter_mappings_updated'] += 1
            except Exception as exc:
                summary['skipped'] += 1
                _bulk_add_issue(summary, 'Products', line_no, exc, product=str(row.get('product_name') or row.get('name') or row.get('product_code') or row.get('code') or ''))

        for line_no, row in enumerate(counter_rows, start=2):
            try:
                company = _resolve_company_from_row(user, selected_company, row, company_lookup)
                if not company:
                    raise ValueError('Company missing or invalid.')
                product = _resolve_product_for_import(company, row)
                if not product:
                    raise ValueError('Product not found. Use product_code or an existing product name.')
                building = _resolve_building_for_company(company, row)
                cafe = _resolve_cafe_for_company(company, row, building=building)
                counter = _resolve_counter_for_company(company, row, cafe=cafe)
                if not counter:
                    raise ValueError('Counter not found.')
                _validate_site_chain(
                    company,
                    building=building,
                    cafe=cafe,
                    counter=counter,
                    state_name=str(row.get('state') or '').strip(),
                    city_name=str(row.get('city') or '').strip(),
                )

                if import_options['replace_counter_mappings'] and product.pk not in cleared_counter_products:
                    ProductCounter.objects.filter(product=product).delete()
                    cleared_counter_products.add(product.pk)

                is_primary = _parse_bool(row.get('is_primary'), False)
                explicit_position = row.get('position_order')
                if explicit_position not in (None, ''):
                    position_order = _parse_int(explicit_position, 0, minimum=0)
                elif is_primary:
                    position_order = 0
                else:
                    position_order = counter_position_tracker.get(product.pk, 1)
                counter_position_tracker[product.pk] = max(counter_position_tracker.get(product.pk, 1), position_order + 1)

                mapping, created = ProductCounter.objects.get_or_create(
                    product=product,
                    counter=counter,
                    defaults={
                        'position_order': position_order,
                        'is_active': _parse_bool(row.get('is_active'), True),
                    },
                )
                if created:
                    summary['counter_mappings_created'] += 1
                else:
                    mapping.position_order = position_order
                    mapping.is_active = _parse_bool(row.get('is_active'), True)
                    mapping.save(update_fields=['position_order', 'is_active'])
                    summary['counter_mappings_updated'] += 1
            except Exception as exc:
                summary['skipped'] += 1
                _bulk_add_issue(summary, 'CounterMapping', line_no, exc, product=str(row.get('product_code') or row.get('code') or row.get('product_name') or row.get('name') or ''))

        if preview_token:
            cache.delete(f'product-bulk-preview:{preview_token}')

        if summary['products_created'] or summary['products_updated'] or summary['counter_mappings_created'] or summary['counter_mappings_updated']:
            messages.success(
                request,
                'Bulk upload finished. '
                f"Products: {summary['products_created']} created / {summary['products_updated']} updated. "
                f"Counter mappings: {summary['counter_mappings_created']} created / {summary['counter_mappings_updated']} updated. "
                f"Skipped: {summary['skipped']}."
            )
        else:
            messages.warning(request, f"No records were imported. Skipped {summary['skipped']} row(s).")

    return render(request, 'dashboard/menu/product_bulk_upload.html', {
        'companies': companies,
        'categories': categories,
        'summary': summary,
        'page_title': 'Bulk Upload Products',
        'selected_company_id': str(selected_company.id) if selected_company else '',
        'import_options': import_options,
        'sheet_names': ['Products', 'CounterMapping'],
    })


# ════════════════════════════════════════════════════════════════
#  PHASE 1 — DRAG-SORT REORDER ENDPOINTS
#  All models already have position_order fields.
#  Each endpoint accepts JSON {"ids": [pk, pk, ...]} and saves
#  position_order = array-index for each pk in the user's scope.
# ════════════════════════════════════════════════════════════════

import json as _json

def _parse_reorder_ids(request):
    """Return list[int] from JSON body {ids:[...]} or raise ValueError."""
    try:
        body = _json.loads(request.body)
        ids = [int(x) for x in body['ids']]
    except Exception:
        raise ValueError('Invalid payload — expected JSON {ids:[...]}')
    return ids


@staff_role_required('superadmin', 'admin')
def product_reorder(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    if not user_can_action(request.user, 'perm_products', 'reorder'):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    try:
        ids = _parse_reorder_ids(request)
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    user = request.user
    qs = Product.objects.filter(is_deleted=False)
    if not user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(user))
    # Verify all ids are in scope (security check)
    found = set(qs.filter(pk__in=ids).values_list('pk', flat=True))
    if len(found) != len(ids):
        return JsonResponse({'success': False, 'error': 'One or more products not accessible'}, status=403)
    for pos, pk in enumerate(ids):
        qs.filter(pk=pk).update(position_order=pos)
    return JsonResponse({'success': True, 'saved': len(ids)})


@staff_role_required('superadmin', 'admin')
def category_reorder(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    if not user_can_action(request.user, 'perm_categories', 'reorder'):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    try:
        ids = _parse_reorder_ids(request)
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    user = request.user
    qs = Category.objects.filter(is_deleted=False)
    if not user.is_superadmin:
        qs = qs.filter(companies__in=_user_companies(user)).distinct()
    found = set(qs.filter(pk__in=ids).values_list('pk', flat=True))
    if len(found) != len(ids):
        return JsonResponse({'success': False, 'error': 'One or more categories not accessible'}, status=403)
    for pos, pk in enumerate(ids):
        Category.objects.filter(pk=pk).update(position_order=pos)
    return JsonResponse({'success': True, 'saved': len(ids)})


@staff_role_required('superadmin')
def offering_reorder(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    if not user_can_action(request.user, 'perm_offerings', 'reorder'):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    try:
        ids = _parse_reorder_ids(request)
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    user = request.user
    qs = Offering.objects.filter(is_deleted=False)
    if not user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(user))
    found = set(qs.filter(pk__in=ids).values_list('pk', flat=True))
    if len(found) != len(ids):
        return JsonResponse({'success': False, 'error': 'One or more offerings not accessible'}, status=403)
    for pos, pk in enumerate(ids):
        Offering.objects.filter(pk=pk).update(position_order=pos)
    return JsonResponse({'success': True, 'saved': len(ids)})


@staff_role_required('superadmin', 'admin')
def counter_reorder(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    if not user_can_action(request.user, 'perm_counters', 'reorder'):
        return JsonResponse({'success': False, 'error': 'Permission denied.'}, status=403)
    try:
        ids = _parse_reorder_ids(request)
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    user = request.user
    qs = Counter.objects.filter(is_deleted=False)
    if not user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(user))
    found = set(qs.filter(pk__in=ids).values_list('pk', flat=True))
    if len(found) != len(ids):
        return JsonResponse({'success': False, 'error': 'One or more counters not accessible'}, status=403)
    for pos, pk in enumerate(ids):
        Counter.objects.filter(pk=pk).update(position_order=pos)
    return JsonResponse({'success': True, 'saved': len(ids)})


@staff_role_required('superadmin', 'admin')
def advertise_reorder(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST required'}, status=405)
    try:
        ids = _parse_reorder_ids(request)
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    user = request.user
    qs = Advertise.objects.all()
    if not user.is_superadmin:
        qs = qs.filter(company__in=_user_companies(user))
    found = set(qs.filter(pk__in=ids).values_list('pk', flat=True))
    if len(found) != len(ids):
        return JsonResponse({'success': False, 'error': 'One or more banners not accessible'}, status=403)
    for pos, pk in enumerate(ids):
        Advertise.objects.filter(pk=pk).update(position_order=pos)
    return JsonResponse({'success': True, 'saved': len(ids)})


# ════════════════════════════════════════════════════════════════
#  PHASE 1 — PRODUCT BULK COPY
#  Select multiple products → choose target company / category /
#  offering → copies each product (skips duplicates by name).
#  Mirrors single-product copy logic exactly.
# ════════════════════════════════════════════════════════════════

@staff_role_required('superadmin', 'admin')
def product_bulk_copy(request):
    """
    POST-only copy-menu flow. Accepts ids[], target_company, optional
    target_category, target_offering, and copy_price_<product id> values.
    Copies one or more selected products into the target company.
    """
    if request.method != 'POST':
        return redirect('dashboard:product_list')

    import uuid as _uuid
    from django.utils.text import slugify as _slugify

    user = request.user
    ids = list(dict.fromkeys([pk for pk in request.POST.getlist('ids') if pk]))
    target_company_id = request.POST.get('target_company')
    target_category_id = request.POST.get('target_category') or None
    target_offering_id = request.POST.get('target_offering') or None

    if not ids:
        messages.warning(request, 'Select at least one product to copy.')
        return redirect('dashboard:product_list')

    if has_any_granular_perms(user):
        can_copy = user_can_action(user, 'perm_products', 'copy')
        can_bulk_copy = user_can_action(user, 'perm_products', 'bulk_copy')
        if len(ids) > 1 and not can_bulk_copy:
            messages.error(request, 'Permission denied for copying multiple products.')
            return redirect('dashboard:no_access')
        if len(ids) == 1 and not (can_copy or can_bulk_copy):
            messages.error(request, 'Permission denied.')
            return redirect('dashboard:no_access')

    target_company = _selected_company_for_user(user, target_company_id)
    if not target_company:
        messages.error(request, 'Target company is required.')
        return redirect('dashboard:product_list')

    price_overrides = {}
    for product_id in ids:
        raw_copy_price = (request.POST.get(f'copy_price_{product_id}') or '').strip()
        if not raw_copy_price:
            continue
        try:
            copy_price = Decimal(raw_copy_price)
            if copy_price < 0:
                raise InvalidOperation()
        except (InvalidOperation, ValueError):
            messages.error(request, 'Copy prices must be valid non-negative numbers.')
            return redirect('dashboard:product_list')
        price_overrides[str(product_id)] = copy_price

    # Scope source products to the user's access
    src_qs = Product.objects.filter(pk__in=ids, is_deleted=False)
    if not user.is_superadmin:
        src_qs = src_qs.filter(company__in=_user_companies(user))

    if not src_qs.exists():
        messages.warning(request, 'No valid products selected to copy.')
        return redirect('dashboard:product_list')

    # Optional overrides
    tgt_category_override = None
    if target_category_id:
        tgt_category_override = _scoped_category_for_company(target_company, target_category_id)
        if not tgt_category_override:
            messages.error(request, 'Override category was not found for the target site.')
            return redirect('dashboard:product_list')

    tgt_offering_override = None
    if target_offering_id:
        tgt_offering_override = _scoped_offering_for_company(target_company, target_offering_id)
        if not tgt_offering_override:
            messages.error(request, 'Override offering was not found for the target site.')
            return redirect('dashboard:product_list')

    copied = 0
    skipped = 0

    for src in src_qs.select_related('category', 'offering').prefetch_related('food_type', 'counter_mappings__counter'):
        dest_category = tgt_category_override or src.category
        dest_offering = tgt_offering_override or src.offering

        # Ensure category is linked to target company
        if dest_category and not dest_category.companies.filter(pk=target_company.pk).exists():
            dest_category.companies.add(target_company)
        if dest_category:
            CategoryCompanyStatus.objects.get_or_create(
                category=dest_category,
                company=target_company,
                defaults={'is_active': dest_category.is_active},
            )

        # Resolve / create matching offering under target company
        if dest_offering and dest_offering.company != target_company:
            dest_offering, _ = Offering.objects.get_or_create(
                company=target_company,
                name=dest_offering.name,
                defaults={
                    'slug': dest_offering.slug,
                    'is_active': dest_offering.is_active,
                    'position_order': dest_offering.position_order,
                    'open_days': dest_offering.open_days,
                    'available_from': dest_offering.available_from,
                    'available_to': dest_offering.available_to,
                }
            )

        # Skip if duplicate (same name, category, company)
        if dest_category and Product.objects.filter(
            company=target_company,
            category=dest_category,
            name=src.name,
            is_deleted=False,
        ).exists():
            skipped += 1
            continue

        new_p = Product(
            company=target_company,
            category=dest_category,
            offering=dest_offering,
            sub_category=None,
            sub_list=None,
            name=src.name,
            slug=f'{_slugify(src.name)}-{_uuid.uuid4().hex[:6]}',
            code=src.code,
            price=price_overrides.get(str(src.pk), src.price),
            company_price=src.company_price,
            room_service_extra_percent=src.room_service_extra_percent,
            packing_price=src.packing_price,
            min_qty=src.min_qty,
            max_qty=src.max_qty,
            web_qty=src.web_qty,
            pos_qty=src.pos_qty,
            description=src.description,
            position_order=src.position_order,
            is_active=src.is_active,
            is_kiosk_active=src.is_kiosk_active,
            preparation_time_minutes=src.preparation_time_minutes,
            calories=src.calories,
        )
        if src.image:
            new_p.image = src.image
        new_p.save()
        new_p.food_type.set(src.food_type.all())
        # Copy counter mappings (best-effort — counter must belong to target company)
        target_counter_ids = set(
            Counter.objects.filter(company=target_company, is_deleted=False).values_list('pk', flat=True)
        )
        for idx, mapping in enumerate(src.counter_mappings.select_related('counter').all()):
            if mapping.counter_id in target_counter_ids:
                try:
                    ProductCounter.objects.create(product=new_p, counter=mapping.counter, position_order=idx)
                except Exception:
                    pass
        copied += 1

    if copied:
        messages.success(request, f'Copy menu complete: {copied} product{"s" if copied != 1 else ""} copied to {target_company.name}.')
    if skipped:
        messages.info(request, f'{skipped} product{"s" if skipped != 1 else ""} skipped (already exist in target).')
    return redirect('dashboard:product_list')


# ════════════════════════════════════════════════════════════════
#  PHASE 2 — SHARED HELPER: save offering schedules
#  Reads indexed POST fields window_day_N / window_start_N / window_end_N
#  Deletes existing schedules for the offering and recreates them.
# ════════════════════════════════════════════════════════════════

def _save_offering_schedules(post_data, offering):
    """
    Parse window_day_0 / window_start_0 / window_end_0 … fields and
    save them as Schedule rows linked to the offering.
    Deletes all prior offering schedules before creating new ones.
    """
    from datetime import time as _time
    offering.schedules.all().delete()
    idx = 0
    while idx <= 20:
        day   = (post_data.get(f'window_day_{idx}') or '').strip()
        start = (post_data.get(f'window_start_{idx}') or '').strip()
        end   = (post_data.get(f'window_end_{idx}') or '').strip()
        if not (day and start and end):
            break
        try:
            Schedule.objects.create(
                offering=offering,
                display_day=day,
                start_time=_time.fromisoformat(start),
                end_time=_time.fromisoformat(end),
            )
        except (ValueError, Exception):
            pass
        idx += 1


# ════════════════════════════════════════════════════════════════
#  PHASE 2 — CASHIER LIMITED PRODUCT EDIT
#  pos role can only change: is_active (availability) + description.
#  This is a deliberately stripped-down view — no price, no qty,
#  no category, no offering, no image controls.
# ════════════════════════════════════════════════════════════════

@staff_role_required('pos')
def dashboard_product_cashier_edit(request, pk):
    """
    Cashier-limited product edit: availability toggle + description only.
    Superadmin/admin have the full product_edit view; this is pos-only.
    """
    user = request.user
    if not user_can_action(user, 'perm_products', 'cashier_edit'):
        return _deny_dashboard_action(request)
    # Scope: cashier can only edit products for their own company
    qs = Product.objects.filter(is_deleted=False)
    if user.company:
        qs = qs.filter(company=user.company)
    product = get_object_or_404(qs, pk=pk)

    if request.method == 'POST':
        product.is_active   = request.POST.get('is_active') == 'on'
        product.description = request.POST.get('description', '').strip()
        product.save(update_fields=['is_active', 'description'])
        messages.success(request, f'"{product.name}" updated.')
        return _dashboard_return_or_default(request, 'dashboard:product_list')

    return render(request, 'dashboard/menu/product_cashier_edit.html', {
        'product': product,
        'page_title': f'Edit — {product.name}',
    })


# ════════════════════════════════════════════════════════════════
#  PHASE 3 — BULK IMAGE UPLOAD
#  Accepts a ZIP of images.  Each image filename (without extension)
#  must match a product_code for the selected company.
#  Supported formats: jpg, jpeg, png, webp, gif.
# ════════════════════════════════════════════════════════════════

@staff_role_required('superadmin', 'admin')
def product_bulk_image_upload(request):
    """
    POST: company (optional if non-superadmin) + zip_file
    Extracts images from the ZIP, matches each by filename stem to
    Product.code within the company, and saves the image field.

    Returns a summary rendered inline (via redirect with session data)
    or a JSON summary for AJAX callers.
    """
    if request.method != 'POST':
        return redirect('dashboard:product_bulk_upload')

    user = request.user
    companies = _user_companies(user)
    company_id = request.POST.get('company')
    company = _selected_company_for_user(user, company_id)
    if not company:
        messages.error(request, 'Please select a company before uploading images.')
        return redirect('dashboard:product_bulk_upload')

    zip_file = request.FILES.get('image_zip')
    if not zip_file:
        messages.error(request, 'Please choose a ZIP file containing product images.')
        return redirect('dashboard:product_bulk_upload')

    import zipfile as _zipfile
    import io as _io
    from PIL import Image as _Image

    ALLOWED_EXTS = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}

    try:
        zf = _zipfile.ZipFile(_io.BytesIO(zip_file.read()))
    except _zipfile.BadZipFile:
        messages.error(request, 'Uploaded file is not a valid ZIP archive.')
        return redirect('dashboard:product_bulk_upload')

    matched = 0
    not_found = []
    errors = []

    # Build a lookup: product_code.lower() → Product (for the selected company)
    product_qs = Product.objects.filter(company=company, is_deleted=False).exclude(code='')
    code_map = {(p.code or '').strip().lower(): p for p in product_qs if p.code}

    for entry in zf.infolist():
        if entry.is_dir():
            continue
        name = entry.filename
        # Strip directory prefix (images/VT-001.jpg → VT-001.jpg)
        basename = name.rsplit('/', 1)[-1]
        stem, ext = (basename.rsplit('.', 1) + [''])[:2]
        ext = ('.' + ext).lower() if ext else ''
        if ext not in ALLOWED_EXTS:
            continue  # skip non-image files silently
        code_key = stem.strip().lower()
        product = code_map.get(code_key)
        if not product:
            not_found.append(stem)
            continue
        try:
            raw = zf.read(entry.filename)
            # Validate it is a real image
            img = _Image.open(_io.BytesIO(raw))
            img.verify()
            # Save via Django's ImageField (stores to product_image_path)
            from django.core.files.base import ContentFile
            ext_clean = ext.lstrip('.')
            fname = f'{uuid.uuid4().hex}.{ext_clean}'
            product.image.save(fname, ContentFile(raw), save=True)
            matched += 1
        except Exception as exc:
            errors.append(f'{basename}: {exc}')

    zf.close()

    if matched:
        messages.success(request, f'Bulk image upload: {matched} product image{"s" if matched != 1 else ""} updated for {company.name}.')
    if not_found:
        messages.warning(request, f'{len(not_found)} filename{"s" if len(not_found) != 1 else ""} had no matching product code: {", ".join(not_found[:10])}{"…" if len(not_found) > 10 else ""}.')
    if errors:
        messages.error(request, f'{len(errors)} image{"s" if len(errors) != 1 else ""} failed to save: {"; ".join(errors[:5])}.')
    if not matched and not not_found and not errors:
        messages.warning(request, 'No image files found in the ZIP. Make sure filenames match product codes (e.g. VT-001.jpg).')

    return redirect('dashboard:product_bulk_upload')
